"""Database paths with URI-significant characters must open the right file.

`f"file:{path}?mode=ro"` interpolation treats a filename as a URI. Characters
that are legal in filenames on Windows and POSIX then change meaning:

* `#` truncates the path at a fragment marker AND swallows `?mode=ro`, so
  SQLite opens (and CREATES) a different, empty database with no read-only
  guarantee — output that is silently wrong rather than an error.
* `%HH` percent-decodes: `50%20off` becomes `50 off`.
* `'` inside an ATTACH terminates the SQL string literal.

These are end-to-end: extract a real workbook under such a directory and read
it back through the product's own connection helpers.
"""

import sqlite3

import openpyxl
import pytest

from xl_marinade.core.api import extract
from xl_marinade.core.db_uri import attach_read_only, connect_read_only, read_only_uri

# Legal on Windows and POSIX; each is URI-significant.
HOSTILE_DIRS = [
    pytest.param("rev#3", id="hash"),
    pytest.param("50%20off", id="percent-escape"),
    pytest.param("100%_report", id="percent-literal"),
    pytest.param("O'Brien", id="apostrophe"),
    pytest.param("my models", id="space"),
    pytest.param("a&b", id="ampersand"),
]


def _make_db(directory, tmp_path):
    workdir = tmp_path / directory
    workdir.mkdir()
    xlsx = workdir / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)
    return extract(xlsx, workdir / "ir.db")


@pytest.mark.parametrize("directory", HOSTILE_DIRS)
def test_connect_read_only_opens_the_intended_file(directory, tmp_path):
    db = _make_db(directory, tmp_path)
    conn = connect_read_only(db)
    try:
        opened = conn.execute("PRAGMA database_list").fetchone()[2]
        assert opened == str(db.resolve()), f"opened {opened!r}, expected {db.resolve()!r}"
        assert conn.execute("SELECT COUNT(*) FROM cells").fetchone()[0] > 0
        with pytest.raises(sqlite3.OperationalError):
            conn.execute("CREATE TABLE canary (x)")  # mode=ro must survive
    finally:
        conn.close()


@pytest.mark.parametrize("directory", HOSTILE_DIRS)
def test_attach_read_only_binds_the_intended_file(directory, tmp_path):
    db = _make_db(directory, tmp_path)
    conn = sqlite3.connect(":memory:", uri=True)
    try:
        attach_read_only(conn, db, "ir")
        attached = conn.execute("PRAGMA database_list").fetchall()[-1][2]
        assert attached == str(db.resolve())
        assert conn.execute("SELECT COUNT(*) FROM ir.cells").fetchone()[0] > 0
    finally:
        conn.close()


def test_hash_path_does_not_create_a_decoy_database(tmp_path):
    """The `#` case specifically: the old interpolation created an empty DB."""
    db = _make_db("rev#3", tmp_path)
    decoy = tmp_path / "rev"
    connect_read_only(db).close()
    assert not decoy.exists(), "URI truncation created a decoy database"


def test_read_only_uri_percent_encodes_significant_characters(tmp_path):
    uri = read_only_uri(tmp_path / "a#b%c" / "ir.db")
    assert "%23" in uri and "%25" in uri
    assert uri.endswith("?mode=ro")


def test_no_raw_uri_interpolation_remains():
    """Call sites must use the helper — the helper alone fixes nothing."""
    import pathlib

    root = pathlib.Path(__file__).resolve().parent.parent / "src"
    offenders = [
        f"{py.relative_to(root)}:{i}"
        for py in sorted(root.rglob("*.py"))
        if py.name != "db_uri.py"
        for i, line in enumerate(py.read_text(encoding="utf-8").splitlines(), 1)
        if 'f"file:{' in line or "f'file:{" in line
    ]
    assert not offenders, (
        "path interpolated into a SQLite URI (use core.db_uri helpers):\n  "
        + "\n  ".join(offenders)
    )


def test_diff_end_to_end_under_hostile_directory(tmp_path):
    """`marinade diff` is a full product path over two read-only URI opens."""
    from xl_marinade.core.api import diff

    a = _make_db("rev#3", tmp_path)
    b = _make_db("50%20off", tmp_path)
    changes = diff(a, b)
    assert changes.get("summary") is not None
