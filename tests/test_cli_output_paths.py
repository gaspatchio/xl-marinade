"""The `-o` path and the CLI's own version surface.

Three papercuts found in clean-room conversion runs, all in how the tool talks
about its own output:

* `marinade --version` did not exist, so the installed version could only be
  read from package metadata (issue #13).
* `-o <dir>/<file>` into a missing directory died with SQLite's bare "unable to
  open database file" — no cause, no path — *after* the CLI had printed the
  output path as though it were fine (issues #17, #31). The telemetry writer
  alongside it already created its own parent, so the two disagreed.
* Telemetry was written to a fixed `telemetry.json` beside the database, so it
  silently overwrote a same-named file and collided between two extractions
  sharing an output directory (issue #31).
"""

import importlib.metadata
import json

import openpyxl
from typer.testing import CliRunner

from xl_marinade.cli.main import app

runner = CliRunner()


def _workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(path)
    return path


def test_version_option_reports_the_installed_distribution():
    result = runner.invoke(app, ["--version"])
    assert result.exit_code == 0, result.output
    assert result.output.strip() == importlib.metadata.version("xl-marinade")


def test_version_short_option_matches():
    assert runner.invoke(app, ["-V"]).output == runner.invoke(app, ["--version"]).output


def test_extract_creates_a_missing_output_directory(tmp_path):
    """The failure only bites the *next* person: the directory usually exists."""
    xlsx = _workbook(tmp_path / "book.xlsx")
    out = tmp_path / "build" / "nested" / "ir.db"

    result = runner.invoke(app, ["extract", str(xlsx), "-o", str(out)])

    assert result.exit_code == 0, result.output
    assert out.exists(), "extraction reported success without writing the database"


def test_telemetry_is_named_after_the_database(tmp_path):
    xlsx = _workbook(tmp_path / "book.xlsx")
    out = tmp_path / "ir.db"

    assert runner.invoke(app, ["extract", str(xlsx), "-o", str(out)]).exit_code == 0

    sidecar = tmp_path / "ir.db.telemetry.json"
    assert sidecar.exists(), (
        f"expected {sidecar.name}, found {[p.name for p in tmp_path.iterdir()]}"
    )
    assert not (tmp_path / "telemetry.json").exists(), "fixed-name telemetry.json came back"
    assert json.loads(sidecar.read_text(encoding="utf-8"))["schema_version"]


def test_two_extractions_in_one_directory_keep_separate_telemetry(tmp_path):
    """The collision the fixed name caused: the second run overwrote the first."""
    xlsx = _workbook(tmp_path / "book.xlsx")

    for name in ("a.db", "b.db"):
        assert runner.invoke(app, ["extract", str(xlsx), "-o", str(tmp_path / name)]).exit_code == 0

    assert (tmp_path / "a.db.telemetry.json").exists()
    assert (tmp_path / "b.db.telemetry.json").exists()
