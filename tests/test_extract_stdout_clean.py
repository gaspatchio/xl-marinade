"""`extract` must keep stdout clean — diagnostics go to stderr.

Regression: the grouping phase printed ~20 progress lines ("Building grouping
data structures...", "Cell address map: N cells", ...) to **stdout**, so a
`marinade extract book.xlsx -o ir.db | ...` pipeline (or the docs examples that
show extraction as silent) got polluted. All progress must go to stderr.
"""

import contextlib
import io


def test_extract_writes_nothing_to_stdout(tmp_path):
    import openpyxl

    from xl_marinade.core.api import extract

    xlsx = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"
    wb.save(xlsx)

    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        extract(xlsx, tmp_path / "ir.db")

    leaked = buf.getvalue()
    assert leaked == "", f"extract leaked to stdout:\n{leaked}"
