"""Two extractions of one workbook must agree, whatever the environment.

Root cause of the intermittent `test_cli_diff_emits_json` failure (#35): the
IR stamped `extractor_git_sha` from `git rev-parse` run in the *caller's*
working directory, with a 5-second timeout. Under load the timeout fired on
one run and not the next, so the two databases disagreed on that key and the
self-diff came back non-empty — an artifact sold on determinism disagreeing
with itself. Running outside a git repository (a pip-installed user, any
sdist) produced the same split.

These tests pin the property under exactly those conditions.
"""

import os
import subprocess
import sys

import openpyxl

from xl_marinade.core.api import diff, extract
from xl_marinade.core.new_arch.fast_extraction_pipeline import get_extractor_version


def _workbook(tmp_path):
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)
    return xlsx


def _extract_subprocess(xlsx, out, env):
    proc = subprocess.run(
        [sys.executable, "-m", "xl_marinade.cli.main", "extract", str(xlsx), "-o", str(out)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        timeout=300,
    )
    assert proc.returncode == 0, proc.stderr[-2000:]
    return out


def test_self_diff_is_empty_without_git_on_path(tmp_path):
    """The reproducer: git resolvable for one run, not the other."""
    xlsx = _workbook(tmp_path)

    with_git = dict(os.environ)
    without_git = dict(os.environ)
    without_git["PATH"] = str(tmp_path / "no-tools")  # git unresolvable

    a = _extract_subprocess(xlsx, tmp_path / "a.db", with_git)
    b = _extract_subprocess(xlsx, tmp_path / "b.db", without_git)

    changes = diff(a, b)["changes"]
    assert changes == [], f"same workbook, different environments, non-empty diff: {changes}"


def test_self_diff_is_empty_from_a_different_working_directory(tmp_path):
    """The stamp must not describe whichever repository the user stands in."""
    xlsx = _workbook(tmp_path)
    elsewhere = tmp_path / "elsewhere"
    elsewhere.mkdir()

    inside_repo = dict(os.environ)
    outside_repo = dict(os.environ)
    outside_repo["PWD"] = str(elsewhere)

    a = _extract_subprocess(xlsx, tmp_path / "c.db", inside_repo)
    cwd = os.getcwd()
    os.chdir(elsewhere)
    try:
        b = _extract_subprocess(xlsx, tmp_path / "d.db", outside_repo)
    finally:
        os.chdir(cwd)

    assert diff(a, b)["changes"] == []


def test_stamp_is_the_installed_version(tmp_path):
    import sqlite3

    db = extract(_workbook(tmp_path), tmp_path / "ir.db")
    conn = sqlite3.connect(db)
    try:
        stored = dict(conn.execute("SELECT key, value FROM ir_metadata").fetchall())
    finally:
        conn.close()
    assert stored["extractor_version"] == get_extractor_version()
    assert "extractor_git_sha" not in stored
