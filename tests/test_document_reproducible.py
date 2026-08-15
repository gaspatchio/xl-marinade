"""`marinade document` is byte-reproducible when the clock is pinned.

The generated artifacts embedded `datetime.now()`, so two runs over the same
IR database differed — noise for anything that diffs or hashes them, and at
odds with the determinism the rest of the pipeline promises.
`SOURCE_DATE_EPOCH` (the reproducible-builds convention) pins it.
"""

import hashlib
import os

import openpyxl
import pytest

from xl_marinade.core.api import extract
from xl_marinade.docs.pipeline import document
from xl_marinade.docs.utils.generation_time import generation_timestamp


@pytest.fixture
def ir_db(tmp_path):
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Premium"
    ws["B1"] = 100
    ws["B2"] = "=B1*1.1"
    wb.save(xlsx)
    return extract(xlsx, tmp_path / "ir.db")


def _digests(out_dir):
    return {
        p.name: hashlib.sha256(p.read_bytes()).hexdigest()
        for p in sorted(out_dir.rglob("*"))
        if p.is_file()
    }


def test_document_is_byte_identical_with_source_date_epoch(ir_db, tmp_path, monkeypatch):
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "1700000000")

    first = tmp_path / "run1"
    second = tmp_path / "run2"
    document(ir_db, first)
    document(ir_db, second)

    d1, d2 = _digests(first), _digests(second)
    assert d1, "document produced no artifacts"
    assert d1 == d2, f"same IR, pinned clock, different bytes: {set(d1.items()) ^ set(d2.items())}"


def test_source_date_epoch_pins_the_timestamp(monkeypatch):
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "1700000000")
    assert generation_timestamp() == "2023-11-14T22:13:20Z"


def test_malformed_source_date_epoch_does_not_fail_the_run(monkeypatch):
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "not-a-number")
    assert generation_timestamp().endswith("Z")


def test_unset_source_date_epoch_uses_the_wall_clock(monkeypatch):
    monkeypatch.delenv("SOURCE_DATE_EPOCH", raising=False)
    assert "SOURCE_DATE_EPOCH" not in os.environ
    assert generation_timestamp().endswith("Z")
