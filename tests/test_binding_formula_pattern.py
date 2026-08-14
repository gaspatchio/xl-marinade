"""`agent_bindings.formula_pattern` must describe the binding, not one cell of it.

When the init merger folds an initialisation cell into the propagation range it
covers, the binding holds two formulas: the init's one-off and the recurrence
that fills the rest. `formula_pattern` was taken from the binding's spatial
top-left, which on exactly this shape is the one-off — 1 occurrence reported for
a range where the other formula holds 40 (issue #12; observed at 1-against-110
on a real UL projection workbook).

The consequence is not cosmetic. On a recurrence the init formula has no
self-reference, so an agent reading `formula_pattern` to decide whether a column
is recursive concludes it is not — and a conversion built on that answer drops
the recursion, or drops a lapse gate that only appears from the second row on.
"""

import sqlite3

import openpyxl
import pytest

from xl_marinade.core.api import extract


def _projection_workbook(path):
    """Init + propagation: the shape the init merger folds into one binding.

    `B1` seeds the column; `B2:B41` each read the cell above. The merger
    requires the propagation to reference the init cell, which is what makes
    this a recurrence rather than two unrelated ranges.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "premium"
    for row in range(2, 42):
        ws[f"A{row}"] = 100
    ws["B1"] = "=A1"
    for row in range(2, 42):
        ws[f"B{row}"] = f"=B{row - 1}+A{row}"
    wb.save(path)
    return path


def _formula_distribution(conn, binding_id):
    return conn.execute(
        """
        SELECT f.formula_r1c1, COUNT(*) AS n
        FROM cell_to_binding ctb
        JOIN cells c ON c.cell_id = ctb.cell_id
        JOIN formulas f ON f.formula_id = c.formula_id
        WHERE ctb.binding_id = ?
        GROUP BY f.formula_r1c1
        ORDER BY n DESC
        """,
        (binding_id,),
    ).fetchall()


@pytest.fixture
def mixed_binding(tmp_path):
    """The merged binding, its reported pattern, and its true distribution."""
    db = extract(_projection_workbook(tmp_path / "proj.xlsx"), tmp_path / "ir.db")
    conn = sqlite3.connect(db)
    try:
        for binding_id, address, pattern in conn.execute(
            "SELECT binding_id, address, formula_pattern FROM agent_bindings "
            "WHERE formula_pattern IS NOT NULL"
        ):
            distribution = _formula_distribution(conn, binding_id)
            if len(distribution) > 1:
                return address, pattern, distribution
    finally:
        conn.close()
    pytest.fail(
        "no mixed binding produced — the init merger did not fire, so this "
        "fixture no longer exercises the case issue #12 is about"
    )
    return None


def test_formula_pattern_reports_the_dominant_formula(mixed_binding):
    address, pattern, distribution = mixed_binding
    dominant, dominant_count = distribution[0]
    minority = [(f, n) for f, n in distribution[1:]]

    assert dominant_count > minority[0][1], (
        "fixture no longer has a clear majority formula, so it cannot "
        f"distinguish dominant from top-left: {distribution}"
    )
    assert pattern == dominant, (
        f"{address}: formula_pattern reports {pattern!r} but the binding's "
        f"dominant formula is {dominant!r} ({dominant_count} of "
        f"{sum(n for _, n in distribution)} cells). Distribution: {distribution}"
    )


def test_formula_pattern_keeps_the_recurrence_visible(mixed_binding):
    """The failure that matters: the init one-off hides the self-reference.

    `R[-1]C` is the recurrence. The init formula does not contain it, so
    reporting the init makes a recursive column look non-recursive to anything
    reading `formula_pattern` — including the SCC-based recursion detection
    conversion tooling runs.
    """
    address, pattern, _ = mixed_binding
    assert "R[-1]C" in pattern, (
        f"{address}: formula_pattern {pattern!r} carries no self-reference, so a "
        "consumer cannot see that this column is a recurrence"
    )
