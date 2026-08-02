"""The dependency-graph backend must behave identically on the dict fallback
and the NetworkX ``DiGraph`` representation.

Both are produced by :func:`build_dependency_graph` and consumed by the
two-pass labeller. ``networkx`` is an optional accelerator, not a runtime
dependency, so these tests opt into the ``DiGraph`` path explicitly and skip
when ``networkx`` is absent.
"""

import pytest

import xl_marinade.docs.dependency_traversal as dt
from xl_marinade.docs.dependency_traversal import (
    get_children,
    get_parents,
    get_topological_order,
)


def test_document_survives_empty_digraph(tmp_path, monkeypatch):
    """``document()`` on an all-orphans model must succeed on the DiGraph path.

    Regression: an empty ``nx.DiGraph`` is falsy, so the ``if not self.graph``
    guard in ``run_pass_1_top_down`` raised ``RuntimeError`` on the DiGraph
    path even though ``build_graph()`` had run — while the always-truthy dict
    fallback (``{'__reverse__': {}}``) worked. Both representations must
    document without crashing.
    """
    pytest.importorskip("networkx")
    monkeypatch.setattr(dt, "HAS_NETWORKX", True)

    from test_workbook_generator.cli import create_comprehensive_test_workbook
    from xl_marinade.core.api import extract

    xlsx = tmp_path / "wb.xlsx"
    create_comprehensive_test_workbook(xlsx)
    ir_db = extract(xlsx, tmp_path / "ir.db")

    from xl_marinade.docs import document

    out = tmp_path / "out"
    md = document(ir_db, out)

    assert md == out / "documentation.md"
    assert md.exists() and md.stat().st_size > 0
    assert (out / "model_spec.json").exists()


def test_full_workbook_extract_produces_populated_graph(tmp_path):
    """The public full-workbook extraction path must yield a non-empty graph.

    Regression: ``extract()`` uses new_arch full-workbook extraction, which
    writes NO ``user_roots`` and persists binding edges in ``binding_edges``.
    ``build_dependency_graph`` was written for the legacy schema — it read
    ``binding_level_edges``/``agent_binding_dependencies`` (absent in new_arch)
    and pruned to reachable-from-``user_roots`` (empty) — so it returned an
    EMPTY graph on every public ``extract()``+``document()`` call, silently
    disabling the two-pass labeller's dependency context. It must read
    ``binding_edges`` and, absent user roots, include the whole edge graph.
    """
    from test_workbook_generator.cli import create_comprehensive_test_workbook
    from xl_marinade.core.api import extract

    xlsx = tmp_path / "wb.xlsx"
    create_comprehensive_test_workbook(xlsx)
    ir_db = extract(xlsx, tmp_path / "ir.db")

    graph = dt.build_dependency_graph(str(ir_db))
    order = get_topological_order(graph)

    assert order, "full-workbook dependency graph is empty (no binding_edges / roots pruning)"
    # the comprehensive fixture has many interdependent bindings, not just one.
    assert len(order) >= 5


def _dict_graph(edges: list[tuple[str, str]]) -> dict[str, set[str]]:
    """Build the dict-backend graph shape (forward adjacency + ``__reverse__``)."""
    graph: dict[str, set[str]] = {}
    reverse: dict[str, set[str]] = {}
    for frm, to in edges:
        graph.setdefault(frm, set()).add(to)
        reverse.setdefault(to, set()).add(frm)
    graph["__reverse__"] = reverse
    return graph


def test_digraph_and_dict_traversal_agree(monkeypatch):
    """``get_topological_order`` / ``get_parents`` / ``get_children`` return
    equivalent results on a populated DiGraph and the equivalent dict graph.

    Edges ``from -> to`` mean "from depends on to". The graph
    ``a->b, b->c, a->c`` has a unique topological order (roots first)::

        a (root, depends on b & c)
        b (depends on c)
        c (leaf)
    """
    nx = pytest.importorskip("networkx")
    edges = [("a", "b"), ("b", "c"), ("a", "c")]

    digraph = nx.DiGraph()
    digraph.add_edges_from(edges)
    dict_graph = _dict_graph(edges)

    # Topological order: identical, roots-first, on both backends.
    monkeypatch.setattr(dt, "HAS_NETWORKX", True)
    di_order = get_topological_order(digraph)
    monkeypatch.setattr(dt, "HAS_NETWORKX", False)
    dict_order = get_topological_order(dict_graph)
    assert di_order == ["a", "b", "c"]
    assert dict_order == ["a", "b", "c"]

    # children = successors (things a node depends on).
    monkeypatch.setattr(dt, "HAS_NETWORKX", True)
    assert get_children(digraph, "a") == ["b", "c"]
    assert get_children(digraph, "b") == ["c"]
    assert get_children(digraph, "c") == []
    monkeypatch.setattr(dt, "HAS_NETWORKX", False)
    assert get_children(dict_graph, "a") == ["b", "c"]
    assert get_children(dict_graph, "b") == ["c"]
    assert get_children(dict_graph, "c") == []

    # parents = predecessors (things that depend on a node).
    monkeypatch.setattr(dt, "HAS_NETWORKX", True)
    assert get_parents(digraph, "c") == ["a", "b"]
    assert get_parents(digraph, "b") == ["a"]
    assert get_parents(digraph, "a") == []
    monkeypatch.setattr(dt, "HAS_NETWORKX", False)
    assert get_parents(dict_graph, "c") == ["a", "b"]
    assert get_parents(dict_graph, "b") == ["a"]
    assert get_parents(dict_graph, "a") == []


def test_cyclic_graph_orders_all_nodes_identically_on_both_backends(monkeypatch):
    """A cyclic binding graph must order every node, identically on both backends.

    Binding-level graphs are legitimately cyclic for recursive models: a
    time-lagged projection column (``pols_if[t] = pols_if[t-1] - pols_death[t-1]``,
    ``pols_death[t] = pols_if[t] * q``) is a DAG cell-by-cell but a cycle once
    collapsed to column bindings.

    Regression: ``nx.topological_sort`` raises ``NetworkXUnfeasible`` on cycles,
    which the ``except nx.NetworkXError`` guard did not catch
    (``NetworkXUnfeasible`` subclasses ``NetworkXAlgorithmError``, not
    ``NetworkXError``) — so ``document()`` crashed on every recursive model when
    networkx was installed, while the dict backend handled the same graph fine.
    """
    nx = pytest.importorskip("networkx")
    edges = [
        ("premiums", "pols_if"),
        ("pols_if", "pols_death"),
        ("pols_if", "pols_lapse"),
        ("pols_death", "pols_if"),
        ("pols_lapse", "pols_if"),
        ("pols_death", "mortality_rate"),
    ]
    digraph = nx.DiGraph()
    digraph.add_edges_from(edges)
    dict_graph = _dict_graph(edges)

    monkeypatch.setattr(dt, "HAS_NETWORKX", True)
    di_order = get_topological_order(digraph)
    monkeypatch.setattr(dt, "HAS_NETWORKX", False)
    dict_order = get_topological_order(dict_graph)

    all_nodes = {node for edge in edges for node in edge}
    assert sorted(di_order) == sorted(all_nodes)
    assert di_order == dict_order


def test_document_succeeds_on_recursive_projection_workbook(tmp_path, monkeypatch):
    """``document()`` must succeed on a workbook with a recursive projection.

    The workbook below is the minimal recursive-projection shape: an in-force
    column seeded from a constant and decremented by last period's deaths, and
    a deaths column computed from the current in-force — mutually dependent
    column bindings, i.e. a cycle in the binding graph. Opts into the DiGraph
    backend, where the cycle crashed ``document()``.
    """
    pytest.importorskip("networkx")
    monkeypatch.setattr(dt, "HAS_NETWORKX", True)

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projection"
    ws["A1"], ws["B1"], ws["C1"] = "t", "pols_if", "pols_death"
    ws["A2"], ws["B2"] = 0, 100.0
    ws["C2"] = "=B2*0.01"
    for row in range(3, 31):
        ws[f"A{row}"] = row - 2
        ws[f"B{row}"] = f"=B{row - 1}-C{row - 1}"
        ws[f"C{row}"] = f"=B{row}*0.01"
    xlsx = tmp_path / "recursive.xlsx"
    wb.save(xlsx)

    from xl_marinade.core.api import extract
    from xl_marinade.docs import document

    ir_db = extract(xlsx, tmp_path / "ir.db")
    out = tmp_path / "out"
    md = document(ir_db, out)

    assert md.exists() and md.stat().st_size > 0
    assert (out / "model_spec.json").exists()
