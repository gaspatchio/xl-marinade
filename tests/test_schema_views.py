"""The documented query surface must exist in a fresh extraction.

Review finding on the atlas_nodes -> marinade_nodes rename: the view had zero
consumers in src/ or tests/, so neither the rename nor a future accidental
removal would fail anything. This pins the agent-facing surface the bundled
skill queries.
"""

import sqlite3

import openpyxl

from xl_marinade.core.api import extract


def test_marinade_nodes_view_selectable(tmp_path):
    xlsx = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)

    db = extract(xlsx, tmp_path / "t.db")

    conn = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    try:
        rows = conn.execute(
            "SELECT node_id, node_kind, display_name FROM marinade_nodes"
        ).fetchall()
    finally:
        conn.close()
    assert rows, "marinade_nodes view missing or empty on a fresh extraction"
