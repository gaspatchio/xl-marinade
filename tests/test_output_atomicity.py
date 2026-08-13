"""A failed extraction must not destroy the previous output database.

`vacuum_into` used to unlink the target before writing, so a VACUUM that
failed part-way left the user with neither the old database nor a new one.
The realistic trigger is a full disk — which is precisely when someone
re-runs an extraction, and the failure mode behind issue #7.
"""

import sqlite3

import openpyxl
import pytest

from xl_marinade.core.api import extract
from xl_marinade.core.new_arch.bulk_loader import BulkLoader


class _FailingConnection:
    """Stands in for the build connection; fails the way a full disk does."""

    def __init__(self, real: sqlite3.Connection) -> None:
        self._real = real
        self.closed = False

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False

    def execute(self, sql, *args, **kwargs):
        if sql.strip().upper().startswith("VACUUM INTO"):
            raise sqlite3.OperationalError("database or disk is full")
        return self._real.execute(sql, *args, **kwargs)

    def close(self) -> None:
        self.closed = True
        self._real.close()


def _workbook(tmp_path):
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)
    return xlsx


def test_previous_database_survives_a_failed_vacuum(tmp_path):
    xlsx = _workbook(tmp_path)
    out = tmp_path / "ir.db"
    extract(xlsx, out)
    good_bytes = out.read_bytes()
    assert good_bytes

    loader = BulkLoader(str(tmp_path / "build.db"))
    loader.open()
    loader.conn = _FailingConnection(loader.conn)
    with pytest.raises(sqlite3.OperationalError):
        loader.vacuum_into(str(out))

    assert out.exists(), "a failed VACUUM deleted the previous database"
    assert out.read_bytes() == good_bytes, "previous database was modified"
    conn = sqlite3.connect(out)
    try:
        assert conn.execute("SELECT COUNT(*) FROM cells").fetchone()[0] > 0
    finally:
        conn.close()
    assert not list(tmp_path.glob("*.tmp-vacuum")), "temp artifact left behind"


def test_successful_extract_overwrites_cleanly(tmp_path):
    xlsx = _workbook(tmp_path)
    out = tmp_path / "ir.db"
    extract(xlsx, out)
    first = out.stat().st_size
    extract(xlsx, out)  # re-extract over an existing database
    assert out.exists() and out.stat().st_size == first
    assert not list(tmp_path.glob("*.tmp-vacuum"))
