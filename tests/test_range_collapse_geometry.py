"""Geometry helpers behind the range-edge collapse (issue #7).

The collapse aggregates per-rect breadths without materializing (rect × cell)
rows — rolling-window models produce hundreds of thousands of DISTINCT but
almost fully overlapping rects, and materializing them spilled >15 GB of
SQLite temp on a real 2.3M-formula forecast model. Union breadths for
multi-rect pairs are computed by decomposing the rect-set into disjoint boxes
and bisecting the target binding's cells against them, so these two helpers
carry the exactness burden.
"""

import random

from xl_marinade.core.new_arch.grouping_native import _disjoint_boxes, _union_membership


def _brute_union(rects):
    return {
        (r, c) for r1, c1, r2, c2 in rects for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)
    }


def test_disjoint_boxes_random_rects_exact_and_disjoint():
    rng = random.Random(7)
    for trial in range(300):
        rects = []
        for _ in range(rng.randint(1, 12)):
            r1, c1 = rng.randint(1, 20), rng.randint(1, 12)
            rects.append((r1, c1, r1 + rng.randint(0, 8), c1 + rng.randint(0, 5)))
        truth = _brute_union(rects)
        covered = set()
        for r1, c1, r2, c2 in _disjoint_boxes(rects):
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    assert (r, c) not in covered, f"boxes overlap (trial {trial})"
                    covered.add((r, c))
        assert covered == truth, f"coverage mismatch (trial {trial})"


def test_disjoint_boxes_collapses_rolling_windows():
    # The issue #7 shape: hundreds of one-step-shifted windows over one column.
    rolling = [(r, 2, r + 119, 2) for r in range(1, 501)]
    assert len(_disjoint_boxes(rolling)) == 1


def test_union_membership_agrees_with_brute_force():
    rng = random.Random(11)
    for trial in range(100):
        rects = []
        for _ in range(rng.randint(1, 10)):
            r1, c1 = rng.randint(1, 15), rng.randint(1, 10)
            rects.append((r1, c1, r1 + rng.randint(0, 6), c1 + rng.randint(0, 4)))
        truth = _brute_union(rects)
        contains = _union_membership(_disjoint_boxes(rects))
        for row in range(0, 25):
            for col in range(0, 18):
                assert contains(row, col) == ((row, col) in truth), (
                    f"membership mismatch at ({row}, {col}) (trial {trial})"
                )


def _build_rolling_workbook(xlsx):
    """The issue #7 shape: distinct, almost fully overlapping shifted windows."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Forecast"
    for row in range(1, 121):
        ws.cell(row=row, column=1, value=row)
    for row in range(1, 81):
        ws.cell(row=row, column=2, value=f"=SUM(A{row}:A{row + 39})")
        ws.cell(
            row=row, column=3, value=f"=SUM(A{row}:A{row + 19})+SUM(B{row}:B{min(row + 9, 80)})"
        )
    wb.save(xlsx)


def _rolling_expected_edges():
    """Hand-derived (from the workbook shape, not the implementation):
    B1:B80 reads A r:r+39 for r=1..80 -> union rows 1..119, and A120 is
    referenced by nothing (hence the A1:A119 constant binding). C1:C71
    (uniform R1C1) reads A r:r+19 -> rows 1..90, and B r:r+9 -> all 80 B
    cells. C72..C80 stay single-cell bindings (the min() clamp changes their
    R1C1 pattern): each reads 20 A cells and 81-r B cells, with C80's B80:B80
    reference landing as a single-cell formula edge."""
    expected = {
        ("B1:B80", "A1:A119", 119, "range_static"),
        ("C1:C71", "A1:A119", 90, "range_static"),
        ("C1:C71", "B1:B80", 80, "range_static"),
        ("C80", "B1:B80", 1, "formula"),
    }
    expected |= {(f"C{r}", "A1:A119", 20, "range_static") for r in range(72, 81)}
    expected |= {(f"C{r}", "B1:B80", 81 - r, "range_static") for r in range(72, 80)}
    return expected


def _addr_edges(db_path):
    import sqlite3

    conn = sqlite3.connect(db_path)
    rows = conn.execute(
        "SELECT bf.address_a1, bt.address_a1, e.edge_count, e.kind "
        "FROM binding_edges e "
        "JOIN bindings bf ON bf.binding_id = e.from_binding_id "
        "JOIN bindings bt ON bt.binding_id = e.to_binding_id"
    ).fetchall()
    conn.close()
    return set(rows)


def test_rolling_window_workbook_extracts_deterministically(tmp_path):
    """End-to-end on the issue #7 shape: overlapping shifted-window ranges.

    Every row's formula references a window shifted one row from its
    neighbour's — distinct, almost fully overlapping rects. Extraction must
    complete and be run-to-run deterministic at the binding-edge level.
    """
    import sqlite3

    from xl_marinade.core.api import extract

    xlsx = tmp_path / "rolling.xlsx"
    _build_rolling_workbook(xlsx)

    def edges(db_path):
        conn = sqlite3.connect(db_path)
        rows = conn.execute(
            "SELECT from_binding_id, to_binding_id, edge_count, kind "
            "FROM binding_edges ORDER BY 1, 2, 4"
        ).fetchall()
        conn.close()
        return rows

    e1 = edges(extract(xlsx, tmp_path / "a.db"))
    e2 = edges(extract(xlsx, tmp_path / "b.db"))
    assert e1, "rolling-window workbook produced no binding edges"
    assert e1 == e2

    # Semantic pin, derived by hand (not from the implementation), so a
    # breadth regression in the SQL/geometry side fails loudly.
    assert _addr_edges(tmp_path / "a.db") == _rolling_expected_edges()


def test_multi_chunk_breadth_aggregation_matches_hand_derived(tmp_path, monkeypatch):
    """A multi-chunk _rect_breadths run must produce the exact same breadths.

    Review finding: the chunking path was entirely uncovered — the e2e
    workbook fits one chunk, `_rect_breadths` has no UNIQUE constraint, and a
    rect landing in two chunks would inflate edge_count with green tests.
    Force many chunks with a tiny budget and assert against the hand-derived
    edge set (not against a same-code single-chunk run, so a bug common to
    both paths cannot cancel out).
    """
    import xl_marinade.core.new_arch.grouping_native as gn
    from xl_marinade.core.api import extract

    xlsx = tmp_path / "rolling.xlsx"
    _build_rolling_workbook(xlsx)
    # Rect estimates for this workbook are ~20-120 cells each; a 500-cell
    # budget forces the ~170 rects into dozens of chunks.
    monkeypatch.setattr(gn, "_CHUNK_CELL_BUDGET", 500)
    db = extract(xlsx, tmp_path / "chunked.db")
    assert _addr_edges(db) == _rolling_expected_edges()
