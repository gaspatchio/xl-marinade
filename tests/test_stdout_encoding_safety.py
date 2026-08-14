"""User-facing output must survive a strict, non-UTF-8 stdout.

Windows encodes redirected/captured stdout with the ANSI code page and
`errors='strict'` (through 3.14 — PEP 686 only defaults to UTF-8 in 3.15), so a
single non-ASCII character in a print() aborts the process with
UnicodeEncodeError. That is worst when it happens *after* the work is committed:
the caller reads a non-zero exit and treats a completed run as a failure.

Two guards: stdout carries no non-ASCII anywhere in the package, and the
extraction CLI survives a cp1252 stdout end to end.
"""

import ast
import os
import subprocess
import sys
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parent.parent / "src"


def _stdout_prints_with_non_ascii(py: Path) -> list[str]:
    """print() calls that go to stdout (no file=) and contain non-ASCII text."""
    tree = ast.parse(py.read_text(encoding="utf-8"), filename=str(py))
    out: list[str] = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        if not (isinstance(node.func, ast.Name) and node.func.id == "print"):
            continue
        if any(kw.arg == "file" for kw in node.keywords):
            continue  # stderr (or an explicit sink) is not the strict-stdout path
        literals = [
            n.value
            for n in ast.walk(node)
            if isinstance(n, ast.Constant) and isinstance(n.value, str)
        ]
        if any(not text.isascii() for text in literals):
            out.append(f"{py.relative_to(SRC)}:{node.lineno}")
    return out


def test_no_non_ascii_on_stdout():
    offenders: list[str] = []
    for py in sorted(SRC.rglob("*.py")):
        offenders.extend(_stdout_prints_with_non_ascii(py))
    assert not offenders, (
        "non-ASCII print() to stdout — crashes with UnicodeEncodeError when "
        "redirected on a non-UTF-8 console (use file=sys.stderr and ASCII):\n  "
        + "\n  ".join(offenders)
    )


def test_extract_survives_cp1252_stdout(tmp_path):
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)

    env = dict(os.environ)
    env["PYTHONIOENCODING"] = "cp1252"  # what Windows does to a redirected stdout
    out_db = tmp_path / "out.db"
    proc = subprocess.run(
        [sys.executable, "-m", "xl_marinade.cli.main", "extract", str(xlsx), "-o", str(out_db)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        timeout=300,
    )
    assert proc.returncode == 0, f"extract failed on a cp1252 console:\n{proc.stderr[-2000:]}"
    assert out_db.exists()
