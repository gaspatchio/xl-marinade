"""Characterization tests locking output-preservation for the ported perf fixes.

These are NOT red-green: each golden is captured from the current (pre-fix) code,
so the test passes before the fix and must STAY green after — proving the
algorithmic change did not alter output.
"""

import json
import sqlite3
from pathlib import Path

from test_workbook_generator.cli import create_comprehensive_test_workbook
from xl_marinade.core.api import extract
from xl_marinade.core.vba.paste_edges import synthesize_paste_edges

GOLDEN = Path(__file__).parent / "golden"
SCHEMA_PATH = (
    Path(__file__).resolve().parent.parent
    / "src"
    / "xl_marinade"
    / "core"
    / "new_arch"
    / "schema.sql"
)


def _binding_edges(db_path) -> list:
    # Identify bindings by (sheet, address) rather than raw binding_id. binding_id
    # is SHA256(workbook_guid + "::" + ...), and workbook_guid is the SHA256 of the
    # fixture .xlsx bytes — which embed an openpyxl docProps/core.xml
    # created/modified wall-clock timestamp. Two otherwise-identical fixture
    # workbooks generated a process (or even just a second) apart hash
    # differently, so comparing raw binding_id strings across a golden-capture
    # run and a later test run is flaky by construction. (sheet, address_a1) is
    # a stable, collision-free identity for a binding within one extraction and
    # is what an auditor actually cares about, so it is used as the comparison
    # key instead.
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        rows = conn.execute(
            """
            SELECT
                fs.sheet_name || '!' || fb.address_a1,
                ts.sheet_name || '!' || tb.address_a1,
                be.kind,
                be.edge_count
            FROM binding_edges be
            JOIN bindings fb ON fb.binding_id = be.from_binding_id
            JOIN bindings tb ON tb.binding_id = be.to_binding_id
            JOIN sheets fs ON fs.sheet_id = fb.sheet_id
            JOIN sheets ts ON ts.sheet_id = tb.sheet_id
            ORDER BY 1, 2, be.kind
            """
        ).fetchall()
    finally:
        conn.close()
    return [list(r) for r in rows]


def test_grouping_binding_edges_unchanged(tmp_path):
    xlsx = tmp_path / "wb.xlsx"
    create_comprehensive_test_workbook(xlsx)
    db = extract(xlsx, tmp_path / "ir.db")
    edges = _binding_edges(db)
    golden = json.loads((GOLDEN / "grouping_binding_edges.json").read_text())
    assert edges == golden


def _paste_tie_db() -> sqlite3.Connection:
    """Minimal IR where procs A::a and Z::z both paste into the same (from,to) edge.

    Built against the production `new_arch/schema.sql` so table shapes match
    what `synthesize_paste_edges` actually reads: `sheets`, `cells`,
    `cell_to_binding`, `vba_modules`, `vba_procedures`, `binding_edges`. Both
    procedures execute the identical `.Value = .Value` assignment
    `Worksheets("Calc").Range("B10").Value = Worksheets("Calc").Range("B2").Value`,
    which resolves — independently, per procedure — to the same single-cell
    source binding ("bind_src") and target binding ("bind_tgt"). Because
    `binding_edges`' PK is `(from_binding_id, to_binding_id)`, the two
    procedures collide on the SAME edge row and only one `provenance_proc` can
    win. 'A::a' sorts before 'Z::z', so a deterministic tie-break must pick
    'A::a'.
    """
    conn = sqlite3.connect(":memory:")
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))

    conn.execute("INSERT INTO sheets (sheet_id, sheet_name) VALUES (1, 'Calc')")
    conn.executemany(
        "INSERT INTO cells (cell_id, sheet_id, row, col, a1) VALUES (?, ?, ?, ?, ?)",
        [
            (1, 1, 2, 2, "B2"),  # source cell
            (2, 1, 10, 2, "B10"),  # target cell
        ],
    )
    conn.executemany(
        "INSERT INTO cell_to_binding (cell_id, binding_id) VALUES (?, ?)",
        [(1, "bind_src"), (2, "bind_tgt")],
    )

    conn.executemany(
        """
        INSERT INTO vba_modules (module_id, name, kind, source_sha256, source_text)
        VALUES (?, ?, 'standard', ?, ?)
        """,
        [(1, "A", "sha_a", "' module A"), (2, "Z", "sha_z", "' module Z")],
    )

    paste_body = 'Worksheets("Calc").Range("B10").Value = Worksheets("Calc").Range("B2").Value'
    conn.executemany(
        """
        INSERT INTO vba_procedures
            (module_id, name, kind, signature, body, normalized_body_hash)
        VALUES (?, ?, 'sub', ?, ?, ?)
        """,
        [
            (1, "a", "Sub a()", paste_body, "hash_a"),
            (2, "z", "Sub z()", paste_body, "hash_z"),
        ],
    )
    conn.commit()
    return conn


def test_vba_paste_provenance_is_deterministic():
    """A (from,to) tie between two procedures must resolve the same way every run.

    Pre-fix, `synthesize_paste_edges` iterated a `set` of (from,to,proc) tuples
    and relied on `INSERT OR IGNORE` (keyed on (from,to)) to pick a winner — so
    the winning `provenance_proc` depended on Python's randomized str hashing
    (PYTHONHASHSEED), not on the data. Post-fix, `sorted(edges)` makes the
    lexicographically-smallest `proc_qual` win, deterministically.
    """
    conn = _paste_tie_db()
    metrics = synthesize_paste_edges(conn)
    assert metrics["events_resolved"] >= 2  # the tie was actually exercised
    row = conn.execute(
        "SELECT provenance_proc FROM binding_edges WHERE kind='via_vba_paste'"
    ).fetchone()
    assert row is not None
    assert row[0] == "A::a"  # sorted() -> smallest proc wins, not set-order roulette


def _create_lookup_dense_workbook(xlsx) -> None:
    """Deterministic workbook exercising every MATCH semantic-resolution path.

    All lookup arrays are constants (semantic resolution reads the value
    snapshot; freshly built openpyxl workbooks carry no cached formula
    values). Covers: exact hit / miss, case-insensitive string match,
    approximate ascending (1) and descending (-1) with type filtering,
    logical-vs-number type classes (TRUE vs 1 must resolve to different
    positions), row arrays, full-column references (the sparse-index path),
    and the same scan repeated across many formulas (the memoized case).
    """
    import openpyxl

    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    for i in range(1, 21):
        data.cell(row=i, column=1, value=i * 10)  # A ascending numbers
        data.cell(row=i, column=2, value=f"Key{i}")  # B strings
        data.cell(row=i, column=6, value=210 - i * 10)  # F descending
    for col, v in enumerate([True, 1, False, 0, 2], start=1):
        data.cell(row=22, column=col, value=v)  # A22:E22 logical/number mix
    for col in range(1, 21):
        data.cell(row=23, column=col, value=col * 5)  # row array A23:T23

    main = wb.create_sheet("Main")
    main["A1"] = 30
    main["A2"] = "key7"  # lowercase on purpose: Excel MATCH is case-insensitive
    fx = [
        "=INDEX(Data!$A$1:$A$20,MATCH(50,Data!$A$1:$A$20,0))",  # exact hit
        "=INDEX(Data!$A$1:$A$20,MATCH(999,Data!$A$1:$A$20,0))",  # exact miss
        "=INDEX(Data!$B$1:$B$20,MATCH($A$2,Data!$B$1:$B$20,0))",  # ci string
        "=INDEX(Data!$A$1:$A$20,MATCH(55,Data!$A$1:$A$20,1))",  # approx asc
        "=INDEX(Data!$F$1:$F$20,MATCH(55,Data!$F$1:$F$20,-1))",  # approx desc
        "=MATCH(TRUE,Data!$A$22:$E$22,0)",  # logical class -> position 1
        "=MATCH(1,Data!$A$22:$E$22,0)",  # number class -> position 2
        "=MATCH(35,Data!$A$23:$T$23,0)",  # row array
        "=MATCH(50,Data!A:A,0)",  # full column (sparse-index path)
    ]
    for i, f in enumerate(fx, start=1):
        main.cell(row=i, column=3, value=f)
    # The memoized case: many formulas repeating a handful of distinct scans.
    for r in range(1, 31):
        main.cell(
            row=r,
            column=5,
            value=f"=INDEX(Data!$A$1:$A$20,MATCH({(r % 3 + 1) * 20},Data!$A$1:$A$20,0))",
        )
    wb.save(xlsx)


def _match_resolution_surface(db_path) -> dict:
    """Everything MATCH resolution can influence, keyed by stable addresses."""
    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    try:
        cell_edges = conn.execute(
            """
            SELECT fs.sheet_name || '!' || fc.a1, ts.sheet_name || '!' || tc.a1
            FROM cell_edges_internal e
            JOIN cells fc ON fc.cell_id = e.from_cell_id
            JOIN cells tc ON tc.cell_id = e.to_cell_id
            JOIN sheets fs ON fs.sheet_id = fc.sheet_id
            JOIN sheets ts ON ts.sheet_id = tc.sheet_id
            ORDER BY 1, 2
            """
        ).fetchall()
        metrics = conn.execute(
            "SELECT function_name, status, count FROM resolution_metrics ORDER BY 1, 2"
        ).fetchall()
    finally:
        conn.close()
    return {
        "binding_edges": _binding_edges(db_path),
        "cell_edges": [list(r) for r in cell_edges],
        "resolution_metrics": [list(r) for r in metrics],
    }


def test_match_resolution_surface_unchanged(tmp_path):
    xlsx = tmp_path / "lookup.xlsx"
    _create_lookup_dense_workbook(xlsx)
    db = extract(xlsx, tmp_path / "ir.db")
    surface = _match_resolution_surface(db)
    golden = json.loads((GOLDEN / "match_resolution_surface.json").read_text())
    assert surface == golden
