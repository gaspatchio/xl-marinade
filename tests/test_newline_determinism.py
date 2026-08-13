"""Generated artifacts must be byte-identical across platforms.

Text mode translates "\\n" to os.linesep on write, so on Windows every
generated file (documentation.md, model_spec.json, the diff changelist,
telemetry, mutation logs) gets CRLF while the same run on Linux/macOS gets LF.
The product's contract is deterministic output, and anything that hashes or
diffs these artifacts — goldens, docs CI, a user comparing two machines — sees
spurious changes.

The AST gate is the load-bearing test: on POSIX the runtime behaviour is
identical with or without the fix, so only a static check can bind it here.
The artifact check earns its keep the moment a Windows CI job exists.
"""

import ast
import json
from pathlib import Path

import openpyxl

from xl_marinade.core.api import diff, extract

SRC = Path(__file__).resolve().parent.parent / "src"


def _text_writes_without_newline(py: Path) -> list[str]:
    tree = ast.parse(py.read_text(encoding="utf-8"), filename=str(py))
    out: list[str] = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        name = (
            node.func.id
            if isinstance(node.func, ast.Name)
            else node.func.attr
            if isinstance(node.func, ast.Attribute)
            else ""
        )
        if name not in {"open", "write_text"}:
            continue
        kwargs = {kw.arg for kw in node.keywords}
        # Only text-mode WRITES translate newlines; reads use universal
        # newlines, and binary handles are unaffected.
        modes = [
            a.value for a in node.args if isinstance(a, ast.Constant) and isinstance(a.value, str)
        ]
        is_write = any(m in {"w", "a", "wt", "at"} for m in modes) or name == "write_text"
        if not is_write or "encoding" not in kwargs:
            continue
        if "newline" not in kwargs:
            out.append(f"{py.relative_to(SRC)}:{node.lineno}")
    return out


def test_text_writes_pin_lf():
    offenders: list[str] = []
    for py in sorted(SRC.rglob("*.py")):
        offenders.extend(_text_writes_without_newline(py))
    assert not offenders, (
        'text write without newline="\\n" — emits CRLF on Windows, so the '
        "artifact is not byte-identical across platforms:\n  " + "\n  ".join(offenders)
    )


def test_generated_artifacts_contain_no_crlf(tmp_path):
    """Byte-level check. Passes trivially on POSIX; binds on a Windows runner."""
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)

    db_a = extract(xlsx, tmp_path / "a.db")
    db_b = extract(xlsx, tmp_path / "b.db")

    changelist = tmp_path / "changes.json"
    changelist.write_text(json.dumps(diff(db_a, db_b)), encoding="utf-8", newline="\n")

    artifacts = [changelist, *tmp_path.rglob("telemetry*.json")]
    for artifact in artifacts:
        assert b"\r\n" not in artifact.read_bytes(), f"{artifact.name} contains CRLF"
