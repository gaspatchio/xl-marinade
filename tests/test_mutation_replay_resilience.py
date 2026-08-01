"""``replay_mutations`` must be resilient to conflicting LLM-proposed mutations.

The enrichment path (sprint7) feeds LLM-proposed structural mutations through
``replay_mutations``. Two merges can consume the same binding, so the later
merge hits an already-superseded source and raises ``MutationConflictError``.
Strict replay (the deterministic path) is all-or-nothing — correct there — but
the enrichment path opts into ``skip_conflicts`` so one bad mutation does not
discard the entire enrichment (previously it fell back to deterministic docs
after burning the LLM calls).
"""

import json
import sqlite3

import pytest

from xl_marinade.core.api import extract
from xl_marinade.core.labelling.mutation_engine import replay_mutations
from xl_marinade.core.labelling.mutation_errors import MutationConflictError


def _tiny_ir_db(tmp_path) -> tuple[str, list[str]]:
    """Extract a tiny workbook and return (ir_db_path, first 3 binding ids)."""
    import openpyxl

    xlsx = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"
    ws["B1"] = "=A3*2"
    ws["B2"] = "=B1+A1"
    wb.save(xlsx)
    db = extract(xlsx, tmp_path / "ir.db")
    conn = sqlite3.connect(f"file:{db}?mode=ro", uri=True)
    try:
        ids = [
            r[0]
            for r in conn.execute("SELECT binding_id FROM bindings ORDER BY binding_id LIMIT 3")
        ]
    finally:
        conn.close()
    return str(db), ids


def _conflicting_mutations(ids: list[str]) -> list[dict]:
    """Two merges that both consume ``ids[1]`` — the second conflicts."""
    b0, b1, b2 = ids
    return [
        {
            "mutation_id": 1,
            "timestamp": "2026-01-01T00:00:00Z",
            "action": "merge_bindings",
            "parameters": {
                "source_binding_ids": [b0, b1],
                "new_binding_id": "comp-1",
                "label": "First",
            },
        },
        {
            # b1 was superseded by mutation 1, so this merge conflicts.
            "mutation_id": 2,
            "timestamp": "2026-01-01T00:00:01Z",
            "action": "merge_bindings",
            "parameters": {
                "source_binding_ids": [b1, b2],
                "new_binding_id": "comp-2",
                "label": "Second",
            },
        },
    ]


def test_strict_replay_aborts_on_conflict(tmp_path):
    """Default (deterministic) replay stays all-or-nothing: the conflict raises."""
    db, ids = _tiny_ir_db(tmp_path)
    assert len(ids) >= 3, "fixture must yield >=3 bindings to exercise the conflict"
    mpath = tmp_path / "mutations.json"
    mpath.write_text(json.dumps(_conflicting_mutations(ids)))
    with pytest.raises(MutationConflictError):
        replay_mutations(db, str(mpath))


def test_skip_conflicts_applies_the_rest(tmp_path):
    """Enrichment mode skips the conflicting mutation and keeps the good one."""
    db, ids = _tiny_ir_db(tmp_path)
    assert len(ids) >= 3, "fixture must yield >=3 bindings to exercise the conflict"
    mpath = tmp_path / "mutations.json"
    mpath.write_text(json.dumps(_conflicting_mutations(ids)))

    overlay = replay_mutations(db, str(mpath), skip_conflicts=True)  # must NOT raise

    # The first, non-conflicting merge applied; the conflicting second did not.
    assert "comp-1" in overlay.bindings
    assert overlay.bindings["comp-1"].is_active
    assert "comp-2" not in overlay.bindings
