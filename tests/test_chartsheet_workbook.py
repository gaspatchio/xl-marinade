"""Workbooks containing chartsheets must not break name/table resolution.

Regression: ``LazyWorkbook.worksheets`` yielded every sheet by name — including
chartsheets, which openpyxl's own ``Workbook.worksheets`` excludes — and
``LazyWorksheet.__init__`` reads ``ws.max_row``, which chartsheets lack. One
chart-only sheet therefore aborted the ENTIRE name/table map build
("Warning: Failed to build name/table map: 'Chartsheet' object has no
attribute 'max_row'"), silently degrading defined-name and table resolution
for the whole workbook. Found on a real workbook
(VBA_Project_PortfolioOpti_FinalV.xlsm) during corpus testing.
"""

import openpyxl
from openpyxl.chart import BarChart, Reference

from xl_marinade.core.lazy_workbook import LazyWorkbook
from xl_marinade.core.names_tables import NameTableMap


def test_name_table_map_survives_chartsheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    wb.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName("my_range", attr_text="Sheet!$A$1:$A$2")
    )
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=2))
    wb.create_chartsheet("MyChart").add_chart(chart)
    xlsx = tmp_path / "with_chartsheet.xlsx"
    wb.save(xlsx)

    with LazyWorkbook(xlsx, data_only=False, keep_vba=False) as lazy:
        name_map = NameTableMap(lazy)
        worksheet_titles = [ws.title for ws in lazy.worksheets]

    assert "MyChart" not in worksheet_titles
    assert any(info.name == "my_range" for info in name_map._names.values())
