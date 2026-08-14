"""Extraction must not depend on the platform's locale encoding.

Field report (Windows, Python 3.14): `marinade extract` died on ANY workbook
with `'charmap' codec can't decode byte 0x90 in position 6904` — schema.sql
was read without an explicit encoding, so Windows decoded it with cp1252 and
choked on the UTF-8 arrow in a comment. The same class reproduces on POSIX by
forcing the C locale, where the default encoding becomes ASCII.

Ruff's PLW1514 (unspecified-encoding) now guards the pattern statically; this
test proves the end-to-end behaviour, subprocess-level, the way the field
failure actually presented.
"""

import os
import subprocess
import sys

import openpyxl


def test_extract_succeeds_under_c_locale(tmp_path):
    xlsx = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = "=A1+1"
    wb.save(xlsx)

    env = dict(os.environ)
    env.update(
        {
            "LC_ALL": "C",
            "LANG": "C",
            "PYTHONUTF8": "0",
            "PYTHONCOERCECLOCALE": "0",
        }
    )
    out_db = tmp_path / "out.db"
    proc = subprocess.run(
        [sys.executable, "-m", "xl_marinade.cli.main", "extract", str(xlsx), "-o", str(out_db)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
        timeout=300,
    )
    assert proc.returncode == 0, (
        f"extract failed under C locale (the Windows cp1252 failure class):\n{proc.stderr[-2000:]}"
    )
    assert out_db.exists()
