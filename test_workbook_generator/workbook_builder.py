# ABOUTME: Basic workbook creation and manipulation utilities for generating test Excel files.
# ABOUTME: Provides foundational functions for creating, populating, and saving Excel workbooks with validation.

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .formula_validator import (
    FormulaValidator,
    create_consistency_report,
    validate_workbook_formulas,
)


def create_workbook() -> Workbook:
    """
    Creates a new empty Excel workbook with deterministic default properties.

    Returns:
        Workbook: A new empty workbook with one active sheet named "Sheet"

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = 42
    """
    wb = Workbook()
    # Ensure deterministic sheet name
    if wb.active is not None:
        wb.active.title = "Sheet"
    return wb


def save_workbook(wb: Workbook, path: Path, overwrite: bool = True) -> None:
    """
    Saves a workbook to disk with validation and error handling.

    Args:
        wb: The workbook to save
        path: Destination file path (must end with .xlsx)
        overwrite: Whether to overwrite existing file (default: True)

    Raises:
        ValueError: If path doesn't end with .xlsx
        FileExistsError: If file exists and overwrite=False
        FileNotFoundError: If parent directory doesn't exist and can't be created
        PermissionError: If path is not writable

    Example:
        >>> wb = create_workbook()
        >>> save_workbook(wb, Path("tests/output/test.xlsx"))
    """
    # Validate file extension
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Path must end with .xlsx extension")

    # Check if file exists when overwrite=False
    if path.exists() and not overwrite:
        raise FileExistsError(f"File already exists and overwrite=False: {path}")

    # Create parent directories if they don't exist
    path.parent.mkdir(parents=True, exist_ok=True)

    # Save the workbook
    try:
        wb.save(path)
    except PermissionError as e:
        raise PermissionError(f"Cannot write to path: {path}") from e


def write_cell_value(ws: Worksheet, address: str, value: Any) -> None:
    """
    Writes a value to a cell with appropriate type handling.

    Args:
        ws: The worksheet to write to
        address: Cell address in A1 notation (e.g., "A1", "B2")
        value: Value to write (string, number, boolean, None, etc.)

    Raises:
        ValueError: If cell address is invalid
        KeyError: If cell address format is incorrect

    Example:
        >>> ws = wb.active
        >>> write_cell_value(ws, "A1", "Hello World")
        >>> write_cell_value(ws, "B1", 42)
    """
    if not address or not isinstance(address, str):
        raise ValueError("Cell address must be a non-empty string")

    try:
        ws[address] = value
    except (ValueError, KeyError) as e:
        raise ValueError(f"Invalid cell address: {address}") from e


def write_cell_formula(ws: Worksheet, address: str, formula: str) -> None:
    """
    Writes a formula to a cell with syntax validation.

    Args:
        ws: The worksheet to write to
        address: Cell address in A1 notation (e.g., "A1", "B2")
        formula: Excel formula (with or without leading '=')

    Raises:
        ValueError: If cell address is invalid or formula syntax is invalid
        KeyError: If cell address format is incorrect

    Example:
        >>> ws = wb.active
        >>> write_cell_formula(ws, "C1", "=A1+B1")
        >>> write_cell_formula(ws, "D1", "SUM(A1:B1)")  # Auto-adds =
    """
    if not formula or not isinstance(formula, str):
        raise ValueError("Formula cannot be empty")

    # Remove leading/trailing whitespace
    formula = formula.strip()

    if not formula or formula == "=":
        raise ValueError("Formula cannot be empty")

    # Automatically add equals sign if missing
    if not formula.startswith("="):
        formula = "=" + formula

    # Basic formula syntax validation
    _validate_formula_syntax(formula)

    # Write the formula to the cell
    try:
        ws[address] = formula
    except (ValueError, KeyError) as e:
        raise ValueError(f"Invalid cell address: {address}") from e


def _validate_formula_syntax(formula: str) -> None:
    """
    Validates basic Excel formula syntax.

    Args:
        formula: Formula string starting with '='

    Raises:
        ValueError: If formula has obvious syntax errors
    """
    # Check for basic syntax issues

    # Must start with =
    if not formula.startswith("="):
        raise ValueError("Invalid formula syntax: formula must start with '='")

    # Check for obvious issues like double operators
    if re.search(r"[+\-*/]{2,}", formula):
        raise ValueError("Invalid formula syntax: consecutive operators found")

    # Check for unmatched parentheses
    open_parens = formula.count("(")
    close_parens = formula.count(")")
    if open_parens != close_parens:
        raise ValueError("Invalid formula syntax: unmatched parentheses")

    # Check for unclosed ranges (basic check for : without proper closing)
    if ":" in formula and re.search(r"[A-Z]+\d+:\s*[^A-Z\d)]", formula):
        raise ValueError("Invalid formula syntax: incomplete range reference")

    # Formula passes basic validation
    pass


def validate_workbook_integrity(wb: Workbook, verbose: bool = False) -> bool:
    """
    Validates the integrity of a workbook using comprehensive formula validation.

    Args:
        wb: The workbook to validate
        verbose: Whether to print detailed validation results

    Returns:
        bool: True if workbook is valid, False otherwise

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = 10
        >>> ws['B1'] = 20
        >>> ws['C1'] = "=A1+B1"
        >>> validate_workbook_integrity(wb)
        True
    """
    try:
        # Validate all formulas in the workbook
        validation_results = validate_workbook_formulas(wb)

        # Check if any worksheet has errors
        has_errors = any(not result.is_valid for result in validation_results.values())

        if verbose:
            # Create and print consistency report
            report = create_consistency_report(validation_results)
            print(report)

        return not has_errors

    except Exception as e:
        if verbose:
            print(f"Validation error: {e}")
        return False


def validate_worksheet_formulas(ws: Worksheet, verbose: bool = False) -> bool:
    """
    Validates all formulas in a specific worksheet.

    Args:
        ws: The worksheet to validate
        verbose: Whether to print detailed validation results

    Returns:
        bool: True if worksheet is valid, False otherwise

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = 10
        >>> ws['B1'] = 20
        >>> ws['C1'] = "=A1+B1"
        >>> validate_worksheet_formulas(ws)
        True
    """
    try:
        validator = FormulaValidator(ws)
        result = validator.validate_all_formulas()

        if verbose:
            print("Worksheet validation results:")
            print(f"  Formulas checked: {result.total_formulas_checked}")
            print(f"  Errors: {len(result.errors)}")
            print(f"  Warnings: {len(result.warnings)}")
            print(f"  Circular references: {len(result.circular_refs)}")
            print(f"  Missing references: {len(result.missing_refs)}")
            print(f"  Valid: {'Yes' if result.is_valid else 'No'}")

            if result.errors:
                print("  Error details:")
                for error in result.errors:
                    print(f"    {error.cell_address}: {error.message}")

        return result.is_valid

    except Exception as e:
        if verbose:
            print(f"Validation error: {e}")
        return False


def check_circular_references(ws: Worksheet) -> list[str]:
    """
    Checks for circular references in a worksheet.

    Args:
        ws: The worksheet to check

    Returns:
        List[str]: List of cell addresses with circular references

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = "=B1+1"
        >>> ws['B1'] = "=A1+1"  # Creates circular reference
        >>> check_circular_references(ws)
        ['A1', 'B1']
    """
    try:
        validator = FormulaValidator(ws)
        result = validator.validate_all_formulas()
        return result.circular_refs
    except Exception:
        return []


def check_missing_references(ws: Worksheet) -> list[str]:
    """
    Checks for references to missing cells in a worksheet.

    Args:
        ws: The worksheet to check

    Returns:
        List[str]: List of missing cell references

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = "=B1+C1"  # B1 and C1 don't exist
        >>> check_missing_references(ws)
        ['B1', 'C1']
    """
    try:
        validator = FormulaValidator(ws)
        result = validator.validate_all_formulas()
        return result.missing_refs
    except Exception:
        return []


def get_volatile_functions(ws: Worksheet) -> list[tuple]:
    """
    Gets all volatile functions used in a worksheet.

    Args:
        ws: The worksheet to check

    Returns:
        List[tuple]: List of (cell_address, formula, volatile_functions) tuples

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> ws['A1'] = "=NOW()"
        >>> ws['B1'] = "=OFFSET(A1,1,1)"
        >>> get_volatile_functions(ws)
        [('A1', '=NOW()', ['NOW']), ('B1', '=OFFSET(A1,1,1)', ['OFFSET'])]
    """
    try:
        validator = FormulaValidator(ws)
        volatile_functions = []

        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula = str(cell.value)
                    volatile_funcs = validator.get_volatile_functions(formula)
                    if volatile_funcs:
                        volatile_functions.append((cell.coordinate, formula, volatile_funcs))

        return volatile_functions
    except Exception:
        return []


def create_validation_report(wb: Workbook, output_path: Path | None = None) -> str:
    """
    Creates a comprehensive validation report for a workbook.

    Args:
        wb: The workbook to validate
        output_path: Optional path to save the report to a file

    Returns:
        str: The validation report as a string

    Example:
        >>> wb = create_workbook()
        >>> # ... populate workbook ...
        >>> report = create_validation_report(wb)
        >>> print(report)
    """
    try:
        # Validate all formulas
        validation_results = validate_workbook_formulas(wb)

        # Create consistency report
        report = create_consistency_report(validation_results)

        # Save to file if path provided
        if output_path:
            output_path.write_text(report, encoding="utf-8")

        return report

    except Exception as e:
        error_report = f"Error creating validation report: {e}"
        if output_path:
            output_path.write_text(error_report, encoding="utf-8")
        return error_report
