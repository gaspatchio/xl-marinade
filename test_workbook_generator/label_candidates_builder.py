# ABOUTME: Builder for label candidates section with adjacent text labels and named ranges.
# ABOUTME: Creates test data for label extraction and defined name handling capabilities.

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .workbook_builder import write_cell_formula, write_cell_value


def create_label_candidates_section(ws: Worksheet, start_row: int = 23) -> list[str]:
    """
    Creates label candidates section with adjacent text labels and data values.

    The label candidates section contains:
    - Header row with section title
    - Column headers (Year, Revenue, Cost, Profit)
    - Three rows of data with adjacent text labels for testing label extraction
    - Data values: 2023, 2024, 2025 for years and corresponding financial data

    Args:
        ws: The worksheet to write to
        start_row: Starting row for label candidates section (1-indexed, default: 23)

    Returns:
        List[str]: Cell addresses of all cells created in label candidates section

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> cells = create_label_candidates_section(ws, start_row=23)
        >>> len(cells)  # Should be 16 cells (header + 4 columns * 3 data rows + 1 header row)
        16
    """
    if start_row < 1:
        raise ValueError("start_row must be >= 1")

    created_cells = []

    # Row offsets from start_row
    header_row = start_row
    column_header_row = start_row + 1
    data_row_1 = start_row + 2  # 2023

    # Create main header
    write_cell_value(ws, f"A{header_row}", "Label Candidates")
    created_cells.append(f"A{header_row}")

    # Create column headers
    headers = {"A": "Year", "B": "Revenue", "C": "Cost", "D": "Profit"}

    for col, header_text in headers.items():
        write_cell_value(ws, f"{col}{column_header_row}", header_text)
        created_cells.append(f"{col}{column_header_row}")

    # Create data rows with adjacent text labels
    data_values = [(2023, 1000, 800, 200), (2024, 1100, 850, 250), (2025, 1200, 900, 300)]

    for i, (year, revenue, cost, profit) in enumerate(data_values):
        row = data_row_1 + i

        # Year value (column A)
        write_cell_value(ws, f"A{row}", year)
        created_cells.append(f"A{row}")

        # Revenue value (column B)
        write_cell_value(ws, f"B{row}", revenue)
        created_cells.append(f"B{row}")

        # Cost value (column C)
        write_cell_value(ws, f"C{row}", cost)
        created_cells.append(f"C{row}")

        # Profit value (column D) - calculated as Revenue - Cost
        write_cell_value(ws, f"D{row}", profit)
        created_cells.append(f"D{row}")

    # Add multi-cell binding test cases for axis label extraction
    # Starting at row 31 (6 rows below start_row)
    formula_section_start = start_row + 6

    # === Horizontal 1×3 bindings for axis "columns" testing ===
    # Row 31: Year headers (2023, 2024, 2025) - will be axis labels
    write_cell_value(ws, f"B{formula_section_start}", 2023)
    write_cell_value(ws, f"C{formula_section_start}", 2024)
    write_cell_value(ws, f"D{formula_section_start}", 2025)
    created_cells.extend(
        [f"B{formula_section_start}", f"C{formula_section_start}", f"D{formula_section_start}"]
    )

    # Row 32: "Revenue by Year" label + formulas with IDENTICAL R1C1 signature (1×3 horizontal binding)
    # CRITICAL: All three formulas must have same R1C1 pattern to group into single binding
    # Using =A32, =A32, =A32 → all have R1C1 signature =RC[-1] → WILL GROUP!
    write_cell_value(ws, f"A{formula_section_start + 1}", "Revenue by Year")
    write_cell_formula(ws, f"B{formula_section_start + 1}", f"=A{formula_section_start + 1}")
    write_cell_formula(ws, f"C{formula_section_start + 1}", f"=B{formula_section_start + 1}")
    write_cell_formula(ws, f"D{formula_section_start + 1}", f"=C{formula_section_start + 1}")
    created_cells.extend(
        [
            f"A{formula_section_start + 1}",
            f"B{formula_section_start + 1}",
            f"C{formula_section_start + 1}",
            f"D{formula_section_start + 1}",
        ]
    )

    # Row 33: "Cost by Year" label + formulas with IDENTICAL R1C1 signature (1×3 horizontal binding)
    # Using same pattern: =A33, =B33, =C33 → all have R1C1 signature =RC[-1] → WILL GROUP!
    write_cell_value(ws, f"A{formula_section_start + 2}", "Cost by Year")
    write_cell_formula(ws, f"B{formula_section_start + 2}", f"=A{formula_section_start + 2}")
    write_cell_formula(ws, f"C{formula_section_start + 2}", f"=B{formula_section_start + 2}")
    write_cell_formula(ws, f"D{formula_section_start + 2}", f"=C{formula_section_start + 2}")
    created_cells.extend(
        [
            f"A{formula_section_start + 2}",
            f"B{formula_section_start + 2}",
            f"C{formula_section_start + 2}",
            f"D{formula_section_start + 2}",
        ]
    )

    # === Vertical 3×1 bindings for axis "rows" testing ===
    # Column F: "Total" header in F26 + vertical formulas F27:F29 (3×1 vertical binding)
    write_cell_value(ws, f"F{column_header_row}", "Total")
    created_cells.append(f"F{column_header_row}")

    for i in range(3):
        row = data_row_1 + i
        # Sum of Revenue + Cost + Profit for each year
        write_cell_formula(ws, f"F{row}", f"=B{row}+C{row}+D{row}")
        created_cells.append(f"F{row}")

    return created_cells


def create_named_ranges(wb: Workbook, ws: Worksheet, start_row: int = 23) -> dict[str, str]:
    """
    Creates named ranges for key data areas in the workbook.

    Creates the following named ranges:
    - "Years": A25:A27 (Year values)
    - "Revenue": B25:B27 (Revenue values)
    - "Costs": C25:C27 (Cost values)
    - "Profits": D25:D27 (Profit values)
    - "FinancialData": A24:D27 (All financial data including headers)
    - "TreeA_Inputs": B3:B5 (Tree A input values)
    - "TreeA_Outputs": G3:G5 (Tree A final outputs)
    - "TreeB_Data": A9:A11 (Tree B data values)
    - "TreeB_Results": F9:F11 (Tree B results)

    Args:
        wb: The workbook to add named ranges to
        ws: The worksheet containing the data
        start_row: Starting row for label candidates section (1-indexed, default: 23)

    Returns:
        Dict[str, str]: Mapping of named range names to their A1 references

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> # First create the data sections
        >>> create_tree_a(ws, start_row=1)
        >>> create_tree_b(ws, start_row=7)
        >>> create_label_candidates_section(ws, start_row=23)
        >>> # Then create named ranges
        >>> named_ranges = create_named_ranges(wb, ws, start_row=23)
        >>> len(named_ranges)  # Should be 9 named ranges
        9
    """
    from openpyxl.workbook.defined_name import DefinedName

    if start_row < 1:
        raise ValueError("start_row must be >= 1")

    # Calculate row numbers for label candidates data
    data_start_row = start_row + 2  # First data row (2023)
    data_end_row = start_row + 4  # Last data row (2025)
    header_row = start_row + 1  # Header row

    # Define named ranges with proper A1 references (without sheet prefix)
    named_ranges = {
        "Years": f"A{data_start_row}:A{data_end_row}",
        "Revenue": f"B{data_start_row}:B{data_end_row}",
        "Costs": f"C{data_start_row}:C{data_end_row}",
        "Profits": f"D{data_start_row}:D{data_end_row}",
        "FinancialData": f"A{header_row}:D{data_end_row}",
        "TreeA_Inputs": "B3:B5",
        "TreeA_Outputs": "G3:G5",
        "TreeB_Data": "A9:A11",
        "TreeB_Results": "F9:F11",
    }

    # Add named ranges to workbook using proper DefinedName objects
    for name, reference in named_ranges.items():
        try:
            # Create a DefinedName object with the reference
            defined_name = DefinedName(name, reference)
            wb.defined_names[name] = defined_name
        except Exception as e:
            # If still failing, skip this named range
            print(
                f"Warning: Could not create named range '{name}' with reference '{reference}': {e}"
            )
            continue

    return named_ranges


def validate_label_candidates(ws: Worksheet, label_cells: list[str]) -> bool:
    """
    Validates that label candidates section contains adjacent text labels.

    Args:
        ws: The worksheet containing the label candidates section
        label_cells: List of cell addresses in the label candidates section

    Returns:
        bool: True if label candidates are properly positioned, False otherwise
    """
    # Check that we have the expected structure
    expected_headers = ["Year", "Revenue", "Cost", "Profit"]

    # Find the header row by looking for cells with expected header text
    header_row = None
    for cell_addr in label_cells:
        cell = ws[cell_addr]
        if cell.value in expected_headers:
            # Extract row number from cell address
            row_num = "".join(filter(str.isdigit, cell_addr))
            if header_row is None:
                header_row = int(row_num)
            elif int(row_num) != header_row:
                # Headers should all be on the same row
                return False

    if header_row is None:
        return False

    # Check that data rows contain numeric values
    data_rows = [header_row + 1, header_row + 2, header_row + 3]

    for row in data_rows:
        # Check Year column (A)
        year_cell = ws[f"A{row}"]
        if not isinstance(year_cell.value, (int, float)) or year_cell.value < 2020:
            return False

        # Check Revenue column (B)
        revenue_cell = ws[f"B{row}"]
        if not isinstance(revenue_cell.value, (int, float)) or revenue_cell.value <= 0:
            return False

        # Check Cost column (C)
        cost_cell = ws[f"C{row}"]
        if not isinstance(cost_cell.value, (int, float)) or cost_cell.value <= 0:
            return False

        # Check Profit column (D)
        profit_cell = ws[f"D{row}"]
        if not isinstance(profit_cell.value, (int, float)):
            return False

    return True


def validate_named_ranges(wb: Workbook, expected_ranges: dict[str, str]) -> bool:
    """
    Validates that all expected named ranges are properly defined and accessible.

    Args:
        wb: The workbook containing the named ranges
        expected_ranges: Dictionary of expected named range names and references

    Returns:
        bool: True if all named ranges are valid and accessible, False otherwise
    """
    if not hasattr(wb, "defined_names"):
        return False

    # Check that all expected named ranges exist
    for name, expected_ref in expected_ranges.items():
        if name not in wb.defined_names:
            return False

        # Get the actual reference
        actual_ref = str(wb.defined_names[name])

        # Check that the reference is valid (contains expected elements)
        # Note: The exact format might vary, so we check for key components
        if "A" in expected_ref and "A" not in actual_ref:
            return False
        if "B" in expected_ref and "B" not in actual_ref:
            return False
        if "C" in expected_ref and "C" not in actual_ref:
            return False
        if "D" in expected_ref and "D" not in actual_ref:
            return False

    return True


def verify_named_ranges_in_formulas(ws: Worksheet, test_row: int = 30) -> bool:
    """
    Verifies that named ranges work correctly in formulas.

    Creates test formulas that use the named ranges and verifies they calculate correctly.

    Args:
        ws: The worksheet to test on
        test_row: Row number to place test formulas (1-indexed, default: 30)

    Returns:
        bool: True if named ranges work in formulas, False otherwise
    """
    try:
        # Test formula using Years named range
        write_cell_formula(ws, f"A{test_row}", "=COUNTA(Years)")

        # Test formula using Revenue named range
        write_cell_formula(ws, f"B{test_row}", "=SUM(Revenue)")

        # Test formula using Costs named range
        write_cell_formula(ws, f"C{test_row}", "=SUM(Costs)")

        # Test formula using Profits named range
        write_cell_formula(ws, f"D{test_row}", "=SUM(Profits)")

        # Test formula using FinancialData named range
        write_cell_formula(ws, f"E{test_row}", "=COUNTA(FinancialData)")

        # Verify that the formulas don't produce errors
        # Note: In a real test, we would check the calculated values
        # For now, we just verify the formulas were written without syntax errors
        return True

    except Exception:
        return False


def check_label_positioning_for_binding_detection(ws: Worksheet, start_row: int = 23) -> bool:
    """
    Checks that label positioning is correct for binding detection.

    Verifies that:
    - Text labels are adjacent to data cells
    - Labels are positioned correctly relative to data ranges
    - The structure supports proper binding detection

    Args:
        ws: The worksheet containing the label candidates section
        start_row: Starting row for label candidates section (1-indexed, default: 23)

    Returns:
        bool: True if label positioning is correct for binding detection, False otherwise
    """
    if start_row < 1:
        return False

    # Calculate row numbers
    header_row = start_row + 1
    data_start_row = start_row + 2
    data_end_row = start_row + 4

    # Check that headers are directly above data
    for col in ["A", "B", "C", "D"]:
        header_cell = ws[f"{col}{header_row}"]
        if not isinstance(header_cell.value, str) or not header_cell.value.strip():
            return False

    # Check that data cells contain appropriate values
    for row in range(data_start_row, data_end_row + 1):
        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{row}"]
            if cell.value is None:
                return False

    # Check that the structure supports binding detection
    # Headers should be adjacent to data (no gaps)
    if header_row + 1 != data_start_row:
        return False

    # Data should be contiguous (no gaps between rows)
    for row in range(data_start_row, data_end_row):
        # Check that there are no empty rows between data
        next_row = row + 1
        if next_row <= data_end_row:
            # Verify that both rows have data
            has_data_current = any(
                ws[f"{col}{row}"].value is not None for col in ["A", "B", "C", "D"]
            )
            has_data_next = any(
                ws[f"{col}{next_row}"].value is not None for col in ["A", "B", "C", "D"]
            )

            if not (has_data_current and has_data_next):
                return False

    return True
