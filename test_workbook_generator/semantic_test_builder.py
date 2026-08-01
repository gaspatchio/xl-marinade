# ABOUTME: Builder for creating comprehensive semantic resolution and constant grouping test workbooks.
# ABOUTME: Generates 78 test cases covering VLOOKUP, HLOOKUP, INDEX, MATCH, XLOOKUP, CHOOSE, ADDRESS, and constant grouping scenarios.

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .workbook_builder import create_workbook, save_workbook, write_cell_formula, write_cell_value


class SemanticTestBuilder:
    """
    Builder for creating comprehensive semantic test workbooks.

    Generates workbook with 78 test cases:
    - 56 lookup resolution tests (VLOOKUP, HLOOKUP, INDEX, MATCH, XLOOKUP, CHOOSE, ADDRESS)
    - 19 constant grouping tests (single-column, single-row, multi-dimensional, edge creation)
    - 3 integration tests (resolution + grouping combined)
    """

    def __init__(self) -> None:
        """Initialize semantic test builder."""
        self.wb: Workbook = create_workbook()

    def create_workbook(self, output_path: Path) -> Workbook:
        """
        Create complete semantic test workbook with all 78 test cases.

        Args:
            output_path: Path to save the workbook

        Returns:
            Workbook: The created workbook
        """
        # Create all sheets
        self._add_lookup_resolution_sheet()
        self._add_constant_grouping_sheet()
        self._add_integration_sheet()
        self._add_assumptions_sheet()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Save workbook
        save_workbook(self.wb, output_path)

        return self.wb

    def _add_lookup_resolution_sheet(self) -> None:
        """Create Lookup_Resolution sheet with all 56 test cases."""
        ws = self.wb.create_sheet("Lookup_Resolution", 0)

        # Add headers
        write_cell_value(ws, "A1", "Test_ID")
        write_cell_value(ws, "B1", "Formula")
        write_cell_value(ws, "C1", "Expected_Status")
        write_cell_value(ws, "D1", "Expected_Result")

        # Populate support data
        self._populate_lookup_tables(ws)
        self._populate_support_cells(ws)

        # Add test formulas
        self._add_vlookup_tests(ws)
        self._add_hlookup_tests(ws)
        self._add_index_tests(ws)
        self._add_match_tests(ws)
        self._add_xlookup_tests(ws)
        self._add_choose_tests(ws)
        self._add_address_tests(ws)

    def _populate_lookup_tables(self, ws: Worksheet) -> None:
        """Populate main lookup tables in E100:G200 and J100:L110."""
        # Main lookup table E100:G200
        for i in range(101):
            row = 100 + i
            write_cell_value(ws, f"E{row}", i + 1)  # ID: 1, 2, 3, ..., 101
            write_cell_value(ws, f"F{row}", f"Name_{i + 1}")  # Names
            write_cell_value(ws, f"G{row}", (i + 1) * 10)  # Values: 10, 20, 30, ..., 1010

        # Small INDEX/MATCH table J100:L110
        for i in range(11):
            row = 100 + i
            write_cell_value(ws, f"J{row}", i + 1)  # Keys: 1-11
            write_cell_value(ws, f"K{row}", 100 + i)  # Values: 100-110
            write_cell_value(ws, f"L{row}", f"Label_{i + 1}")  # Labels

    def _populate_support_cells(self, ws: Worksheet) -> None:
        """Populate support cells for column/row indices (H10:H30)."""
        write_cell_value(ws, "H10", 2)  # For VLOOKUP_02
        write_cell_value(ws, "H11", 1)  # For VLOOKUP_04 (expression test: H11+1)
        write_cell_value(ws, "H12", 2)  # For HLOOKUP_02
        write_cell_value(ws, "H13", 1)  # For HLOOKUP_04 (expression test: H13+1)
        write_cell_value(ws, "H20", 5)  # For MATCH_02
        write_cell_value(ws, "H21", 50)  # For MATCH_04
        write_cell_value(ws, "H22", 2)  # For CHOOSE_02
        write_cell_value(ws, "H23", 5)  # For ADDRESS_02 (row)
        write_cell_value(ws, "H24", 3)  # For ADDRESS_02 (column)

        # Support cells for INDEX tests
        write_cell_value(ws, "D35", "=1+1")  # Formula for INDEX_06 (unresolvable)
        write_cell_value(ws, "E35", "=2*1")  # Formula for INDEX_06 (unresolvable)

    def _add_vlookup_tests(self, ws: Worksheet) -> None:
        """Create all 11 VLOOKUP test formulas (rows 10-20)."""
        # VLOOKUP_01: Literal column index
        write_cell_value(ws, "A10", "VLOOKUP_01_Literal_ColIndex")
        write_cell_formula(ws, "B10", "=VLOOKUP(E101, E100:G200, 2, 0)")
        write_cell_value(ws, "C10", "resolved")
        write_cell_value(ws, "D10", "F101")

        # VLOOKUP_02: Cell reference column index
        write_cell_value(ws, "A11", "VLOOKUP_02_CellRef_ColIndex")
        write_cell_formula(ws, "B11", "=VLOOKUP(E101, E100:G200, H10, 0)")
        write_cell_value(ws, "C11", "resolved")
        write_cell_value(ws, "D11", "F101")

        # VLOOKUP_03: MATCH column index
        write_cell_value(ws, "A12", "VLOOKUP_03_MATCH_ColIndex")
        write_cell_formula(ws, "B12", '=VLOOKUP(E101, E100:G200, MATCH("Name", E100:G100, 0), 0)')
        write_cell_value(ws, "C12", "resolved")
        write_cell_value(ws, "D12", "F101")

        # VLOOKUP_04: Expression column index
        write_cell_value(ws, "A13", "VLOOKUP_04_Expression_ColIndex")
        write_cell_formula(ws, "B13", "=VLOOKUP(E101, E100:G200, H11+1, 0)")
        write_cell_value(ws, "C13", "resolved")
        write_cell_value(ws, "D13", "F101")

        # VLOOKUP_05: Nested VLOOKUP
        write_cell_value(ws, "A14", "VLOOKUP_05_Nested_VLOOKUP")
        write_cell_formula(ws, "B14", "=VLOOKUP(VLOOKUP(1, E100:F105, 2, 0), E100:G200, 2, 0)")
        write_cell_value(ws, "C14", "resolved")
        write_cell_value(ws, "D14", "F100:F200")

        # VLOOKUP_06: Cross-sheet reference
        write_cell_value(ws, "A15", "VLOOKUP_06_CrossSheet")
        write_cell_formula(ws, "B15", "=VLOOKUP(E101, Assumptions!E10:G110, 2, 0)")
        write_cell_value(ws, "C15", "resolved")
        write_cell_value(ws, "D15", "Assumptions!F10:F110")

        # VLOOKUP_07: IF column index (unresolvable)
        write_cell_value(ws, "A16", "VLOOKUP_07_IF_ColIndex")
        write_cell_formula(ws, "B16", "=VLOOKUP(E101, E100:G200, IF(H10>0, 2, 3), 0)")
        write_cell_value(ws, "C16", "conservative_fallback")
        write_cell_value(ws, "D16", "E100:G200")

        # VLOOKUP_08: Complex unresolvable (RAND)
        write_cell_value(ws, "A17", "VLOOKUP_08_Complex_Unresolvable")
        write_cell_formula(ws, "B17", "=VLOOKUP(E101, E100:G200, RAND()*3, 0)")
        write_cell_value(ws, "C17", "conservative_fallback")
        write_cell_value(ws, "D17", "E100:G200")

        # VLOOKUP_09: Invalid column index (0)
        write_cell_value(ws, "A18", "VLOOKUP_09_Invalid_ColIndex")
        write_cell_formula(ws, "B18", "=VLOOKUP(E101, E100:G200, 0, 0)")
        write_cell_value(ws, "C18", "conservative_fallback")
        write_cell_value(ws, "D18", "E100:G200")

        # VLOOKUP_10: Column index out of range
        write_cell_value(ws, "A19", "VLOOKUP_10_ColIndex_OutOfRange")
        write_cell_formula(ws, "B19", "=VLOOKUP(E101, E100:G200, 5, 0)")
        write_cell_value(ws, "C19", "conservative_fallback")
        write_cell_value(ws, "D19", "E100:G200")

        # VLOOKUP_11: Invalid context (formula with error)
        write_cell_value(ws, "A20", "VLOOKUP_11_Invalid_Context")
        write_cell_formula(ws, "B20", "=VLOOKUP(E101, E100:G200, 2, 0) + #REF!")
        write_cell_value(ws, "C20", "invalid_context")
        write_cell_value(ws, "D20", "N/A")

    def _add_hlookup_tests(self, ws: Worksheet) -> None:
        """Create all 5 HLOOKUP test formulas (rows 25-29)."""
        # First, create horizontal lookup table in A95:Z97 (avoiding E100:G200 range)
        for col_idx in range(26):  # A-Z
            col_letter = chr(65 + col_idx)
            write_cell_value(ws, f"{col_letter}95", col_idx + 1)  # Row 1: IDs
            write_cell_value(ws, f"{col_letter}96", f"Name_{col_idx + 1}")  # Row 2: Names
            write_cell_value(ws, f"{col_letter}97", (col_idx + 1) * 10)  # Row 3: Values

        # HLOOKUP_01: Literal row index
        write_cell_value(ws, "A25", "HLOOKUP_01_Literal_RowIndex")
        write_cell_formula(ws, "B25", "=HLOOKUP(5, A95:Z97, 2, 0)")
        write_cell_value(ws, "C25", "resolved")
        write_cell_value(ws, "D25", "E96")

        # HLOOKUP_02: Cell reference row index
        write_cell_value(ws, "A26", "HLOOKUP_02_CellRef_RowIndex")
        write_cell_formula(ws, "B26", "=HLOOKUP(5, A95:Z97, H12, 0)")
        write_cell_value(ws, "C26", "resolved")
        write_cell_value(ws, "D26", "E96")

        # HLOOKUP_03: MATCH row index
        write_cell_value(ws, "A27", "HLOOKUP_03_MATCH_RowIndex")
        write_cell_formula(ws, "B27", '=HLOOKUP(5, A95:Z97, MATCH("Name_5", E95:E97, 0), 0)')
        write_cell_value(ws, "C27", "resolved")
        write_cell_value(ws, "D27", "E96")

        # HLOOKUP_04: Expression row index
        write_cell_value(ws, "A28", "HLOOKUP_04_Expression_RowIndex")
        write_cell_formula(ws, "B28", "=HLOOKUP(5, A95:Z97, H13+1, 0)")
        write_cell_value(ws, "C28", "resolved")
        write_cell_value(ws, "D28", "E96")

        # HLOOKUP_05: Cross-sheet reference
        write_cell_value(ws, "A29", "HLOOKUP_05_CrossSheet")
        write_cell_formula(ws, "B29", "=HLOOKUP(5, Assumptions!A120:Z122, 2, 0)")
        write_cell_value(ws, "C29", "resolved")
        write_cell_value(ws, "D29", "Assumptions!E121")

    def _add_index_tests(self, ws: Worksheet) -> None:
        """Create all 6 INDEX test formulas (rows 35-40)."""
        # INDEX_01: Both row and column resolved
        write_cell_value(ws, "A35", "INDEX_01_Both_Resolved")
        write_cell_formula(ws, "B35", "=INDEX(J100:L103, 2, 2)")
        write_cell_value(ws, "C35", "resolved")
        write_cell_value(ws, "D35", "K101")

        # INDEX_02: Row only (column 0 means entire row)
        write_cell_value(ws, "A36", "INDEX_02_Row_Only")
        write_cell_formula(ws, "B36", "=INDEX(J100:L103, 2, 0)")
        write_cell_value(ws, "C36", "partial_resolved")
        write_cell_value(ws, "D36", "J101:L101")

        # INDEX_03: Column only (row 0 means entire column)
        write_cell_value(ws, "A37", "INDEX_03_Col_Only")
        write_cell_formula(ws, "B37", "=INDEX(J100:L103, 0, 2)")
        write_cell_value(ws, "C37", "partial_resolved")
        write_cell_value(ws, "D37", "K100:K103")

        # INDEX_04: MATCH for row
        write_cell_value(ws, "A38", "INDEX_04_MATCH_Row")
        write_cell_formula(ws, "B38", "=INDEX(J100:L110, MATCH(5, J100:J110, 0), 2)")
        write_cell_value(ws, "C38", "resolved")
        write_cell_value(ws, "D38", "K104")

        # INDEX_05: MATCH for both row and column
        write_cell_value(ws, "A39", "INDEX_05_MATCH_Both")
        write_cell_formula(
            ws, "B39", '=INDEX(J100:L110, MATCH(5, J100:J110, 0), MATCH("Label_5", J100:L100, 0))'
        )
        write_cell_value(ws, "C39", "resolved")
        write_cell_value(ws, "D39", "L104")

        # INDEX_06: Neither resolved (both are formulas)
        write_cell_value(ws, "A40", "INDEX_06_Neither_Resolved")
        write_cell_formula(ws, "B40", "=INDEX(J100:L103, D35, E35)")
        write_cell_value(ws, "C40", "unresolved")
        write_cell_value(ws, "D40", "N/A")

    def _add_match_tests(self, ws: Worksheet) -> None:
        """Create all 5 MATCH test formulas (rows 45-49)."""
        # MATCH_01: Literal value
        write_cell_value(ws, "A45", "MATCH_01_Literal_Value")
        write_cell_formula(ws, "B45", "=MATCH(5, J100:J110, 0)")
        write_cell_value(ws, "C45", "resolved")
        write_cell_value(ws, "D45", "5")

        # MATCH_02: Cell reference value
        write_cell_value(ws, "A46", "MATCH_02_CellRef_Value")
        write_cell_formula(ws, "B46", "=MATCH(H20, J100:J110, 0)")
        write_cell_value(ws, "C46", "resolved")
        write_cell_value(ws, "D46", "5")

        # MATCH_03: Nested in INDEX
        write_cell_value(ws, "A47", "MATCH_03_Nested_In_INDEX")
        write_cell_formula(ws, "B47", "=INDEX(J100:J110, MATCH(H20, K100:K110, 0))")
        write_cell_value(ws, "C47", "resolved")
        write_cell_value(ws, "D47", "J104")

        # MATCH_04: Approximate match (match_type = 1)
        write_cell_value(ws, "A48", "MATCH_04_Approximate_Match")
        write_cell_formula(ws, "B48", "=MATCH(H21, J100:J110, 1)")
        write_cell_value(ws, "C48", "resolved")
        write_cell_value(ws, "D48", "Position")

        # MATCH_05: Unresolvable value (RAND)
        write_cell_value(ws, "A49", "MATCH_05_Unresolvable_Value")
        write_cell_formula(ws, "B49", "=MATCH(RAND(), J100:J110, 0)")
        write_cell_value(ws, "C49", "conservative_fallback")
        write_cell_value(ws, "D49", "N/A")

    def _add_xlookup_tests(self, ws: Worksheet) -> None:
        """Create all 2 XLOOKUP test formulas (rows 55-56)."""
        # XLOOKUP_01: Basic
        write_cell_value(ws, "A55", "XLOOKUP_01_Basic")
        write_cell_formula(ws, "B55", "=XLOOKUP(5, J100:J110, K100:K110)")
        write_cell_value(ws, "C55", "resolved")
        write_cell_value(ws, "D55", "K100:K110")

        # XLOOKUP_02: With if_not_found
        write_cell_value(ws, "A56", "XLOOKUP_02_With_IfNotFound")
        write_cell_formula(ws, "B56", '=XLOOKUP(5, J100:J110, K100:K110, "Not Found")')
        write_cell_value(ws, "C56", "resolved")
        write_cell_value(ws, "D56", "K100:K110")

    def _add_choose_tests(self, ws: Worksheet) -> None:
        """Create all 3 CHOOSE test formulas (rows 60-62)."""
        # CHOOSE_01: Literal index
        write_cell_value(ws, "A60", "CHOOSE_01_Literal_Index")
        write_cell_formula(ws, "B60", "=CHOOSE(2, J100, K100, L100)")
        write_cell_value(ws, "C60", "resolved")
        write_cell_value(ws, "D60", "K100")

        # CHOOSE_02: Cell reference index
        write_cell_value(ws, "A61", "CHOOSE_02_CellRef_Index")
        write_cell_formula(ws, "B61", "=CHOOSE(H22, J100, K100, L100)")
        write_cell_value(ws, "C61", "resolved")
        write_cell_value(ws, "D61", "K100")

        # CHOOSE_03: Unresolvable index
        write_cell_value(ws, "A62", "CHOOSE_03_Unresolvable_Index")
        write_cell_formula(ws, "B62", "=CHOOSE(RAND()*3, J100, K100, L100)")
        write_cell_value(ws, "C62", "conservative_fallback")
        write_cell_value(ws, "D62", "J100,K100,L100")

    def _add_address_tests(self, ws: Worksheet) -> None:
        """Create all 3 ADDRESS test formulas (rows 65-67)."""
        # ADDRESS_01: Literal both
        write_cell_value(ws, "A65", "ADDRESS_01_Literal_Both")
        write_cell_formula(ws, "B65", "=ADDRESS(5, 3)")
        write_cell_value(ws, "C65", "resolved")
        write_cell_value(ws, "D65", "C5")

        # ADDRESS_02: Cell reference both
        write_cell_value(ws, "A66", "ADDRESS_02_CellRef_Both")
        write_cell_formula(ws, "B66", "=ADDRESS(H23, H24)")
        write_cell_value(ws, "C66", "resolved")
        write_cell_value(ws, "D66", "C5")

        # ADDRESS_03: With sheet
        write_cell_value(ws, "A67", "ADDRESS_03_With_Sheet")
        write_cell_formula(ws, "B67", '=ADDRESS(5, 3, 1, 1, "Sheet2")')
        write_cell_value(ws, "C67", "resolved")
        write_cell_value(ws, "D67", "Sheet2!C5")

    def _add_constant_grouping_sheet(self) -> None:
        """Create Constant_Grouping sheet with all 19 test cases."""
        ws = self.wb.create_sheet("Constant_Grouping")

        # Add headers
        write_cell_value(ws, "A1", "Test_ID")
        write_cell_value(ws, "B1", "Test_Type")
        write_cell_value(ws, "C1", "Expected_Merge")
        write_cell_value(ws, "D1", "Reason")

        # Single-column tests (F6:F25)
        self._add_single_column_grouping_tests(ws)

        # Single-row tests (Row 40)
        self._add_single_row_grouping_tests(ws)

        # Multi-dimensional tests (E50:G60, A80:B84)
        self._add_multi_dim_grouping_tests(ws)

        # Edge creation tests (use formulas referencing constant ranges)
        self._add_edge_creation_tests(ws)

    def _add_single_column_grouping_tests(self, ws: Worksheet) -> None:
        """Create single-column grouping test data (column H and I to avoid F conflict)."""
        # H6:H10 - Numbers
        for i, val in enumerate([10, 20, 30, 40, 50], start=6):
            write_cell_value(ws, f"H{i}", val)

        # H11:H20 - Numbers
        for i, val in enumerate([60, 70, 80, 90, 100, 110, 120, 130, 140, 150], start=11):
            write_cell_value(ws, f"H{i}", val)

        # H21:H25 - Text
        for i, val in enumerate(["Text_1", "Text_2", "Text_3", "Text_4", "Text_5"], start=21):
            write_cell_value(ws, f"H{i}", val)

        # I8:I12 - Numbers (misaligned test)
        for i, val in enumerate([100, 200, 300, 400, 500], start=8):
            write_cell_value(ws, f"I{i}", val)

        # Test metadata
        write_cell_value(ws, "A6", "CONST_SC_01_Adjacent_Same_Dtype")
        write_cell_value(ws, "B6", "Single_Column")
        write_cell_value(ws, "C6", "H6:H20")
        write_cell_value(ws, "D6", "Adjacent, same dtype")

        write_cell_value(ws, "A7", "CONST_SC_02_Overlapping")
        write_cell_value(ws, "B7", "Single_Column")
        write_cell_value(ws, "C7", "H6:H20")
        write_cell_value(ws, "D7", "Overlapping")

        write_cell_value(ws, "A8", "CONST_SC_03_Contained")
        write_cell_value(ws, "B8", "Single_Column")
        write_cell_value(ws, "C8", "H6:H20")
        write_cell_value(ws, "D8", "Containment")

        write_cell_value(ws, "A9", "CONST_SC_04_Different_Dtype")
        write_cell_value(ws, "B9", "Single_Column")
        write_cell_value(ws, "C9", "No merge")
        write_cell_value(ws, "D9", "Different dtypes")

        write_cell_value(ws, "A10", "CONST_SC_05_With_Blanks")
        write_cell_value(ws, "B10", "Single_Column")
        write_cell_value(ws, "C10", "H6:H20")
        write_cell_value(ws, "D10", "Blanks allowed")

        write_cell_value(ws, "A11", "CONST_SC_06_Misaligned")
        write_cell_value(ws, "B11", "Single_Column")
        write_cell_value(ws, "C11", "Multi-dim test")
        write_cell_value(ws, "D11", "Misaligned columns")

    def _add_single_row_grouping_tests(self, ws: Worksheet) -> None:
        """Create single-row grouping test data (row 40)."""
        # A40:C40 - Numbers
        for col, val in zip(["A", "B", "C"], [10, 20, 30], strict=False):
            write_cell_value(ws, f"{col}40", val)

        # D40:E40 - Numbers
        for col, val in zip(["D", "E"], [40, 50], strict=False):
            write_cell_value(ws, f"{col}40", val)

        # G40:I40 - Text
        for col, val in zip(["G", "H", "I"], ["A", "B", "C"], strict=False):
            write_cell_value(ws, f"{col}40", val)

        # Test metadata
        write_cell_value(ws, "A41", "CONST_SR_01_Adjacent_Same_Dtype")
        write_cell_value(ws, "B41", "Single_Row")
        write_cell_value(ws, "C41", "A40:E40")
        write_cell_value(ws, "D41", "Adjacent, same dtype")

        write_cell_value(ws, "A42", "CONST_SR_02_Overlapping")
        write_cell_value(ws, "B42", "Single_Row")
        write_cell_value(ws, "C42", "A40:F40")
        write_cell_value(ws, "D42", "Overlapping")

        write_cell_value(ws, "A43", "CONST_SR_03_Different_Dtype")
        write_cell_value(ws, "B43", "Single_Row")
        write_cell_value(ws, "C43", "No merge")
        write_cell_value(ws, "D43", "Different dtypes")

        write_cell_value(ws, "A44", "CONST_SR_04_With_Blanks")
        write_cell_value(ws, "B44", "Single_Row")
        write_cell_value(ws, "C44", "A40:J40")
        write_cell_value(ws, "D44", "Blanks allowed")

    def _add_multi_dim_grouping_tests(self, ws: Worksheet) -> None:
        """Create multi-dimensional grouping test data."""
        # N50:N60 - Numbers (moved from E to avoid conflicts)
        for i, val in enumerate(range(10, 120, 10), start=50):
            write_cell_value(ws, f"N{i}", val)

        # O50:O60 - Numbers (moved from F to avoid F6:F607 conflict)
        for i, val in enumerate(range(15, 125, 10), start=50):
            write_cell_value(ws, f"O{i}", val)

        # P50:P60 - Text (moved from G)
        for i in range(50, 61):
            write_cell_value(ws, f"P{i}", f"Row_{i}")

        # Row 56 is fully blank for boundary test (already blank by default)

        # K80:L84 - Numbers with sparse blanks (moved from A80:B84 to avoid edge test metadata conflict)
        for i in range(80, 85):
            if i != 81:  # K81 will be blank
                write_cell_value(ws, f"K{i}", i * 10)
            if i != 83:  # L83 will be blank
                write_cell_value(ws, f"L{i}", i * 10 + 5)

        # Test metadata (updated column references N:P)
        write_cell_value(ws, "A51", "CONST_MD_01_Adjacent_Columns")
        write_cell_value(ws, "B51", "Multi_Dim")
        write_cell_value(ws, "C51", "N50:O60")
        write_cell_value(ws, "D51", "Adjacent columns")

        write_cell_value(ws, "A52", "CONST_MD_02_Misaligned")
        write_cell_value(ws, "B52", "Multi_Dim")
        write_cell_value(ws, "C52", "Bbox merge if same dtype")
        write_cell_value(ws, "D52", "Misaligned")

        write_cell_value(ws, "A53", "CONST_MD_03_Blank_Row_Boundary")
        write_cell_value(ws, "B53", "Multi_Dim")
        write_cell_value(ws, "C53", "No merge")
        write_cell_value(ws, "D53", "Fully blank row 56")

        write_cell_value(ws, "A54", "CONST_MD_04_Blank_Column_Boundary")
        write_cell_value(ws, "B54", "Multi_Dim")
        write_cell_value(ws, "C54", "No merge")
        write_cell_value(ws, "D54", "Fully blank column")

        write_cell_value(ws, "A55", "CONST_MD_05_Different_Dtype")
        write_cell_value(ws, "B55", "Multi_Dim")
        write_cell_value(ws, "C55", "No merge")
        write_cell_value(ws, "D55", "Different dtypes")

        write_cell_value(ws, "A56", "CONST_MD_06_Sparse_Blanks")
        write_cell_value(ws, "B56", "Multi_Dim")
        write_cell_value(ws, "C56", "K80:L84")
        write_cell_value(ws, "D56", "Individual blanks OK")

    def _add_edge_creation_tests(self, ws: Worksheet) -> None:
        """Create edge creation test formulas (rows 80-82)."""
        # Create large constant range F6:F607 for edge tests
        for i in range(6, 608):
            write_cell_value(ws, f"F{i}", i * 10)

        # Edge test metadata
        write_cell_value(ws, "A80", "CONST_EDGE_01_Exact_Match")
        write_cell_value(ws, "B80", "Edge_Creation")
        write_cell_value(ws, "C80", "1 edge to F6:F607")
        write_cell_value(ws, "D80", "Exact match")

        write_cell_value(ws, "A81", "CONST_EDGE_02_Multiple_Overlaps")
        write_cell_value(ws, "B81", "Edge_Creation")
        write_cell_value(ws, "C81", "2 edges")
        write_cell_value(ws, "D81", "Overlapping bindings")

        write_cell_value(ws, "A82", "CONST_EDGE_03_Larger_Binding")
        write_cell_value(ws, "B82", "Edge_Creation")
        write_cell_value(ws, "C82", "Edge to F6:F607")
        write_cell_value(ws, "D82", "Formula refs subset")

    def _add_integration_sheet(self) -> None:
        """Create Integration sheet with 3 combined resolution + grouping tests."""
        ws = self.wb.create_sheet("Integration")

        # Headers
        write_cell_value(ws, "A1", "Test_ID")
        write_cell_value(ws, "B1", "Formula")
        write_cell_value(ws, "C1", "Semantic_Ref")
        write_cell_value(ws, "D1", "Grouped_Binding")

        # Test values for formulas
        write_cell_value(ws, "A10", 1)
        write_cell_value(ws, "A11", 10)
        write_cell_value(ws, "A12", 1)
        write_cell_value(ws, "B12", 1)

        # INTEGRATION_01: VLOOKUP with grouping
        write_cell_value(ws, "A5", "INTEGRATION_01_VLOOKUP_With_Grouping")
        write_cell_formula(ws, "B5", "=VLOOKUP(A10, Constant_Grouping!E6:G607, 2, 0)")
        write_cell_value(ws, "C5", "Constant_Grouping!F6:F607")
        write_cell_value(ws, "D5", "F6:F607")

        # INTEGRATION_02: INDEX with grouping
        write_cell_value(ws, "A6", "INTEGRATION_02_INDEX_With_Grouping")
        write_cell_formula(ws, "B6", "=INDEX(Constant_Grouping!F6:F607, A11)")
        write_cell_value(ws, "C6", "Constant_Grouping!F6:F607")
        write_cell_value(ws, "D6", "F6:F607")

        # INTEGRATION_03: Fallback with grouping (should NOT merge different dtypes)
        write_cell_value(ws, "A7", "INTEGRATION_03_Fallback_With_Grouping")
        write_cell_formula(ws, "B7", "=VLOOKUP(A12, Constant_Grouping!E6:G607, IF(B12>0,2,3), 0)")
        write_cell_value(ws, "C7", "Constant_Grouping!E6:G607")
        write_cell_value(ws, "D7", "No merge (fallback)")

    def _add_assumptions_sheet(self) -> None:
        """Create Assumptions sheet for cross-sheet reference tests."""
        ws = self.wb.create_sheet("Assumptions")

        # Duplicate main lookup table E10:G110
        for i in range(101):
            row = 10 + i
            write_cell_value(ws, f"E{row}", i + 1)  # ID: 1-101
            write_cell_value(ws, f"F{row}", f"Name_{i + 1}")  # Names
            write_cell_value(ws, f"G{row}", (i + 1) * 10)  # Values

        # Duplicate horizontal lookup table A120:Z122 for HLOOKUP cross-sheet test (fully separate from vertical table)
        for col_idx in range(26):  # A-Z
            col_letter = chr(65 + col_idx)
            write_cell_value(ws, f"{col_letter}120", col_idx + 1)  # Row 1: IDs
            write_cell_value(ws, f"{col_letter}121", f"Name_{col_idx + 1}")  # Row 2: Names
            write_cell_value(ws, f"{col_letter}122", (col_idx + 1) * 10)  # Row 3: Values
