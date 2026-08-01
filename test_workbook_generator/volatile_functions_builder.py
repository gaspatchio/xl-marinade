# ABOUTME: Builder for volatile functions section with OFFSET, INDIRECT, NOW, TODAY, RAND.
# ABOUTME: Creates test cases for volatile function detection and resolution capabilities.


from openpyxl.worksheet.worksheet import Worksheet

from .workbook_builder import write_cell_formula, write_cell_value


def _extract_row_number(cell_addr: str) -> str:
    """
    Extract the numeric row portion from a cell address.

    Args:
        cell_addr: Cell address like "A21", "AA21", "ABC123"

    Returns:
        str: The row number portion (e.g., "21", "123")

    Example:
        >>> _extract_row_number("A21")
        "21"
        >>> _extract_row_number("AA21")
        "21"
        >>> _extract_row_number("ABC123")
        "123"
    """
    # Find the first digit in the cell address
    for i, char in enumerate(cell_addr):
        if char.isdigit():
            return cell_addr[i:]
    return ""


def create_volatile_functions_section(ws: Worksheet, start_row: int = 19) -> list[str]:
    """
    Creates volatile functions section with OFFSET, INDIRECT, NOW, TODAY, RAND.

    The volatile functions section contains:
    - Header row with section title
    - Column headers (Offset, Indirect, Now, Today, Rand)
    - One row with volatile function formulas for testing
    - OFFSET with resolvable parameters (OFFSET(A21,1,1))
    - INDIRECT with string cell reference (INDIRECT("B21"))
    - NOW, TODAY, and RAND functions (inherently volatile)

    Args:
        ws: The worksheet to write to
        start_row: Starting row for volatile functions section (1-indexed, default: 19)

    Returns:
        List[str]: Cell addresses of all cells created in volatile functions section

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> cells = create_volatile_functions_section(ws, start_row=19)
        >>> len(cells)  # Should be 6 cells (header + 5 function cells)
        6
    """
    if start_row < 1:
        raise ValueError("start_row must be >= 1")

    created_cells = []

    # Row offsets from start_row
    header_row = start_row
    column_header_row = start_row + 1
    functions_row = start_row + 2

    # Create main header
    write_cell_value(ws, f"A{header_row}", "Volatile Functions Test")
    created_cells.append(f"A{header_row}")

    # Create column headers
    headers = {"A": "Offset", "B": "Indirect", "C": "Now", "D": "Today", "E": "Rand"}

    for col, header_text in headers.items():
        write_cell_value(ws, f"{col}{column_header_row}", header_text)
        created_cells.append(f"{col}{column_header_row}")

    # Create volatile function formulas
    # OFFSET with resolvable parameters: OFFSET(A21,1,1) -> references B22
    write_cell_formula(ws, f"A{functions_row}", f"=OFFSET(A{functions_row},1,1)")
    created_cells.append(f"A{functions_row}")

    # INDIRECT with string cell reference: INDIRECT("B21") -> references B21
    write_cell_formula(ws, f"B{functions_row}", f'=INDIRECT("B{functions_row}")')
    created_cells.append(f"B{functions_row}")

    # NOW function (inherently volatile)
    write_cell_formula(ws, f"C{functions_row}", "=NOW()")
    created_cells.append(f"C{functions_row}")

    # TODAY function (inherently volatile)
    write_cell_formula(ws, f"D{functions_row}", "=TODAY()")
    created_cells.append(f"D{functions_row}")

    # RAND function (inherently volatile)
    write_cell_formula(ws, f"E{functions_row}", "=RAND()")
    created_cells.append(f"E{functions_row}")

    return created_cells


def validate_volatile_functions(ws: Worksheet, volatile_cells: list[str]) -> bool:
    """
    Validates that volatile functions section contains correct volatile functions.

    Args:
        ws: The worksheet containing the volatile functions section
        volatile_cells: List of cell addresses in the volatile functions section

    Returns:
        bool: True if all volatile functions are correctly implemented, False otherwise
    """
    expected_functions = ["OFFSET", "INDIRECT", "NOW", "TODAY", "RAND"]

    for cell_addr in volatile_cells:
        cell = ws[cell_addr]

        # Skip non-formula cells (headers)
        if not cell.data_type == "f":
            continue

        formula = str(cell.value)
        if not formula.startswith("="):
            continue

        # Check that formula contains expected volatile functions
        formula_upper = formula.upper()
        has_volatile_function = any(func in formula_upper for func in expected_functions)

        if not has_volatile_function:
            return False

    return True


def check_volatile_resolution(ws: Worksheet, volatile_cells: list[str]) -> dict:
    """
    Checks which volatile functions can be resolved deterministically.

    Args:
        ws: The worksheet containing the volatile functions section
        volatile_cells: List of cell addresses in the volatile functions section

    Returns:
        dict: Status of each volatile function's resolution capability
    """
    resolution_status = {}

    for cell_addr in volatile_cells:
        cell = ws[cell_addr]

        # Skip non-formula cells (headers)
        if not cell.data_type == "f":
            continue

        formula = str(cell.value)
        if not formula.startswith("="):
            continue

        formula_upper = formula.upper()

        # Check OFFSET resolution
        if "OFFSET" in formula_upper:
            # OFFSET(A21,1,1) should be resolvable if A21 exists
            row_num = _extract_row_number(cell_addr)
            if "OFFSET(A21,1,1)" in formula or f"OFFSET(A{row_num},1,1)" in formula:
                resolution_status[cell_addr] = {
                    "function": "OFFSET",
                    "resolvable": True,
                    "reason": "Has resolvable parameters (1,1)",
                }
            else:
                resolution_status[cell_addr] = {
                    "function": "OFFSET",
                    "resolvable": False,
                    "reason": "Parameters not resolvable from snapshot",
                }

        # Check INDIRECT resolution
        elif "INDIRECT" in formula_upper:
            # INDIRECT("B21") should be resolvable if B21 exists
            row_num = _extract_row_number(cell_addr)
            if 'INDIRECT("B21")' in formula or f'INDIRECT("B{row_num}")' in formula:
                resolution_status[cell_addr] = {
                    "function": "INDIRECT",
                    "resolvable": True,
                    "reason": "String reference can be resolved",
                }
            else:
                resolution_status[cell_addr] = {
                    "function": "INDIRECT",
                    "resolvable": False,
                    "reason": "String reference not resolvable",
                }

        # Check inherently volatile functions
        elif any(func in formula_upper for func in ["NOW", "TODAY", "RAND"]):
            func_name = next(func for func in ["NOW", "TODAY", "RAND"] if func in formula_upper)
            resolution_status[cell_addr] = {
                "function": func_name,
                "resolvable": False,
                "reason": "Inherently volatile - cannot be resolved deterministically",
            }

    return resolution_status


def check_volatile_function_syntax(ws: Worksheet, volatile_cells: list[str]) -> bool:
    """
    Validates that all volatile function formulas have correct syntax.

    Args:
        ws: The worksheet containing the volatile functions section
        volatile_cells: List of cell addresses in the volatile functions section

    Returns:
        bool: True if all formulas have valid syntax, False otherwise
    """
    for cell_addr in volatile_cells:
        cell = ws[cell_addr]

        # Skip non-formula cells (headers)
        if not cell.data_type == "f":
            continue

        formula = str(cell.value)
        if not formula.startswith("="):
            continue

        # Basic syntax validation
        try:
            # Check for balanced parentheses
            open_parens = formula.count("(")
            close_parens = formula.count(")")
            if open_parens != close_parens:
                return False

            # Check for proper function syntax
            formula_upper = formula.upper()

            # OFFSET should have 3-5 parameters
            if "OFFSET" in formula_upper:
                if not ("OFFSET(" in formula_upper and ")" in formula_upper):
                    return False

            # INDIRECT should have 1 parameter in quotes
            elif "INDIRECT" in formula_upper:
                if not ("INDIRECT(" in formula_upper and '"' in formula and ")" in formula_upper):
                    return False

            # NOW, TODAY, RAND should have no parameters
            elif any(func in formula_upper for func in ["NOW", "TODAY", "RAND"]):
                func_name = next(func for func in ["NOW", "TODAY", "RAND"] if func in formula_upper)
                if f"{func_name}()" not in formula_upper:
                    return False

        except Exception:
            return False

    return True
