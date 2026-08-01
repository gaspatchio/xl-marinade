# ABOUTME: Command-line interface for generating test workbooks and validating IR extractor integration.
# ABOUTME: Provides CLI commands for creating test workbooks and running integration tests.

import argparse
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .formula_validator import create_consistency_report, validate_workbook_formulas
from .label_candidates_builder import create_label_candidates_section, create_named_ranges
from .overlap_region_builder import create_overlap_region
from .tree_a_builder import create_tree_a
from .tree_b_builder import create_tree_b
from .volatile_functions_builder import create_volatile_functions_section
from .workbook_builder import (
    create_workbook,
    save_workbook,
    validate_workbook_integrity,
)


def create_comprehensive_test_workbook(
    output_path: Path,
    include_tree_a: bool = True,
    include_tree_b: bool = True,
    include_overlap: bool = True,
    include_volatile: bool = True,
    include_labels: bool = True,
    verbose: bool = False,
) -> bool:
    """
    Creates a comprehensive test workbook with all components.

    Args:
        output_path: Path where to save the workbook
        include_tree_a: Whether to include Tree A (financial calculations)
        include_tree_b: Whether to include Tree B (statistical analysis)
        include_overlap: Whether to include overlap region
        include_volatile: Whether to include volatile functions section
        include_labels: Whether to include label candidates section
        verbose: Whether to print detailed progress

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        if verbose:
            print("Creating comprehensive test workbook...")

        # Create workbook
        wb = create_workbook()
        ws = wb.active

        if ws is None:
            print("Error: No active worksheet found")
            return False

        # Set worksheet title
        ws.title = "TestSheet"

        created_cells = []

        # Create Tree A (Financial Calculations)
        if include_tree_a:
            if verbose:
                print("  Adding Tree A (Financial Calculations)...")
            tree_a_cells = create_tree_a(ws, start_row=1)
            created_cells.extend(tree_a_cells)

        # Create Tree B (Statistical Analysis)
        if include_tree_b:
            if verbose:
                print("  Adding Tree B (Statistical Analysis)...")
            tree_b_cells = create_tree_b(ws, start_row=7)
            created_cells.extend(tree_b_cells)

        # Create Overlap Region
        if include_overlap and include_tree_a and include_tree_b:
            if verbose:
                print("  Adding Overlap Region...")
            overlap_cells = create_overlap_region(ws, start_row=13)
            created_cells.extend(overlap_cells)

        # Create Volatile Functions Section
        if include_volatile:
            if verbose:
                print("  Adding Volatile Functions Section...")
            volatile_cells = create_volatile_functions_section(ws, start_row=19)
            created_cells.extend(volatile_cells)

        # Create Label Candidates Section
        if include_labels:
            if verbose:
                print("  Adding Label Candidates Section...")
            label_cells = create_label_candidates_section(ws, start_row=23)
            created_cells.extend(label_cells)

            # Add named ranges
            if verbose:
                print("  Adding Named Ranges...")
            create_named_ranges(wb, ws)

        # Validate workbook integrity
        if verbose:
            print("  Validating workbook integrity...")

        is_valid = validate_workbook_integrity(wb, verbose=verbose)

        if not is_valid:
            print("Warning: Workbook validation failed, but continuing...")

        # Save workbook
        if verbose:
            print(f"  Saving workbook to {output_path}...")

        save_workbook(wb, output_path, overwrite=True)

        if verbose:
            print(f"Successfully created test workbook with {len(created_cells)} cells")
            print(f"Workbook saved to: {output_path}")

        return True

    except Exception as e:
        print(f"Error creating test workbook: {e}")
        if verbose:
            import traceback

            traceback.print_exc()
        return False


def validate_workbook_with_ir_extractor(
    workbook_path: Path, verbose: bool = False, return_data: bool = False
) -> tuple[bool, dict[str, Any] | None]:
    """
    Validates a workbook using mock IR extractor (since real one is in Sprint 2).

    This function simulates what the IR extractor would do:
    1. Load the workbook
    2. Extract formulas and dependencies
    3. Validate structure
    4. Generate mock SQLite output

    Args:
        workbook_path: Path to the workbook to validate
        verbose: Whether to print detailed progress

    Returns:
        bool: True if validation successful, False otherwise
    """
    try:
        if verbose:
            print(f"Validating workbook with mock IR extractor: {workbook_path}")

        # Load workbook
        wb = load_workbook(workbook_path, data_only=False)

        if verbose:
            print(f"  Loaded workbook with {len(wb.worksheets)} worksheets")

        # Validate each worksheet
        all_valid = True
        for ws in wb.worksheets:
            if verbose:
                print(f"  Validating worksheet: {ws.title}")

            # Validate formulas
            validation_results = validate_workbook_formulas(wb)

            for sheet_name, result in validation_results.items():
                if not result.is_valid:
                    print(f"    Warning: {sheet_name} has validation issues")
                    all_valid = False

                if verbose:
                    print(f"    Formulas checked: {result.total_formulas_checked}")
                    print(f"    Errors: {len(result.errors)}")
                    print(f"    Warnings: {len(result.warnings)}")

        # Generate consistency report
        if verbose:
            print("  Generating consistency report...")
            report = create_consistency_report(validation_results)
            print("  Consistency Report:")
            print("  " + "\n  ".join(report.split("\n")))

        # Mock SQLite database creation
        if verbose:
            print("  Simulating SQLite database creation...")

        # This would be replaced with actual IR extractor in Sprint 2
        mock_sqlite_data = {
            "bindings": [],
            "cells": [],
            "dependencies": [],
            "volatile_functions": [],
            "label_candidates": [],
            "consistency_report": [],
        }

        # Extract basic information (mock implementation)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_data = {
                            "address": cell.coordinate,
                            "value": str(cell.value),
                            "data_type": cell.data_type,
                            "is_formula": cell.data_type == "f",
                        }
                        mock_sqlite_data["cells"].append(cell_data)

        if verbose:
            print(f"  Extracted {len(mock_sqlite_data['cells'])} cells")
            print("  Mock SQLite database structure created")

        if return_data:
            return all_valid, mock_sqlite_data
        else:
            return all_valid, None

    except Exception as e:
        print(f"Error validating workbook with IR extractor: {e}")
        if verbose:
            import traceback

            traceback.print_exc()
        if return_data:
            return False, None
        else:
            return False, None


def run_end_to_end_test(workbook_path: Path, verbose: bool = False) -> bool:
    """
    Runs end-to-end test using Simple actuarial model.xlsx.

    Args:
        workbook_path: Path to the Simple actuarial model.xlsx file
        verbose: Whether to print detailed progress

    Returns:
        bool: True if test successful, False otherwise
    """
    try:
        if not workbook_path.exists():
            print(f"Error: Test workbook not found: {workbook_path}")
            return False

        if verbose:
            print(f"Running end-to-end test with: {workbook_path}")

        # Load the test workbook
        wb = load_workbook(workbook_path, data_only=False)

        if verbose:
            print(f"  Loaded workbook with {len(wb.worksheets)} worksheets")
            for ws in wb.worksheets:
                print(f"    Worksheet: {ws.title}")

        # Validate with mock IR extractor
        success, _ = validate_workbook_with_ir_extractor(workbook_path, verbose=verbose)

        if success:
            if verbose:
                print("  End-to-end test completed successfully")
            return True
        else:
            print("  End-to-end test failed")
            return False

    except Exception as e:
        print(f"Error running end-to-end test: {e}")
        if verbose:
            import traceback

            traceback.print_exc()
        return False


def run_round_trip_test(workbook_path: Path, verbose: bool = False) -> bool:
    """
    Runs round-trip test to ensure identical extracted data across runs.

    Args:
        workbook_path: Path to the workbook to test
        verbose: Whether to print detailed progress

    Returns:
        bool: True if round-trip test successful, False otherwise
    """
    try:
        if verbose:
            print("Running round-trip test...")

        # First pass - capture extracted data
        if verbose:
            print("  First pass...")
        result1, data1 = validate_workbook_with_ir_extractor(
            workbook_path, verbose=False, return_data=True
        )

        # Second pass - capture extracted data
        if verbose:
            print("  Second pass...")
        result2, data2 = validate_workbook_with_ir_extractor(
            workbook_path, verbose=False, return_data=True
        )

        # Compare boolean results
        if result1 != result2:
            print("  Round-trip test failed: boolean results differ")
            return False

        # Compare extracted data
        if data1 != data2:
            print("  Round-trip test failed: extracted data differs")
            if verbose:
                print("  Data comparison details:")
                print(f"    First run cells: {len(data1.get('cells', []))}")
                print(f"    Second run cells: {len(data2.get('cells', []))}")
                print(f"    First run bindings: {len(data1.get('bindings', []))}")
                print(f"    Second run bindings: {len(data2.get('bindings', []))}")
            return False

        if verbose:
            print("  Round-trip test passed: identical results and data")
        return True

    except Exception as e:
        print(f"Error running round-trip test: {e}")
        if verbose:
            import traceback

            traceback.print_exc()
        return False


def main() -> int:
    """
    Main CLI entry point.

    Returns:
        int: Exit code (0 for success, 1 for failure)
    """
    parser = argparse.ArgumentParser(
        description="Test Workbook Generator CLI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate a basic test workbook
  python -m test_workbook_generator.cli generate --output test.xlsx
  
  # Generate with specific components
  python -m test_workbook_generator.cli generate --output test.xlsx --no-volatile --no-labels
  
  # Validate an existing workbook
  python -m test_workbook_generator.cli validate --workbook test.xlsx
  
  # Run end-to-end test with Simple actuarial model.xlsx
  python -m test_workbook_generator.cli test-e2e --workbook "test_spreadsheets/Simple actuarial model.xlsx"
  
  # Run round-trip test
  python -m test_workbook_generator.cli test-roundtrip --workbook test.xlsx
        """,
    )

    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # Generate command
    generate_parser = subparsers.add_parser("generate", help="Generate a test workbook")
    generate_parser.add_argument(
        "--output", "-o", type=Path, required=True, help="Output path for the workbook (.xlsx)"
    )
    generate_parser.add_argument(
        "--no-tree-a", action="store_true", help="Skip Tree A (financial calculations)"
    )
    generate_parser.add_argument(
        "--no-tree-b", action="store_true", help="Skip Tree B (statistical analysis)"
    )
    generate_parser.add_argument("--no-overlap", action="store_true", help="Skip overlap region")
    generate_parser.add_argument(
        "--no-volatile", action="store_true", help="Skip volatile functions section"
    )
    generate_parser.add_argument(
        "--no-labels", action="store_true", help="Skip label candidates section"
    )
    generate_parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    # Validate command
    validate_parser = subparsers.add_parser(
        "validate", help="Validate a workbook with IR extractor"
    )
    validate_parser.add_argument(
        "--workbook", "-w", type=Path, required=True, help="Path to workbook to validate"
    )
    validate_parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    # Test e2e command
    test_e2e_parser = subparsers.add_parser("test-e2e", help="Run end-to-end test")
    test_e2e_parser.add_argument(
        "--workbook",
        "-w",
        type=Path,
        required=True,
        help="Path to test workbook (e.g., Simple actuarial model.xlsx)",
    )
    test_e2e_parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")

    # Test roundtrip command
    test_roundtrip_parser = subparsers.add_parser("test-roundtrip", help="Run round-trip test")
    test_roundtrip_parser.add_argument(
        "--workbook", "-w", type=Path, required=True, help="Path to workbook to test"
    )
    test_roundtrip_parser.add_argument(
        "--verbose", "-v", action="store_true", help="Verbose output"
    )

    # Parse arguments
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    try:
        if args.command == "generate":
            success = create_comprehensive_test_workbook(
                output_path=args.output,
                include_tree_a=not args.no_tree_a,
                include_tree_b=not args.no_tree_b,
                include_overlap=not args.no_overlap,
                include_volatile=not args.no_volatile,
                include_labels=not args.no_labels,
                verbose=args.verbose,
            )
            return 0 if success else 1

        elif args.command == "validate":
            success = validate_workbook_with_ir_extractor(
                workbook_path=args.workbook, verbose=args.verbose
            )
            return 0 if success else 1

        elif args.command == "test-e2e":
            success = run_end_to_end_test(workbook_path=args.workbook, verbose=args.verbose)
            return 0 if success else 1

        elif args.command == "test-roundtrip":
            success = run_round_trip_test(workbook_path=args.workbook, verbose=args.verbose)
            return 0 if success else 1

        else:
            print(f"Unknown command: {args.command}")
            return 1

    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
        return 1
    except Exception as e:
        print(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
