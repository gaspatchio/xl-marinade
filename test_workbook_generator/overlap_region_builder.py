# ABOUTME: Builder for overlap region with shared dependencies between Tree A and Tree B.
# ABOUTME: Creates cells that reference both trees, establishing complex dependency chains for testing.


from openpyxl.worksheet.worksheet import Worksheet

from .workbook_builder import write_cell_formula, write_cell_value


def create_overlap_region(ws: Worksheet, start_row: int = 13) -> list[str]:
    """
    Creates overlap region with shared dependencies between Tree A and Tree B.

    The overlap region contains cells that reference both Tree A (rows 3-5) and Tree B (rows 9-11),
    creating complex dependency chains that test binding detection and formula validation.

    Args:
        ws: The worksheet to write to
        start_row: Starting row for overlap region (1-indexed, default: 13)

    Returns:
        List[str]: Cell addresses of all cells created in overlap region

    Example:
        >>> wb = create_workbook()
        >>> ws = wb.active
        >>> # First create Tree A and Tree B
        >>> tree_a_cells = create_tree_a(ws, start_row=1)
        >>> tree_b_cells = create_tree_b(ws, start_row=7)
        >>> # Then create overlap region
        >>> overlap_cells = create_overlap_region(ws, start_row=13)
        >>> len(overlap_cells)  # Should be 5-8 cells
        8
    """
    if start_row < 1:
        raise ValueError("start_row must be >= 1")

    created_cells = []

    # Row offsets from start_row
    header_row = start_row
    column_header_row = start_row + 1
    shared_row = start_row + 2  # Main shared formulas row

    # Create main header
    write_cell_value(ws, f"A{header_row}", "Overlap Region (Shared by both trees)")
    created_cells.append(f"A{header_row}")

    # Create column headers for shared dependencies
    headers = {"A": "Shared1", "B": "Shared2", "C": "Shared3", "D": "Shared4", "E": "Shared5"}

    for col, header_text in headers.items():
        write_cell_value(ws, f"{col}{column_header_row}", header_text)
        created_cells.append(f"{col}{column_header_row}")

    # Create shared formulas that combine Tree A and Tree B cells
    # These formulas reference both trees to create dependency chains

    # Shared1: Tree A B3 (Rate Value) + Tree B B9 (VLOOKUP result for 100)
    write_cell_formula(ws, f"A{shared_row}", "=B3+B9")
    created_cells.append(f"A{shared_row}")

    # Shared2: Tree A C3 (Rate Calc1) + Tree B C9 (INDEX result for 100)
    write_cell_formula(ws, f"B{shared_row}", "=C3+C9")
    created_cells.append(f"B{shared_row}")

    # Shared3: Tree A D3 (Rate Calc2) + Tree B D9 (MATCH result for 100)
    write_cell_formula(ws, f"C{shared_row}", "=D3+D9")
    created_cells.append(f"C{shared_row}")

    # Shared4: Tree A E3 (Rate Time) + Tree B E9 (Array result for 100)
    write_cell_formula(ws, f"D{shared_row}", "=E3+E9")
    created_cells.append(f"D{shared_row}")

    # Shared5: Tree A F3 (Rate Growth) + Tree B F9 (Result for 100)
    write_cell_formula(ws, f"E{shared_row}", "=F3+F9")
    created_cells.append(f"E{shared_row}")

    return created_cells


def validate_dependency_chain(ws: Worksheet, overlap_cells: list[str]) -> bool:
    """
    Validates that overlap region formulas reference both Tree A and Tree B cells.

    Args:
        ws: The worksheet containing the overlap region
        overlap_cells: List of cell addresses in the overlap region

    Returns:
        bool: True if all formulas reference both trees, False otherwise
    """
    tree_a_rows = {3, 4, 5}  # Tree A data rows
    tree_b_rows = {9, 10, 11}  # Tree B data rows

    for cell_addr in overlap_cells:
        cell = ws[cell_addr]

        # Skip non-formula cells
        if not cell.data_type == "f":
            continue

        formula = str(cell.value)
        if not formula.startswith("="):
            continue

        # Check if formula references Tree A cells (rows 3-5)
        has_tree_a_ref = any(
            f"B{row}" in formula
            or f"C{row}" in formula
            or f"D{row}" in formula
            or f"E{row}" in formula
            or f"F{row}" in formula
            or f"G{row}" in formula
            for row in tree_a_rows
        )

        # Check if formula references Tree B cells (rows 9-11)
        has_tree_b_ref = any(
            f"A{row}" in formula
            or f"B{row}" in formula
            or f"C{row}" in formula
            or f"D{row}" in formula
            or f"E{row}" in formula
            or f"F{row}" in formula
            for row in tree_b_rows
        )

        # Both trees must be referenced
        if not (has_tree_a_ref and has_tree_b_ref):
            return False

        # Additionally, verify that the referenced cells actually exist and have values
        # This catches cases where Tree A or Tree B wasn't created
        for row in tree_a_rows:
            for col in ["B", "C", "D", "E", "F", "G"]:
                if f"{col}{row}" in formula:
                    if ws[f"{col}{row}"].value is None:
                        return False

        for row in tree_b_rows:
            for col in ["A", "B", "C", "D", "E", "F"]:
                if f"{col}{row}" in formula:
                    if ws[f"{col}{row}"].value is None:
                        return False

    return True


def check_circular_references(ws: Worksheet, overlap_cells: list[str]) -> bool:
    """
    Checks that no overlap region formulas create circular references.

    Args:
        ws: The worksheet containing the overlap region
        overlap_cells: List of cell addresses in the overlap region

    Returns:
        bool: True if no circular references detected, False if circular refs found
    """
    for cell_addr in overlap_cells:
        cell = ws[cell_addr]

        # Skip non-formula cells
        if not cell.data_type == "f":
            continue

        formula = str(cell.value)
        if not formula.startswith("="):
            continue

        # Check if formula references its own cell
        if cell_addr in formula:
            return False

        # Check for obvious circular patterns (e.g., A15 references A15)
        cell_col = "".join(filter(str.isalpha, cell_addr))
        cell_row = "".join(filter(str.isdigit, cell_addr))

        # Look for references to the same cell address
        if f"{cell_col}{cell_row}" in formula:
            return False

    return True
