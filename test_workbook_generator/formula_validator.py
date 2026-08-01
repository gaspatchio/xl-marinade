# ABOUTME: Comprehensive formula validation and error handling for Excel workbooks.
# ABOUTME: Provides syntax validation, cell reference checking, circular reference detection, and consistency reporting.

import re
from dataclasses import dataclass
from enum import Enum

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ValidationErrorType(Enum):
    """Types of validation errors that can occur."""

    SYNTAX_ERROR = "syntax_error"
    INVALID_REFERENCE = "invalid_reference"
    CIRCULAR_REFERENCE = "circular_reference"
    MISSING_CELL = "missing_cell"
    INVALID_FUNCTION = "invalid_function"
    UNBALANCED_PARENTHESES = "unbalanced_parentheses"
    INVALID_RANGE = "invalid_range"


@dataclass
class ValidationError:
    """Represents a validation error with details."""

    error_type: ValidationErrorType
    cell_address: str
    formula: str
    message: str
    severity: str = "error"  # error, warning, info


@dataclass
class ValidationResult:
    """Result of formula validation containing errors and warnings."""

    is_valid: bool
    errors: list[ValidationError]
    warnings: list[ValidationError]
    circular_refs: list[str]
    missing_refs: list[str]
    total_formulas_checked: int


class FormulaValidator:
    """Comprehensive formula validator for Excel workbooks."""

    # Excel function names (subset of most common functions)
    EXCEL_FUNCTIONS = {
        "SUM",
        "AVERAGE",
        "COUNT",
        "COUNTA",
        "MAX",
        "MIN",
        "IF",
        "VLOOKUP",
        "HLOOKUP",
        "INDEX",
        "MATCH",
        "OFFSET",
        "INDIRECT",
        "NOW",
        "TODAY",
        "RAND",
        "RANDBETWEEN",
        "ROUND",
        "ROUNDUP",
        "ROUNDDOWN",
        "ABS",
        "SQRT",
        "POWER",
        "EXP",
        "LN",
        "LOG",
        "SIN",
        "COS",
        "TAN",
        "ASIN",
        "ACOS",
        "ATAN",
        "PI",
        "E",
        "CONCATENATE",
        "LEFT",
        "RIGHT",
        "MID",
        "LEN",
        "FIND",
        "SEARCH",
        "SUBSTITUTE",
        "REPLACE",
        "TRIM",
        "UPPER",
        "LOWER",
        "PROPER",
        "VALUE",
        "TEXT",
        "DATE",
        "TIME",
        "YEAR",
        "MONTH",
        "DAY",
        "HOUR",
        "MINUTE",
        "SECOND",
        "WEEKDAY",
        "WEEKNUM",
        "ISNUMBER",
        "ISTEXT",
        "ISBLANK",
        "ISERROR",
        "ISNA",
        "IFERROR",
        "IFNA",
        "AND",
        "OR",
        "NOT",
        "TRUE",
        "FALSE",
        "NA",
        "ERROR",
        "CHOOSE",
        "SWITCH",
        "XLOOKUP",
        "XMATCH",
        "UNIQUE",
        "SORT",
        "FILTER",
        "SORTBY",
        "UNSORT",
        "RANDARRAY",
        "SEQUENCE",
        "LET",
    }

    # Volatile functions that change on recalculation
    VOLATILE_FUNCTIONS = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "OFFSET", "INDIRECT", "RANDARRAY"}

    def __init__(self, worksheet: Worksheet):
        """Initialize validator with a worksheet."""
        self.ws = worksheet
        self.validation_errors: list[ValidationError] = []
        self.validation_warnings: list[ValidationError] = []
        self.circular_refs: set[str] = set()
        self.missing_refs: set[str] = set()
        self.visited_cells: set[str] = set()
        self.current_path: list[str] = []

    def validate_all_formulas(self) -> ValidationResult:
        """
        Validate all formulas in the worksheet.

        Returns:
            ValidationResult: Complete validation results
        """
        self.validation_errors.clear()
        self.validation_warnings.clear()
        self.circular_refs.clear()
        self.missing_refs.clear()

        formula_count = 0

        # Get all cells with formulas
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula_count += 1
                    self._validate_single_formula(cell.coordinate, str(cell.value))

        # Check for circular references
        self._detect_circular_references()

        # Check for missing cell references
        self._check_missing_references()

        is_valid = len(self.validation_errors) == 0

        return ValidationResult(
            is_valid=is_valid,
            errors=self.validation_errors.copy(),
            warnings=self.validation_warnings.copy(),
            circular_refs=list(self.circular_refs),
            missing_refs=list(self.missing_refs),
            total_formulas_checked=formula_count,
        )

    def _validate_single_formula(self, cell_address: str, formula: str) -> None:
        """Validate a single formula."""
        if not formula.startswith("="):
            self.validation_errors.append(
                ValidationError(
                    error_type=ValidationErrorType.SYNTAX_ERROR,
                    cell_address=cell_address,
                    formula=formula,
                    message="Formula must start with '='",
                )
            )
            return

        # Basic syntax validation
        self._validate_syntax(cell_address, formula)

        # Extract and validate cell references
        references = self._extract_cell_references(formula)
        for ref in references:
            self._validate_cell_reference(cell_address, formula, ref)

    def _validate_syntax(self, cell_address: str, formula: str) -> None:
        """Validate basic formula syntax."""
        # Check for balanced parentheses
        open_parens = formula.count("(")
        close_parens = formula.count(")")
        if open_parens != close_parens:
            self.validation_errors.append(
                ValidationError(
                    error_type=ValidationErrorType.UNBALANCED_PARENTHESES,
                    cell_address=cell_address,
                    formula=formula,
                    message=f"Unbalanced parentheses: {open_parens} open, {close_parens} close",
                )
            )

        # Check for consecutive operators (but allow some valid cases)
        if re.search(r"[+\-*/]{2,}", formula):
            # Three or more consecutive operators are always invalid
            if re.search(r"[+\-*/]{3,}", formula):
                self.validation_errors.append(
                    ValidationError(
                        error_type=ValidationErrorType.SYNTAX_ERROR,
                        cell_address=cell_address,
                        formula=formula,
                        message="Too many consecutive operators found",
                    )
                )
            # Check for invalid double operators (but allow double minus)
            elif re.search(r"[+\*/]{2,}", formula):
                # Double plus, multiply, or divide are invalid
                self.validation_errors.append(
                    ValidationError(
                        error_type=ValidationErrorType.SYNTAX_ERROR,
                        cell_address=cell_address,
                        formula=formula,
                        message="Invalid consecutive operators found",
                    )
                )
            # Note: Double minus (--) is valid in Excel for subtraction from negative numbers
            # Examples: =A1--B1 (subtract negative B1 from A1), =A1--5 (subtract negative 5 from A1)

        # Check for invalid characters (be more permissive)
        # Allow common Excel operators and characters
        # Note: The character class needs to be properly escaped
        if re.search(r'[^A-Za-z0-9+\-*/=().,:\s"\'&<>!=]', formula):
            self.validation_errors.append(
                ValidationError(
                    error_type=ValidationErrorType.SYNTAX_ERROR,
                    cell_address=cell_address,
                    formula=formula,
                    message="Invalid characters in formula",
                )
            )

        # Check for incomplete ranges (be more specific)
        if ":" in formula:
            # Check if there are any incomplete ranges
            if re.search(r"[A-Z]+\d+:\s*[^A-Z\d)]", formula):
                # Only flag if it's clearly incomplete (missing closing cell reference)
                if not re.search(r"[A-Z]+\d+:[A-Z]+\d+", formula):
                    self.validation_errors.append(
                        ValidationError(
                            error_type=ValidationErrorType.INVALID_RANGE,
                            cell_address=cell_address,
                            formula=formula,
                            message="Incomplete range reference",
                        )
                    )

    def _extract_cell_references(self, formula: str) -> list[str]:
        """Extract all cell references from a formula."""
        # Pattern to match cell references like A1, B2, AA10, etc.
        cell_ref_pattern = r"\b[A-Z]+\d+\b"
        references = re.findall(cell_ref_pattern, formula)

        # Also extract range references like A1:B10
        range_pattern = r"\b[A-Z]+\d+:[A-Z]+\d+\b"
        range_refs = re.findall(range_pattern, formula)

        # Split range references into individual cells
        all_refs = references.copy()
        for range_ref in range_refs:
            start, end = range_ref.split(":")
            all_refs.extend([start, end])

        return list(set(all_refs))  # Remove duplicates

    def _validate_cell_reference(self, cell_address: str, formula: str, reference: str) -> None:
        """Validate that a cell reference exists and is valid."""
        try:
            # Check if the referenced cell exists in the worksheet
            # First check if the cell is in the worksheet's used range
            cell_exists = False
            cell_value = None

            # Check if the referenced cell exists and has been explicitly set
            # We need to check if the cell has been written to, not just if it's in the used range
            try:
                cell_value = self.ws[reference].value
                # Check if cell has been explicitly set (not just default None)
                # In openpyxl, cells that haven't been written to return None
                # but we need to distinguish between "not written to" and "written to with None value"

                # Check if the cell is in the worksheet's used range
                if hasattr(self.ws, "max_row") and hasattr(self.ws, "max_column"):
                    import re

                    match = re.match(r"([A-Z]+)(\d+)", reference)
                    if match:
                        col_letter = match.group(1)
                        row_num = int(match.group(2))

                        # Convert column letter to number
                        col_num = 0
                        for char in col_letter:
                            col_num = col_num * 26 + (ord(char) - ord("A") + 1)

                        # Check if cell is within used range
                        if row_num <= self.ws.max_row and col_num <= self.ws.max_column:
                            # Cell is in used range, check if it has a value
                            if cell_value is not None:
                                cell_exists = True
                            else:
                                # Cell is in used range but has no value - this means it was never written to
                                cell_exists = False
                        else:
                            cell_exists = False
                    else:
                        cell_exists = False
                else:
                    # Fallback: if cell has a value, it exists
                    cell_exists = cell_value is not None
            except (KeyError, ValueError):
                cell_exists = False

            if not cell_exists:
                # Cell doesn't exist
                self.missing_refs.add(reference)
                self.validation_errors.append(
                    ValidationError(
                        error_type=ValidationErrorType.MISSING_CELL,
                        cell_address=cell_address,
                        formula=formula,
                        message=f"Referenced cell {reference} does not exist",
                    )
                )
            elif cell_value is None:
                # Cell exists but is empty
                self.validation_warnings.append(
                    ValidationError(
                        error_type=ValidationErrorType.INVALID_REFERENCE,
                        cell_address=cell_address,
                        formula=formula,
                        message=f"Referenced cell {reference} is empty",
                        severity="warning",
                    )
                )

        except Exception as e:
            self.validation_errors.append(
                ValidationError(
                    error_type=ValidationErrorType.INVALID_REFERENCE,
                    cell_address=cell_address,
                    formula=formula,
                    message=f"Invalid cell reference {reference}: {str(e)}",
                )
            )

    def _detect_circular_references(self) -> None:
        """Detect circular references in the worksheet."""
        self.visited_cells.clear()
        self.current_path.clear()

        # Get all formula cells
        formula_cells = []
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and cell.value:
                    formula_cells.append(cell.coordinate)

        # Check each formula cell for circular references
        for cell_address in formula_cells:
            if cell_address not in self.visited_cells:
                self._dfs_circular_check(cell_address, [])

    def _dfs_circular_check(self, cell_address: str, path: list[str]) -> None:
        """Depth-first search to detect circular references."""
        if cell_address in path:
            # Found a circular reference
            self.circular_refs.add(cell_address)
            cycle_start = path.index(cell_address)
            cycle = path[cycle_start:] + [cell_address]

            self.validation_errors.append(
                ValidationError(
                    error_type=ValidationErrorType.CIRCULAR_REFERENCE,
                    cell_address=cell_address,
                    formula=str(self.ws[cell_address].value),
                    message=f"Circular reference detected: {' -> '.join(cycle)}",
                )
            )
            return

        if cell_address in self.visited_cells:
            return

        self.visited_cells.add(cell_address)
        path.append(cell_address)

        # Get formula and extract references
        cell = self.ws[cell_address]
        if cell.data_type == "f" and cell.value:
            formula = str(cell.value)
            references = self._extract_cell_references(formula)

            # Recursively check referenced cells
            for ref in references:
                try:
                    cell = self.ws[ref]
                    if cell.data_type == "f" and cell.value:
                        self._dfs_circular_check(ref, path.copy())
                except (KeyError, ValueError):
                    # Cell doesn't exist, skip
                    pass

        path.pop()

    def _check_missing_references(self) -> None:
        """Check for references to cells that don't exist."""
        # This is already handled in _validate_cell_reference
        # but we can add additional checks here if needed
        pass

    def validate_formula_syntax(self, formula: str) -> bool:
        """
        Validate formula syntax without checking cell references.

        Args:
            formula: Formula string to validate

        Returns:
            bool: True if syntax is valid, False otherwise
        """
        if not formula.startswith("="):
            return False

        # Check balanced parentheses
        open_parens = formula.count("(")
        close_parens = formula.count(")")
        if open_parens != close_parens:
            return False

        # Check for consecutive operators (but allow double minus)
        if re.search(r"[+\-*/]{2,}", formula):
            # Three or more consecutive operators are always invalid
            if re.search(r"[+\-*/]{3,}", formula) or re.search(r"[+\*/]{2,}", formula):
                return False
            # Note: Double minus (--) is valid in Excel for subtraction from negative numbers

        # Check for invalid characters
        if re.search(r'[^A-Za-z0-9+\-*/=().,:\s"\'&<>!=]', formula):
            return False

        # Check for missing operators between cell references
        # Pattern: cell reference followed by space and another cell reference
        if re.search(r"[A-Z]+\d+\s+[A-Z]+\d+", formula):
            return False

        return True

    def get_volatile_functions(self, formula: str) -> list[str]:
        """
        Extract volatile functions from a formula.

        Args:
            formula: Formula string to analyze

        Returns:
            List[str]: List of volatile functions found
        """
        if not formula.startswith("="):
            return []

        formula_upper = formula.upper()
        volatile_found = []

        for func in self.VOLATILE_FUNCTIONS:
            if func in formula_upper:
                volatile_found.append(func)

        return volatile_found

    def check_function_syntax(self, formula: str) -> list[str]:
        """
        Check for invalid function names in a formula.

        Args:
            formula: Formula string to check

        Returns:
            List[str]: List of invalid function names found
        """
        if not formula.startswith("="):
            return []

        # Extract function names (words followed by parentheses)
        function_pattern = r"\b([A-Z][A-Z0-9]*)\s*\("
        functions = re.findall(function_pattern, formula.upper())

        invalid_functions = []
        for func in functions:
            if func not in self.EXCEL_FUNCTIONS:
                invalid_functions.append(func)

        return invalid_functions


def validate_workbook_formulas(workbook: Workbook) -> dict[str, ValidationResult]:
    """
    Validate all formulas in a workbook across all worksheets.

    Args:
        workbook: The workbook to validate

    Returns:
        Dict[str, ValidationResult]: Results for each worksheet
    """
    results = {}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        validator = FormulaValidator(ws)
        results[sheet_name] = validator.validate_all_formulas()

    return results


def create_consistency_report(validation_results: dict[str, ValidationResult]) -> str:
    """
    Create a human-readable consistency report from validation results.

    Args:
        validation_results: Results from validate_workbook_formulas

    Returns:
        str: Formatted consistency report
    """
    report_lines = []
    report_lines.append("=== FORMULA VALIDATION CONSISTENCY REPORT ===")
    report_lines.append("")

    total_errors = 0
    total_warnings = 0
    total_circular_refs = 0
    total_missing_refs = 0
    total_formulas = 0

    for sheet_name, result in validation_results.items():
        report_lines.append(f"Sheet: {sheet_name}")
        report_lines.append(f"  Formulas checked: {result.total_formulas_checked}")
        report_lines.append(f"  Errors: {len(result.errors)}")
        report_lines.append(f"  Warnings: {len(result.warnings)}")
        report_lines.append(f"  Circular references: {len(result.circular_refs)}")
        report_lines.append(f"  Missing references: {len(result.missing_refs)}")
        report_lines.append(f"  Valid: {'Yes' if result.is_valid else 'No'}")
        report_lines.append("")

        total_errors += len(result.errors)
        total_warnings += len(result.warnings)
        total_circular_refs += len(result.circular_refs)
        total_missing_refs += len(result.missing_refs)
        total_formulas += result.total_formulas_checked

        # Add detailed error information
        if result.errors:
            report_lines.append("  ERRORS:")
            for error in result.errors:
                report_lines.append(f"    {error.cell_address}: {error.message}")
            report_lines.append("")

        if result.warnings:
            report_lines.append("  WARNINGS:")
            for warning in result.warnings:
                report_lines.append(f"    {warning.cell_address}: {warning.message}")
            report_lines.append("")

        if result.circular_refs:
            report_lines.append("  CIRCULAR REFERENCES:")
            for ref in result.circular_refs:
                report_lines.append(f"    {ref}")
            report_lines.append("")

        if result.missing_refs:
            report_lines.append("  MISSING REFERENCES:")
            for ref in result.missing_refs:
                report_lines.append(f"    {ref}")
            report_lines.append("")

    # Summary
    report_lines.append("=== SUMMARY ===")
    report_lines.append(f"Total formulas checked: {total_formulas}")
    report_lines.append(f"Total errors: {total_errors}")
    report_lines.append(f"Total warnings: {total_warnings}")
    report_lines.append(f"Total circular references: {total_circular_refs}")
    report_lines.append(f"Total missing references: {total_missing_refs}")
    report_lines.append(f"Overall valid: {'Yes' if total_errors == 0 else 'No'}")

    return "\n".join(report_lines)
