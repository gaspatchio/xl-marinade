"""Build the BasicTerm_S workbook — an Excel twin of lifelib's BasicTerm_S model.

Every formula cell mirrors one lifelib cell (see column/name mapping in
verification/verify_basicterm_s.py), except the pv_net_cf_t presentation
column (net_cf x disc_factor per month), which has no single lifelib
counterpart and is verified against the product of the lifelib vectors.
Conventions:
  blue font  = hardcoded input
  black font = in-sheet formula
  green font = formula pulling from another sheet

Output: BasicTerm_S.xlsm when vbaProject.bin is present next to this script
(zip surgery injects the macro project); BasicTerm_S.xlsx otherwise.
"""

import shutil
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parents[2]
MODEL = REPO / "models" / "basiclife" / "BasicTerm_S"
REF_ALL = REPO / "verification" / "reference" / "basicterm_all_points.csv"
HERE = Path(__file__).resolve().parent
VBA_BIN = HERE / "vbaProject.bin"
OUT_XLSX = HERE / "BasicTerm_S.xlsx"
OUT_XLSM = HERE / "BasicTerm_S.xlsm"

MAX_T = 240  # engine rows t = 0..240 cover the longest policy term (20y)

# ---------------------------------------------------------------- styles
NAVY = "1F3864"
LIGHT = "D9E2F3"
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")
CALC_FONT = Font(name="Calibri", size=10, color="000000")
LINK_FONT = Font(name="Calibri", size=10, color="006100")  # cross-sheet
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color=NAVY)
SUB_FILL = PatternFill("solid", start_color=LIGHT)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")

FMT_COUNT = "0.000000"
FMT_RATE = "0.0000%"
FMT_MONEY = "#,##0.00"
FMT_INT = "0"


def hdr(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def label(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=bold)
    return c


def name_cell(wb, name, sheet, addr):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{addr}"))


def item_row(ws, r, item, value, note, fmt=None, formula=False, link=False):
    """One Item / Value / Notes row in the standard 3-column block layout."""
    label(ws, r, 2, item)
    c = ws.cell(row=r, column=3, value=value)
    c.font = LINK_FONT if link else (CALC_FONT if formula else INPUT_FONT)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = Font(size=9, italic=True, color="7F7F7F")
    return c


# ---------------------------------------------------------------- build
def build():
    mp = pd.read_excel(MODEL / "model_point_table.xlsx", index_col=0)
    mort = pd.read_excel(MODEL / "mort_table.xlsx", index_col=0)
    disc = pd.read_excel(MODEL / "disc_rate_ann.xlsx", index_col=0)
    ref_all = pd.read_csv(REF_ALL, index_col=0)

    wb = Workbook()

    # ---------------------------------------------------------- Cover
    cv = wb.active
    cv.title = "Cover"
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["B"].width = 28
    cv.column_dimensions["C"].width = 78
    t = cv.cell(row=2, column=2, value="BasicTerm_S — Monthly Term Assurance Cash Flow Projection")
    t.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    cv.cell(row=3, column=2, value="Single model point projection of a level-premium term assurance portfolio"
            ).font = Font(size=11, italic=True)

    rows = [
        ("Purpose", "Projects monthly decrements and cash flows for the model point selected on the "
                    "Control sheet, and derives the net and office premium on an equivalence basis. "
                    "The workbook is a formula-for-formula replication of the open-source lifelib "
                    "model 'BasicTerm_S' (MIT licence) and is verified cell-by-cell against it."),
        ("Product", "Level-premium term assurance; monthly premiums payable throughout the policy term; "
                    "sum assured payable immediately on death; no surrender value."),
        ("Basis", "Mortality per select-and-ultimate style table (6 duration bands); lapses by policy year; "
                  "expenses per policy with 1% p.a. inflation; discounting at the annual spot rates on the "
                  "Discount sheet, converted to monthly effective."),
        ("Projection grain", "Monthly time steps t = 0 … 240 (policy anniversaries at t = 12k). Cash flows "
                             "beyond the maturity month are structurally zero."),
        ("Batch run", "The RunAllModelPoints macro (Batch_Results sheet) projects every model point in "
                      "the file, and reconciles premium and PV results against the lifelib values stored "
                      "on the Lifelib_Reference sheet."),
        ("Source model", "lifelib basiclife.BasicTerm_S v0.13.0 — https://lifelib.io (MIT licence)"),
        ("Verification", "verification/verify_basicterm_s.py compares every mapped cell against the Python "
                         "model for 10 model points spanning ages, sexes and terms; the in-workbook batch "
                         "reconciliation covers all 10,000 points. See Checks sheet for consistency tests."),
    ]
    r = 5
    for k, v in rows:
        label(cv, r, 2, k, bold=True)
        c = cv.cell(row=r, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cv.row_dimensions[r].height = 42
        r += 1

    r += 1
    label(cv, r, 2, "Version control", bold=True)
    for i, (a, b) in enumerate([("Version", "0.3.0"), ("Date", "2026-07-14"),
                                ("Prepared by", "XL Marinade project"),
                                ("Review status", "Verified against lifelib; adversarial actuarial audit remediated (see verification/reports)")]):
        label(cv, r + 1 + i, 2, a)
        cv.cell(row=r + 1 + i, column=3, value=b).font = Font(size=10)

    r += 6
    label(cv, r, 2, "Sheet index", bold=True)
    for i, (s, d) in enumerate([
            ("Control", "Run control — model point selection and batch run settings"),
            ("Assumptions", "Basis assumptions (expenses, inflation, loading, lapses)"),
            ("Model_Points", "Model point file — 10,000 policies (input data)"),
            ("Mortality", "Annual mortality rates by attained age and duration band (input data)"),
            ("Discount", "Annual spot discount rates by year (input data)"),
            ("Projection", "Monthly projection engine — decrements, cash flows, discounting"),
            ("Summary", "Present values, premium basis and results for the selected model point"),
            ("Checks", "Internal consistency checks (all must be TRUE)"),
            ("Lifelib_Reference", "lifelib results for all model points (reconciliation target)"),
            ("Batch_Results", "All-points batch run output and reconciliation vs lifelib (macro)")]):
        label(cv, r + 1 + i, 2, s)
        cv.cell(row=r + 1 + i, column=3, value=d).font = Font(size=10)

    r += 12
    label(cv, r, 2, "Conventions", bold=True)
    for i, (txt, font) in enumerate([
            ("Blue — hardcoded input", INPUT_FONT),
            ("Black — calculated on this sheet", CALC_FONT),
            ("Green — link from another sheet", LINK_FONT)]):
        c = cv.cell(row=r + 1 + i, column=2, value=txt)
        c.font = font

    # ---------------------------------------------------------- Control
    ct = wb.create_sheet("Control")
    ct.sheet_view.showGridLines = False
    ct.column_dimensions["B"].width = 34
    ct.column_dimensions["C"].width = 14
    ct.column_dimensions["D"].width = 62

    label(ct, 2, 2, "Model point selection", bold=True)
    hdr(ct, 3, 2, "Item"); hdr(ct, 3, 3, "Value"); hdr(ct, 3, 4, "Notes / lifelib cell")
    item_row(ct, 4, "Selected model point ID", 1, "point_id — choose 1 to 10,000", FMT_INT)
    name_cell(wb, "PointID", "Control", "$C$4")
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="10000",
                        allow_blank=False)
    dv.error = "Model point ID must be between 1 and 10,000"
    ct.add_data_validation(dv)
    dv.add(ct["C4"])

    label(ct, 6, 2, "Selected model point (from Model_Points)", bold=True)
    hdr(ct, 7, 2, "Item"); hdr(ct, 7, 3, "Value"); hdr(ct, 7, 4, "Notes / lifelib cell")
    item_row(ct, 8, "Age at entry", "=INDEX(MP_Age,MATCH(PointID,MP_ID,0))", "age_at_entry()", FMT_INT, link=True)
    item_row(ct, 9, "Sex", "=INDEX(MP_Sex,MATCH(PointID,MP_ID,0))", "sex()", None, link=True)
    item_row(ct, 10, "Policy term (years)", "=INDEX(MP_Term,MATCH(PointID,MP_ID,0))", "policy_term()", FMT_INT, link=True)
    item_row(ct, 11, "Sum assured", "=INDEX(MP_SumAssured,MATCH(PointID,MP_ID,0))", "sum_assured() = claim_pp(t)", FMT_MONEY, link=True)
    item_row(ct, 12, "Policies in force at t = 0", 1, "pols_if_init()", FMT_COUNT)
    item_row(ct, 13, "Projection length (months)", "=12*PolicyTerm+1", "proj_len() = 12 × policy_term + 1", FMT_INT, formula=True)
    for nm, addr in [("AgeAtEntry", "$C$8"), ("PolicyTerm", "$C$10"), ("SumAssured", "$C$11"),
                     ("PolsIfInit", "$C$12"), ("ProjLen", "$C$13")]:
        name_cell(wb, nm, "Control", addr)

    label(ct, 15, 2, "Batch run (RunAllModelPoints macro)", bold=True)
    hdr(ct, 16, 2, "Item"); hdr(ct, 16, 3, "Value"); hdr(ct, 16, 4, "Notes")
    item_row(ct, 17, "Model points to run (blank = all)", None,
             "Cap for quick test runs; leave blank to run all 10,000 points", FMT_INT)
    item_row(ct, 18, "Reconciliation tolerance (relative)", 0.000001,
             "A point reconciles when |workbook − lifelib| ≤ tol × max(1, |lifelib|)", "0.0000000")
    name_cell(wb, "BatchLimit", "Control", "$C$17")
    name_cell(wb, "BatchTol", "Control", "$C$18")
    ct.cell(row=20, column=2, value="Run the batch from Developer ▸ Macros ▸ RunAllModelPoints; results and "
            "the reconciliation summary appear on Batch_Results.").font = Font(size=9, italic=True, color="7F7F7F")

    # ---------------------------------------------------------- Assumptions
    asm = wb.create_sheet("Assumptions")
    asm.sheet_view.showGridLines = False
    asm.column_dimensions["B"].width = 34
    asm.column_dimensions["C"].width = 14
    asm.column_dimensions["D"].width = 60

    label(asm, 2, 2, "Basis assumptions", bold=True)
    hdr(asm, 3, 2, "Item"); hdr(asm, 3, 3, "Value"); hdr(asm, 3, 4, "Notes / lifelib cell")
    item_row(asm, 4, "Acquisition expense per policy", 300, "expense_acq() — incurred at t = 0", FMT_MONEY)
    item_row(asm, 5, "Maintenance expense p.a. per policy", 60, "expense_maint() — incurred monthly as 1/12", FMT_MONEY)
    item_row(asm, 6, "Expense inflation p.a.", 0.01, "inflation_rate() — compound, applied as (1+i)^(t/12)", FMT_RATE)
    item_row(asm, 7, "Premium loading", 0.50, "loading_prem() — office premium = (1+loading) × net premium", FMT_RATE)
    label(asm, 9, 2, "Lapse basis (annual rate by policy year)", bold=True)
    hdr(asm, 10, 2, "Item"); hdr(asm, 10, 3, "Value"); hdr(asm, 10, 4, "Notes / lifelib cell")
    item_row(asm, 11, "Lapse rate, year 1", 0.10, "lapse_rate(t) = MAX(base − step × duration, floor)", FMT_RATE)
    item_row(asm, 12, "Reduction per policy year", 0.02, "", FMT_RATE)
    item_row(asm, 13, "Minimum lapse rate", 0.02, "", FMT_RATE)
    for nm, addr in [("ExpenseAcq", "$C$4"), ("ExpenseMaint", "$C$5"), ("InflationRate", "$C$6"),
                     ("LoadingPrem", "$C$7"), ("LapseBase", "$C$11"), ("LapseStep", "$C$12"),
                     ("LapseFloor", "$C$13")]:
        name_cell(wb, nm, "Assumptions", addr)

    # ---------------------------------------------------------- Model_Points
    mps = wb.create_sheet("Model_Points")
    for j, (h, w) in enumerate([("point_id", 10), ("age_at_entry", 13), ("sex", 7),
                                ("policy_term", 12), ("policy_count", 13), ("sum_assured", 13)], start=1):
        hdr(mps, 1, j, h, w)
    for i, (pid, row) in enumerate(mp.iterrows(), start=2):
        vals = [pid, row["age_at_entry"], row["sex"], row["policy_term"],
                row["policy_count"], row["sum_assured"]]
        for j, v in enumerate(vals, start=1):
            c = mps.cell(row=i, column=j, value=v)
            c.font = INPUT_FONT
    mps.freeze_panes = "A2"
    n = len(mp)
    for nm, col in [("MP_ID", "A"), ("MP_Age", "B"), ("MP_Sex", "C"),
                    ("MP_Term", "D"), ("MP_Count", "E"), ("MP_SumAssured", "F")]:
        wb.defined_names.add(DefinedName(nm, attr_text=f"Model_Points!${col}$2:${col}${n + 1}"))

    # ---------------------------------------------------------- Mortality
    mt = wb.create_sheet("Mortality")
    hdr(mt, 1, 1, "Attained age", 12)
    for j, cname in enumerate(mort.columns, start=2):
        # last band applies to durations 5 and over (lifelib clamps MIN(d,5))
        title = f"Duration {cname}+" if j == len(mort.columns) + 1 else f"Duration {cname}"
        hdr(mt, 1, j, title, 11)
    for i, (age_, row) in enumerate(mort.iterrows(), start=2):
        c = mt.cell(row=i, column=1, value=age_)
        c.font = INPUT_FONT
        for j, v in enumerate(row, start=2):
            c = mt.cell(row=i, column=j, value=float(v))
            c.font = INPUT_FONT
            c.number_format = "0.000000"
    mt.freeze_panes = "B2"
    na = len(mort)
    wb.defined_names.add(DefinedName("MortAges", attr_text=f"Mortality!$A$2:$A${na + 1}"))
    wb.defined_names.add(DefinedName("MortRates", attr_text=f"Mortality!$B$2:$G${na + 1}"))

    # ---------------------------------------------------------- Discount
    ds = wb.create_sheet("Discount")
    hdr(ds, 1, 1, "Year", 8)
    hdr(ds, 1, 2, "Annual spot rate", 15)
    for i, (yr, row) in enumerate(disc.iterrows(), start=2):
        ds.cell(row=i, column=1, value=yr).font = INPUT_FONT
        c = ds.cell(row=i, column=2, value=float(row.iloc[0]))
        c.font = INPUT_FONT
        c.number_format = FMT_RATE
    ds.freeze_panes = "A2"
    ny = len(disc)
    wb.defined_names.add(DefinedName("DiscRateAnn", attr_text=f"Discount!$B$2:$B${ny + 1}"))

    # ---------------------------------------------------------- Projection
    pj = wb.create_sheet("Projection")
    cols = [
        ("t", "Month", 6, FMT_INT),
        ("duration", "Duration (yrs)", 9, FMT_INT),
        ("age", "Attained age", 9, FMT_INT),
        ("mort_rate", "Mortality rate p.a.", 12, "0.000000"),
        ("mort_rate_mth", "Mortality rate p.m.", 12, "0.00000000"),
        ("lapse_rate", "Lapse rate p.a.", 10, FMT_RATE),
        ("pols_if", "Policies in force", 12, FMT_COUNT),
        ("pols_death", "Deaths", 12, "0.00000000"),
        ("pols_lapse", "Lapses", 12, FMT_COUNT),
        ("pols_maturity", "Maturities", 12, FMT_COUNT),
        ("premiums", "Premium income", 12, FMT_MONEY),
        ("claims", "Death claims", 12, FMT_MONEY),
        ("commissions", "Commission", 12, FMT_MONEY),
        ("inflation_factor", "Inflation factor", 11, "0.000000"),
        ("expenses", "Expenses", 11, FMT_MONEY),
        ("net_cf", "Net cash flow", 12, FMT_MONEY),
        ("disc_rate_mth", "Disc rate p.m.", 11, "0.00000000"),
        ("disc_factor", "Discount factor v^t", 12, "0.00000000"),
        ("pv_net_cf_t", "PV net cash flow", 12, FMT_MONEY),
    ]
    pj.sheet_view.showGridLines = False
    hdr_row = 2
    label(pj, 1, 1, "Monthly projection — all rows formula-driven; flows are structurally zero after maturity", bold=True)
    for j, (key, title, w, fmt) in enumerate(cols, start=1):
        hdr(pj, hdr_row, j, title, w)
        pj.cell(row=hdr_row + 1, column=j, value=key).font = Font(size=8, italic=True, color="7F7F7F")
    first = hdr_row + 2  # first data row (t = 0)

    C = {key: get_column_letter(j) for j, (key, *_r) in enumerate(cols, start=1)}

    for t in range(MAX_T + 1):
        r = first + t
        p = r - 1  # previous row
        f = {}
        f["t"] = t if t == 0 else f"={C['t']}{p}+1"
        f["duration"] = f"=INT({C['t']}{r}/12)"
        f["age"] = f"=AgeAtEntry+{C['duration']}{r}"
        f["mort_rate"] = (f"=INDEX(MortRates,MATCH({C['age']}{r},MortAges,0),"
                          f"MIN({C['duration']}{r},5)+1)")
        f["mort_rate_mth"] = f"=1-(1-{C['mort_rate']}{r})^(1/12)"
        f["lapse_rate"] = f"=MAX(LapseBase-LapseStep*{C['duration']}{r},LapseFloor)"
        if t == 0:
            f["pols_if"] = "=PolsIfInit"
            f["pols_maturity"] = 0
        else:
            f["pols_maturity"] = (f"=IF({C['t']}{r}=12*PolicyTerm,"
                                  f"{C['pols_if']}{p}-{C['pols_lapse']}{p}-{C['pols_death']}{p},0)")
            f["pols_if"] = (f"=IF({C['t']}{r}>12*PolicyTerm,0,"
                            f"{C['pols_if']}{p}-{C['pols_lapse']}{p}-{C['pols_death']}{p}"
                            f"-{C['pols_maturity']}{r})")
        f["pols_death"] = f"={C['pols_if']}{r}*{C['mort_rate_mth']}{r}"
        f["pols_lapse"] = (f"=({C['pols_if']}{r}-{C['pols_death']}{r})"
                           f"*(1-(1-{C['lapse_rate']}{r})^(1/12))")
        f["premiums"] = f"=PremiumPP*{C['pols_if']}{r}"
        f["claims"] = f"=SumAssured*{C['pols_death']}{r}"
        f["commissions"] = f"=IF({C['duration']}{r}=0,{C['premiums']}{r},0)"
        f["inflation_factor"] = f"=(1+InflationRate)^({C['t']}{r}/12)"
        f["expenses"] = (f"={C['pols_if']}{r}*ExpenseMaint/12*{C['inflation_factor']}{r}"
                         f"+IF({C['t']}{r}=0,ExpenseAcq,0)")
        f["net_cf"] = (f"={C['premiums']}{r}-{C['claims']}{r}"
                       f"-{C['expenses']}{r}-{C['commissions']}{r}")
        f["disc_rate_mth"] = f"=(1+INDEX(DiscRateAnn,{C['duration']}{r}+1))^(1/12)-1"
        f["disc_factor"] = f"=(1+{C['disc_rate_mth']}{r})^-{C['t']}{r}"
        f["pv_net_cf_t"] = f"={C['net_cf']}{r}*{C['disc_factor']}{r}"

        for key, _title, _w, fmt in cols:
            c = pj.cell(row=r, column=cols.index((key, _title, _w, fmt)) + 1, value=f[key])
            c.number_format = fmt
            if isinstance(f[key], str) and f[key].startswith("="):
                uses_names = any(nm in f[key] for nm in
                                 ("AgeAtEntry", "PolicyTerm", "PolsIfInit", "PremiumPP",
                                  "SumAssured", "ExpenseAcq", "ExpenseMaint", "InflationRate",
                                  "LapseBase", "LapseStep", "LapseFloor", "MortRates",
                                  "MortAges", "DiscRateAnn"))
                c.font = LINK_FONT if uses_names else CALC_FONT
            else:
                c.font = CALC_FONT
    pj.freeze_panes = pj.cell(row=first, column=3).coordinate
    last = first + MAX_T
    for nm, key in [("Proj_T", "t"), ("Proj_PolsIf", "pols_if"), ("Proj_Premiums", "premiums"),
                    ("Proj_Claims", "claims"), ("Proj_Commissions", "commissions"),
                    ("Proj_Expenses", "expenses"), ("Proj_NetCF", "net_cf"),
                    ("Proj_DiscFactor", "disc_factor"), ("Proj_PVNetCF", "pv_net_cf_t"),
                    ("Proj_Deaths", "pols_death"), ("Proj_Lapses", "pols_lapse"),
                    ("Proj_Maturities", "pols_maturity")]:
        col = C[key]
        wb.defined_names.add(DefinedName(nm, attr_text=f"Projection!${col}${first}:${col}${last}"))

    # ---------------------------------------------------------- Summary
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["B"].width = 34
    sm.column_dimensions["C"].width = 16
    sm.column_dimensions["D"].width = 14
    sm.column_dimensions["E"].width = 56

    label(sm, 2, 2, "Model point", bold=True)
    hdr(sm, 3, 2, "Item"); hdr(sm, 3, 3, "Value"); hdr(sm, 3, 4, ""); hdr(sm, 3, 5, "lifelib cell")

    def sm_row(r, item, formula, note, fmt=FMT_MONEY, col3="", name=None):
        label(sm, r, 2, item)
        c = sm.cell(row=r, column=3, value=formula)
        c.font = LINK_FONT
        if fmt:
            c.number_format = fmt
        c.border = BORDER
        if col3:
            c2 = sm.cell(row=r, column=4, value=col3)
            c2.font = LINK_FONT
            c2.number_format = "0.0%"
            c2.border = BORDER
        sm.cell(row=r, column=5, value=note).font = Font(size=9, italic=True, color="7F7F7F")
        if name:
            name_cell(wb, name, "Summary", f"$C${r}")

    sm_row(4, "Model point ID", "=PointID", "point_id", FMT_INT)
    sm_row(5, "Age at entry / sex / term", '=AgeAtEntry&" / "&Control!C9&" / "&PolicyTerm&"y"', "", None)
    sm_row(6, "Sum assured", "=SumAssured", "sum_assured()", FMT_MONEY)

    label(sm, 8, 2, "Premium basis (equivalence principle)", bold=True)
    hdr(sm, 9, 2, "Item"); hdr(sm, 9, 3, "Value"); hdr(sm, 9, 4, ""); hdr(sm, 9, 5, "lifelib cell")
    sm_row(10, "PV policies in force (annuity factor)", "=SUMPRODUCT(Proj_PolsIf,Proj_DiscFactor)",
           "pv_pols_if() — monthly annuity-due factor", FMT_COUNT, name="PV_PolsIf")
    sm_row(11, "PV death claims", "=SUMPRODUCT(Proj_Claims,Proj_DiscFactor)",
           "pv_claims()", FMT_MONEY, name="PV_Claims")
    sm_row(12, "Net premium per policy (monthly)", "=PV_Claims/PV_PolsIf",
           "net_premium_pp() = pv_claims / pv_pols_if", FMT_MONEY, name="NetPremiumPP")
    sm_row(13, "Office premium per policy (monthly)", "=ROUND((1+LoadingPrem)*NetPremiumPP,2)",
           "premium_pp() — rounded to cents", FMT_MONEY, name="PremiumPP")

    label(sm, 15, 2, "Present value of cash flows", bold=True)
    hdr(sm, 16, 2, "Item"); hdr(sm, 16, 3, "PV"); hdr(sm, 16, 4, "% premium"); hdr(sm, 16, 5, "lifelib cell")
    sm_row(17, "Premiums", "=SUMPRODUCT(Proj_Premiums,Proj_DiscFactor)", "pv_premiums()",
           FMT_MONEY, col3="=C17/PV_Premiums", name="PV_Premiums")
    sm_row(18, "Death claims", "=PV_Claims", "pv_claims()", FMT_MONEY, col3="=C18/PV_Premiums")
    sm_row(19, "Expenses", "=SUMPRODUCT(Proj_Expenses,Proj_DiscFactor)", "pv_expenses()",
           FMT_MONEY, col3="=C19/PV_Premiums", name="PV_Expenses")
    sm_row(20, "Commission", "=SUMPRODUCT(Proj_Commissions,Proj_DiscFactor)", "pv_commissions()",
           FMT_MONEY, col3="=C20/PV_Premiums", name="PV_Commissions")
    sm_row(21, "Net cash flow", "=PV_Premiums-PV_Claims-PV_Expenses-PV_Commissions",
           "pv_net_cf()", FMT_MONEY, col3="=C21/PV_Premiums", name="PV_NetCF")
    sm.cell(row=21, column=2).font = Font(size=10, bold=True)
    sm.cell(row=21, column=3).font = Font(name="Calibri", size=10, bold=True, color="006100")

    # ---------------------------------------------------------- Checks
    ck = wb.create_sheet("Checks")
    ck.sheet_view.showGridLines = False
    ck.column_dimensions["B"].width = 46
    ck.column_dimensions["C"].width = 10
    ck.column_dimensions["D"].width = 72
    label(ck, 2, 2, "Internal consistency checks", bold=True)
    hdr(ck, 3, 2, "Check"); hdr(ck, 3, 3, "Result"); hdr(ck, 3, 4, "Description")

    checks = [
        ("PV identity (check_pv_net_cf)",
         "=ABS(SUM(Proj_PVNetCF)-PV_NetCF)<0.000001",
         "Sum of the per-month discounted net cash flow column equals PV premiums − claims "
         "− expenses − commission (lifelib check_pv_net_cf)."),
        ("In-force never negative",
         "=MIN(Proj_PolsIf)>=0",
         "Policies in force are non-negative at every projection month."),
        ("Decrements exhaust the in-force",
         "=ABS(PolsIfInit-SUM(Proj_Deaths)-SUM(Proj_Lapses)-SUM(Proj_Maturities))<0.000001",
         "Deaths + lapses + maturities over the projection equal the initial policy count."),
        ("In-force zero after maturity",
         "=INDEX(Proj_PolsIf,12*PolicyTerm+1)=0",
         "The in-force count is exactly zero in the maturity month (t = 12 × policy term)."),
        ("Office premium covers net premium",
         "=PremiumPP>=NetPremiumPP",
         "Loaded premium is at least the equivalence-principle net premium."),
    ]
    for i, (name, formula, desc) in enumerate(checks):
        r = 4 + i
        label(ck, r, 2, name)
        c = ck.cell(row=r, column=3, value=formula)
        c.font = LINK_FONT
        c.border = BORDER
        c.alignment = CENTER
        ck.cell(row=r, column=4, value=desc).font = Font(size=9, color="7F7F7F")
    rall = 4 + len(checks) + 1
    label(ck, rall, 2, "ALL CHECKS PASS", bold=True)
    c = ck.cell(row=rall, column=3, value=f"=AND(C4:C{3 + len(checks)})")
    c.font = Font(name="Calibri", size=10, bold=True, color="006100")
    c.alignment = CENTER
    name_cell(wb, "AllChecksPass", "Checks", f"$C${rall}")

    # ---------------------------------------------------------- Lifelib_Reference
    lr = wb.create_sheet("Lifelib_Reference")
    label(lr, 1, 1, "lifelib results for all model points (basiclife.BasicTerm_M — vectorised twin of "
          "BasicTerm_S; the S-vs-M equivalence was cross-checked to ~1e-14 on the 10 deep-"
          "verification points, and the batch reconciles the workbook against all 10,000 M rows). "
          "Generated by verification/generate_reference_all_points.py.", bold=True)
    ref_cols = ["point_id", "premium_pp", "pv_premiums", "pv_claims",
                "pv_expenses", "pv_commissions", "pv_net_cf"]
    for j, h in enumerate(ref_cols, start=1):
        hdr(lr, 2, j, h, 14)
    for i, (pid, row) in enumerate(ref_all.iterrows(), start=3):
        lr.cell(row=i, column=1, value=int(pid)).font = INPUT_FONT
        for j, colname in enumerate(ref_cols[1:], start=2):
            c = lr.cell(row=i, column=j, value=float(row[colname]))
            c.font = INPUT_FONT
            c.number_format = FMT_MONEY
    lr.freeze_panes = "A3"

    # ---------------------------------------------------------- Batch_Results
    br = wb.create_sheet("Batch_Results")
    br.sheet_view.showGridLines = False
    br.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGHI", (30, 16, 16, 14, 16, 16, 14, 10)):
        br.column_dimensions[col].width = w
    label(br, 1, 2, "All model points — batch run and reconciliation vs lifelib", bold=True)
    br.cell(row=2, column=2, value="Filled by the RunAllModelPoints macro (see Control sheet for settings)."
            ).font = Font(size=9, italic=True, color="7F7F7F")

    summary_rows = [
        ("Model points run", "B4"),
        ("Reconciled within tolerance", "B5"),
        ("Mismatches", "B6"),
        ("Max |diff| office premium", "B7"),
        ("Max relative diff PV (all components)", "B8"),
        ("Run time (seconds)", "B9"),
        ("RECONCILIATION RESULT", "B10"),
    ]
    for i, (txt, _) in enumerate(summary_rows):
        label(br, 4 + i, 2, txt, bold=(i == len(summary_rows) - 1))
        c = br.cell(row=4 + i, column=3)
        c.border = BORDER

    hdrs = ["point_id", "premium_pp (wb)", "premium_pp (lifelib)", "|diff| premium",
            "rel pv_premiums", "rel pv_claims", "rel pv_expenses", "rel pv_commissions",
            "pv_net_cf (wb)", "pv_net_cf (lifelib)", "rel pv_net_cf", "reconciles"]
    for j, h in enumerate(hdrs, start=2):
        hdr(br, 12, j, h)
    br.freeze_panes = "A13"

    # Bind the VBA project's document modules: the harvested vbaProject.bin
    # contains doc modules Sheet1..Sheet11 (creation order) + ThisWorkbook.
    # Without matching codeNames Excel cannot attach them, and the first
    # VBA reference to ThisWorkbook fails with run-time error 429.
    wb.code_name = "ThisWorkbook"
    for i, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_properties.codeName = f"Sheet{i}"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    if VBA_BIN.exists():
        make_xlsm(OUT_XLSX, OUT_XLSM, VBA_BIN)
        print(f"wrote {OUT_XLSM} (macro project injected)")
    else:
        print("vbaProject.bin not found - .xlsm not produced. "
              "Bootstrap it once via workbooks/BasicTerm_S/bootstrap_vba.py")


def make_xlsm(src_xlsx: Path, dst_xlsm: Path, vba_bin: Path):
    """Produce an .xlsm from the .xlsx by injecting vbaProject.bin (zip surgery)."""
    shutil.copy(src_xlsx, dst_xlsm)
    # Read all parts, rewrite with modified content types / rels + the bin
    with zipfile.ZipFile(src_xlsx) as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}

    ct = parts["[Content_Types].xml"].decode()
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    if "vnd.ms-office.vbaProject" not in ct:
        ct = ct.replace(
            "</Types>",
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
        )
    parts["[Content_Types].xml"] = ct.encode()

    rels = parts["xl/_rels/workbook.xml.rels"].decode()
    if "vbaProject" not in rels:
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdVBA" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>',
        )
    parts["xl/_rels/workbook.xml.rels"] = rels.encode()
    parts["xl/vbaProject.bin"] = vba_bin.read_bytes()

    with zipfile.ZipFile(dst_xlsm, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)


if __name__ == "__main__":
    build()
