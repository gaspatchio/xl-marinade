"""Build the Solvency2 workbook — an Excel twin of lifelib's solvency2 model.

Life SCR standard formula for a selectable policy: the annual life engine
(BaseProj + PV recursions) is projected under the base basis and six life
stresses (mortality, longevity, expense, lapse up/down/mass) on seven
structurally identical scenario sheets; the SCR sheet aggregates the risk
charges through the 7x7 correlation matrix.

Scope: the workbook maps the cells lifelib actually evaluates for
SCR_life() — the SCR subgraph. Statutory reserve/profit cells
(ChangeRsrv, ProfitBefTax, InvstIncome, AccumCF, ...) and the VAL rate
basis are never evaluated by SCR_life and are not replicated. The three
risks without lifelib overrides (disab, rev, cat) produce a stressed
projection identical to base, so Life(risk) = 0 is entered directly with
that reasoning documented on the SCR sheet.

Lookup semantics: lifelib resolves most items with modelx `match`, which
masks argument combinations from most to least specific, skipping keys
whose value is None — but resolves CnsmpTax/InflRate and the LoadAcqSA /
LoadMaintPrem params with a plain direct dict get (no cascade). The
workbook mirrors both: match-items use an 8-step IFERROR/MATCH cascade
over concatenated keys replicating modelx's mask order for the
(prod, polt, gen) positions (modelx additionally enumerates masks that
blank the item key itself; those are unreachable because no lookup table
has item-less rows), direct-items use a single exact-key MATCH. Rows with
None values are omitted from the workbook tables (present-with-None and
absent are equivalent to both lookup styles).

Conventions: blue = input, black = in-sheet formula, green = cross-sheet
link; machine-key row (grey) under each header carries the lifelib cell
name. Output: Solvency2.xlsm when vbaProject.bin is present (zip surgery),
Solvency2.xlsx otherwise.
"""

import shutil
import zipfile
from pathlib import Path

import modelx as mx
from openpyxl import Workbook
from openpyxl.chart import RadarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

REPO = Path(__file__).resolve().parents[2]
MODEL = REPO / "models" / "solvency2" / "model"
REF_CSV = REPO / "verification" / "reference" / "solvency2_all_policies.csv"
HERE = Path(__file__).resolve().parent
VBA_BIN = HERE / "vbaProject.bin"
OUT_XLSX = HERE / "Solvency2.xlsx"
OUT_XLSM = HERE / "Solvency2.xlsm"

MAX_T = 103          # projection rows t = 0..103 (max last_t = 102)
MAX_X = 130          # commutation ages 0..130
N_POL = 300
FIRST = 7            # first data row on time-indexed sheets
LAST = FIRST + MAX_T
CFIRST = 4           # first data row on the Commutation sheet
RISKS = ["mort", "longev", "disab", "lapse", "exps", "rev", "cat"]

STRESS_SHEETS = [
    # (sheet, risk, shock, risk_sign)
    ("Proj_Base", "base", "", 0),
    ("Proj_Mort", "mort", "", 1),
    ("Proj_Longev", "longev", "", -1),
    ("Proj_Exps", "exps", "", 0),
    ("Proj_LapseUp", "lapse", "up", 0),
    ("Proj_LapseDown", "lapse", "down", 0),
    ("Proj_LapseMass", "lapse", "mass", 0),
]

# ---------------------------------------------------------------- styles
NAVY = "1F3864"
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")
CALC_FONT = Font(name="Calibri", size=10, color="000000")
LINK_FONT = Font(name="Calibri", size=10, color="006100")
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color=NAVY)
KEY_FONT = Font(size=8, italic=True, color="7F7F7F")
NOTE_FONT = Font(size=9, italic=True, color="7F7F7F")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")

FMT_INT = "0"
FMT_RATE = "0.0000%"
FMT_MONEY = "#,##0.00"
FMT_CNT = "0.00000000"
FMT_FACT = "0.00000000"


def hdr(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def label(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=bold)
    return c


def name_cell(wb, name, sheet, addr):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{addr}"))


def item_row(ws, r, item, value, note, fmt=None, formula=False, link=False):
    label(ws, r, 2, item)
    c = ws.cell(row=r, column=3, value=value)
    c.font = LINK_FONT if link else (CALC_FONT if formula else INPUT_FONT)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = NOTE_FONT
    return c


def cascade(key, keys_name, vals_name):
    """The 8-step modelx-match lookup cascade as one Excel formula.

    Masking order matches modelx CellsImpl.find_match combinations for a
    (key, prod, polt, gen) node: exact, then 3-arg, 2-arg, 1-arg masks in
    lexicographic order of retained positions.
    """
    p, t, g = "Product", "PolicyType", "Gen"
    cands = [
        f'"{key}|"&{p}&"|"&{t}&"|"&{g}',
        f'"{key}|"&{p}&"|"&{t}&"|"',
        f'"{key}|"&{p}&"||"&{g}',
        f'"{key}||"&{t}&"|"&{g}',
        f'"{key}|"&{p}&"||"',
        f'"{key}||"&{t}&"|"',
        f'"{key}|||"&{g}',
        f'"{key}|||"',
    ]
    expr = f"MATCH({cands[-1]},{keys_name},0)"
    for cand in reversed(cands[:-1]):
        expr = f"IFERROR(MATCH({cand},{keys_name},0),{expr})"
    return f"=IFERROR(INDEX({vals_name},{expr}),\"NA\")"


def load_inputs():
    model = mx.read_model(str(MODEL))
    inp = model.Input
    pol_attrs = ["Product", "PolicyType", "Gen", "Sex", "IssueAge",
                 "PremFreq", "PolicyTerm", "PolicyCount", "SumAssured"]
    data = {
        "policies": [[pid] + [inp.PolicyData[pid, a] for a in pol_attrs]
                     for pid in range(1, N_POL + 1)],
        "assumption": [(k, v) for k in inp.Assumption.keys()
                       if (v := inp.Assumption[k]) is not None],
        "spec": [(k, v) for k in inp.ProductSpec.keys()
                 if (v := inp.ProductSpec[k]) is not None],
        # some unused (table, sex, age) entries hold None - left as blank cells
        "mort": {(tid, sex): [None if (v := inp.MortalityTables[tid, sex, x]) is None
                              else float(v) for x in range(MAX_X + 1)]
                 for tid in range(1, 7) for sex in "MF"},
        "asmp_tables": {},
        "scen": {sid: [float(inp.Scenarios[sid, "IntRate", t])
                       for t in range(151)] for sid in range(1, 11)},
        "corr": {k: float(inp.CorrData[k]) for k in inp.CorrData.keys()},
        "factors": [(k, float(inp.FactorData[k]))
                    for k in inp.FactorData.keys()],
    }
    tbls = sorted({k[0] for k in inp.AssumptionTables.keys()})
    for tbl in tbls:
        data["asmp_tables"][tbl] = [float(inp.AssumptionTables[tbl, y])
                                    for y in range(20)]
    return data


def concat_key(parts):
    return "|".join("" if x is None else str(x) for x in parts)


# ---------------------------------------------------------------- build
def build():
    data = load_inputs()
    wb = Workbook()

    # ---------------------------------------------------------- Cover
    cv = wb.active
    cv.title = "Cover"
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["B"].width = 28
    cv.column_dimensions["C"].width = 80
    t = cv.cell(row=2, column=2, value="Solvency2 — Life Underwriting Risk SCR (Standard Formula)")
    t.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    cv.cell(row=3, column=2, value="Life SCR for a selected policy: base and stressed liability "
            "projections aggregated through the standard-formula correlation matrix"
            ).font = Font(size=11, italic=True)
    rows = [
        ("Purpose", "Calculates the Solvency II life underwriting risk SCR for the policy selected "
                    "on the Control sheet. The annual liability cashflow engine is projected on "
                    "seven scenario sheets — the base basis plus mortality, longevity, expense and "
                    "three lapse stresses — and the SCR sheet aggregates the resulting risk charges "
                    "with the standard-formula correlation matrix. The workbook is a formula-for-"
                    "formula replication of the open-source lifelib model 'solvency2' (MIT licence), "
                    "verified cell-by-cell against it."),
        ("Method", "SCR_life = sqrt( sum_ij Corr(i,j) * Life_i * Life_j ). Each Life charge is "
                   "max(NAV_base − NAV_stressed, 0), where NAV is the present value of net liability "
                   "cashflows at time t0 (assets assumed insensitive to life stresses, as in lifelib). "
                   "Lapse risk takes the worst of the up / down / mass shocks."),
        ("Engine", "Net level premium engine: gross premium rates and reserves from PREM-basis "
                   "commutation functions (Commutation sheet); decrements from the base mortality "
                   "table, mortality factor and surrender rate tables; expense and commission "
                   "assumptions resolved per product via the Basis_Lookup cascade."),
        ("Stresses", "Scenario sheets share one engine layout; each carries a parameter block "
                     "(risk, shock, factors from the Factors table) that reproduces the lifelib "
                     "Override spaces: mortality/longevity rate factors from t0, lapse up/down rate "
                     "shocks with limits, one-off mass lapse at t0, expense level and inflation "
                     "shocks. Disability, revision and catastrophe risks have no lifelib override "
                     "(stressed projection would equal base), so their charges are zero."),
        ("Scope", "Only the cells lifelib evaluates for SCR_life() are replicated. Statutory "
                  "reserve roll-forward and profit cells (ChangeRsrv, ProfitBefTax, InvstIncome, "
                  "AccumCF) and the VAL rate basis lie outside that subgraph."),
        ("Reconciliation", "The RunAllPolicies macro (Batch_Results) computes the SCR and its risk "
                           "decomposition for every one of the 300 policies and reconciles them "
                           "against the lifelib values on Lifelib_Reference."),
        ("Source model", "lifelib solvency2 — https://lifelib.io (MIT licence)"),
        ("Verification", "verification/verify_solvency2.py compares every mapped cell against the "
                         "Python model for deep profiles across products, sexes, t0 and scenarios; "
                         "see the Checks sheet for in-workbook consistency tests."),
    ]
    r = 5
    for k, v in rows:
        label(cv, r, 2, k, bold=True)
        c = cv.cell(row=r, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cv.row_dimensions[r].height = 54
        r += 1
    r += 1
    label(cv, r, 2, "Version control", bold=True)
    for i, (a, b) in enumerate([("Version", "0.2.0"), ("Date", "2026-07-14"),
                                ("Prepared by", "XL Marinade project"),
                                ("Review status", "Verified against lifelib; adversarial "
                                 "actuarial audit remediated (see verification/reports)")]):
        label(cv, r + 1 + i, 2, a)
        cv.cell(row=r + 1 + i, column=3, value=b).font = Font(size=10)
    r += 6
    label(cv, r, 2, "Sheet index", bold=True)
    for i, (s, d) in enumerate([
            ("Control", "Run selection (policy, scenario, t0) and batch settings"),
            ("Policy_Basis", "Selected policy, resolved lookups, premium/reserve rates"),
            ("Policy_Data", "Policy file — 300 policies (input data)"),
            ("Basis_Lookup", "Assumption and product-spec lookup tables (input data)"),
            ("Basis_Tables", "Year-indexed assumption tables and shock factors (input data)"),
            ("Mortality", "Mortality tables qx by age, 6 tables x M/F (input data)"),
            ("Scenarios", "Discount rate paths by scenario (input data)"),
            ("Commutation", "PREM-basis commutation functions lx/Dx/Cx/Nx/Mx"),
            ("Policy_Sched", "Per-policy schedule shared by all stress projections"),
            ("Proj_Base … Proj_LapseMass", "Liability projections — base + 6 life stresses"),
            ("SCR", "NAVs, risk charges, correlation matrix, SCR aggregation"),
            ("Summary", "Run details, SCR decomposition, risk radar"),
            ("Checks", "Internal consistency checks (all must be TRUE)"),
            ("Lifelib_Reference", "lifelib results for all policies (reconciliation target)"),
            ("Batch_Results", "All-policies batch reconciliation vs lifelib (macro)")]):
        label(cv, r + 1 + i, 2, s)
        cv.cell(row=r + 1 + i, column=3, value=d).font = Font(size=10)
    r += 17
    label(cv, r, 2, "Conventions", bold=True)
    for i, (txt, font) in enumerate([
            ("Blue — hardcoded input", INPUT_FONT),
            ("Black — calculated on this sheet", CALC_FONT),
            ("Green — link from another sheet", LINK_FONT)]):
        cv.cell(row=r + 1 + i, column=2, value=txt).font = font

    # ---------------------------------------------------------- Control
    ct = wb.create_sheet("Control")
    ct.sheet_view.showGridLines = False
    ct.column_dimensions["B"].width = 34
    ct.column_dimensions["C"].width = 14
    ct.column_dimensions["D"].width = 64
    label(ct, 2, 2, "Run selection", bold=True)
    hdr(ct, 3, 2, "Item"); hdr(ct, 3, 3, "Value"); hdr(ct, 3, 4, "Notes / lifelib parameter")
    item_row(ct, 4, "Policy ID", 1, "PolicyID — choose 1 to 300", FMT_INT)
    item_row(ct, 5, "Economic scenario", 1, "ScenID — discount rate path, choose 1 to 10", FMT_INT)
    item_row(ct, 6, "Valuation time t0 (years)", 0,
             "t0 — stresses apply from t0; NAV is PV of net cashflows at t0", FMT_INT)
    for nm, addr, lo, hi in [("PolicyID", "$C$4", 1, N_POL), ("ScenID", "$C$5", 1, 10),
                             ("T0", "$C$6", 0, MAX_T - 1)]:
        name_cell(wb, nm, "Control", addr)
        dv = DataValidation(type="whole", operator="between",
                            formula1=str(lo), formula2=str(hi), allow_blank=False)
        dv.error = f"Value must be between {lo} and {hi}"
        ct.add_data_validation(dv)
        dv.add(ct[addr.replace("$", "")])
    label(ct, 8, 2, "Batch run (RunAllPolicies macro)", bold=True)
    hdr(ct, 9, 2, "Item"); hdr(ct, 9, 3, "Value"); hdr(ct, 9, 4, "Notes")
    item_row(ct, 10, "Policies to run (blank = all)", None,
             "Cap for quick test runs; leave blank to run all 300 policies", FMT_INT)
    item_row(ct, 11, "Reconciliation tolerance (relative)", 0.000001,
             "A value reconciles when |workbook − lifelib| ≤ tol × max(1, |lifelib|)", "0.0000000")
    name_cell(wb, "BatchLimit", "Control", "$C$10")
    name_cell(wb, "BatchTol", "Control", "$C$11")
    ct.cell(row=13, column=2, value="The batch runs at t0 = 0, scenario 1 (the Lifelib_Reference "
            "basis) and restores the run selection afterwards.").font = NOTE_FONT

    # ---------------------------------------------------------- Policy_Data
    pdt = wb.create_sheet("Policy_Data")
    label(pdt, 1, 1, "Policy data — 300 policies (lifelib Input.PolicyData; unused attributes "
          "Channel/Duration/PaymentMode/MaxPolicyTerm omitted)", bold=True)
    pcols = [("policy_id", 9), ("Product", 9), ("PolicyType", 10), ("Gen", 6),
             ("Sex", 6), ("IssueAge", 9), ("PremFreq", 9), ("PolicyTerm", 10),
             ("PolicyCount", 11), ("SumAssured", 12)]
    for j, (h, w) in enumerate(pcols, start=1):
        hdr(pdt, 2, j, h, w)
    for j, key in enumerate(["PolicyID", "Product", "PolicyType", "Gen", "Sex",
                             "IssueAge", "PremFreq", "PolicyTerm", "PolicyCount",
                             "SumAssured"], start=1):
        pdt.cell(row=3, column=j, value=key).font = KEY_FONT
    for i, row in enumerate(data["policies"]):
        for j, v in enumerate(row, start=1):
            pdt.cell(row=4 + i, column=j, value=v).font = INPUT_FONT
    pdt.freeze_panes = "A4"
    for nm, col in [("MP_ID", "A"), ("MP_Product", "B"), ("MP_PolType", "C"),
                    ("MP_Gen", "D"), ("MP_Sex", "E"), ("MP_IssueAge", "F"),
                    ("MP_PremFreq", "G"), ("MP_Term", "H"), ("MP_Count", "I"),
                    ("MP_SumAssured", "J")]:
        wb.defined_names.add(DefinedName(
            nm, attr_text=f"Policy_Data!${col}$4:${col}${3 + N_POL}"))

    # ---------------------------------------------------------- Basis_Lookup
    bl = wb.create_sheet("Basis_Lookup")
    bl.sheet_view.showGridLines = False
    label(bl, 1, 1, "Assumption and product-spec lookup tables (rows with empty values omitted — "
          "modelx match() skips them). Lookup key = item|product|policy_type|generation; "
          "the Policy_Basis cascade tries masks from most to least specific.", bold=True)
    for j, h in enumerate(["item", "prod", "polt", "gen", "value", "lookup key"], start=1):
        hdr(bl, 2, j, h, 16 if h in ("item", "lookup key") else 9)
        hdr(bl, 2, j + 7, h, 16 if h in ("item", "lookup key") else 9)
    bl.cell(row=1, column=1).font = Font(size=10, bold=True)
    label(bl, 2, 7, "")
    for j, key in enumerate(["", "", "", "", "Assumption", ""], start=1):
        bl.cell(row=3, column=j, value=key).font = KEY_FONT
    for j, key in enumerate(["", "", "", "", "ProductSpec", ""], start=1):
        bl.cell(row=3, column=j + 7, value=key).font = KEY_FONT

    def lookup_block(rows, col0):
        for i, ((k, prod, polt, gen), v) in enumerate(rows):
            r = 4 + i
            for j, x in enumerate([k, prod, polt, gen, v], start=col0):
                c = bl.cell(row=r, column=j, value=x)
                c.font = INPUT_FONT
            cl = [get_column_letter(col0 + j) for j in range(4)]
            c = bl.cell(row=r, column=col0 + 5,
                        value=f'={cl[0]}{r}&"|"&{cl[1]}{r}&"|"&{cl[2]}{r}&"|"&{cl[3]}{r}')
            c.font = CALC_FONT
        return 4, 3 + len(rows)

    a_lo, a_hi = lookup_block(data["assumption"], 1)
    s_lo, s_hi = lookup_block(data["spec"], 8)
    wb.defined_names.add(DefinedName("AsmpKeys", attr_text=f"Basis_Lookup!$F${a_lo}:$F${a_hi}"))
    wb.defined_names.add(DefinedName("AsmpVals", attr_text=f"Basis_Lookup!$E${a_lo}:$E${a_hi}"))
    wb.defined_names.add(DefinedName("SpecKeys", attr_text=f"Basis_Lookup!$M${s_lo}:$M${s_hi}"))
    wb.defined_names.add(DefinedName("SpecVals", attr_text=f"Basis_Lookup!$L${s_lo}:$L${s_hi}"))

    # ---------------------------------------------------------- Basis_Tables
    bt = wb.create_sheet("Basis_Tables")
    bt.sheet_view.showGridLines = False
    label(bt, 1, 1, "Assumption tables by policy year (values carry forward beyond year 19, "
          "as in lifelib), and the standard-formula shock factors", bold=True)
    tbl_names = sorted(data["asmp_tables"])
    hdr(bt, 2, 1, "year", 6)
    for j, tn in enumerate(tbl_names, start=2):
        hdr(bt, 2, j, tn, 11)
    bt.cell(row=3, column=1, value="y").font = KEY_FONT
    bt.cell(row=3, column=2, value="AssumptionTables").font = KEY_FONT
    for y in range(20):
        bt.cell(row=4 + y, column=1, value=y).font = INPUT_FONT
        for j, tn in enumerate(tbl_names, start=2):
            c = bt.cell(row=4 + y, column=j, value=data["asmp_tables"][tn][y])
            c.font = INPUT_FONT
            c.number_format = "0.000000"
    ncols = len(tbl_names)
    lastc = get_column_letter(1 + ncols)
    wb.defined_names.add(DefinedName("AsmpTblGrid", attr_text=f"Basis_Tables!$B$4:${lastc}$23"))
    wb.defined_names.add(DefinedName("AsmpTblNames", attr_text=f"Basis_Tables!$B$2:${lastc}$2"))

    fcol = ncols + 3   # factors block
    hdr(bt, 2, fcol, "shock factor key", 22)
    hdr(bt, 2, fcol + 1, "value", 9)
    bt.cell(row=3, column=fcol, value="FactorData").font = KEY_FONT
    for i, (k, v) in enumerate(sorted(data["factors"], key=lambda kv: str(kv[0]))):
        bt.cell(row=4 + i, column=fcol, value=concat_key(k)).font = INPUT_FONT
        c = bt.cell(row=4 + i, column=fcol + 1, value=v)
        c.font = INPUT_FONT
        c.number_format = "0.0000"
    fk, fv = get_column_letter(fcol), get_column_letter(fcol + 1)
    nfac = len(data["factors"])
    wb.defined_names.add(DefinedName("FactorKeys", attr_text=f"Basis_Tables!${fk}$4:${fk}${3 + nfac}"))
    wb.defined_names.add(DefinedName("FactorVals", attr_text=f"Basis_Tables!${fv}$4:${fv}${3 + nfac}"))

    # ---------------------------------------------------------- Mortality
    mt = wb.create_sheet("Mortality")
    label(mt, 1, 1, "Mortality tables qx by age — 6 tables x M/F (lifelib Input.MortalityTables). "
          "Tables 5 and 6 are empty in the lifelib data set (never referenced by any policy's "
          "basis); the blank columns are kept so table ids map 1:1 to grid columns.",
          bold=True)
    hdr(mt, 2, 1, "age", 6)
    combos = [(tid, sex) for tid in range(1, 7) for sex in "MF"]
    for j, (tid, sex) in enumerate(combos, start=2):
        hdr(mt, 2, j, f"T{tid} {sex}", 9)
    mt.cell(row=3, column=1, value="x").font = KEY_FONT
    mt.cell(row=3, column=2, value="MortalityTables").font = KEY_FONT
    for x in range(MAX_X + 1):
        mt.cell(row=4 + x, column=1, value=x).font = INPUT_FONT
        for j, combo in enumerate(combos, start=2):
            c = mt.cell(row=4 + x, column=j, value=data["mort"][combo][x])
            c.font = INPUT_FONT
            c.number_format = "0.000000"
    mt.freeze_panes = "B4"
    wb.defined_names.add(DefinedName("MortGrid", attr_text=f"Mortality!$B$4:$M${4 + MAX_X}"))
    wb.defined_names.add(DefinedName("MortAges", attr_text=f"Mortality!$A$4:$A${4 + MAX_X}"))

    # ---------------------------------------------------------- Scenarios
    sc = wb.create_sheet("Scenarios")
    label(sc, 1, 1, "Economic scenarios — flat annual discount/earned rate paths by year "
          "(lifelib Input.Scenarios, IntRate)", bold=True)
    hdr(sc, 2, 1, "t", 6)
    for sid in range(1, 11):
        hdr(sc, 2, 1 + sid, f"Scen {sid}", 9)
    sc.cell(row=3, column=1, value="t").font = KEY_FONT
    sc.cell(row=3, column=2, value="DiscRate").font = KEY_FONT
    for t_ in range(151):
        sc.cell(row=4 + t_, column=1, value=t_).font = INPUT_FONT
        for sid in range(1, 11):
            c = sc.cell(row=4 + t_, column=1 + sid, value=data["scen"][sid][t_])
            c.font = INPUT_FONT
            c.number_format = FMT_RATE
    sc.freeze_panes = "B4"
    wb.defined_names.add(DefinedName("ScenGrid", attr_text="Scenarios!$B$4:$K$154"))

    # ---------------------------------------------------------- Policy_Basis
    pb = wb.create_sheet("Policy_Basis")
    pb.sheet_view.showGridLines = False
    pb.column_dimensions["B"].width = 36
    pb.column_dimensions["C"].width = 15
    pb.column_dimensions["D"].width = 70
    r = 2
    label(pb, r, 2, "Selected policy (from Policy_Data)", bold=True)
    hdr(pb, r + 1, 2, "Item"); hdr(pb, r + 1, 3, "Value"); hdr(pb, r + 1, 4, "Notes / lifelib cell")
    r += 2
    pol_rows = [
        ("Product", "Product", None, "Product() — TERM / WL / ENDW"),
        ("Policy type", "PolicyType", FMT_INT, "PolicyType()"),
        ("Generation", "Gen", FMT_INT, "Gen()"),
        ("Sex", "Sex", None, "Sex()"),
        ("Issue age", "IssueAge", FMT_INT, "IssueAge() — x"),
        ("Premium frequency p.a.", "PremFreq", FMT_INT, "PremFreq()"),
        ("Policy term (years)", "PolicyTerm", FMT_INT, "PolicyTerm() — n = m (premium term alias)"),
        ("Policy count", "PolicyCount", FMT_INT, "PolicyCount()"),
        ("Sum assured", "SumAssured", FMT_MONEY, "SumAssured()"),
    ]
    mp_col = {"Product": "MP_Product", "PolicyType": "MP_PolType", "Gen": "MP_Gen",
              "Sex": "MP_Sex", "IssueAge": "MP_IssueAge", "PremFreq": "MP_PremFreq",
              "PolicyTerm": "MP_Term", "PolicyCount": "MP_Count", "SumAssured": "MP_SumAssured"}
    for itm, nm, fmt, note in pol_rows:
        item_row(pb, r, itm, f"=INDEX({mp_col[nm]},MATCH(PolicyID,MP_ID,0))",
                 note, fmt, link=True)
        name_cell(wb, nm, "Policy_Basis", f"$C${r}")
        r += 1
    r += 1
    label(pb, r, 2, "Resolved lookups (modelx match cascade)", bold=True)
    hdr(pb, r + 1, 2, "Item"); hdr(pb, r + 1, 3, "Value"); hdr(pb, r + 1, 4, "Notes / lifelib cell")
    r += 2
    lk_rows = [
        ("BaseMortTable", "BaseMort", "A", FMT_INT, "asmp.BaseMortRate — base mortality table id"),
        ("MortFactorTable", "MortFactor", "A", None, "asmp.MortFactor — factor table name"),
        ("SurrTable", "Surrender", "A", None, "asmp.SurrRate — surrender table name"),
        ("CnsmpTax", "CnsmpTax", "A0", FMT_RATE,
         "asmp.CnsmpTax() — direct global lookup in lifelib, no cascade"),
        ("CommInitPrem", "CommInitPrem", "A", FMT_RATE, "asmp.CommInitPrem()"),
        ("CommRenPrem", "CommRenPrem", "A", FMT_RATE, "asmp.CommRenPrem()"),
        ("CommRenTerm", "CommRenTerm", "A", FMT_INT, "asmp.CommRenTerm()"),
        ("ExpsAcqAnnPrem", "ExpsAcqAnnPrem", "A", "0.0000", "asmp.ExpsAcqAnnPrem()"),
        ("ExpsAcqPol", "ExpsAcqPol", "A", FMT_MONEY, "asmp.ExpsAcqPol()"),
        ("ExpsAcqSA", "ExpsAcqSA", "A", "0.000000", "asmp.ExpsAcqSA()"),
        ("ExpsMaintAnnPrem", "ExpsMaintPrem", "A", "0.0000", "asmp.ExpsMaintAnnPrem()"),
        ("ExpsMaintPol", "ExpsMaintPol", "A", FMT_MONEY, "asmp.ExpsMaintPol()"),
        ("ExpsMaintSA", "ExpsMaintSA", "A", "0.000000", "asmp.ExpsMaintSA()"),
        ("InflRate", "InflRate", "A0", FMT_RATE,
         "asmp.InflRate() — direct global lookup in lifelib, no cascade"),
        ("IntRatePrem", "IntRatePrem", "S", FMT_RATE, "pol.IntRate('PREM')"),
        ("TableIDPrem", "MortTablePrem", "S", FMT_INT, "pol.TableID('PREM')"),
        ("LoadAcqP1", "LoadAcqSAParam1", "SP", "0.000000",
         "pol.LoadAcqSA — param 1; direct (item, product) lookup in lifelib"),
        ("LoadAcqP2", "LoadAcqSAParam2", "SP", "0.000000",
         "pol.LoadAcqSA — param 2; direct (item, product) lookup"),
        ("LoadMaintPremP1", "LoadMaintPremParam1", "SP", "0.000000",
         "pol.LoadMaintPrem — param 1; direct (item, product) lookup"),
        ("LoadMaintPremP2", "LoadMaintPremParam2", "SP", "0.000000",
         "pol.LoadMaintPrem — param 2; direct lookup — NA when absent, as in lifelib"),
        ("LoadMaintSA", "LoadMaintSA", "S", "0.000000", "pol.LoadMaintSA() — gamma"),
        ("LoadMaintSA2", "LoadMaintSA2", "S", "0.000000", "pol.LoadMaintSA2() — gamma2"),
        ("SurrP1", "SurrChargeParam1", "S", "0.000000", "pol.InitSurrCharge — param 1"),
        ("SurrP2", "SurrChargeParam2", "S", "0.000000", "pol.InitSurrCharge — param 2"),
    ]
    for nm, key, kind, fmt, note in lk_rows:
        # "A"/"S": modelx match() cascade; "A0"/"SP": lifelib calls the lookup
        # cell directly (plain dict get) - a single exact-key MATCH, no cascade
        if kind == "A0":
            f = f'=IFERROR(INDEX(AsmpVals,MATCH("{key}|||",AsmpKeys,0)),"NA")'
        elif kind == "SP":
            f = (f'=IFERROR(INDEX(SpecVals,MATCH("{key}|"&Product&"||",'
                 f'SpecKeys,0)),"NA")')
        else:
            keys, vals = ("AsmpKeys", "AsmpVals") if kind == "A" else ("SpecKeys", "SpecVals")
            f = cascade(key, keys, vals)
        item_row(pb, r, nm, f, note, fmt, link=True)
        name_cell(wb, nm, "Policy_Basis", f"$C${r}")
        r += 1
    r += 1
    label(pb, r, 2, "Derived policy quantities", bold=True)
    hdr(pb, r + 1, 2, "Item"); hdr(pb, r + 1, 3, "Value"); hdr(pb, r + 1, 4, "Notes / lifelib cell")
    r += 2
    ia1 = "IssueAge+1"
    ian1 = "IssueAge+PolicyTerm+1"
    der_rows = [
        ("BaseMortCol", "=(BaseMortTable-1)*2+IF(Sex=\"M\",1,2)", FMT_INT,
         "column of the base mortality table in the Mortality grid"),
        ("PremMortCol", "=(TableIDPrem-1)*2+IF(Sex=\"M\",1,2)", FMT_INT,
         "column of the PREM-basis mortality table"),
        ("MortFactorCol", "=MATCH(MortFactorTable,AsmpTblNames,0)", FMT_INT, ""),
        ("SurrCol", "=MATCH(SurrTable,AsmpTblNames,0)", FMT_INT, ""),
        ("LastAge", "=INDEX(MortAges,MATCH(1,INDEX(MortGrid,0,BaseMortCol),0))", FMT_INT,
         "asmp.LastAge() — first age with qx = 1 in the base table"),
        ("LastT", "=MIN(LastAge-IssueAge,PolicyTerm)", FMT_INT,
         "last_t() = min(LastAge − IssueAge, PolicyTerm)"),
        ("DxX", f"=INDEX(Comm_Dx,{ia1})", "0.00", "Dx(x)"),
        ("MxX", f"=INDEX(Comm_Mx,{ia1})", "0.00", "Mx(x)"),
        ("NxX", f"=INDEX(Comm_Nx,{ia1})", "0.00", "Nx(x)"),
        ("Dx_xn", f"=INDEX(Comm_Dx,{ian1})", "0.00", "Dx(x+n)"),
        ("Mx_xn", f"=INDEX(Comm_Mx,{ian1})", "0.00", "Mx(x+n)"),
        ("Nx_xn", f"=INDEX(Comm_Nx,{ian1})", "0.00", "Nx(x+n) = Nx(x+m), premium term m = n"),
        ("Axn", "=(MxX-Mx_xn)/DxX", FMT_FACT, "lt.Axn(x, n)"),
        ("Exn", "=Dx_xn/DxX", FMT_FACT, "lt.Exn(x, n) — endowment pure survival factor"),
        ("AnnDue_xm_k", "=(NxX-Nx_xn)/DxX-(PremFreq-1)/(2*PremFreq)*(1-Dx_xn/DxX)", FMT_FACT,
         "lt.AnnDuenx(x, m, PremFreq) = AnnDuenx(x, n, PremFreq) since m = n"),
        ("AnnDue_xn_1", "=(NxX-Nx_xn)/DxX", FMT_FACT, "lt.AnnDuenx(x, n)"),
        ("AnnDue_gamma2", "=(Nx_xn-Nx_xn)/DxX", FMT_FACT,
         "lt.AnnDuenx(x, n−m, 1, m) — identically 0 while m = n"),
        ("LoadAcqSA", "=LoadAcqP1+LoadAcqP2*MIN(PolicyTerm/10,1)", "0.000000",
         "pol.LoadAcqSA() — alpha"),
        ("LoadMaintPrem",
         "=IF(ISNUMBER(LoadMaintPremP1),LoadMaintPremP1,(LoadMaintPremP2+MIN(10,PolicyTerm))/100)",
         "0.000000", "pol.LoadMaintPrem() — beta; param 1 takes priority, as in lifelib"),
        ("LoadWaiver", "=IF(PolicyTerm<5,0.0005,IF(PolicyTerm<10,0.001,0.002))", "0.000000",
         "pol.LoadMaintPremWaiverPrem() — delta"),
        ("InitSurrCharge", "=SurrP1+SurrP2*MIN(PolicyTerm/10,1)", "0.000000",
         "pol.InitSurrCharge()"),
        ("GrossPremRate",
         '=(IF(Product="ENDW",Exn,0)+Axn+LoadAcqSA+LoadMaintSA*AnnDue_xm_k'
         "+LoadMaintSA2*AnnDue_gamma2)/(1-LoadMaintPrem-LoadWaiver)/PremFreq/AnnDue_xm_k",
         FMT_FACT, "pol.GrossPremRate() — per sum assured per payment"),
        ("NetPremRatePrem", "=(Axn+LoadMaintSA2*AnnDue_gamma2)/AnnDue_xn_1", FMT_FACT,
         "pol.NetPremRate('PREM')"),
        ("AnnPremRate", "=GrossPremRate*IF(PremFreq=0,0.1,PremFreq)", FMT_FACT,
         "pol.AnnPremRate()"),
        ("SizePremium", "=SumAssured*GrossPremRate*PremFreq", FMT_MONEY,
         "SizePremium(t) — constant over t"),
        ("SizeAnnPrem", "=SumAssured*AnnPremRate", FMT_MONEY, "SizeAnnPrem(t)"),
        ("SizeExpsAcq0", "=SizeAnnPrem*ExpsAcqAnnPrem+SumAssured*ExpsAcqSA+ExpsAcqPol", FMT_MONEY,
         "SizeExpsAcq(0) — InflFactor(0)/InflFactor(0) = 1"),
        ("SizeExpsCommInit0", "=SizePremium*CommInitPrem*(1+CnsmpTax)", FMT_MONEY,
         "SizeExpsCommInit(0)"),
    ]
    for nm, f, fmt, note in der_rows:
        link = any(x in f for x in ("Comm_", "MortGrid", "MortAges", "AsmpTblNames"))
        item_row(pb, r, nm, f, note, fmt, formula=not link, link=link)
        name_cell(wb, nm, "Policy_Basis", f"$C${r}")
        r += 1

    # ---------------------------------------------------------- Commutation
    cm = wb.create_sheet("Commutation")
    label(cm, 1, 1, "Commutation functions, PREM basis (the only basis in the SCR subgraph): "
          "qx from the PREM mortality table, i = IntRatePrem. Nx/Mx recursion floors at age 110, "
          "as in lifelib (LifeTable.Nx/Mx).", bold=True)
    ccols = [("x", "x", 6, FMT_INT), ("qx", "qx", 10, "0.000000"),
             ("lx", "lx", 12, "0.000"), ("dx", "dx", 10, "0.000"),
             ("Dx", "Dx", 12, "0.000"), ("Cx", "Cx", 10, "0.0000"),
             ("Nx", "Nx", 14, "0.000"), ("Mx", "Mx", 12, "0.0000")]
    for j, (h, key, w, _f) in enumerate(ccols, start=1):
        hdr(cm, 2, j, h, w)
        cm.cell(row=3, column=j, value=key).font = KEY_FONT
    for x in range(MAX_X + 1):
        rr = CFIRST + x
        nx = rr + 1
        p = rr - 1
        f = {
            "A": x if x == 0 else f"=A{p}+1",
            "B": f"=INDEX(MortGrid,A{rr}+1,PremMortCol)",
            "C": 100000 if x == 0 else f"=C{p}-D{p}",
            "D": f"=C{rr}*B{rr}",
            "E": f"=C{rr}*(1/(1+IntRatePrem))^A{rr}",
            "F": f"=D{rr}*(1/(1+IntRatePrem))^(A{rr}+0.5)",
            "G": f"=IF(A{rr}>=110,E{rr},G{nx}+E{rr})",
            "H": f"=IF(A{rr}>=110,E{rr},H{nx}+F{rr})",
        }
        for (hh, key, w, fmt), col in zip(ccols, "ABCDEFGH"):
            c = cm[f"{col}{rr}"]
            c.value = f[col]
            c.number_format = fmt
            c.font = LINK_FONT if col == "B" or "IntRatePrem" in str(f[col]) else CALC_FONT
    cm.freeze_panes = "A4"
    for nm, col in [("Comm_Dx", "E"), ("Comm_Nx", "G"), ("Comm_Mx", "H"),
                    ("Comm_lx", "C")]:
        wb.defined_names.add(DefinedName(
            nm, attr_text=f"Commutation!${col}${CFIRST}:${col}${CFIRST + MAX_X}"))

    # ---------------------------------------------------------- Policy_Sched
    ps = wb.create_sheet("Policy_Sched")
    label(ps, 1, 1, "Per-policy schedule shared by all stress projections (lifelib pol/asmp/scen "
          "cells are policy-level and common to every Projection[risk, shock])", bold=True)
    scols = [
        ("t", "t", 6, FMT_INT),
        ("Attained age", "AttAge", 9, FMT_INT),
        ("Base mortality qx", "BaseMortRate", 12, "0.000000"),
        ("Mortality factor", "MortFactor", 11, "0.000000"),
        ("Surrender rate (base)", "SurrRate", 12, FMT_RATE),
        ("NLP reserve rate (PREM)", "ReserveNLP_Rate", 14, FMT_FACT),
        ("Surrender charge rate", "SurrCharge", 12, FMT_FACT),
        ("Cash value rate", "CashValueRate", 12, FMT_FACT),
        ("Surrender benefit pp", "SizeBenefitSurr", 13, FMT_MONEY),
        ("Renewal commission pp", "SizeExpsCommRen", 13, FMT_MONEY),
        ("Discount rate", "DiscRate", 10, FMT_RATE),
    ]
    for j, (h, key, w, _f) in enumerate(scols, start=1):
        hdr(ps, 5, j, h, w)
        ps.cell(row=6, column=j, value=key).font = KEY_FONT
    for t_ in range(MAX_T + 1):
        rr = FIRST + t_
        nx = rr + 1
        p = rr - 1
        guard = f"$A{rr}>LastT"
        guard1 = f"$A{rr}>LastT+1"
        f = {
            "A": t_ if t_ == 0 else f"=A{p}+1",
            "B": f"=IssueAge+A{rr}",
            "C": f"=IF({guard},0,INDEX(MortGrid,B{rr}+1,BaseMortCol))",
            "D": f"=IF({guard},0,INDEX(AsmpTblGrid,MIN(A{rr},19)+1,MortFactorCol))",
            "E": f"=IF({guard},0,INDEX(AsmpTblGrid,MIN(A{rr},19)+1,SurrCol))",
            # nested IFs, not OR(): OR() evaluates both arguments, and beyond
            # the horizon B+1 runs past the commutation grid -> #REF!
            "F": (f"=IF({guard1},0,IF(INDEX(Comm_Dx,B{rr}+1)=0,0,"
                  f"(INDEX(Comm_Mx,B{rr}+1)-Mx_xn+LoadMaintSA2*(Nx_xn-Nx_xn)"
                  f"-IF(A{rr}<=PolicyTerm,NetPremRatePrem*(INDEX(Comm_Nx,B{rr}+1)-Nx_xn),0))"
                  f"/INDEX(Comm_Dx,B{rr}+1)))"),
            "G": (f"=IF({guard1},0,InitSurrCharge"
                  f"*MAX((MIN(PolicyTerm,10)-A{rr})/MIN(PolicyTerm,10),0))"),
            "H": f"=IF({guard1},0,MAX(F{rr}-G{rr},0))",
            "I": f"=IF({guard},0,SumAssured*(H{rr}+H{nx})/2)",
            "J": (f"=IF(OR($A{rr}=0,{guard}),0,IF(A{rr}<CommRenTerm,"
                  f"SizePremium*CommRenPrem*(1+CnsmpTax),0))"),
            "K": f"=IF({guard},0,INDEX(ScenGrid,A{rr}+1,ScenID))",
        }
        for (hh, key, w, fmt), col in zip(scols, "ABCDEFGHIJK"):
            c = ps[f"{col}{rr}"]
            c.value = f[col]
            c.number_format = fmt
            c.font = CALC_FONT if col in ("A", "B") else LINK_FONT
    ps.freeze_panes = ps.cell(row=FIRST, column=3).coordinate

    # ---------------------------------------------------------- stress sheets
    pj_cols = [
        ("t", "t", 5, FMT_INT),
        ("Inflation factor", "InflFactor", 10, "0.000000"),
        ("Mortality shock factor", "MortRateFactor", 11, "0.000000"),
        ("Surrender rate", "SurrRate", 11, FMT_RATE),
        ("In force: end", "PolsIF_End", 11, FMT_CNT),
        ("Maturities", "PolsMaturity", 10, FMT_CNT),
        ("In force: after maturity", "PolsIF_Beg", 11, FMT_CNT),
        ("New business", "PolsNewBiz", 10, FMT_CNT),
        ("Mass lapse", "PolsSurrMass", 10, FMT_CNT),
        ("In force: beginning", "PolsIF_Beg1", 11, FMT_CNT),
        ("Deaths", "PolsDeath", 11, FMT_CNT),
        ("Surrenders", "PolsSurr", 11, FMT_CNT),
        ("Maint expense pp", "SizeExpsMaint", 11, FMT_MONEY),
        ("Premium income", "PremIncome", 12, FMT_MONEY),
        ("Death benefits", "BenefitDeath", 12, FMT_MONEY),
        ("Surrender benefits", "BenefitSurr", 12, FMT_MONEY),
        ("Total benefits", "BenefitTotal", 12, FMT_MONEY),
        ("Acquisition expenses", "ExpsAcq", 11, FMT_MONEY),
        ("Initial commission", "ExpsCommInit", 11, FMT_MONEY),
        ("Renewal commission", "ExpsCommRen", 11, FMT_MONEY),
        ("Maintenance expenses", "ExpsMaint", 11, FMT_MONEY),
        ("Total expenses", "ExpsTotal", 12, FMT_MONEY),
        ("PV premiums", "PV_PremIncome", 13, FMT_MONEY),
        ("PV benefits", "PV_BenefitTotal", 13, FMT_MONEY),
        ("PV expenses", "PV_ExpsTotal", 13, FMT_MONEY),
        ("PV net cashflow", "PV_NetCashflow", 13, FMT_MONEY),
        ("PV net cf (fwd check)", "PV_NetCashflowForCheck", 13, FMT_MONEY),
    ]
    LC = {key: get_column_letter(j) for j, (_h, key, _w, _f) in enumerate(pj_cols, start=1)}

    for sheet, risk, shock, sign in STRESS_SHEETS:
        pj = wb.create_sheet(sheet)
        stress_lbl = {"base": "base basis (no shock)",
                      "mort": "mortality stress (+15% rates from t0)",
                      "longev": "longevity stress (−25% rates from t0)",
                      "exps": "expense stress (+10% level, +1% inflation from t0)",
                      "lapse": {"up": "lapse up stress", "down": "lapse down stress",
                                "mass": "mass lapse stress (one-off at t0)"}.get(shock)}[risk]
        label(pj, 1, 1, f"Liability projection — {stress_lbl}. "
              "lifelib SCR_life.Projection[{!r}, {!r}, None]".format(risk if risk != "base" else "base",
                                                                     shock or None), bold=True)
        for j, (lbl, key) in enumerate([("Risk", "Risk"), ("Shock", "Shock"),
                                        ("Risk sign", ""), ("Shock factor", "Factor"),
                                        ("Shock limit", ""), ("Inflation shock", "")], start=2):
            c = pj.cell(row=2, column=j, value=lbl)
            c.font = Font(size=9, bold=True)
            if key:
                pj.cell(row=4, column=j, value=key).font = KEY_FONT
        pj.cell(row=3, column=2, value=risk if risk != "base" else "base").font = INPUT_FONT
        pj.cell(row=3, column=3, value=shock).font = INPUT_FONT
        pj.cell(row=3, column=4, value=sign).font = INPUT_FONT
        for col, formula in [
                (5, '=IFERROR(INDEX(FactorVals,MATCH($B$3&"|"&$C$3&"||",FactorKeys,0)),0)'),
                (6, '=IFERROR(INDEX(FactorVals,MATCH($B$3&"|"&$C$3&"||limit",FactorKeys,0)),0)'),
                (7, '=IFERROR(INDEX(FactorVals,MATCH($B$3&"|||inflation",FactorKeys,0)),0)')]:
            c = pj.cell(row=3, column=col, value=formula)
            c.font = LINK_FONT
            c.number_format = "0.0000"
        for j, (h, key, w, _f) in enumerate(pj_cols, start=1):
            hdr(pj, 5, j, h, w)
            pj.cell(row=6, column=j, value=key).font = KEY_FONT

        S = "Policy_Sched!"
        for t_ in range(MAX_T + 1):
            rr = FIRST + t_
            nx = rr + 1
            p = rr - 1
            g = f"$A{rr}>LastT"
            disc = f"{S}$K{rr}"
            f = {
                "t": t_ if t_ == 0 else f"=A{p}+1",
                "InflFactor": ("=1" if t_ == 0 else
                               f'=IF({g},0,IF($B$3="exps",B{p}*(1+InflRate'
                               f"+IF($A{rr}>=T0,$G$3,0)),B{p}/(1+InflRate)))"),
                "MortRateFactor": (f"=IF({g},0,IF(AND($D$3<>0,$A{rr}>=T0,"
                                   f"{S}$B{rr}<LastAge),1+$D$3*$E$3,1))"),
                "SurrRate": (f'=IF({g},0,IF($C$3="up",MIN({S}$E{rr}*(1+IF($A{rr}>=T0,$E$3,0)),$F$3),'
                             f'IF($C$3="down",MAX({S}$E{rr}*(1-IF($A{rr}>=T0,$E$3,0)),'
                             f"{S}$E{rr}-$F$3),{S}$E{rr})))"),
                "PolsIF_End": ("=0" if t_ == 0 else
                               f"=IF({g},0,{LC['PolsIF_Beg1']}{p}-{LC['PolsDeath']}{p}"
                               f"-{LC['PolsSurr']}{p})"),
                "PolsMaturity": (f"=IF({g},0,IF($A{rr}=PolicyTerm,"
                                 f"{LC['PolsIF_End']}{rr},0))"),
                "PolsIF_Beg": f"=IF({g},0,{LC['PolsIF_End']}{rr}-{LC['PolsMaturity']}{rr})",
                "PolsNewBiz": f"=IF($A{rr}=0,PolicyCount,0)",
                "PolsSurrMass": (f'=IF({g},0,IF(AND($C$3="mass",$A{rr}=T0),'
                                 f"({LC['PolsIF_Beg']}{rr}+{LC['PolsNewBiz']}{rr})*$E$3,0))"),
                "PolsIF_Beg1": (f"=IF({g},0,{LC['PolsIF_Beg']}{rr}+{LC['PolsNewBiz']}{rr}"
                                f"-{LC['PolsSurrMass']}{rr})"),
                "PolsDeath": (f"=IF({g},0,{LC['PolsIF_Beg1']}{rr}*{S}$C{rr}*{S}$D{rr}"
                              f"*{LC['MortRateFactor']}{rr})"),
                "PolsSurr": f"={LC['PolsIF_Beg1']}{rr}*{LC['SurrRate']}{rr}",
                "SizeExpsMaint": (f"=IF({g},0,(SizeAnnPrem*ExpsMaintAnnPrem"
                                  f"+(SumAssured*ExpsMaintSA+ExpsMaintPol)*{LC['InflFactor']}{rr})"
                                  f'*(1+IF($B$3="exps",$E$3,0)))'),
                "PremIncome": f"=SizePremium*{LC['PolsIF_Beg1']}{rr}",
                "BenefitDeath": f"=SumAssured*{LC['PolsDeath']}{rr}",
                "BenefitSurr": (f"={S}$I{rr}*({LC['PolsSurr']}{rr}"
                                f"+{LC['PolsSurrMass']}{rr})"),
                "BenefitTotal": f"={LC['BenefitDeath']}{rr}+{LC['BenefitSurr']}{rr}",
                "ExpsAcq": f"=IF($A{rr}=0,SizeExpsAcq0*{LC['PolsNewBiz']}{rr},0)",
                "ExpsCommInit": f"=IF($A{rr}=0,SizeExpsCommInit0*{LC['PolsIF_Beg1']}{rr},0)",
                "ExpsCommRen": f"={S}$J{rr}*{LC['PolsIF_Beg1']}{rr}",
                "ExpsMaint": f"={LC['SizeExpsMaint']}{rr}*{LC['PolsIF_Beg1']}{rr}",
                "ExpsTotal": (f"={LC['ExpsCommInit']}{rr}+{LC['ExpsCommRen']}{rr}"
                              f"+{LC['ExpsAcq']}{rr}+{LC['ExpsMaint']}{rr}"),
                "PV_PremIncome": (f"=IF({g},0,{LC['PremIncome']}{rr}"
                                  f"+{LC['PV_PremIncome']}{nx}/(1+{disc}))"),
                "PV_BenefitTotal": (f"=IF({g},0,(-{LC['BenefitTotal']}{rr}"
                                    f"+{LC['PV_BenefitTotal']}{nx})/(1+{disc}))"),
                "PV_ExpsTotal": (f"=IF({g},0,-{LC['ExpsTotal']}{rr}"
                                 f"+{LC['PV_ExpsTotal']}{nx}/(1+{disc}))"),
                "PV_NetCashflow": (f"={LC['PV_PremIncome']}{rr}+{LC['PV_ExpsTotal']}{rr}"
                                   f"+{LC['PV_BenefitTotal']}{rr}"),
                "PV_NetCashflowForCheck": (f"=IF({g},0,{LC['PremIncome']}{rr}"
                                           f"-{LC['ExpsTotal']}{rr}"
                                           f"-{LC['BenefitTotal']}{rr}/(1+{disc})"
                                           f"+{LC['PV_NetCashflow']}{nx}/(1+{disc}))"),
            }
            for (hh, key, w, fmt) in pj_cols:
                c = pj[f"{LC[key]}{rr}"]
                c.value = f[key]
                c.number_format = fmt
                cross = isinstance(f[key], str) and (
                    "Policy_Sched" in f[key] or "Size" in f[key].replace("SizeExpsMaint", "")
                    or any(nm2 in f[key] for nm2 in
                           ("LastT", "T0", "PolicyCount", "PolicyTerm", "SumAssured",
                            "InflRate", "LastAge", "SizePremium", "SizeAnnPrem",
                            "ExpsMaint", "SizeExpsAcq0", "SizeExpsCommInit0")))
                c.font = LINK_FONT if cross else CALC_FONT
        pj.freeze_panes = pj.cell(row=FIRST, column=2).coordinate
        for nm, key in [("PVNCF", "PV_NetCashflow"), ("PVCHK", "PV_NetCashflowForCheck"),
                        ("Beg1", "PolsIF_Beg1")]:
            col = LC[key]
            wb.defined_names.add(DefinedName(
                f"{nm}_{sheet.split('_', 1)[1]}",
                attr_text=f"{sheet}!${col}${FIRST}:${col}${LAST}"))

    # ---------------------------------------------------------- SCR
    scr = wb.create_sheet("SCR")
    scr.sheet_view.showGridLines = False
    scr.column_dimensions["B"].width = 30
    scr.column_dimensions["C"].width = 15
    scr.column_dimensions["D"].width = 66
    label(scr, 2, 2, "Net asset value by scenario — PV of net liability cashflows at t0", bold=True)
    hdr(scr, 3, 2, "Scenario"); hdr(scr, 3, 3, "NAV"); hdr(scr, 3, 4, "Notes / lifelib cell")
    scr.cell(row=4, column=2, value="NetAstValue").font = KEY_FONT
    nav_rows = [("Base", "Base", "NetAstValue() — Projection['base'].PV_NetCashflow(t0)"),
                ("Mortality stress", "Mort", "NetAstValue('mort')"),
                ("Longevity stress", "Longev", "NetAstValue('longev')"),
                ("Expense stress", "Exps", "NetAstValue('exps')"),
                ("Lapse up", "LapseUp", "NetAstValue('lapse','up')"),
                ("Lapse down", "LapseDown", "NetAstValue('lapse','down')"),
                ("Mass lapse", "LapseMass", "NetAstValue('lapse','mass')")]
    r = 5
    for lbl, sfx, note in nav_rows:
        label(scr, r, 2, lbl)
        c = scr.cell(row=r, column=3, value=f"=INDEX(PVNCF_{sfx},T0+1)")
        c.font = LINK_FONT
        c.number_format = FMT_MONEY
        c.border = BORDER
        scr.cell(row=r, column=4, value=note).font = NOTE_FONT
        name_cell(wb, f"NAV_{sfx}", "SCR", f"$C${r}")
        r += 1
    r += 1
    label(scr, r, 2, "Lapse risk by shock", bold=True)
    hdr(scr, r + 1, 2, "Shock"); hdr(scr, r + 1, 3, "Charge"); hdr(scr, r + 1, 4, "Notes / lifelib cell")
    scr.cell(row=r + 2, column=2, value="LapseRisk").font = KEY_FONT
    r += 2
    for lbl, nm, f in [("Up", "LapseUp_Risk", "=MAX(NAV_Base-NAV_LapseUp,0)"),
                       ("Down", "LapseDown_Risk", "=MAX(NAV_Base-NAV_LapseDown,0)"),
                       ("Mass", "LapseMass_Risk",
                        "=MAX(NAV_Base-NAV_LapseMass,0)")]:
        label(scr, r, 2, lbl)
        c = scr.cell(row=r, column=3, value=f)
        c.font = CALC_FONT
        c.number_format = FMT_MONEY
        c.border = BORDER
        note = ("LapseRisk(shock); mass: retail_share = 1 in lifelib, so the non-retail "
                "leg vanishes" if lbl == "Mass" else "LapseRisk(shock)")
        scr.cell(row=r, column=4, value=note).font = NOTE_FONT
        name_cell(wb, nm, "SCR", f"$C${r}")
        r += 1
    r += 1
    label(scr, r, 2, "Life underwriting risk charges", bold=True)
    hdr(scr, r + 1, 2, "Risk"); hdr(scr, r + 1, 3, "Life(risk)"); hdr(scr, r + 1, 4, "Notes / lifelib cell")
    scr.cell(row=r + 2, column=2, value="Life").font = KEY_FONT
    r += 2
    life_first = r
    life_rows = [
        ("mort", "=MAX(NAV_Base-NAV_Mort,0)", "Life('mort')"),
        ("longev", "=MAX(NAV_Base-NAV_Longev,0)", "Life('longev')"),
        ("disab", 0, "Life('disab') = 0 — no lifelib override: the stressed projection "
                     "is identical to base by construction"),
        ("lapse", "=MAX(LapseUp_Risk,LapseDown_Risk,LapseMass_Risk)", "Life('lapse')"),
        ("exps", "=MAX(NAV_Base-NAV_Exps,0)", "Life('exps')"),
        ("rev", 0, "Life('rev') = 0 — no lifelib override"),
        ("cat", 0, "Life('cat') = 0 — no lifelib override"),
    ]
    for risk, f, note in life_rows:
        label(scr, r, 2, risk)
        c = scr.cell(row=r, column=3, value=f)
        c.font = CALC_FONT if isinstance(f, str) else INPUT_FONT
        c.number_format = FMT_MONEY
        c.border = BORDER
        scr.cell(row=r, column=4, value=note).font = NOTE_FONT
        name_cell(wb, f"Life_{risk.capitalize()}", "SCR", f"$C${r}")
        r += 1
    wb.defined_names.add(DefinedName(
        "LifeVec", attr_text=f"SCR!$C${life_first}:$C${life_first + 6}"))
    r += 1
    label(scr, r, 2, "Correlation matrix (standard formula, lifelib CorrData)", bold=True)
    corr_first = r + 3
    scr.cell(row=r + 1, column=2, value="Corr").font = KEY_FONT
    hdr(scr, r + 2, 1, "")
    for j, risk in enumerate(RISKS):
        hdr(scr, r + 2, 2 + j, risk, 8)
        hdr(scr, corr_first + j, 1, risk)
    for i, ri in enumerate(RISKS):
        for j, rj in enumerate(RISKS):
            c = scr.cell(row=corr_first + i, column=2 + j, value=data["corr"][(ri, rj)])
            c.font = INPUT_FONT
            c.number_format = "0.00"
            c.border = BORDER
    lastc = get_column_letter(1 + len(RISKS))
    wb.defined_names.add(DefinedName(
        "CorrGrid", attr_text=f"SCR!$B${corr_first}:${lastc}${corr_first + 6}"))
    r = corr_first + 7 + 1
    label(scr, r, 2, "Correlated products Life(i)·Life(j)·Corr(i,j)", bold=True)
    prod_first = r + 3
    scr.cell(row=r + 1, column=2, value="SCR_life").font = KEY_FONT
    for j, risk in enumerate(RISKS):
        hdr(scr, r + 2, 2 + j, risk, 8)
        hdr(scr, prod_first + j, 1, risk)
    for i in range(7):
        for j in range(7):
            c = scr.cell(row=prod_first + i, column=2 + j,
                         value=f"=INDEX(LifeVec,{i + 1})*INDEX(LifeVec,{j + 1})"
                               f"*INDEX(CorrGrid,{i + 1},{j + 1})")
            c.font = CALC_FONT
            c.number_format = "#,##0"
            c.border = BORDER
    wb.defined_names.add(DefinedName(
        "ProdGrid", attr_text=f"SCR!$B${prod_first}:${lastc}${prod_first + 6}"))
    r = prod_first + 7 + 1
    label(scr, r, 2, "SCR life underwriting risk", bold=True)
    c = scr.cell(row=r, column=3, value="=SQRT(SUM(ProdGrid))")
    c.font = Font(name="Calibri", size=11, bold=True, color="006100")
    c.number_format = FMT_MONEY
    c.border = BORDER
    scr.cell(row=r, column=4,
             value="SCR_life() = sqrt( sum_ij Corr(i,j)·Life(i)·Life(j) )").font = NOTE_FONT
    name_cell(wb, "SCR_Life", "SCR", f"$C${r}")

    # ---------------------------------------------------------- Summary
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["B"].width = 32
    sm.column_dimensions["C"].width = 16
    sm.column_dimensions["D"].width = 56
    label(sm, 2, 2, "Run", bold=True)
    hdr(sm, 3, 2, "Item"); hdr(sm, 3, 3, "Value"); hdr(sm, 3, 4, "")
    item_row(sm, 4, "Policy ID", "=PolicyID", "calculation sentinel for the batch macro",
             FMT_INT, link=True)
    item_row(sm, 5, "Scenario", "=ScenID", "", FMT_INT, link=True)
    item_row(sm, 6, "Valuation time t0", "=T0", "", FMT_INT, link=True)
    item_row(sm, 7, "Product / sex / age / term",
             '=Product&" / "&Sex&" / "&IssueAge&" / "&PolicyTerm&"y"', "", None, link=True)
    label(sm, 9, 2, "SCR results", bold=True)
    hdr(sm, 10, 2, "Item"); hdr(sm, 10, 3, "Value"); hdr(sm, 10, 4, "")
    item_row(sm, 11, "SCR life underwriting risk", "=SCR_Life", "", FMT_MONEY, link=True)
    item_row(sm, 12, "Mortality risk", "=Life_Mort", "", FMT_MONEY, link=True)
    item_row(sm, 13, "Longevity risk", "=Life_Longev", "", FMT_MONEY, link=True)
    item_row(sm, 14, "Lapse risk", "=Life_Lapse", "", FMT_MONEY, link=True)
    item_row(sm, 15, "Expense risk", "=Life_Exps", "", FMT_MONEY, link=True)
    item_row(sm, 16, "Sum of undiversified charges", "=SUM(LifeVec)", "", FMT_MONEY, link=True)
    item_row(sm, 17, "Diversification benefit", "=SUM(LifeVec)-SCR_Life", "", FMT_MONEY, link=True)
    item_row(sm, 18, "Base net asset value", "=NAV_Base", "PV net liability cashflows at t0",
             FMT_MONEY, link=True)

    chart = RadarChart()
    chart.type = "filled"
    chart.title = "Life SCR risk profile"
    chart.height = 9
    chart.width = 12
    chart.add_data(Reference(scr, min_col=3, min_row=life_first, max_row=life_first + 6),
                   titles_from_data=False)
    chart.set_categories(Reference(scr, min_col=2, min_row=life_first, max_row=life_first + 6))
    chart.legend = None
    sm.add_chart(chart, "F3")

    # ---------------------------------------------------------- Checks
    ck = wb.create_sheet("Checks")
    ck.sheet_view.showGridLines = False
    ck.column_dimensions["B"].width = 48
    ck.column_dimensions["C"].width = 10
    ck.column_dimensions["D"].width = 78
    label(ck, 2, 2, "Internal consistency checks", bold=True)
    hdr(ck, 3, 2, "Check"); hdr(ck, 3, 3, "Result"); hdr(ck, 3, 4, "Description")
    checks = []
    for sheet, *_rest in STRESS_SHEETS:
        sfx = sheet.split("_", 1)[1]
        checks.append((
            f"PV identity — {sheet}",
            f"=ABS(INDEX(PVNCF_{sfx},T0+1)-INDEX(PVCHK_{sfx},T0+1))"
            f"<=0.000001*MAX(1,ABS(INDEX(PVNCF_{sfx},T0+1)))",
            "lifelib PV_Check: the backward PV recursion equals the forward "
            "check recursion at t0.", False))
    for sheet, *_rest in STRESS_SHEETS:
        sfx = sheet.split("_", 1)[1]
        checks.append((
            f"In-force non-negative — {sheet}",
            f"=MIN(Beg1_{sfx})>=-0.000000001",
            "Policies in force never go negative.", False))
    checks += [
        ("Decrements exhaust in-force — base",
         "=ABS(PolicyCount-SUM(Proj_Base!K7:K110)-SUM(Proj_Base!L7:L110)"
         "-SUM(Proj_Base!F7:F110)-SUM(Proj_Base!I7:I110)"
         "-(INDEX(Beg1_Base,LastT+1)-INDEX(Proj_Base!K7:K110,LastT+1)"
         "-INDEX(Proj_Base!L7:L110,LastT+1)))<=0.000001",
         "New business equals deaths + surrenders + maturities + mass lapses "
         "+ survivors at the end of the projection.", False),
        ("Decrements exhaust in-force — mass lapse",
         "=ABS(PolicyCount-SUM(Proj_LapseMass!K7:K110)-SUM(Proj_LapseMass!L7:L110)"
         "-SUM(Proj_LapseMass!F7:F110)-SUM(Proj_LapseMass!I7:I110)"
         "-(INDEX(Beg1_LapseMass,LastT+1)-INDEX(Proj_LapseMass!K7:K110,LastT+1)"
         "-INDEX(Proj_LapseMass!L7:L110,LastT+1)))<=0.000001",
         "Same identity under the mass-lapse stress (exercises PolsSurrMass).", False),
        ("Net level premium reserve zero at issue",
         "=ABS(INDEX(Policy_Sched!F7:F110,1))<=0.00000001",
         "ReserveNLP_Rate(PREM, 0) = 0 by the equivalence principle when the "
         "premium term equals the policy term.", False),
        ("Radix and survivorship",
         "=AND(INDEX(Comm_lx,1)=100000,MIN(Comm_lx)>=0)",
         "Commutation lx starts at the 100,000 radix and never goes negative.", False),
        ("Risk charges non-negative", "=MIN(LifeVec)>=0",
         "Life(risk) = max(·, 0) for every risk.", False),
        ("Diversification", "=SCR_Life<=SUM(LifeVec)+0.000001",
         "The correlated SCR never exceeds the sum of standalone charges.", False),
        ("Correlated product grid symmetric",
         "=SUMPRODUCT(--(ABS(ProdGrid-TRANSPOSE(ProdGrid))"
         ">0.000001*(1+ABS(ProdGrid))))=0",
         "Life(i)·Life(j)·Corr(i,j) grid is symmetric (the correlation matrix is).", True),
        ("Shock factors resolved",
         "=AND(Proj_Mort!$E$3>0,Proj_Longev!$E$3>0,Proj_Exps!$E$3>0,"
         "Proj_Exps!$G$3>0,Proj_LapseUp!$E$3>0,Proj_LapseUp!$F$3>0,"
         "Proj_LapseDown!$E$3>0,Proj_LapseDown!$F$3>0,Proj_LapseMass!$E$3>0)",
         "Every stress sheet found its factor in the FactorData table — the "
         "IFERROR default of 0 would silently degrade a stress to base.", False),
    ]
    for i, (name, formula, desc, cse) in enumerate(checks):
        rr = 4 + i
        label(ck, rr, 2, name)
        c = ck.cell(row=rr, column=3)
        c.value = ArrayFormula(f"C{rr}", formula) if cse else formula
        c.font = LINK_FONT
        c.border = BORDER
        c.alignment = CENTER
        ck.cell(row=rr, column=4, value=desc).font = Font(size=9, color="7F7F7F")
    rall = 4 + len(checks) + 1
    label(ck, rall, 2, "ALL CHECKS PASS", bold=True)
    c = ck.cell(row=rall, column=3, value=f"=AND(C4:C{3 + len(checks)})")
    c.font = Font(name="Calibri", size=10, bold=True, color="006100")
    c.alignment = CENTER
    name_cell(wb, "AllChecksPass", "Checks", f"$C${rall}")

    # ---------------------------------------------------------- Lifelib_Reference
    lr = wb.create_sheet("Lifelib_Reference")
    label(lr, 1, 1, "lifelib solvency2 results for all policies (t0 = 0, scenario 1). Generated "
          "by verification/generate_reference_solvency2.py.", bold=True)
    ref_cols = ["policy_id", "scr_life", "life_mort", "life_longev", "life_lapse",
                "life_exps", "lapse_up", "lapse_down", "lapse_mass", "nav_base"]
    for j, h in enumerate(ref_cols, start=1):
        hdr(lr, 2, j, h, 13)
    if REF_CSV.exists():
        import csv as _csv
        with REF_CSV.open() as fcsv:
            rows = list(_csv.DictReader(fcsv))
        for i, row in enumerate(rows):
            lr.cell(row=3 + i, column=1, value=int(row["policy_id"])).font = INPUT_FONT
            for j, cn in enumerate(ref_cols[1:], start=2):
                c = lr.cell(row=3 + i, column=j, value=float(row[cn]))
                c.font = INPUT_FONT
                c.number_format = FMT_MONEY
    else:
        print("WARNING: reference CSV missing - Lifelib_Reference left empty")
    lr.freeze_panes = "A3"

    # ---------------------------------------------------------- Batch_Results
    br = wb.create_sheet("Batch_Results")
    br.sheet_view.showGridLines = False
    br.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGHIJKLMN", (34, 15, 15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 11)):
        br.column_dimensions[col].width = w
    label(br, 1, 2, "All policies — SCR batch run and reconciliation vs lifelib", bold=True)
    br.cell(row=2, column=2, value="Filled by the RunAllPolicies macro (t0 = 0, scenario 1; "
            "see Control for settings).").font = NOTE_FONT
    for i, txt in enumerate(["Policies run", "Reconciled within tolerance", "Mismatches",
                             "Max rel diff SCR", "Max rel diff Life charges",
                             "Max rel diff NAV", "Run time (seconds)",
                             "RECONCILIATION RESULT"]):
        label(br, 4 + i, 2, txt, bold=(i == 7))
        br.cell(row=4 + i, column=3).border = BORDER
    hdrs = ["policy_id", "SCR (wb)", "SCR (lifelib)", "rel diff SCR", "rel mort",
            "rel longev", "rel lapse", "rel lapse up", "rel lapse down",
            "rel lapse mass", "rel exps", "rel NAV", "reconciles"]
    for j, h in enumerate(hdrs, start=2):
        hdr(br, 13, j, h)
    br.freeze_panes = "A14"

    # codeName plumbing for the VBA project (see marinade-notes P5)
    wb.code_name = "ThisWorkbook"
    for i, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_properties.codeName = f"Sheet{i}"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    if VBA_BIN.exists():
        make_xlsm(OUT_XLSX, OUT_XLSM, VBA_BIN)
        print(f"wrote {OUT_XLSM} (macro project injected)")
    else:
        print("vbaProject.bin not found - .xlsm not produced. "
              "Bootstrap it once via workbooks/Solvency2/bootstrap_vba.py")


def make_xlsm(src_xlsx: Path, dst_xlsm: Path, vba_bin: Path):
    """Produce an .xlsm from the .xlsx by injecting vbaProject.bin (zip surgery)."""
    shutil.copy(src_xlsx, dst_xlsm)
    with zipfile.ZipFile(src_xlsx) as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}
    ct = parts["[Content_Types].xml"].decode()
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    if "vnd.ms-office.vbaProject" not in ct:
        ct = ct.replace(
            "</Types>",
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
        )
    parts["[Content_Types].xml"] = ct.encode()
    rels = parts["xl/_rels/workbook.xml.rels"].decode()
    if "vbaProject" not in rels:
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdVBA" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>',
        )
    parts["xl/_rels/workbook.xml.rels"] = rels.encode()
    parts["xl/vbaProject.bin"] = vba_bin.read_bytes()
    with zipfile.ZipFile(dst_xlsm, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, dd in parts.items():
            zout.writestr(name, dd)


if __name__ == "__main__":
    build()
