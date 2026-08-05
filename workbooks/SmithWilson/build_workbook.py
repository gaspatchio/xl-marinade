"""Build the SmithWilson workbook — an Excel twin of lifelib's smithwilson model.

Smith-Wilson extrapolation of risk-free spot rates (CEIOPS/EIOPA QIS5
methodology): fit zeta parameters to N=25 observed EUR spot rates by
inverting the Wilson-function matrix, then extrapolate bond prices and
spot rates to 65 years converging to the ultimate forward rate (UFR).

The matrix algebra is deliberately kept in native Excel array formulas
(MINVERSE/MMULT entered as CSE arrays) — this workbook is the array-formula
test case for XL Marinade. Conventions:
  blue font  = hardcoded input
  black font = in-sheet formula
  green font = formula pulling from another sheet

Output: SmithWilson.xlsm when vbaProject.bin is present next to this script
(zip surgery injects the macro project); SmithWilson.xlsx otherwise.
"""

import json
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

REPO = Path(__file__).resolve().parents[2]
REF_JSON = REPO / "verification" / "reference" / "smithwilson.json"
HERE = Path(__file__).resolve().parent
VBA_BIN = HERE / "vbaProject.bin"
OUT_XLSX = HERE / "SmithWilson.xlsx"
OUT_XLSM = HERE / "SmithWilson.xlsm"

N_OBS = 25   # observed maturities (lifelib N)
T_MAX = 65   # extrapolation horizon in years

# ---------------------------------------------------------------- styles
NAVY = "1F3864"
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")
CALC_FONT = Font(name="Calibri", size=10, color="000000")
LINK_FONT = Font(name="Calibri", size=10, color="006100")  # cross-sheet
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color=NAVY)
KEY_FONT = Font(size=8, italic=True, color="7F7F7F")
NOTE_FONT = Font(size=9, italic=True, color="7F7F7F")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")

FMT_RATE = "0.0000%"
FMT_PRICE = "0.00000000"
FMT_ZETA = "0.000000"
FMT_INT = "0"


def hdr(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def label(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=bold)
    return c


def name_cell(wb, name, sheet, addr):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{addr}"))


def item_row(ws, r, item, value, note, fmt=None, formula=False, link=False):
    label(ws, r, 2, item)
    c = ws.cell(row=r, column=3, value=value)
    c.font = LINK_FONT if link else (CALC_FONT if formula else INPUT_FONT)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = NOTE_FONT
    return c


# ---------------------------------------------------------------- build
def build():
    ref = json.loads(REF_JSON.read_text())
    spot_rates = ref["spot_rates"]
    ufr = ref["params"]["UFR"]
    alpha = ref["params"]["alpha"]
    assert ref["params"]["N"] == N_OBS and ref["params"]["T_MAX"] == T_MAX

    wb = Workbook()

    # ---------------------------------------------------------- Cover
    cv = wb.active
    cv.title = "Cover"
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["B"].width = 28
    cv.column_dimensions["C"].width = 78
    t = cv.cell(row=2, column=2, value="SmithWilson — Risk-Free Yield Curve Extrapolation")
    t.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    cv.cell(row=3, column=2, value="Smith-Wilson extrapolation of CHF risk-free spot rates to the "
            "ultimate forward rate (Solvency II methodology per the CEIOPS/EIOPA QIS5 paper)"
            ).font = Font(size=11, italic=True)

    rows = [
        ("Purpose", "Fits the Smith-Wilson pricing function to 25 observed CHF spot rates and "
                    "extrapolates zero-coupon bond prices and annual-compound spot rates to 65 years, "
                    "converging to the ultimate forward rate (UFR). The workbook is a formula-for-"
                    "formula replication of the open-source lifelib model 'smithwilson' (MIT licence) "
                    "and is verified cell-by-cell against it."),
        ("Method", "CEIOPS/EIOPA, 'QIS 5 Risk-free interest rates — Extrapolation method' (2010). "
                   "Observed bond prices m(u) are fitted exactly by P(t) = e^(−UFR·t) + Σ ζj·W(t,uj); "
                   "the ζ vector solves the linear system W·ζ = m − μ, inverted on the Calibration "
                   "sheet with native Excel MINVERSE/MMULT array formulas."),
        ("Inputs", "25 observed spot rates (annual compound; Switzerland EIOPA risk-free term "
                   "structure as at 31 May 2019, RFR_spot_no_VA — negative short end) on "
                   "Spot_Rates; UFR (continuous compound) and convergence parameter α on "
                   "Assumptions."),
        ("Reconciliation", "The RunCurveReconciliation macro (Batch_Results sheet) recalculates the "
                           "workbook and reconciles the full ζ vector and the extrapolated P and R "
                           "curves against the lifelib values stored on Lifelib_Reference."),
        ("Source model", "lifelib smithwilson — https://lifelib.io (MIT licence)"),
        ("Verification", "verification/verify_smithwilson.py compares every mapped cell (including "
                         "the full 65×25 Wilson matrix) against the Python model; see the Checks "
                         "sheet for in-workbook consistency tests."),
    ]
    r = 5
    for k, v in rows:
        label(cv, r, 2, k, bold=True)
        c = cv.cell(row=r, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cv.row_dimensions[r].height = 42
        r += 1

    r += 1
    label(cv, r, 2, "Version control", bold=True)
    for i, (a, b) in enumerate([("Version", "0.2.0"), ("Date", "2026-07-14"),
                                ("Prepared by", "XL Marinade project"),
                                ("Review status", "Verified against lifelib; adversarial "
                                 "actuarial audit remediated (see verification/reports)")]):
        label(cv, r + 1 + i, 2, a)
        cv.cell(row=r + 1 + i, column=3, value=b).font = Font(size=10)

    r += 6
    label(cv, r, 2, "Sheet index", bold=True)
    for i, (s, d) in enumerate([
            ("Control", "Run control — reporting maturity and reconciliation settings"),
            ("Assumptions", "UFR, convergence parameter α, curve dimensions"),
            ("Spot_Rates", "Observed EUR spot rates by maturity (input data)"),
            ("Wilson", "Wilson function W(t, uj) — 65 × 25 grid"),
            ("Calibration", "Observed bond prices and the ζ vector (MINVERSE/MMULT array solve)"),
            ("Extrapolation", "Extrapolated bond prices P(t) and spot rates R(t), t = 1 … 65"),
            ("Summary", "Curve at key durations, convergence to UFR, chart"),
            ("Checks", "Internal consistency checks (all must be TRUE)"),
            ("Lifelib_Reference", "lifelib ζ, P and R values (reconciliation target)"),
            ("Batch_Results", "Full-curve reconciliation vs lifelib (macro output)")]):
        label(cv, r + 1 + i, 2, s)
        cv.cell(row=r + 1 + i, column=3, value=d).font = Font(size=10)

    r += 12
    label(cv, r, 2, "Conventions", bold=True)
    for i, (txt, font) in enumerate([
            ("Blue — hardcoded input", INPUT_FONT),
            ("Black — calculated on this sheet", CALC_FONT),
            ("Green — link from another sheet", LINK_FONT)]):
        cv.cell(row=r + 1 + i, column=2, value=txt).font = font

    # ---------------------------------------------------------- Control
    ct = wb.create_sheet("Control")
    ct.sheet_view.showGridLines = False
    ct.column_dimensions["B"].width = 34
    ct.column_dimensions["C"].width = 14
    ct.column_dimensions["D"].width = 62

    label(ct, 2, 2, "Reporting", bold=True)
    hdr(ct, 3, 2, "Item"); hdr(ct, 3, 3, "Value"); hdr(ct, 3, 4, "Notes")
    item_row(ct, 4, "Reporting maturity (years)", 30,
             "Maturity highlighted on the Summary sheet — choose 1 to 65", FMT_INT)
    name_cell(wb, "SelMaturity", "Control", "$C$4")
    dv = DataValidation(type="whole", operator="between", formula1="1",
                        formula2=str(T_MAX), allow_blank=False)
    dv.error = f"Reporting maturity must be between 1 and {T_MAX}"
    ct.add_data_validation(dv)
    dv.add(ct["C4"])

    label(ct, 6, 2, "Reconciliation (RunCurveReconciliation macro)", bold=True)
    hdr(ct, 7, 2, "Item"); hdr(ct, 7, 3, "Value"); hdr(ct, 7, 4, "Notes")
    item_row(ct, 8, "Reconciliation tolerance (relative)", 0.000001,
             "A value reconciles when |workbook − lifelib| ≤ tol × max(1, |lifelib|)", "0.0000000")
    name_cell(wb, "BatchTol", "Control", "$C$8")
    ct.cell(row=10, column=2, value="Run the reconciliation from Developer ▸ Macros ▸ "
            "RunCurveReconciliation; results appear on Batch_Results.").font = NOTE_FONT

    # ---------------------------------------------------------- Assumptions
    asm = wb.create_sheet("Assumptions")
    asm.sheet_view.showGridLines = False
    asm.column_dimensions["B"].width = 34
    asm.column_dimensions["C"].width = 16
    asm.column_dimensions["D"].width = 62

    label(asm, 2, 2, "Smith-Wilson parameters", bold=True)
    hdr(asm, 3, 2, "Item"); hdr(asm, 3, 3, "Value"); hdr(asm, 3, 4, "Notes / lifelib cell")
    item_row(asm, 4, "Ultimate forward rate (continuous)", ufr,
             "UFR — ln(1.029); the 2019 EIOPA UFR for CHF", "0.000000000000")
    item_row(asm, 5, "Ultimate forward rate (annual compound)", "=EXP(UFR)-1",
             "e^UFR − 1; the long-end anchor of the extrapolated curve", FMT_RATE, formula=True)
    item_row(asm, 6, "Convergence parameter α", alpha,
             "alpha — speed of convergence towards the UFR", "0.000000")
    item_row(asm, 7, "Observed maturities N", N_OBS,
             "N — number of observed spot rates (Spot_Rates sheet)", FMT_INT)
    item_row(asm, 8, "Extrapolation horizon (years)", T_MAX,
             "Length of the Wilson and Extrapolation grids (structural)", FMT_INT)
    for nm, addr in [("UFR", "$C$4"), ("UFR_Annual", "$C$5"), ("Alpha", "$C$6"),
                     ("N_Obs", "$C$7"), ("T_Max", "$C$8")]:
        name_cell(wb, nm, "Assumptions", addr)

    # ---------------------------------------------------------- Spot_Rates
    sr = wb.create_sheet("Spot_Rates")
    sr.sheet_view.showGridLines = False
    label(sr, 1, 1, "Observed CHF spot rates (annual compound) — Switzerland EIOPA risk-free "
          "term structure as at 2019-05-31 (RFR_spot_no_VA), shipped with lifelib", bold=True)
    hdr(sr, 2, 1, "i", 6)
    hdr(sr, 2, 2, "Maturity (years)", 14)
    hdr(sr, 2, 3, "Spot rate p.a.", 12)
    for j, key in enumerate(["", "u", "spot_rates"], start=1):
        sr.cell(row=3, column=j, value=key).font = KEY_FONT
    for i, rate in enumerate(spot_rates, start=1):
        rr = 3 + i
        sr.cell(row=rr, column=1, value=i).font = INPUT_FONT
        sr.cell(row=rr, column=2, value=i).font = INPUT_FONT
        c = sr.cell(row=rr, column=3, value=float(rate))
        c.font = INPUT_FONT
        c.number_format = FMT_RATE
    wb.defined_names.add(DefinedName("SpotRates", attr_text=f"Spot_Rates!$C$4:$C${3 + N_OBS}"))
    wb.defined_names.add(DefinedName("SpotMaturities", attr_text=f"Spot_Rates!$B$4:$B${3 + N_OBS}"))

    # ---------------------------------------------------------- Wilson
    wsn = wb.create_sheet("Wilson")
    label(wsn, 1, 1, "Wilson functions W(t, uj) — formula (2), QIS5 technical paper; "
          "rows t = 1 … 65, columns uj = observed maturities 1 … 25", bold=True)
    hdr(wsn, 2, 1, "t", 6)
    for j in range(1, N_OBS + 1):
        c = hdr(wsn, 2, 1 + j, j, 11)
        c.number_format = '"u="0'
    wsn.cell(row=3, column=1, value="u").font = KEY_FONT
    for j in range(1, N_OBS + 1):
        wsn.cell(row=3, column=1 + j, value="W").font = KEY_FONT
    first = 4
    for i in range(1, T_MAX + 1):
        rr = first + i - 1
        c = wsn.cell(row=rr, column=1, value=1 if i == 1 else f"=A{rr - 1}+1")
        c.font = CALC_FONT
        c.number_format = FMT_INT
        for j in range(1, N_OBS + 1):
            col = get_column_letter(1 + j)
            f = (f"=EXP(-UFR*($A{rr}+{col}$2))*(Alpha*MIN($A{rr},{col}$2)"
                 f"-0.5*EXP(-Alpha*MAX($A{rr},{col}$2))"
                 f"*(EXP(Alpha*MIN($A{rr},{col}$2))-EXP(-Alpha*MIN($A{rr},{col}$2))))")
            c = wsn.cell(row=rr, column=1 + j, value=f)
            c.font = LINK_FONT
            c.number_format = FMT_PRICE
    wsn.freeze_panes = "B4"
    last_col = get_column_letter(1 + N_OBS)
    wb.defined_names.add(DefinedName(
        "WMatrix", attr_text=f"Wilson!$B${first}:${last_col}${first + N_OBS - 1}"))
    wb.defined_names.add(DefinedName(
        "WGrid", attr_text=f"Wilson!$B${first}:${last_col}${first + T_MAX - 1}"))

    # ---------------------------------------------------------- Calibration
    cb = wb.create_sheet("Calibration")
    cb.sheet_view.showGridLines = False
    label(cb, 1, 1, "Calibration of the ζ vector: ζ = W⁻¹ (m − μ), solved with a native "
          "MINVERSE/MMULT array formula (formula (5), QIS5 technical paper)", bold=True)
    cal_cols = [
        ("j", "j", 6, FMT_INT),
        ("Maturity u_j", "u", 12, FMT_INT),
        ("Spot rate p.a.", "spot_rates", 12, FMT_RATE),
        ("Bond price m_j", "m", 14, FMT_PRICE),
        ("UFR price μ_j", "mu", 14, FMT_PRICE),
        ("m_j − μ_j", "", 14, FMT_PRICE),
        ("ζ_j", "zeta", 14, FMT_ZETA),
    ]
    for j, (title, key, w, _fmt) in enumerate(cal_cols, start=1):
        hdr(cb, 2, j, title, w)
        cb.cell(row=3, column=j, value=key).font = KEY_FONT
    first = 4
    last = first + N_OBS - 1
    for i in range(1, N_OBS + 1):
        rr = first + i - 1
        vals = {
            "A": (1 if i == 1 else f"=A{rr - 1}+1", CALC_FONT),
            "B": (f"=INDEX(SpotMaturities,A{rr})", LINK_FONT),
            "C": (f"=INDEX(SpotRates,A{rr})", LINK_FONT),
            "D": (f"=(1+C{rr})^(-B{rr})", CALC_FONT),
            "E": (f"=EXP(-UFR*B{rr})", LINK_FONT),
            "F": (f"=D{rr}-E{rr}", CALC_FONT),
        }
        for col, (v, font) in vals.items():
            c = cb[f"{col}{rr}"]
            c.value = v
            c.font = font
    cb[f"G{first}"] = ArrayFormula(f"G{first}:G{last}", "=MMULT(MINVERSE(WMatrix),DVec)")
    for i in range(first, last + 1):
        for col, (_t, _k, _w, fmt) in zip("ABCDEFG", cal_cols):
            cb[f"{col}{i}"].number_format = fmt
        cb[f"G{i}"].font = LINK_FONT
    for nm, col in [("MVec", "D"), ("MuVecObs", "E"), ("DVec", "F"), ("ZetaVec", "G")]:
        wb.defined_names.add(DefinedName(
            nm, attr_text=f"Calibration!${col}${first}:${col}${last}"))

    # ---------------------------------------------------------- Extrapolation
    ex = wb.create_sheet("Extrapolation")
    ex.sheet_view.showGridLines = False
    label(ex, 1, 1, "Extrapolated curve: P(t) = μ(t) + Σj ζj·W(t,uj) (formula (6)); "
          "R(t) = (1/P(t))^(1/t) − 1", bold=True)
    ext_cols = [
        ("Time t (years)", "u", 12, FMT_INT),
        ("UFR price μ(t)", "mu", 14, FMT_PRICE),
        ("Bond price P(t)", "P", 14, FMT_PRICE),
        ("Spot rate R(t)", "R", 12, FMT_RATE),
        ("Observed spot rate", "spot_rates", 14, FMT_RATE),
    ]
    for j, (title, key, w, _fmt) in enumerate(ext_cols, start=1):
        hdr(ex, 2, j, title, w)
        ex.cell(row=3, column=j, value=key).font = KEY_FONT
    first = 4
    last = first + T_MAX - 1
    for i in range(1, T_MAX + 1):
        rr = first + i - 1
        vals = {
            "A": (1 if i == 1 else f"=A{rr - 1}+1", CALC_FONT),
            "B": (f"=EXP(-UFR*A{rr})", LINK_FONT),
            # CSE array formula: openpyxl-written plain formulas get legacy
            # semantics, and Excel's implicit intersection (@) would collapse
            # ZetaVec to a scalar and return #VALUE!.
            "C": (ArrayFormula(f"C{rr}",
                               f"=B{rr}+MMULT(INDEX(WGrid,A{rr},0),ZetaVec)"), LINK_FONT),
            "D": (f"=(1/C{rr})^(1/A{rr})-1", CALC_FONT),
            "E": (f'=IF(A{rr}<=N_Obs,INDEX(SpotRates,A{rr}),"")', LINK_FONT),
        }
        for col, (v, font) in vals.items():
            c = ex[f"{col}{rr}"]
            c.value = v
            c.font = font
            c.number_format = dict(zip("ABCDE", (f[3] for f in ext_cols)))[col]
    ex.freeze_panes = "A4"
    for nm, col, lo, hi in [("Ext_T", "A", first, last), ("Ext_Mu", "B", first, last),
                            ("Ext_P", "C", first, last), ("Ext_R", "D", first, last),
                            ("P_Fitted_Obs", "C", first, first + N_OBS - 1),
                            ("R_Fitted_Obs", "D", first, first + N_OBS - 1)]:
        wb.defined_names.add(DefinedName(
            nm, attr_text=f"Extrapolation!${col}${lo}:${col}${hi}"))

    # ---------------------------------------------------------- Summary
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["B"].width = 34
    sm.column_dimensions["C"].width = 14
    sm.column_dimensions["D"].width = 14
    sm.column_dimensions["E"].width = 14

    label(sm, 2, 2, "Curve at the reporting maturity", bold=True)
    hdr(sm, 3, 2, "Item"); hdr(sm, 3, 3, "Value"); hdr(sm, 3, 4, "")
    item_row(sm, 4, "Reporting maturity (years)", "=SelMaturity", "", FMT_INT, link=True)
    item_row(sm, 5, "Bond price P(t)", "=INDEX(Ext_P,SelMaturity)", "", FMT_PRICE, link=True)
    item_row(sm, 6, "Spot rate R(t)", "=INDEX(Ext_R,SelMaturity)", "", FMT_RATE, link=True)
    item_row(sm, 7, "UFR (annual compound)", "=UFR_Annual", "", FMT_RATE, link=True)
    item_row(sm, 8, "Gap to UFR", "=UFR_Annual-INDEX(Ext_R,SelMaturity)",
             "Convergence towards the UFR is gradual at the calibrated α (see Assumptions)",
             FMT_RATE, link=True)

    label(sm, 10, 2, "Curve at key durations", bold=True)
    hdr(sm, 11, 2, "Maturity"); hdr(sm, 11, 3, "Fitted R(t)")
    hdr(sm, 11, 4, "Observed"); hdr(sm, 11, 5, "Bond price P(t)")
    key_ts = [1, 5, 10, 15, 20, 25, 30, 40, 50, 60, 65]
    for i, kt in enumerate(key_ts):
        rr = 12 + i
        c = sm.cell(row=rr, column=2, value=kt)
        c.font = INPUT_FONT
        c.number_format = FMT_INT
        c.border = BORDER
        for col, f, fmt in [
                (3, f"=INDEX(Ext_R,B{rr})", FMT_RATE),
                (4, f'=IF(B{rr}<=N_Obs,INDEX(SpotRates,B{rr}),"")', FMT_RATE),
                (5, f"=INDEX(Ext_P,B{rr})", FMT_PRICE)]:
            c = sm.cell(row=rr, column=col, value=f)
            c.font = LINK_FONT
            c.number_format = fmt
            c.border = BORDER

    chart = LineChart()
    chart.title = "Extrapolated spot curve vs observed rates"
    chart.style = 2
    chart.height = 9
    chart.width = 20
    chart.y_axis.numFmt = "0.0%"
    chart.x_axis.title = "Maturity (years)"
    ext_ws = wb["Extrapolation"]
    s1 = Series(Reference(ext_ws, min_col=4, min_row=4, max_row=68), title="Fitted R(t)")
    # observed series stops at row 28 (t = 25): the "" cells beyond the
    # observed range would otherwise chart as zeros
    s2 = Series(Reference(ext_ws, min_col=5, min_row=4, max_row=28), title="Observed spot rates")
    chart.series.append(s1)
    chart.series.append(s2)
    chart.set_categories(Reference(ext_ws, min_col=1, min_row=4, max_row=68))
    sm.add_chart(chart, "G3")

    # ---------------------------------------------------------- Checks
    ck = wb.create_sheet("Checks")
    ck.sheet_view.showGridLines = False
    ck.column_dimensions["B"].width = 46
    ck.column_dimensions["C"].width = 10
    ck.column_dimensions["D"].width = 76
    label(ck, 2, 2, "Internal consistency checks", bold=True)
    hdr(ck, 3, 2, "Check"); hdr(ck, 3, 3, "Result"); hdr(ck, 3, 4, "Description")

    checks = [
        ("Fitted rates reproduce observed spot rates",
         "=SUMPRODUCT(--(ABS(R_Fitted_Obs-SpotRates)>0.00000001))=0",
         "R(t) equals the observed spot rate at every observed maturity t = 1 … 25 "
         "(within 1e-8 absolute — the fit is exact up to matrix-inversion rounding).", True),
        ("Fitted prices reproduce observed bond prices",
         "=SUMPRODUCT(--(ABS(P_Fitted_Obs-MVec)>0.00000001))=0",
         "P(t) equals the observed zero-coupon bond price m(t) at every observed maturity.", True),
        ("Wilson matrix is symmetric",
         "=SUMPRODUCT(--(ABS(WMatrix-TRANSPOSE(WMatrix))>0.000000000001))=0",
         "W(t, u) = W(u, t) by construction; a violation indicates a copy/paste error "
         "in the Wilson grid.", True),
        ("ζ solves the calibration system",
         "=SUMPRODUCT(--(ABS(MMULT(WMatrix,ZetaVec)-DVec)>0.00000001))=0",
         "W·ζ reproduces m − μ: the array solve on Calibration is internally consistent.", True),
        ("Extrapolated bond prices are positive",
         "=MIN(Ext_P)>0",
         "P(t) > 0 at every projected maturity.", False),
        ("Grid length matches the stated horizon",
         "=ROWS(Ext_R)=T_Max",
         "The Extrapolation grid covers exactly T_Max years (structural check).", False),
    ]
    for i, (name, formula, desc, cse) in enumerate(checks):
        rr = 4 + i
        label(ck, rr, 2, name)
        c = ck.cell(row=rr, column=3)
        if cse:
            c.value = ArrayFormula(f"C{rr}", formula)
        else:
            c.value = formula
        c.font = LINK_FONT
        c.border = BORDER
        c.alignment = CENTER
        ck.cell(row=rr, column=4, value=desc).font = Font(size=9, color="7F7F7F")
    rall = 4 + len(checks) + 1
    label(ck, rall, 2, "ALL CHECKS PASS", bold=True)
    c = ck.cell(row=rall, column=3, value=f"=AND(C4:C{3 + len(checks)})")
    c.font = Font(name="Calibri", size=10, bold=True, color="006100")
    c.alignment = CENTER
    name_cell(wb, "AllChecksPass", "Checks", f"$C${rall}")

    # ---------------------------------------------------------- Lifelib_Reference
    lr = wb.create_sheet("Lifelib_Reference")
    label(lr, 1, 1, "lifelib smithwilson results (reconciliation target). Generated by "
          "verification/generate_reference_smithwilson.py.", bold=True)
    hdr(lr, 2, 1, "t", 6); hdr(lr, 2, 2, "P(t)", 14); hdr(lr, 2, 3, "R(t)", 12)
    hdr(lr, 2, 5, "j", 6); hdr(lr, 2, 6, "ζ_j", 14)
    for i in range(1, T_MAX + 1):
        rr = 2 + i
        lr.cell(row=rr, column=1, value=i).font = INPUT_FONT
        c = lr.cell(row=rr, column=2, value=float(ref["P"][i - 1]))
        c.font = INPUT_FONT
        c.number_format = FMT_PRICE
        c = lr.cell(row=rr, column=3, value=float(ref["R"][i - 1]))
        c.font = INPUT_FONT
        c.number_format = FMT_RATE
    for j in range(1, N_OBS + 1):
        rr = 2 + j
        lr.cell(row=rr, column=5, value=j).font = INPUT_FONT
        c = lr.cell(row=rr, column=6, value=float(ref["zeta"][j - 1]))
        c.font = INPUT_FONT
        c.number_format = FMT_ZETA
    lr.freeze_panes = "A3"
    wb.defined_names.add(DefinedName("Ref_P", attr_text=f"Lifelib_Reference!$B$3:$B${2 + T_MAX}"))
    wb.defined_names.add(DefinedName("Ref_R", attr_text=f"Lifelib_Reference!$C$3:$C${2 + T_MAX}"))
    wb.defined_names.add(DefinedName("Ref_Zeta", attr_text=f"Lifelib_Reference!$F$3:$F${2 + N_OBS}"))

    # ---------------------------------------------------------- Batch_Results
    br = wb.create_sheet("Batch_Results")
    br.sheet_view.showGridLines = False
    br.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGH", (32, 16, 16, 16, 16, 14, 12)):
        br.column_dimensions[col].width = w
    label(br, 1, 2, "Full-curve reconciliation vs lifelib", bold=True)
    br.cell(row=2, column=2, value="Filled by the RunCurveReconciliation macro "
            "(see Control sheet for the tolerance).").font = NOTE_FONT

    for i, txt in enumerate(["Values compared", "Reconciled within tolerance", "Mismatches",
                             "Max |diff| ζ", "Max |diff| P ÷ max(1,|lifelib|)",
                             "Max |diff| R ÷ max(1,|lifelib|)",
                             "Run time (seconds)", "RECONCILIATION RESULT"]):
        label(br, 4 + i, 2, txt, bold=(i == 7))
        br.cell(row=4 + i, column=3).border = BORDER

    for j, h in enumerate(["quantity", "index", "workbook", "lifelib",
                           "|diff|", "rel diff", "reconciles"], start=2):
        hdr(br, 13, j, h)
    br.freeze_panes = "A14"

    # Bind the VBA project's document modules (codeName plumbing — without
    # this the first ThisWorkbook reference dies with run-time error 429).
    wb.code_name = "ThisWorkbook"
    for i, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_properties.codeName = f"Sheet{i}"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    if VBA_BIN.exists():
        make_xlsm(OUT_XLSX, OUT_XLSM, VBA_BIN)
        print(f"wrote {OUT_XLSM} (macro project injected)")
    else:
        print("vbaProject.bin not found - .xlsm not produced. "
              "Bootstrap it once via workbooks/SmithWilson/bootstrap_vba.py")


def make_xlsm(src_xlsx: Path, dst_xlsm: Path, vba_bin: Path):
    """Produce an .xlsm from the .xlsx by injecting vbaProject.bin (zip surgery)."""
    shutil.copy(src_xlsx, dst_xlsm)
    with zipfile.ZipFile(src_xlsx) as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}

    ct = parts["[Content_Types].xml"].decode()
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    if "vnd.ms-office.vbaProject" not in ct:
        ct = ct.replace(
            "</Types>",
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
        )
    parts["[Content_Types].xml"] = ct.encode()

    rels = parts["xl/_rels/workbook.xml.rels"].decode()
    if "vbaProject" not in rels:
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdVBA" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>',
        )
    parts["xl/_rels/workbook.xml.rels"] = rels.encode()
    parts["xl/vbaProject.bin"] = vba_bin.read_bytes()

    with zipfile.ZipFile(dst_xlsm, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)


if __name__ == "__main__":
    build()
