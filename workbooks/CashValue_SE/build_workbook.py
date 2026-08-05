"""Build the CashValue_SE workbook — Excel twin of lifelib's CashValue_SE model.

Universal-life / savings product: per-policy account value rolled forward
through four within-month timings (BEF_PREM -> BEF_FEE -> BEF_INV ->
MID_MTH), premium loading, cost of insurance, maintenance fees, stochastic
lognormal investment returns (scenario selectable), surrender charges by
product spec, and whole-of-life terms. Verified formula-for-formula
against lifelib (MIT).

Conventions as the BasicTerm workbooks. Output: .xlsm when vbaProject.bin
exists next to this script.
"""

import shutil
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(__file__).resolve().parents[2]
MODEL = REPO / "models" / "savings" / "CashValue_SE"
REF_ALL = REPO / "verification" / "reference" / "cashvalue_se_all_runs.csv"
HERE = Path(__file__).resolve().parent
VBA_BIN = HERE / "vbaProject.bin"
OUT_XLSX = HERE / "CashValue_SE.xlsx"
OUT_XLSM = HERE / "CashValue_SE.xlsm"

# t = 0..1141: max proj_len-1 is 1140 (whole-life issued at age 20, last
# mortality age 115), plus one extra row so av_change / inv_income forward
# references at the final projected month resolve to real (zero) cells.
MAX_T = 1141

NAVY = "1F3864"
INPUT_FONT = Font(name="Calibri", size=10, color="0000FF")
CALC_FONT = Font(name="Calibri", size=10, color="000000")
LINK_FONT = Font(name="Calibri", size=10, color="006100")
HDR_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", start_color=NAVY)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")

FMT_COUNT = "0.000000"
FMT_RATE = "0.0000%"
FMT_MONEY = "#,##0.00"
FMT_INT = "0"

NAMES_IN_FORMULAS = (
    "AgeAtEntry", "PolicyTerm", "PolsIfInit", "SumAssured", "PolicyCount",
    "Duration0", "PremiumPPInput", "AvPPInit", "PremiumType", "HasSurrCharge",
    "SurrChargeID", "LoadPremRate", "IsWL", "MortLastAge", "ScenID",
    "ExpenseAcq", "ExpenseMaint", "InflationRate", "LapseBase", "LapseStep",
    "LapseFloor", "CoiMultiple", "MaintFeeRateAnn", "CommissionRate",
    "Mu", "Sigma", "MortRates", "MortAges", "DiscRateAnn", "NormRand",
    "SurrRates", "SurrTypes", "SurrDurMax",
)


def hdr(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def label(ws, row, col, text, bold=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=10, bold=bold)
    return c


def name_cell(wb, name, sheet, addr):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{addr}"))


def item_row(ws, r, item, value, note, fmt=None, formula=False, link=False):
    label(ws, r, 2, item)
    c = ws.cell(row=r, column=3, value=value)
    c.font = LINK_FONT if link else (CALC_FONT if formula else INPUT_FONT)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = Font(size=9, italic=True, color="7F7F7F")
    return c


def build():
    mp = pd.read_excel(MODEL / "model_point_samples.xlsx", index_col=0)
    specs = pd.read_excel(MODEL / "product_spec_table.xlsx", index_col=0)
    mort = pd.read_excel(MODEL / "mort_table.xlsx", index_col=0)
    disc = pd.read_excel(MODEL / "disc_rate_ann.xlsx", index_col=0)
    surr = pd.read_excel(MODEL / "surr_charge_table.xlsx", index_col=0)
    snr = pd.read_csv(MODEL / "std_norm_rand.csv", index_col=0)
    snr_grid = snr.pivot_table(index=snr.index, columns="t",
                               values="std_norm_rand")
    ref_all = pd.read_csv(REF_ALL)

    wb = Workbook()

    # ---------------------------------------------------------- Cover
    cv = wb.active
    cv.title = "Cover"
    cv.sheet_view.showGridLines = False
    cv.column_dimensions["B"].width = 28
    cv.column_dimensions["C"].width = 80
    t = cv.cell(row=2, column=2, value="CashValue_SE — Universal Life / Savings Cash Flow Projection")
    t.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    cv.cell(row=3, column=2, value="Monthly account-value roll-forward with stochastic investment "
            "returns, product-spec tariffs and margin analysis").font = Font(size=11, italic=True)

    rows = [
        ("Purpose", "Projects the account value, decrements, cash flows and margins for the model "
                    "point and investment scenario selected on the Control sheet. A formula-for-"
                    "formula replication of the open-source lifelib model 'CashValue_SE' (MIT "
                    "licence), verified cell-by-cell against it."),
        ("Products", "Four sample products via the Product_Specs join: A single-premium 10-year "
                     "endowment; B single-premium 20-year with surrender charges; C level-premium "
                     "whole-of-life; D level-premium whole-of-life with surrender charges and 5% "
                     "premium loading."),
        ("Account value", "Per-policy AV at four within-month timings: BEF_PREM (after prior month's "
                          "investment) → + premium into AV → BEF_FEE → − maintenance fee − cost of "
                          "insurance → BEF_INV → + month's investment income (MID_MTH holds AV plus "
                          "half of it, the death/surrender basis)."),
        ("Investment returns", "Lognormal monthly returns exp((μ−½σ²)Δt + σ√Δt·Z) − 1 with μ = 2%, "
                               "σ = 3% p.a., Δt = 1/12, over the stored standard-normal draws on the "
                               "Scenarios sheet (10 scenarios × months; scenario chosen on Control)."),
        ("Margins", "net cash flow decomposes exactly into expense margin (premium loading + "
                    "surrender charges + maintenance fees − commissions − expenses) and mortality "
                    "margin (cost of insurance − death strain) — asserted per month on Checks."),
        ("Projection grain", "Monthly steps t = 0 … 1140 (whole-of-life runs to the mortality "
                             "table's terminal age, 115); row t = 1141 is a helper row so "
                             "forward references at the final month resolve, and is excluded "
                             "from every named range. Flows are structurally zero after "
                             "maturity."),
        ("Batch run", "The RunAllModelPoints macro reconciles every model point × scenario "
                      "combination (40 runs) against lifelib values on Lifelib_Reference."),
        ("Source model", "lifelib savings.CashValue_SE v0.13.0 — https://lifelib.io (MIT licence)"),
        ("Verification", "verification/verify_cashvalue_se.py compares every mapped cell against "
                         "the Python model for six (point, scenario) profiles including both "
                         "whole-of-life products; the batch covers all 40 combinations."),
    ]
    r = 5
    for k, v in rows:
        label(cv, r, 2, k, bold=True)
        c = cv.cell(row=r, column=3, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        cv.row_dimensions[r].height = 54
        r += 1

    r += 1
    label(cv, r, 2, "Sheet index", bold=True)
    for i, (s, d) in enumerate([
            ("Control", "Run control — model point and scenario selection, batch settings"),
            ("Assumptions", "Basis assumptions (expenses, inflation, COI, fees, returns)"),
            ("Model_Points", "Model point samples (input data)"),
            ("Product_Specs", "Product specification table (input data)"),
            ("Mortality", "Mortality table by age and duration (input data)"),
            ("Discount", "Annual spot discount rates (input data)"),
            ("Surr_Charges", "Surrender charge table (input data)"),
            ("Scenarios", "Standard-normal draws by scenario and month (input data)"),
            ("Projection", "Monthly engine — AV roll-forward, decrements, cash flows, margins"),
            ("Summary", "Present values and results for the selected run"),
            ("Checks", "Internal consistency checks (all must be TRUE)"),
            ("Lifelib_Reference", "lifelib results for all runs (reconciliation target)"),
            ("Batch_Results", "All-runs batch reconciliation vs lifelib (macro)")]):
        label(cv, r + 1 + i, 2, s)
        cv.cell(row=r + 1 + i, column=3, value=d).font = Font(size=10)
    r += 15
    label(cv, r, 2, "Version control", bold=True)
    for i, (a, b) in enumerate([("Version", "0.2.0"), ("Date", "2026-07-14"),
                                ("Prepared by", "XL Marinade project"),
                                ("Review status", "Verified against lifelib; adversarial actuarial audit remediated (see verification/reports)")]):
        label(cv, r + 1 + i, 2, a)
        cv.cell(row=r + 1 + i, column=3, value=b).font = Font(size=10)

    r += 6
    label(cv, r, 2, "Conventions", bold=True)
    for i, (txt, font) in enumerate([
            ("Blue — hardcoded input", INPUT_FONT),
            ("Black — calculated on this sheet", CALC_FONT),
            ("Green — link from another sheet", LINK_FONT)]):
        cv.cell(row=r + 1 + i, column=2, value=txt).font = font

    # ---------------------------------------------------------- Control
    ct = wb.create_sheet("Control")
    ct.sheet_view.showGridLines = False
    ct.column_dimensions["B"].width = 38
    ct.column_dimensions["C"].width = 14
    ct.column_dimensions["D"].width = 66

    label(ct, 2, 2, "Run selection", bold=True)
    hdr(ct, 3, 2, "Item"); hdr(ct, 3, 3, "Value"); hdr(ct, 3, 4, "Notes / lifelib cell")
    item_row(ct, 4, "Selected model point ID", 1, "point_id — choose 1 to 4", FMT_INT)
    item_row(ct, 5, "Selected investment scenario", 1, "scen_id — choose 1 to 10", FMT_INT)
    name_cell(wb, "PointID", "Control", "$C$4")
    name_cell(wb, "ScenID", "Control", "$C$5")
    dv1 = DataValidation(type="whole", operator="between", formula1="1", formula2="4")
    dv2 = DataValidation(type="whole", operator="between", formula1="1", formula2="10")
    ct.add_data_validation(dv1); ct.add_data_validation(dv2)
    dv1.add(ct["C4"]); dv2.add(ct["C5"])

    label(ct, 7, 2, "Selected model point (Model_Points ⋈ Product_Specs)", bold=True)
    hdr(ct, 8, 2, "Item"); hdr(ct, 8, 3, "Value"); hdr(ct, 8, 4, "Notes / lifelib cell")
    item_row(ct, 9, "Product spec", "=INDEX(MP_Spec,MATCH(PointID,MP_ID,0))", "spec_id — join key into Product_Specs", None, link=True)
    item_row(ct, 10, "Age at entry", "=INDEX(MP_Age,MATCH(PointID,MP_ID,0))", "age_at_entry()", FMT_INT, link=True)
    item_row(ct, 11, "Sex", "=INDEX(MP_Sex,MATCH(PointID,MP_ID,0))", "sex()", None, link=True)
    item_row(ct, 12, "Policy term input (years)", "=INDEX(MP_Term,MATCH(PointID,MP_ID,0))",
             "model point policy_term; ignored when the product is whole-of-life (is_wl flag "
             "on Product_Specs — the sample data uses 9999 as a placeholder here)", FMT_INT, link=True)
    item_row(ct, 13, "Policies in model point", "=INDEX(MP_Count,MATCH(PointID,MP_ID,0))", "policy_count", FMT_INT, link=True)
    item_row(ct, 14, "Sum assured", "=INDEX(MP_SumAssured,MATCH(PointID,MP_ID,0))", "sum_assured()", FMT_MONEY, link=True)
    item_row(ct, 15, "Months in force at valuation date", "=INDEX(MP_DurMth,MATCH(PointID,MP_ID,0))", "duration_mth(0)", FMT_INT, link=True)
    item_row(ct, 16, "Premium per policy", "=INDEX(MP_Premium,MATCH(PointID,MP_ID,0))", "model point premium_pp (single or monthly level)", FMT_MONEY, link=True)
    item_row(ct, 17, "Initial account value per policy", "=INDEX(MP_AvInit,MATCH(PointID,MP_ID,0))", "av_pp_init()", FMT_MONEY, link=True)
    for nm, addr in [("SpecID", "$C$9"), ("AgeAtEntry", "$C$10"), ("TermInput", "$C$12"),
                     ("PolicyCount", "$C$13"), ("SumAssured", "$C$14"), ("Duration0", "$C$15"),
                     ("PremiumPPInput", "$C$16"), ("AvPPInit", "$C$17")]:
        name_cell(wb, nm, "Control", addr)

    label(ct, 19, 2, "Product spec (from Product_Specs)", bold=True)
    hdr(ct, 20, 2, "Item"); hdr(ct, 20, 3, "Value"); hdr(ct, 20, 4, "Notes / lifelib cell")
    item_row(ct, 21, "Premium type", "=INDEX(PS_PremType,MATCH(SpecID,PS_ID,0))", "premium_type() — SINGLE or LEVEL", None, link=True)
    item_row(ct, 22, "Has surrender charge", "=INDEX(PS_HasSurr,MATCH(SpecID,PS_ID,0))", "has_surr_charge()", None, link=True)
    item_row(ct, 23, "Surrender charge table", "=INDEX(PS_SurrID,MATCH(SpecID,PS_ID,0))", "surr_charge_id()", None, link=True)
    item_row(ct, 24, "Premium loading rate", "=INDEX(PS_LoadPrem,MATCH(SpecID,PS_ID,0))", "load_prem_rate() — share of premium NOT credited to AV", FMT_RATE, link=True)
    item_row(ct, 25, "Whole of life", "=INDEX(PS_IsWL,MATCH(SpecID,PS_ID,0))", "is_wl()", None, link=True)
    for nm, addr in [("PremiumType", "$C$21"), ("HasSurrCharge", "$C$22"),
                     ("SurrChargeID", "$C$23"), ("LoadPremRate", "$C$24"), ("IsWL", "$C$25")]:
        name_cell(wb, nm, "Control", addr)

    label(ct, 27, 2, "Derived", bold=True)
    hdr(ct, 28, 2, "Item"); hdr(ct, 28, 3, "Value"); hdr(ct, 28, 4, "Notes / lifelib cell")
    item_row(ct, 29, "Mortality table terminal age", "=Mortality!$J$2",
             "mort_table_last_age() — first age with q = 1 in every duration band", FMT_INT, link=True)
    item_row(ct, 30, "Policy term (years)", "=IF(IsWL,MortLastAge-AgeAtEntry,TermInput)",
             "policy_term() — whole-of-life runs to the terminal age", FMT_INT, formula=True)
    item_row(ct, 31, "Policies in force at t = 0", "=IF(Duration0>0,PolicyCount,0)", "pols_if_init()", FMT_COUNT, formula=True)
    item_row(ct, 32, "Projection length (months)", "=MAX(12*PolicyTerm-Duration0+1,0)", "proj_len()", FMT_INT, formula=True)
    name_cell(wb, "MortLastAge", "Control", "$C$29")
    name_cell(wb, "PolicyTerm", "Control", "$C$30")
    name_cell(wb, "PolsIfInit", "Control", "$C$31")
    name_cell(wb, "ProjLen", "Control", "$C$32")

    label(ct, 34, 2, "Batch run (RunAllModelPoints macro)", bold=True)
    hdr(ct, 35, 2, "Item"); hdr(ct, 35, 3, "Value"); hdr(ct, 35, 4, "Notes")
    item_row(ct, 36, "Runs to execute (blank = all 40)", None, "Model point × scenario combinations", FMT_INT)
    item_row(ct, 37, "Reconciliation tolerance (relative)", 0.000001,
             "A run reconciles when |workbook − lifelib| ≤ tol × max(1, |lifelib|)", "0.0000000")
    name_cell(wb, "BatchLimit", "Control", "$C$36")
    name_cell(wb, "BatchTol", "Control", "$C$37")

    # ---------------------------------------------------------- Assumptions
    asm = wb.create_sheet("Assumptions")
    asm.sheet_view.showGridLines = False
    asm.column_dimensions["B"].width = 36
    asm.column_dimensions["C"].width = 14
    asm.column_dimensions["D"].width = 64

    label(asm, 2, 2, "Basis assumptions", bold=True)
    hdr(asm, 3, 2, "Item"); hdr(asm, 3, 3, "Value"); hdr(asm, 3, 4, "Notes / lifelib cell")
    item_row(asm, 4, "Acquisition expense per policy", 5000, "expense_acq() — incurred at issue", FMT_MONEY)
    item_row(asm, 5, "Maintenance expense p.a. per policy", 500, "expense_maint() — incurred monthly as 1/12", FMT_MONEY)
    item_row(asm, 6, "Expense inflation p.a.", 0.01, "inflation_rate()", FMT_RATE)
    item_row(asm, 7, "Commission rate on premium", 0.05, "commissions(t) = 5% × premiums(t)", FMT_RATE)
    item_row(asm, 8, "Cost-of-insurance multiple", 1.1, "coi_rate(t) = 1.1 × monthly mortality rate", "0.00")
    item_row(asm, 9, "Maintenance fee p.a. of account value", 0.01, "maint_fee_rate() = 1% / 12 monthly", FMT_RATE)
    label(asm, 11, 2, "Lapse basis (annual rate by policy year)", bold=True)
    hdr(asm, 12, 2, "Item"); hdr(asm, 12, 3, "Value"); hdr(asm, 12, 4, "Notes / lifelib cell")
    item_row(asm, 13, "Lapse rate, year 1", 0.10, "lapse_rate(t) = MAX(base − step × duration, floor)", FMT_RATE)
    item_row(asm, 14, "Reduction per policy year", 0.02, "", FMT_RATE)
    item_row(asm, 15, "Minimum lapse rate", 0.02, "", FMT_RATE)
    label(asm, 17, 2, "Investment return model (lognormal)", bold=True)
    hdr(asm, 18, 2, "Item"); hdr(asm, 18, 3, "Value"); hdr(asm, 18, 4, "Notes / lifelib cell")
    item_row(asm, 19, "Expected return μ p.a.", 0.02, "inv_return_table(): exp((μ−½σ²)Δt + σ√Δt·Z) − 1", FMT_RATE)
    item_row(asm, 20, "Volatility σ p.a.", 0.03, "Δt = 1/12; Z from the Scenarios sheet", FMT_RATE)
    for nm, addr in [("ExpenseAcq", "$C$4"), ("ExpenseMaint", "$C$5"), ("InflationRate", "$C$6"),
                     ("CommissionRate", "$C$7"), ("CoiMultiple", "$C$8"), ("MaintFeeRateAnn", "$C$9"),
                     ("LapseBase", "$C$13"), ("LapseStep", "$C$14"), ("LapseFloor", "$C$15"),
                     ("Mu", "$C$19"), ("Sigma", "$C$20")]:
        name_cell(wb, nm, "Assumptions", addr)

    # ---------------------------------------------------------- Model_Points
    mps = wb.create_sheet("Model_Points")
    cols_mp = [("point_id", 9), ("spec_id", 8), ("age_at_entry", 12), ("sex", 6),
               ("policy_term", 11), ("policy_count", 12), ("sum_assured", 12),
               ("duration_mth", 12), ("premium_pp", 12), ("av_pp_init", 11)]
    for j, (h, w) in enumerate(cols_mp, start=1):
        hdr(mps, 1, j, h, w)
    for i, (pid, row) in enumerate(mp.iterrows(), start=2):
        vals = [pid, row["spec_id"], row["age_at_entry"], row["sex"], row["policy_term"],
                row["policy_count"], row["sum_assured"], row["duration_mth"],
                row["premium_pp"], row["av_pp_init"]]
        for j, v in enumerate(vals, start=1):
            mps.cell(row=i, column=j, value=v).font = INPUT_FONT
    nmp = len(mp)
    for nm, col in [("MP_ID", "A"), ("MP_Spec", "B"), ("MP_Age", "C"), ("MP_Sex", "D"),
                    ("MP_Term", "E"), ("MP_Count", "F"), ("MP_SumAssured", "G"),
                    ("MP_DurMth", "H"), ("MP_Premium", "I"), ("MP_AvInit", "J")]:
        wb.defined_names.add(DefinedName(nm, attr_text=f"Model_Points!${col}$2:${col}${nmp + 1}"))

    # ---------------------------------------------------------- Product_Specs
    pss = wb.create_sheet("Product_Specs")
    for j, (h, w) in enumerate([("spec_id", 8), ("premium_type", 12), ("has_surr_charge", 13),
                                ("surr_charge_id", 13), ("load_prem_rate", 13), ("is_wl", 8)], start=1):
        hdr(pss, 1, j, h, w)
    for i, (sid, row) in enumerate(specs.iterrows(), start=2):
        vals = [sid, row["premium_type"], bool(row["has_surr_charge"]),
                "" if pd.isna(row["surr_charge_id"]) else row["surr_charge_id"],
                float(row["load_prem_rate"]), bool(row["is_wl"])]
        for j, v in enumerate(vals, start=1):
            c = pss.cell(row=i, column=j, value=v)
            c.font = INPUT_FONT
            if j == 5:
                c.number_format = FMT_RATE
    nsp = len(specs)
    for nm, col in [("PS_ID", "A"), ("PS_PremType", "B"), ("PS_HasSurr", "C"),
                    ("PS_SurrID", "D"), ("PS_LoadPrem", "E"), ("PS_IsWL", "F")]:
        wb.defined_names.add(DefinedName(nm, attr_text=f"Product_Specs!${col}$2:${col}${nsp + 1}"))

    # ---------------------------------------------------------- Mortality
    mt = wb.create_sheet("Mortality")
    hdr(mt, 1, 1, "Attained age", 12)
    for j, cname in enumerate(mort.columns, start=2):
        hdr(mt, 1, j, f"Duration {cname}", 11)
    hdr(mt, 1, 9, "All q = 1?", 10)
    hdr(mt, 1, 10, "Terminal age", 11)
    for i, (age_, row) in enumerate(mort.iterrows(), start=2):
        mt.cell(row=i, column=1, value=age_).font = INPUT_FONT
        for j, v in enumerate(row, start=2):
            c = mt.cell(row=i, column=j, value=float(v))
            c.font = INPUT_FONT
            c.number_format = "0.000000"
        c = mt.cell(row=i, column=9, value=f"=IF(COUNTIF(B{i}:G{i},1)=6,A{i},\"\")")
        c.font = CALC_FONT
    c = mt.cell(row=2, column=10, value=f"=MIN(I2:I{len(mort) + 1})")
    c.font = CALC_FONT
    mt.cell(row=3, column=10, value="mort_table_last_age()").font = Font(size=8, italic=True, color="7F7F7F")
    mt.freeze_panes = "B2"
    na = len(mort)
    wb.defined_names.add(DefinedName("MortAges", attr_text=f"Mortality!$A$2:$A${na + 1}"))
    wb.defined_names.add(DefinedName("MortRates", attr_text=f"Mortality!$B$2:$G${na + 1}"))

    # ---------------------------------------------------------- Discount
    ds = wb.create_sheet("Discount")
    hdr(ds, 1, 1, "Year", 8)
    hdr(ds, 1, 2, "Annual spot rate", 15)
    for i, (yr, row) in enumerate(disc.iterrows(), start=2):
        ds.cell(row=i, column=1, value=yr).font = INPUT_FONT
        c = ds.cell(row=i, column=2, value=float(row.iloc[0]))
        c.font = INPUT_FONT
        c.number_format = FMT_RATE
    ny = len(disc)
    wb.defined_names.add(DefinedName("DiscRateAnn", attr_text=f"Discount!$B$2:$B${ny + 1}"))

    # ---------------------------------------------------------- Surr_Charges
    sch = wb.create_sheet("Surr_Charges")
    hdr(sch, 1, 1, "Policy year", 11)
    for j, cname in enumerate(surr.columns, start=2):
        hdr(sch, 1, j, cname, 10)
    for i, (dur, row) in enumerate(surr.iterrows(), start=2):
        sch.cell(row=i, column=1, value=int(dur)).font = INPUT_FONT
        for j, v in enumerate(row, start=2):
            c = sch.cell(row=i, column=j, value=float(v))
            c.font = INPUT_FONT
            c.number_format = FMT_RATE
    nsc = len(surr)
    wb.defined_names.add(DefinedName("SurrTypes", attr_text=f"Surr_Charges!$B$1:$D$1"))
    wb.defined_names.add(DefinedName("SurrRates", attr_text=f"Surr_Charges!$B$2:$D${nsc + 1}"))
    wb.defined_names.add(DefinedName("SurrDurMax", attr_text=f"Surr_Charges!$A${nsc + 1}"))

    # ---------------------------------------------------------- Scenarios
    sc = wb.create_sheet("Scenarios")
    label(sc, 1, 1, "Standard normal draws Z by scenario and month — lifelib std_norm_rand. "
          "Monthly return = exp((μ−½σ²)/12 + σ√(1/12)·Z) − 1 (computed in the Projection engine).",
          bold=True)
    hdr(sc, 2, 1, "scen_id \\ t", 10)
    n_t = snr_grid.shape[1]
    for j in range(n_t):
        c = sc.cell(row=2, column=j + 2, value=j)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    for i, (sid, row) in enumerate(snr_grid.iterrows(), start=3):
        sc.cell(row=i, column=1, value=int(sid)).font = INPUT_FONT
        for j, v in enumerate(row, start=2):
            sc.cell(row=i, column=j, value=float(v)).font = INPUT_FONT
    nscen = len(snr_grid)
    last_col = get_column_letter(n_t + 1)
    wb.defined_names.add(DefinedName(
        "NormRand", attr_text=f"Scenarios!$B$3:${last_col}${nscen + 2}"))
    sc.freeze_panes = "B3"

    # ---------------------------------------------------------- Projection
    pj = wb.create_sheet("Projection")
    cols = [
        ("t", "Month", 6, FMT_INT),
        ("duration_mth", "Months in force", 8, FMT_INT),
        ("duration", "Duration (yrs)", 8, FMT_INT),
        ("age", "Age", 7, FMT_INT),
        ("premium_pp", "Premium pp", 11, FMT_MONEY),
        ("mort_rate", "Mort rate p.a.", 11, "0.000000"),
        ("mort_rate_mth", "Mort rate p.m.", 11, "0.00000000"),
        ("lapse_rate", "Lapse rate p.a.", 10, FMT_RATE),
        ("surr_charge_rate", "Surr charge rate", 10, FMT_RATE),
        ("pols_if", "In force (BEF_MAT)", 12, FMT_COUNT),
        ("pols_maturity", "Maturities", 11, FMT_COUNT),
        ("pols_if_bef_nb", "In force (BEF_NB)", 12, FMT_COUNT),
        ("pols_new_biz", "New business", 11, FMT_COUNT),
        ("pols_if_bef_decr", "In force (BEF_DECR)", 12, FMT_COUNT),
        ("pols_death", "Deaths", 11, "0.00000000"),
        ("pols_lapse", "Lapses", 11, FMT_COUNT),
        ("av_pp_bef_prem", "AV pp BEF_PREM", 13, FMT_MONEY),
        ("prem_to_av_pp", "Premium to AV pp", 12, FMT_MONEY),
        ("av_pp_bef_fee", "AV pp BEF_FEE", 13, FMT_MONEY),
        ("maint_fee_pp", "Maint fee pp", 11, FMT_MONEY),
        ("net_amt_at_risk", "Net amount at risk", 13, FMT_MONEY),
        ("coi_rate", "COI rate", 11, "0.00000000"),
        ("coi_pp", "COI pp", 11, FMT_MONEY),
        ("av_pp_bef_inv", "AV pp BEF_INV", 13, FMT_MONEY),
        ("inv_return_mth", "Inv return p.m.", 11, "0.000000%"),
        ("inv_income_pp", "Inv income pp", 12, FMT_MONEY),
        ("av_pp_mid_mth", "AV pp MID_MTH", 13, FMT_MONEY),
        ("av_bef_mat", "AV total (BEF_MAT)", 14, FMT_MONEY),
        ("av_bef_nb", "AV total (BEF_NB)", 14, FMT_MONEY),
        ("av_bef_fee", "AV total (BEF_FEE)", 14, FMT_MONEY),
        ("premiums", "Premium income", 13, FMT_MONEY),
        ("prem_to_av", "Premium to AV", 13, FMT_MONEY),
        ("claim_pp_death", "Death claim pp", 12, FMT_MONEY),
        ("claims_death", "Death claims", 13, FMT_MONEY),
        ("cfav_death", "AV released: death", 13, FMT_MONEY),
        ("cfav_lapse", "AV released: lapse", 13, FMT_MONEY),
        ("surr_charge", "Surrender charge", 12, FMT_MONEY),
        ("claims_lapse", "Surrender claims", 13, FMT_MONEY),
        ("cfav_maturity", "AV released: maturity", 13, FMT_MONEY),
        ("claims_total", "Claims total", 13, FMT_MONEY),
        ("claims_over_av", "Death strain over AV", 12, FMT_MONEY),
        ("coi", "COI charge", 12, FMT_MONEY),
        ("maint_fee", "Maintenance fee", 12, FMT_MONEY),
        ("inv_income", "Investment income", 13, FMT_MONEY),
        ("inflation_factor", "Inflation factor", 10, "0.000000"),
        ("expenses", "Expenses", 12, FMT_MONEY),
        ("commissions", "Commission", 11, FMT_MONEY),
        ("av_change", "Change in AV", 13, FMT_MONEY),
        ("net_cf", "Net cash flow", 13, FMT_MONEY),
        ("margin_expense", "Expense margin", 13, FMT_MONEY),
        ("margin_mortality", "Mortality margin", 13, FMT_MONEY),
        ("margin_diff", "Margin identity rel diff", 11, "0.000000000"),
        ("av_roll_diff", "AV roll-fwd rel diff", 11, "0.000000000"),
        ("disc_rate_mth", "Disc rate p.m.", 11, "0.00000000"),
        ("disc_factor", "Discount factor", 12, "0.00000000"),
        ("pv_net_cf_t", "PV net cash flow", 13, FMT_MONEY),
    ]
    pj.sheet_view.showGridLines = False
    hdr_row = 2
    label(pj, 1, 1, "Monthly projection — AV per policy rolls BEF_PREM → +premium → BEF_FEE → "
          "−fees/COI → BEF_INV → +investment → next month; counts as in BasicTerm_SE; "
          "margin and AV roll-forward identities checked per row", bold=True)
    for j, (key, title, w, fmt) in enumerate(cols, start=1):
        hdr(pj, hdr_row, j, title, w)
        pj.cell(row=hdr_row + 1, column=j, value=key).font = Font(size=8, italic=True, color="7F7F7F")
    first = hdr_row + 2

    C = {key: get_column_letter(j) for j, (key, *_r) in enumerate(cols, start=1)}
    colfmt = {key: fmt for key, _t, _w, fmt in cols}

    def cell_ref(key, row):
        return f"{C[key]}{row}"

    for t in range(MAX_T + 1):
        r = first + t
        p = r - 1
        nx = r + 1
        f = {}
        f["t"] = t if t == 0 else f"={cell_ref('t', p)}+1"
        f["duration_mth"] = f"=Duration0+{cell_ref('t', r)}"
        f["duration"] = f"=INT({cell_ref('duration_mth', r)}/12)"
        f["age"] = f"=AgeAtEntry+{cell_ref('duration', r)}"
        f["premium_pp"] = (f"=IF(PremiumType=\"SINGLE\","
                           f"IF({cell_ref('duration_mth', r)}=0,PremiumPPInput,0),"
                           f"IF({cell_ref('duration_mth', r)}<12*PolicyTerm,PremiumPPInput,0))")
        # age clamped at the table's last row: rows beyond the projection
        # horizon would otherwise look up ages past 120 (#N/A); flows there
        # are structurally zero so the clamped rate is never used
        f["mort_rate"] = (f"=INDEX(MortRates,MATCH(MIN({cell_ref('age', r)},120),MortAges,0),"
                          f"MAX(MIN({cell_ref('duration', r)},5),0)+1)")
        f["mort_rate_mth"] = f"=1-(1-{cell_ref('mort_rate', r)})^(1/12)"
        f["lapse_rate"] = f"=MAX(LapseBase-LapseStep*{cell_ref('duration', r)},LapseFloor)"
        f["surr_charge_rate"] = (f"=IF(HasSurrCharge,"
                                 f"INDEX(SurrRates,MIN({cell_ref('duration', r)},SurrDurMax)+1,"
                                 f"MATCH(SurrChargeID,SurrTypes,0)),0)")
        if t == 0:
            f["pols_if"] = "=PolsIfInit"
            f["av_pp_bef_prem"] = "=AvPPInit"
        else:
            f["pols_if"] = (f"={cell_ref('pols_if_bef_decr', p)}"
                            f"-{cell_ref('pols_lapse', p)}-{cell_ref('pols_death', p)}")
            f["av_pp_bef_prem"] = f"={cell_ref('av_pp_bef_inv', p)}+{cell_ref('inv_income_pp', p)}"
        f["pols_maturity"] = (f"=IF({cell_ref('duration_mth', r)}=12*PolicyTerm,"
                              f"{cell_ref('pols_if', r)},0)")
        f["pols_if_bef_nb"] = f"={cell_ref('pols_if', r)}-{cell_ref('pols_maturity', r)}"
        f["pols_new_biz"] = f"=IF({cell_ref('duration_mth', r)}=0,PolicyCount,0)"
        f["pols_if_bef_decr"] = f"={cell_ref('pols_if_bef_nb', r)}+{cell_ref('pols_new_biz', r)}"
        f["pols_death"] = f"={cell_ref('pols_if_bef_decr', r)}*{cell_ref('mort_rate_mth', r)}"
        f["pols_lapse"] = (f"=({cell_ref('pols_if_bef_decr', r)}-{cell_ref('pols_death', r)})"
                           f"*(1-(1-{cell_ref('lapse_rate', r)})^(1/12))")
        f["prem_to_av_pp"] = f"=(1-LoadPremRate)*{cell_ref('premium_pp', r)}"
        f["av_pp_bef_fee"] = f"={cell_ref('av_pp_bef_prem', r)}+{cell_ref('prem_to_av_pp', r)}"
        f["maint_fee_pp"] = f"=MaintFeeRateAnn/12*{cell_ref('av_pp_bef_fee', r)}"
        f["net_amt_at_risk"] = f"=MAX(SumAssured-{cell_ref('av_pp_bef_fee', r)},0)"
        f["coi_rate"] = f"=CoiMultiple*{cell_ref('mort_rate_mth', r)}"
        f["coi_pp"] = f"={cell_ref('coi_rate', r)}*{cell_ref('net_amt_at_risk', r)}"
        f["av_pp_bef_inv"] = (f"={cell_ref('av_pp_bef_fee', r)}-{cell_ref('maint_fee_pp', r)}"
                              f"-{cell_ref('coi_pp', r)}")
        f["inv_return_mth"] = (f"=EXP((Mu-0.5*Sigma^2)/12+Sigma*SQRT(1/12)"
                               f"*INDEX(NormRand,ScenID,{cell_ref('t', r)}+1))-1")
        f["inv_income_pp"] = f"={cell_ref('inv_return_mth', r)}*{cell_ref('av_pp_bef_inv', r)}"
        f["av_pp_mid_mth"] = f"={cell_ref('av_pp_bef_inv', r)}+0.5*{cell_ref('inv_income_pp', r)}"
        f["av_bef_mat"] = f"={cell_ref('av_pp_bef_prem', r)}*{cell_ref('pols_if', r)}"
        f["av_bef_nb"] = f"={cell_ref('av_pp_bef_prem', r)}*{cell_ref('pols_if_bef_nb', r)}"
        f["av_bef_fee"] = f"={cell_ref('av_pp_bef_fee', r)}*{cell_ref('pols_if_bef_decr', r)}"
        f["premiums"] = f"={cell_ref('premium_pp', r)}*{cell_ref('pols_if_bef_decr', r)}"
        f["prem_to_av"] = f"={cell_ref('prem_to_av_pp', r)}*{cell_ref('pols_if_bef_decr', r)}"
        f["claim_pp_death"] = f"=MAX(SumAssured,{cell_ref('av_pp_mid_mth', r)})"
        f["claims_death"] = f"={cell_ref('claim_pp_death', r)}*{cell_ref('pols_death', r)}"
        f["cfav_death"] = f"={cell_ref('av_pp_mid_mth', r)}*{cell_ref('pols_death', r)}"
        f["cfav_lapse"] = f"={cell_ref('av_pp_mid_mth', r)}*{cell_ref('pols_lapse', r)}"
        f["surr_charge"] = (f"={cell_ref('surr_charge_rate', r)}*{cell_ref('av_pp_mid_mth', r)}"
                            f"*{cell_ref('pols_lapse', r)}")
        f["claims_lapse"] = f"={cell_ref('cfav_lapse', r)}-{cell_ref('surr_charge', r)}"
        f["cfav_maturity"] = f"={cell_ref('av_pp_bef_prem', r)}*{cell_ref('pols_maturity', r)}"
        f["claims_total"] = (f"={cell_ref('claims_death', r)}+{cell_ref('claims_lapse', r)}"
                             f"+{cell_ref('cfav_maturity', r)}")
        f["claims_over_av"] = (f"=({cell_ref('claim_pp_death', r)}-{cell_ref('av_pp_mid_mth', r)})"
                               f"*{cell_ref('pols_death', r)}")
        f["coi"] = f"={cell_ref('coi_pp', r)}*{cell_ref('pols_if_bef_decr', r)}"
        f["maint_fee"] = f"={cell_ref('maint_fee_pp', r)}*{cell_ref('pols_if_bef_decr', r)}"
        f["inv_income"] = (f"={cell_ref('inv_income_pp', r)}*{cell_ref('pols_if', nx)}"
                           f"+0.5*{cell_ref('inv_income_pp', r)}"
                           f"*({cell_ref('pols_death', r)}+{cell_ref('pols_lapse', r)})")
        f["inflation_factor"] = f"=(1+InflationRate)^({cell_ref('t', r)}/12)"
        f["expenses"] = (f"=ExpenseAcq*{cell_ref('pols_new_biz', r)}"
                         f"+{cell_ref('pols_if_bef_decr', r)}*ExpenseMaint/12"
                         f"*{cell_ref('inflation_factor', r)}")
        f["commissions"] = f"=CommissionRate*{cell_ref('premiums', r)}"
        f["av_change"] = f"={cell_ref('av_bef_mat', nx)}-{cell_ref('av_bef_mat', r)}"
        f["net_cf"] = (f"={cell_ref('premiums', r)}+{cell_ref('inv_income', r)}"
                       f"-{cell_ref('claims_total', r)}-{cell_ref('expenses', r)}"
                       f"-{cell_ref('commissions', r)}-{cell_ref('av_change', r)}")
        f["margin_expense"] = (f"=LoadPremRate*{cell_ref('premium_pp', r)}"
                               f"*{cell_ref('pols_if_bef_decr', r)}"
                               f"+{cell_ref('surr_charge', r)}+{cell_ref('maint_fee', r)}"
                               f"-{cell_ref('commissions', r)}-{cell_ref('expenses', r)}")
        f["margin_mortality"] = f"={cell_ref('coi', r)}-{cell_ref('claims_over_av', r)}"
        # identity residuals RELATIVE to the quantity's own scale, so the
        # Checks-sheet tolerance is scale-independent (absolute residuals at
        # 50M account values legitimately exceed cents in float arithmetic)
        f["margin_diff"] = (f"=({cell_ref('net_cf', r)}-{cell_ref('margin_expense', r)}"
                            f"-{cell_ref('margin_mortality', r)})"
                            f"/MAX(1,ABS({cell_ref('net_cf', r)}),"
                            f"ABS({cell_ref('margin_expense', r)}),"
                            f"ABS({cell_ref('margin_mortality', r)}))")
        # the per-row identity does not hold across the maturity boundary
        # (the maturity release is booked one row after BEF_NB empties; it
        # telescopes to zero over the two rows) - lifelib's own
        # check_av_roll_fwd returns False there for single-premium
        # endowments, so those rows are excluded, matching the identity's
        # actual domain of validity
        f["av_roll_diff"] = (f"=IF(OR({cell_ref('pols_maturity', r)}>0,"
                             f"{cell_ref('pols_maturity', nx)}>0),0,"
                             f"({cell_ref('av_bef_nb', nx)}-({cell_ref('av_bef_nb', r)}"
                             f"+{cell_ref('prem_to_av', r)}-{cell_ref('maint_fee', r)}"
                             f"-{cell_ref('coi', r)}+{cell_ref('inv_income', r)}"
                             f"-{cell_ref('cfav_death', r)}-{cell_ref('cfav_lapse', r)}"
                             f"-{cell_ref('cfav_maturity', r)}))"
                             f"/MAX(1,ABS({cell_ref('av_bef_nb', r)}),"
                             f"ABS({cell_ref('av_bef_nb', nx)})))")
        f["disc_rate_mth"] = f"=(1+INDEX(DiscRateAnn,INT({cell_ref('t', r)}/12)+1))^(1/12)-1"
        f["disc_factor"] = f"=(1+{cell_ref('disc_rate_mth', r)})^-{cell_ref('t', r)}"
        f["pv_net_cf_t"] = f"={cell_ref('net_cf', r)}*{cell_ref('disc_factor', r)}"

        for key, _title, _w, fmt in cols:
            c = pj.cell(row=r, column=list(C).index(key) + 1, value=f[key])
            if fmt:
                c.number_format = fmt
            if isinstance(f[key], str) and f[key].startswith("="):
                c.font = LINK_FONT if any(nm in f[key] for nm in NAMES_IN_FORMULAS) else CALC_FONT
            else:
                c.font = CALC_FONT
    pj.freeze_panes = pj.cell(row=first, column=3).coordinate
    # named columns end at the last PROJECTED month (MAX_T - 1); the final
    # grid row exists only to resolve forward references, and lifelib's PV
    # sums stop at proj_len <= MAX_T
    last = first + MAX_T - 1
    for nm, key in [("Proj_T", "t"), ("Proj_DurMth", "duration_mth"),
                    ("Proj_PolsIf", "pols_if"), ("Proj_Maturities", "pols_maturity"),
                    ("Proj_NewBiz", "pols_new_biz"), ("Proj_Deaths", "pols_death"),
                    ("Proj_Lapses", "pols_lapse"), ("Proj_Premiums", "premiums"),
                    ("Proj_ClaimsTotal", "claims_total"), ("Proj_ClaimsDeath", "claims_death"),
                    ("Proj_ClaimsLapse", "claims_lapse"), ("Proj_ClaimsMat", "cfav_maturity"),
                    ("Proj_Expenses", "expenses"), ("Proj_Commissions", "commissions"),
                    ("Proj_InvIncome", "inv_income"), ("Proj_AvChange", "av_change"),
                    ("Proj_NetCF", "net_cf"), ("Proj_MarginDiff", "margin_diff"),
                    ("Proj_AvRollDiff", "av_roll_diff"),
                    ("Proj_DiscFactor", "disc_factor"), ("Proj_PVNetCF", "pv_net_cf_t")]:
        col = C[key]
        wb.defined_names.add(DefinedName(nm, attr_text=f"Projection!${col}${first}:${col}${last}"))

    # ---------------------------------------------------------- Summary
    sm = wb.create_sheet("Summary")
    sm.sheet_view.showGridLines = False
    sm.column_dimensions["B"].width = 36
    sm.column_dimensions["C"].width = 18
    sm.column_dimensions["D"].width = 60

    label(sm, 2, 2, "Run", bold=True)
    hdr(sm, 3, 2, "Item"); hdr(sm, 3, 3, "Value"); hdr(sm, 3, 4, "lifelib cell")

    def sm_row(r, item, formula, note, fmt=FMT_MONEY, name=None):
        label(sm, r, 2, item)
        c = sm.cell(row=r, column=3, value=formula)
        c.font = LINK_FONT
        if fmt:
            c.number_format = fmt
        c.border = BORDER
        sm.cell(row=r, column=4, value=note).font = Font(size=9, italic=True, color="7F7F7F")
        if name:
            name_cell(wb, name, "Summary", f"$C${r}")

    sm_row(4, "Model point ID", "=PointID", "point_id (batch sentinel)", FMT_INT)
    sm_row(5, "Investment scenario", "=ScenID", "scen_id (batch sentinel)", FMT_INT)
    sm_row(6, "Product / term / policies",
           '=SpecID&" / "&PolicyTerm&"y / "&PolicyCount', "", None)

    label(sm, 8, 2, "Present value of cash flows", bold=True)
    hdr(sm, 9, 2, "Item"); hdr(sm, 9, 3, "PV"); hdr(sm, 9, 4, "lifelib cell")
    sm_row(10, "Premiums", "=SUMPRODUCT(Proj_Premiums,Proj_DiscFactor)", "pv_premiums()", FMT_MONEY, name="PV_Premiums")
    sm_row(11, "Claims — death", "=SUMPRODUCT(Proj_ClaimsDeath,Proj_DiscFactor)", 'pv_claims("DEATH")', FMT_MONEY, name="PV_ClaimsDeath")
    sm_row(12, "Claims — surrender", "=SUMPRODUCT(Proj_ClaimsLapse,Proj_DiscFactor)", 'pv_claims("LAPSE")', FMT_MONEY, name="PV_ClaimsLapse")
    sm_row(13, "Claims — maturity", "=SUMPRODUCT(Proj_ClaimsMat,Proj_DiscFactor)", 'pv_claims("MATURITY")', FMT_MONEY, name="PV_ClaimsMat")
    sm_row(14, "Claims — total", "=PV_ClaimsDeath+PV_ClaimsLapse+PV_ClaimsMat", "pv_claims()", FMT_MONEY, name="PV_Claims")
    sm_row(15, "Expenses", "=SUMPRODUCT(Proj_Expenses,Proj_DiscFactor)", "pv_expenses()", FMT_MONEY, name="PV_Expenses")
    sm_row(16, "Commission", "=SUMPRODUCT(Proj_Commissions,Proj_DiscFactor)", "pv_commissions()", FMT_MONEY, name="PV_Commissions")
    sm_row(17, "Investment income", "=SUMPRODUCT(Proj_InvIncome,Proj_DiscFactor)", "pv_inv_income()", FMT_MONEY, name="PV_InvIncome")
    sm_row(18, "Change in account value", "=SUMPRODUCT(Proj_AvChange,Proj_DiscFactor)", "pv_av_change()", FMT_MONEY, name="PV_AvChange")
    sm_row(19, "Policies in force", "=SUMPRODUCT(Proj_PolsIf,Proj_DiscFactor)", "pv_pols_if()", FMT_COUNT, name="PV_PolsIf")
    sm_row(20, "Net cash flow",
           "=PV_Premiums+PV_InvIncome-PV_Claims-PV_Expenses-PV_Commissions-PV_AvChange",
           "pv_net_cf()", FMT_MONEY, name="PV_NetCF")
    sm.cell(row=20, column=2).font = Font(size=10, bold=True)
    sm.cell(row=20, column=3).font = Font(name="Calibri", size=10, bold=True, color="006100")

    # ---------------------------------------------------------- Checks
    ck = wb.create_sheet("Checks")
    ck.sheet_view.showGridLines = False
    ck.column_dimensions["B"].width = 46
    ck.column_dimensions["C"].width = 10
    ck.column_dimensions["D"].width = 76
    label(ck, 2, 2, "Internal consistency checks", bold=True)
    hdr(ck, 3, 2, "Check"); hdr(ck, 3, 3, "Result"); hdr(ck, 3, 4, "Description")

    checks = [
        ("AV roll-forward (check_av_roll_fwd)",
         "=SUMPRODUCT(--(ABS(Proj_AvRollDiff)>0.000000001))=0",
         "Within 1e-9 relative, for every month EXCEPT the two maturity-boundary rows: "
         "AV(t+1, BEF_NB) equals AV(t, BEF_NB) + premium into AV − fees − COI + investment "
         "income − AV released on death, surrender and maturity. At the maturity boundary the "
         "per-month identity does not hold in lifelib itself (the maturity release is booked "
         "one row after BEF_NB empties; lifelib's own check_av_roll_fwd returns False there "
         "for single-premium endowments), so those two rows are excluded — matching the "
         "identity's actual domain of validity."),
        ("Margin decomposition (check_margin)",
         "=SUMPRODUCT(--(ABS(Proj_MarginDiff)>0.000000001))=0",
         "Every month, within 1e-9 relative: net cash flow equals expense margin + mortality margin."),
        ("PV identity (check_pv_net_cf)",
         "=ABS(SUMPRODUCT(Proj_NetCF,Proj_DiscFactor)-PV_NetCF)<0.01",
         "PV of the net cash flow column equals the sum of the component PVs."),
        ("In-force never negative",
         "=MIN(Proj_PolsIf)>=0",
         "Policies in force are non-negative at every projection month."),
        ("Entrants equal leavers",
         "=ABS(PolsIfInit+SUM(Proj_NewBiz)-SUM(Proj_Deaths)-SUM(Proj_Lapses)-SUM(Proj_Maturities))<0.000001",
         "Initial in-force plus new business equals deaths + lapses + maturities."),
        ("No in-force after maturity",
         "=SUMPRODUCT((Proj_DurMth>12*PolicyTerm)*ABS(Proj_PolsIf))=0",
         "The in-force count is exactly zero in every month after the maturity month."),
    ]
    for i, (name, formula, desc) in enumerate(checks):
        r = 4 + i
        label(ck, r, 2, name)
        c = ck.cell(row=r, column=3, value=formula)
        c.font = LINK_FONT
        c.border = BORDER
        c.alignment = CENTER
        ck.cell(row=r, column=4, value=desc).font = Font(size=9, color="7F7F7F")
    rall = 4 + len(checks) + 1
    label(ck, rall, 2, "ALL CHECKS PASS", bold=True)
    c = ck.cell(row=rall, column=3, value=f"=AND(C4:C{3 + len(checks)})")
    c.font = Font(name="Calibri", size=10, bold=True, color="006100")
    c.alignment = CENTER
    name_cell(wb, "AllChecksPass", "Checks", f"$C${rall}")

    # ---------------------------------------------------------- Lifelib_Reference
    lr = wb.create_sheet("Lifelib_Reference")
    label(lr, 1, 1, "lifelib results for every model point × scenario (savings.CashValue_SE). "
          "Generated by verification/generate_reference_cashvalue_se.py.", bold=True)
    ref_cols = list(ref_all.columns)
    for j, h in enumerate(ref_cols, start=1):
        hdr(lr, 2, j, h, 14)
    for i, (_, row) in enumerate(ref_all.iterrows(), start=3):
        lr.cell(row=i, column=1, value=int(row.iloc[0])).font = INPUT_FONT
        lr.cell(row=i, column=2, value=int(row.iloc[1])).font = INPUT_FONT
        for j in range(3, len(ref_cols) + 1):
            c = lr.cell(row=i, column=j, value=float(row.iloc[j - 1]))
            c.font = INPUT_FONT
            c.number_format = FMT_MONEY
    lr.freeze_panes = "A3"

    # ---------------------------------------------------------- Batch_Results
    br = wb.create_sheet("Batch_Results")
    br.sheet_view.showGridLines = False
    br.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGHIJKLM", (30, 16, 10, 16, 16, 14, 14, 14, 14, 14, 14, 10)):
        br.column_dimensions[col].width = w
    label(br, 1, 2, "All model points × scenarios — batch run and reconciliation vs lifelib", bold=True)
    br.cell(row=2, column=2, value="Filled by the RunAllModelPoints macro (see Control sheet for settings)."
            ).font = Font(size=9, italic=True, color="7F7F7F")
    for i, txt in enumerate(["Runs executed", "Reconciled within tolerance", "Mismatches",
                             "Max relative diff PV premiums", "Max relative diff PV (all components)",
                             "Run time (seconds)", "RECONCILIATION RESULT"]):
        label(br, 4 + i, 2, txt, bold=(i == 6))
        br.cell(row=4 + i, column=3).border = BORDER
    hdrs = ["point_id", "scen_id", "pv_net_cf (wb)", "pv_net_cf (lifelib)", "rel pv_premiums",
            "rel pv_claims", "rel pv_expenses", "rel pv_commissions", "rel pv_inv_income",
            "rel pv_av_change", "rel pv_net_cf", "reconciles"]
    for j, h in enumerate(hdrs, start=2):
        hdr(br, 12, j, h)
    br.freeze_panes = "A13"

    wb.code_name = "ThisWorkbook"
    for i, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_properties.codeName = f"Sheet{i}"

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")

    if VBA_BIN.exists():
        make_xlsm(OUT_XLSX, OUT_XLSM, VBA_BIN)
        print(f"wrote {OUT_XLSM} (macro project injected)")
    else:
        print("vbaProject.bin not found - bootstrap via workbooks/CashValue_SE/bootstrap_vba.py")


def make_xlsm(src_xlsx: Path, dst_xlsm: Path, vba_bin: Path):
    shutil.copy(src_xlsx, dst_xlsm)
    with zipfile.ZipFile(src_xlsx) as zin:
        parts = {n: zin.read(n) for n in zin.namelist()}
    ct = parts["[Content_Types].xml"].decode()
    ct = ct.replace(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    if "vnd.ms-office.vbaProject" not in ct:
        ct = ct.replace(
            "</Types>",
            '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
        )
    parts["[Content_Types].xml"] = ct.encode()
    rels = parts["xl/_rels/workbook.xml.rels"].decode()
    if "vbaProject" not in rels:
        rels = rels.replace(
            "</Relationships>",
            '<Relationship Id="rIdVBA" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
            'Target="vbaProject.bin"/></Relationships>',
        )
    parts["xl/_rels/workbook.xml.rels"] = rels.encode()
    parts["xl/vbaProject.bin"] = vba_bin.read_bytes()
    with zipfile.ZipFile(dst_xlsm, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in parts.items():
            zout.writestr(name, data)


if __name__ == "__main__":
    build()
