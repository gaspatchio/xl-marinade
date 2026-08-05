# ABOUTME: Snapshot-only resolution for volatile/lookup functions and special data sources
# ABOUTME: Implements deterministic resolution per ADR-004 without recalculation

from __future__ import annotations

from dataclasses import dataclass, field
from typing import TYPE_CHECKING, Any, Protocol, runtime_checkable

from openpyxl import Workbook


@runtime_checkable
class ValueSource(Protocol):
    """Protocol for value access in resolution engine."""

    sheetnames: list[str]
    active_sheet: str | None

    def get_value_at(self, sheet: str, coord: str) -> Any:
        """Get evaluated value for cell."""
        ...


if TYPE_CHECKING:
    from xl_marinade.core.manual_resolution import ManualResolutionProvider

from xl_marinade.core.ref_converter import (
    EXCEL_MAX_ROWS,
    col_num_to_letter,
    format_cell_address,
    parse_a1_reference,
    parse_cell_address,
)

# Volatile function names (case-insensitive)
VOLATILE_FUNCTIONS = {
    "OFFSET",
    "INDIRECT",
    "NOW",
    "TODAY",
    "RAND",
    "RANDBETWEEN",
    "CELL",
    "INFO",
    "RTD",
    "HYPERLINK",
}

# Lookup function names
LOOKUP_FUNCTIONS = {"INDEX", "MATCH", "XLOOKUP", "VLOOKUP", "HLOOKUP", "LOOKUP"}

# Excel max column count (XFD)
EXCEL_MAX_COLS = 16384

# Special data source functions
PIVOT_FUNCTIONS = {"GETPIVOTDATA"}
CUBE_FUNCTIONS = {
    "CUBEVALUE",
    "CUBEMEMBER",
    "CUBESET",
    "CUBEKPIMEMBER",
    "CUBERANKEDMEMBER",
    "CUBEMEMBERPROPERTY",
}


@dataclass
class ResolutionResult:
    """
    Result of volatile/lookup function resolution.

    Attributes:
        status: "resolved" | "context_dependent" | "unresolved" |
            "conservative_fallback" | "partial_resolved" | "invalid_context"
        resolved_volatile_ref: Resolved A1 reference for OFFSET/INDIRECT
            (if resolved)
        resolved_lookup_ref: Resolved A1 reference for lookup functions
            (if resolved)
        resolved_value: Resolved primitive value (e.g., MATCH position integer)
            for non-reference results
        volatile_drivers: List of cell addresses that drive volatile resolution
        lookup_drivers: List of cell addresses that drive lookup resolution
        volatile_kind: Type of volatility
            ("time" | "random" | "workbook_meta" | "external_link" | "address_computed")
        resolution_source: Source of resolution ("automatic" | "manual")
        notes: Additional context or error messages
        partial_info: Rich metadata for partial resolutions (resolution level,
            component resolutions, etc.)
    """

    status: str = "unresolved"
    resolved_volatile_ref: str | None = None
    resolved_lookup_ref: str | None = None
    resolved_value: Any = None
    volatile_drivers: list[str] = field(default_factory=list)
    lookup_drivers: list[str] = field(default_factory=list)
    volatile_kind: str | None = None
    resolution_source: str = "automatic"
    notes: str = ""
    partial_info: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class ArgumentResolutionResult:
    """
    Rich result type for argument resolution.

    Attributes:
        value: Resolved value (concrete value or None if failed)
        drivers: Tuple of cell addresses that contributed to resolution
        success: Whether resolution succeeded
        failure_reason: Why resolution failed (if applicable)
        attempted_strategies: Strategies that were tried (in order)
    """

    value: Any
    drivers: tuple[str, ...]
    success: bool
    failure_reason: str | None = None
    attempted_strategies: tuple[str, ...] = ()


@dataclass
class SpecialDataSourceInfo:
    """
    Information about special data sources (pivot tables, cube queries, etc.).

    Attributes:
        ref_kinds: List of classification strings
        extras: Dictionary of metadata fields
    """

    ref_kinds: list[str] = field(default_factory=list)
    extras: dict[str, Any] = field(default_factory=dict)


class ResolutionEngine:
    """
    Engine for snapshot-only resolution of volatile/lookup functions and special data sources.

    Implements deterministic resolution per ADR-004:
    - Uses only value_snapshot and pure semantics
    - No recalculation
    - Deterministic across runs
    """

    def __init__(
        self,
        value_source: ValueSource | Workbook | None,
        manual_provider: ManualResolutionProvider | None = None,
    ) -> None:
        """
        Initialize resolution engine.

        Args:
            value_source: Value source (LazyValueFetcher or Workbook)
            manual_provider: Optional ManualResolutionProvider for manual overrides
        """
        self.value_source = value_source
        self.manual_provider = manual_provider
        # Classify the source ONCE: isinstance() against a runtime_checkable
        # Protocol resolves structurally via inspect.getattr_static (~10us per
        # call), and _get_cell_value runs once per scanned cell of every MATCH
        # lookup range — per-call checks dominated extraction on lookup-dense
        # sheets. value_source is only ever assigned here.
        self._vs_is_value_source = isinstance(value_source, ValueSource)
        self._vs_is_workbook = not self._vs_is_value_source and isinstance(
            value_source, Workbook
        )
        # MATCH scans are pure functions of (array, lookup_value, match_type)
        # over the immutable snapshot; formulas across a row repeat the same
        # scan verbatim. Entries are ints/None, so no eviction needed.
        self._match_scan_cache: dict[tuple[Any, ...], int | None] = {}
        self._value_index_cache: dict[
            str, tuple[dict[int, list[tuple[int, Any]]], dict[int, list[tuple[int, Any]]]]
        ] = {}

    @property
    def _sheetnames(self) -> list[str]:
        if self._vs_is_value_source or self._vs_is_workbook:
            return self.value_source.sheetnames
        return []

    @property
    def _active_sheet_name(self) -> str | None:
        if self._vs_is_value_source:
            return self.value_source.active_sheet
        elif self._vs_is_workbook:
            return self.value_source.active.title if self.value_source.active else None
        return None

    def _get_cell_value(self, sheet: str, row: int, col: int) -> Any:
        """Get cell value using value source."""
        if self._vs_is_value_source:
            coord = format_cell_address("", row, col)
            return self.value_source.get_value_at(sheet, coord)
        elif self._vs_is_workbook:
            try:
                return self.value_source[sheet].cell(row=row, column=col).value
            except (KeyError, AttributeError, IndexError):
                return None
        return None

    def _get_sheet_value_index(
        self, sheet: str
    ) -> tuple[dict[int, list[tuple[int, Any]]], dict[int, list[tuple[int, Any]]]]:
        """
        Build (and cache) sparse row/column indexes for sheet values.
        """
        cached = self._value_index_cache.get(sheet)
        if cached:
            return cached

        if self.value_source is None or not hasattr(self.value_source, "get_sheet_values"):
            empty: tuple[dict[int, list[tuple[int, Any]]], dict[int, list[tuple[int, Any]]]] = (
                {},
                {},
            )
            return empty

        values = self.value_source.get_sheet_values(sheet)
        col_index: dict[int, list[tuple[int, Any]]] = {}
        row_index: dict[int, list[tuple[int, Any]]] = {}

        for coord, value in values.items():
            parsed = parse_a1_reference(coord)
            if not parsed:
                continue
            col_index.setdefault(parsed.col, []).append((parsed.row, value))
            row_index.setdefault(parsed.row, []).append((parsed.col, value))

        for entries in col_index.values():
            entries.sort(key=lambda item: item[0])
        for entries in row_index.values():
            entries.sort(key=lambda item: item[0])

        self._value_index_cache[sheet] = (col_index, row_index)
        return col_index, row_index

    def detect_volatile(self, ast: dict[str, Any]) -> tuple[bool, list[str]]:
        """
        Detect if AST contains volatile functions.

        Args:
            ast: AST dictionary from parser

        Returns:
            Tuple of (is_volatile, volatile_function_names)

        Example:
            >>> detect_volatile(parse_formula("=OFFSET(A1,1,1)"))
            (True, ['OFFSET'])
        """
        volatile_funcs: list[str] = []
        self._collect_volatile_functions(ast, volatile_funcs)
        return (len(volatile_funcs) > 0, volatile_funcs)

    def _collect_volatile_functions(self, node: dict[str, Any], result: list[str]) -> None:
        """Recursively collect volatile function names from AST"""
        if not isinstance(node, dict):
            return

        node_type = node.get("type")

        if node_type == "Function":
            func_name = node.get("name", "").upper()
            if func_name in VOLATILE_FUNCTIONS:
                result.append(func_name)

            # Recurse into arguments
            for arg in node.get("args", []):
                self._collect_volatile_functions(arg, result)

        elif node_type == "Binary":
            left = node.get("left")
            right = node.get("right")
            if isinstance(left, dict):
                self._collect_volatile_functions(left, result)
            if isinstance(right, dict):
                self._collect_volatile_functions(right, result)

        elif node_type == "Unary":
            operand = node.get("operand")
            if isinstance(operand, dict):
                self._collect_volatile_functions(operand, result)

    def resolve_offset(
        self,
        base_ref: str,
        row_offset: int | None,
        col_offset: int | None,
        height: int | None = None,
        width: int | None = None,
        current_sheet: str = "",
    ) -> ResolutionResult:
        """
        Resolve OFFSET function to concrete range.

        OFFSET(BaseRef, r, c, h?, w?)

        Args:
            base_ref: Base reference (A1 notation)
            row_offset: Row offset (can be None if unresolved)
            col_offset: Column offset (can be None if unresolved)
            height: Optional height
            width: Optional width
            current_sheet: Current sheet name for relative references

        Returns:
            ResolutionResult with resolved reference or context_dependent status

        Example:
            >>> resolve_offset("A1", 1, 1, current_sheet="Sheet1")
            ResolutionResult(status="resolved", resolved_volatile_ref="Sheet1!B2", ...)
        """
        result = ResolutionResult()
        result.volatile_drivers = [base_ref]
        result.volatile_kind = "address_computed"

        # Check if all required parameters are resolved
        if row_offset is None or col_offset is None:
            result.status = "context_dependent"
            result.notes = "OFFSET parameters are context-dependent"
            return result

        try:
            # Parse base reference
            parsed_base = parse_cell_address(base_ref)
            if not parsed_base:
                result.status = "unresolved"
                result.notes = f"Could not parse base reference: {base_ref}"
                return result

            # Check if parsing failed (row=0, col=0 indicates unparseable)
            base_row = parsed_base.get("row", 0)
            base_col = parsed_base.get("col", 0)
            # Ensure row/col are integers for type safety
            if not isinstance(base_row, int) or not isinstance(base_col, int):
                result.status = "unresolved"
                result.notes = f"Invalid row/col types in base reference: {base_ref}"
                return result
            if base_row == 0 or base_col == 0:
                result.status = "unresolved"
                result.notes = f"Could not parse base reference: {base_ref}"
                return result

            # Get base sheet (use current sheet if not qualified)
            base_sheet_raw = parsed_base.get("sheet", "")
            base_sheet: str = base_sheet_raw if isinstance(base_sheet_raw, str) else ""
            if not base_sheet:
                base_sheet = current_sheet

            # Apply offsets
            new_row = base_row + row_offset
            new_col = base_col + col_offset

            # Validate bounds
            if new_row < 1 or new_col < 1:
                result.status = "unresolved"
                result.notes = f"OFFSET result out of bounds: row={new_row}, col={new_col}"
                return result

            # Determine result dimensions
            result_height_raw = parsed_base.get("height", 1) if height is None else height
            result_height = (
                int(result_height_raw) if isinstance(result_height_raw, (int, float)) else 1
            )

            result_width_raw = parsed_base.get("width", 1) if width is None else width
            result_width = (
                int(result_width_raw) if isinstance(result_width_raw, (int, float)) else 1
            )

            # Excel returns #REF! when OFFSET height or width is <= 0
            if result_height < 1 or result_width < 1:
                result.status = "unresolved"
                result.notes = (
                    f"OFFSET height/width out of bounds (Excel #REF!): "
                    f"height={result_height}, width={result_width}"
                )
                return result

            # Build result reference
            if result_height == 1 and result_width == 1:
                # Single cell
                resolved_ref = format_cell_address(base_sheet, new_row, new_col)
            else:
                # Range
                end_row = new_row + result_height - 1
                end_col = new_col + result_width - 1
                start_addr = format_cell_address("", new_row, new_col)
                end_addr = format_cell_address("", end_row, end_col)
                if base_sheet:
                    resolved_ref = f"{base_sheet}!{start_addr}:{end_addr}"
                else:
                    resolved_ref = f"{start_addr}:{end_addr}"

            result.status = "resolved"
            result.resolved_volatile_ref = resolved_ref
            result.notes = f"OFFSET resolved to {resolved_ref}"

        except (ValueError, KeyError, TypeError) as e:
            result.status = "unresolved"
            result.notes = f"OFFSET resolution error: {str(e)}"

        return result

    def resolve_indirect(self, ref_text: str | None, current_sheet: str = "") -> ResolutionResult:
        """
        Resolve INDIRECT function to concrete reference.

        INDIRECT(text_address)

        Args:
            ref_text: Reference text (can be None if unresolved)
            current_sheet: Current sheet name

        Returns:
            ResolutionResult with resolved reference or context_dependent status

        Example:
            >>> resolve_indirect("A1", current_sheet="Sheet1")
            ResolutionResult(status="resolved", resolved_volatile_ref="Sheet1!A1", ...)
        """
        result = ResolutionResult()
        result.volatile_kind = "address_computed"

        if ref_text is None:
            result.status = "context_dependent"
            result.notes = "INDIRECT argument is context-dependent"
            return result

        try:
            # Try to parse as A1 reference
            parsed = parse_cell_address(ref_text)
            if parsed:
                # Check if parsing failed (row=0, col=0 indicates unparseable)
                row = parsed.get("row", 0)
                col = parsed.get("col", 0)

                # Ensure row/col are integers for type safety
                if not isinstance(row, int) or not isinstance(col, int):
                    result.status = "unresolved"
                    result.notes = f"Invalid row/col types in INDIRECT: {ref_text}"
                    return result

                if row == 0 or col == 0:
                    result.status = "unresolved"
                    result.notes = f"INDIRECT could not parse: {ref_text}"
                    return result

                # Get sheet (use current sheet if not qualified)
                sheet_raw = parsed.get("sheet", "")
                sheet: str = sheet_raw if isinstance(sheet_raw, str) else ""
                if not sheet:
                    sheet = current_sheet

                height = parsed.get("height")
                width = parsed.get("width")
                if height is not None or width is not None:
                    # Range reference (e.g. "Sheet!A1:B10")
                    end_row = row + (height or 1) - 1
                    end_col = col + (width or 1) - 1
                    start_addr = format_cell_address("", row, col)
                    end_addr = format_cell_address("", end_row, end_col)
                    if sheet:
                        resolved_ref = f"{sheet}!{start_addr}:{end_addr}"
                    else:
                        resolved_ref = f"{start_addr}:{end_addr}"
                else:
                    # Single cell
                    resolved_ref = format_cell_address(sheet, row, col)
                result.status = "resolved"
                result.resolved_volatile_ref = resolved_ref
                result.notes = f"INDIRECT resolved to {resolved_ref}"
            else:
                result.status = "unresolved"
                result.notes = f"INDIRECT could not parse: {ref_text}"

        except Exception as e:
            result.status = "unresolved"
            result.notes = f"INDIRECT resolution error: {str(e)}"

        return result

    def resolve_index_match(
        self, lookup_array: str | None, match_position: int | None, current_sheet: str = ""
    ) -> ResolutionResult:
        """
        Resolve INDEX/MATCH combination to specific cell.

        Args:
            lookup_array: Lookup array reference (A1 notation)
            match_position: 1-based position from MATCH (can be None if unresolved)
            current_sheet: Current sheet name

        Returns:
            ResolutionResult with resolved cell reference

        Example:
            >>> resolve_index_match("A1:A10", 3, current_sheet="Sheet1")
            ResolutionResult(status="resolved", resolved_lookup_ref="Sheet1!A3", ...)
        """
        result = ResolutionResult()

        if lookup_array:
            result.lookup_drivers = [lookup_array]

        if lookup_array is None or match_position is None:
            result.status = "context_dependent"
            result.notes = "INDEX/MATCH parameters are context-dependent"
            return result

        try:
            # Parse lookup array
            parsed = parse_cell_address(lookup_array)
            if not parsed:
                result.status = "unresolved"
                result.notes = f"Could not parse lookup array: {lookup_array}"
                return result

            # Check if parsing failed
            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)
            # Ensure row/col are integers for type safety
            if not isinstance(start_row, int) or not isinstance(start_col, int):
                result.status = "unresolved"
                result.notes = f"Invalid row/col types in lookup array: {lookup_array}"
                return result
            if start_row == 0 or start_col == 0:
                result.status = "unresolved"
                result.notes = f"Could not parse lookup array: {lookup_array}"
                return result

            # Get sheet (use current sheet if not qualified)
            sheet_raw = parsed.get("sheet", "")
            sheet: str = sheet_raw if isinstance(sheet_raw, str) else ""
            if not sheet:
                sheet = current_sheet

            height_raw = parsed.get("height", 1)
            width_raw = parsed.get("width", 1)
            # Ensure height/width are integers
            height = int(height_raw) if isinstance(height_raw, (int, float)) else 1
            width = int(width_raw) if isinstance(width_raw, (int, float)) else 1

            # Validate position
            if match_position < 1:
                result.status = "unresolved"
                result.notes = f"Invalid MATCH position: {match_position}"
                return result

            # Determine if column or row lookup
            if height > 1 and width == 1:
                # Column lookup
                target_row = start_row + match_position - 1
                if target_row > start_row + height - 1:
                    result.status = "unresolved"
                    result.notes = f"MATCH position {match_position} out of range"
                    return result
                resolved_ref = format_cell_address(sheet, target_row, start_col)

            elif width > 1 and height == 1:
                # Row lookup
                target_col = start_col + match_position - 1
                if target_col > start_col + width - 1:
                    result.status = "unresolved"
                    result.notes = f"MATCH position {match_position} out of range"
                    return result
                resolved_ref = format_cell_address(sheet, start_row, target_col)

            else:
                # 2D array - need row and column position (not implemented here)
                result.status = "context_dependent"
                result.notes = "INDEX with 2D array requires both row and column"
                return result

            result.status = "resolved"
            result.resolved_lookup_ref = resolved_ref
            result.notes = f"INDEX/MATCH resolved to {resolved_ref}"

        except (ValueError, KeyError, TypeError) as e:
            result.status = "unresolved"
            result.notes = f"INDEX/MATCH resolution error: {str(e)}"

        return result

    def _resolve_argument(
        self, arg_node: dict[str, Any], current_sheet: str = ""
    ) -> ArgumentResolutionResult:
        """
        Resolve argument node to concrete value using snapshot.

        Handles:
        - Literals (numbers, strings): Return value directly
        - Cell references: Get value from snapshot
        - Simple expressions (A1+1): Evaluate using snapshot
        - MATCH function calls: Recursively resolve MATCH to position
        - Complex expressions: Return failure

        Args:
            arg_node: AST node of argument
            current_sheet: Current sheet name

        Returns:
            ArgumentResolutionResult with:
            - value: Concrete value (int, str, float) or None if failed
            - drivers: List of cell addresses used in resolution
            - success: True if resolved, False if unresolvable
            - failure_reason: Why resolution failed (if applicable)
            - attempted_strategies: Strategies that were tried

        Example:
            >>> _resolve_argument({"type": "Const", "value": 2})
            ArgumentResolutionResult(value=2, drivers=(), success=True)
            >>> _resolve_argument({"type": "Ref", "ref": "H10"})
            ArgumentResolutionResult(value=2, drivers=("H10",), success=True)
        """
        if not isinstance(arg_node, dict):
            return ArgumentResolutionResult(
                value=None,
                drivers=(),
                success=False,
                failure_reason="Argument is not a dict node",
                attempted_strategies=("type_check",),
            )

        node_type = arg_node.get("type")

        # Handle literals (constants)
        if node_type == "Const":
            value = arg_node.get("value")
            return ArgumentResolutionResult(
                value=value, drivers=(), success=True, attempted_strategies=("literal",)
            )

        # Handle negative literals: parser emits Unary(-, Const(n)) for e.g. -2
        # (mirrors new_arch/offset_edges._is_const_int)
        if node_type == "Unary":
            operator = arg_node.get("operator", "")
            operand = arg_node.get("operand")
            if (
                operator in ("-", "+")
                and isinstance(operand, dict)
                and operand.get("type") == "Const"
            ):
                operand_value = operand.get("value")
                if isinstance(operand_value, (int, float)) and not isinstance(operand_value, bool):
                    return ArgumentResolutionResult(
                        value=-operand_value if operator == "-" else operand_value,
                        drivers=(),
                        success=True,
                        attempted_strategies=("literal",),
                    )

        # Handle cell references
        if node_type == "Ref":
            ref = arg_node.get("ref", "")
            try:
                # Parse reference to extract sheet, row, col
                parsed = parse_cell_address(ref)
                if not parsed:
                    return ArgumentResolutionResult(
                        value=None,
                        drivers=(ref,),
                        success=False,
                        failure_reason=f"Failed to parse cell reference: {ref}",
                        attempted_strategies=("cell_ref",),
                    )

                sheet_name = parsed.get("sheet", "")
                row = parsed.get("row", 0)
                col = parsed.get("col", 0)

                if row == 0 or col == 0:
                    return ArgumentResolutionResult(
                        value=None,
                        drivers=(ref,),
                        success=False,
                        failure_reason=f"Invalid row or column in reference: {ref}",
                        attempted_strategies=("cell_ref",),
                    )

                # Get worksheet
                if sheet_name:
                    if sheet_name not in self._sheetnames:
                        return ArgumentResolutionResult(
                            value=None,
                            drivers=(ref,),
                            success=False,
                            failure_reason=f"Sheet not found: {sheet_name}",
                            attempted_strategies=("cell_ref",),
                        )
                    target_sheet = sheet_name
                elif current_sheet and current_sheet in self._sheetnames:
                    target_sheet = current_sheet
                else:
                    target_sheet = self._active_sheet_name
                    if not target_sheet:
                        return ArgumentResolutionResult(
                            value=None,
                            drivers=(ref,),
                            success=False,
                            failure_reason="No active sheet available",
                            attempted_strategies=("cell_ref",),
                        )

                # Get cell value from snapshot
                value = self._get_cell_value(target_sheet, row, col)

                return ArgumentResolutionResult(
                    value=value, drivers=(ref,), success=True, attempted_strategies=("cell_ref",)
                )

            except (ValueError, KeyError, IndexError, AttributeError) as e:
                return ArgumentResolutionResult(
                    value=None,
                    drivers=(ref,),
                    success=False,
                    failure_reason=f"Exception resolving cell reference: {str(e)}",
                    attempted_strategies=("cell_ref",),
                )

        # Handle simple binary expressions (e.g., A1+1, H10-1)
        if node_type == "Binary":
            operator = arg_node.get("operator", "")
            left_raw = arg_node.get("left")
            right_raw = arg_node.get("right")

            # Recursively resolve operands only if they're dicts
            if not isinstance(left_raw, dict) or not isinstance(right_raw, dict):
                return ArgumentResolutionResult(
                    value=None,
                    drivers=(),
                    success=False,
                    failure_reason="Binary expression operands are not dict nodes",
                    attempted_strategies=("expression",),
                )

            left_result = self._resolve_argument(left_raw, current_sheet)
            right_result = self._resolve_argument(right_raw, current_sheet)

            drivers = tuple(left_result.drivers) + tuple(right_result.drivers)

            if not (left_result.success and right_result.success):
                return ArgumentResolutionResult(
                    value=None,
                    drivers=drivers,
                    success=False,
                    failure_reason=(
                        f"Failed to resolve operands: "
                        f"left={left_result.failure_reason}, right={right_result.failure_reason}"
                    ),
                    attempted_strategies=("expression",),
                )

            # Try to evaluate simple arithmetic
            try:
                if operator == "+":
                    result = left_result.value + right_result.value
                elif operator == "-":
                    result = left_result.value - right_result.value
                elif operator == "*":
                    result = left_result.value * right_result.value
                elif operator == "/":
                    if right_result.value == 0:
                        return ArgumentResolutionResult(
                            value=None,
                            drivers=drivers,
                            success=False,
                            failure_reason="Division by zero",
                            attempted_strategies=("expression",),
                        )
                    result = left_result.value / right_result.value
                elif operator == "&":
                    # Handle string concatenation
                    left_str = str(left_result.value) if left_result.value is not None else ""
                    right_str = str(right_result.value) if right_result.value is not None else ""
                    result = left_str + right_str
                else:
                    # Unsupported operator
                    return ArgumentResolutionResult(
                        value=None,
                        drivers=drivers,
                        success=False,
                        failure_reason=f"Unsupported operator: {operator}",
                        attempted_strategies=("expression",),
                    )

                return ArgumentResolutionResult(
                    value=result,
                    drivers=drivers,
                    success=True,
                    attempted_strategies=("expression",),
                )
            except (TypeError, ValueError, ZeroDivisionError) as e:
                return ArgumentResolutionResult(
                    value=None,
                    drivers=drivers,
                    success=False,
                    failure_reason=f"Expression evaluation error: {str(e)}",
                    attempted_strategies=("expression",),
                )

        # Handle nested MATCH function calls
        if node_type == "Function":
            func_name = arg_node.get("name", "").upper()

            if func_name == "CONCATENATE":
                # Resolve all arguments and join
                resolved_args = []
                drivers: tuple[str, ...] = ()
                success = True
                failure_reasons = []
                for arg in arg_node.get("args", []):
                    arg_result = self._resolve_argument(arg, current_sheet)
                    drivers = drivers + arg_result.drivers
                    if not arg_result.success:
                        success = False
                        failure_reasons.append(arg_result.failure_reason or "unknown")
                        break
                    arg_str = str(arg_result.value) if arg_result.value is not None else ""
                    resolved_args.append(arg_str)

                if success:
                    return ArgumentResolutionResult(
                        value="".join(resolved_args),
                        drivers=drivers,
                        success=True,
                        attempted_strategies=("function_concatenate",),
                    )
                else:
                    reasons_str = ", ".join(failure_reasons)
                    return ArgumentResolutionResult(
                        value=None,
                        drivers=drivers,
                        success=False,
                        failure_reason=f"CONCATENATE argument resolution failed: {reasons_str}",
                        attempted_strategies=("function_concatenate",),
                    )

            if func_name == "MATCH":
                # Recursively resolve MATCH
                match_result = self.resolve_match_semantic(arg_node, current_sheet)
                if match_result.status == "resolved" and match_result.resolved_value is not None:
                    return ArgumentResolutionResult(
                        value=match_result.resolved_value,
                        drivers=match_result.lookup_drivers,
                        success=True,
                        attempted_strategies=("function_match",),
                    )
                else:
                    return ArgumentResolutionResult(
                        value=None,
                        drivers=match_result.lookup_drivers,
                        success=False,
                        failure_reason=f"MATCH resolution failed: {match_result.notes}",
                        attempted_strategies=("function_match",),
                    )

        # Unresolvable (complex expression, IF statement, etc.)
        return ArgumentResolutionResult(
            value=None,
            drivers=(),
            success=False,
            failure_reason=f"Unresolvable node type: {node_type}",
            attempted_strategies=("fallback",),
        )

    def _match_scan_position(
        self,
        target_sheet: str,
        start_row: int,
        start_col: int,
        height: int,
        width: int,
        lookup_value: Any,
        match_type: int,
    ) -> int | None:
        """Scan a MATCH lookup array; return the 1-based match position or None.

        The scan is a pure function of the immutable value snapshot, so results
        are memoized: lookup-dense sheets repeat the identical scan for every
        formula in a row, and each scan reads O(height) cells. The cache key
        carries the lookup value's type name because Python hashes True == 1
        while Excel MATCH treats logicals and numbers as distinct type classes;
        strings are lowered to mirror Excel's case-insensitive collation.
        """
        key = (
            target_sheet,
            start_row,
            start_col,
            height,
            width,
            match_type,
            type(lookup_value).__name__,
            lookup_value.lower() if isinstance(lookup_value, str) else lookup_value,
        )
        cacheable = True
        try:
            return self._match_scan_cache[key]
        except KeyError:
            pass
        except TypeError:  # unhashable lookup value — scan without caching
            cacheable = False

        position: int | None = None
        end_row = start_row + height - 1
        end_col = start_col + width - 1

        def _type_key(value: Any) -> str | None:
            """Excel type class for MATCH comparison (logical is not a number)."""
            if isinstance(value, bool):
                return "bool"
            if isinstance(value, (int, float)):
                return "number"
            if isinstance(value, str):
                return "str"
            return None

        lookup_type = _type_key(lookup_value)

        def _update_match_position(cell_value: Any, index: int) -> bool:
            """
            Update position based on match_type. Returns True if search should stop.
            """
            nonlocal position
            if match_type == 0:  # Exact match (case-insensitive for strings, as Excel)
                if isinstance(cell_value, str) and isinstance(lookup_value, str):
                    if cell_value.lower() == lookup_value.lower():
                        position = index  # 1-based position
                        return True
                elif cell_value == lookup_value:
                    position = index  # 1-based position
                    return True
            elif match_type in (1, -1):
                # Excel's approximate MATCH ignores cells whose type class
                # differs from the lookup_value's (text vs number vs logical)
                if cell_value is None or _type_key(cell_value) != lookup_type:
                    return False
                cell_cmp, lookup_cmp = cell_value, lookup_value
                if lookup_type == "str":
                    # Excel text collation is case-insensitive
                    cell_cmp, lookup_cmp = cell_cmp.lower(), lookup_cmp.lower()
                if match_type == 1:  # Largest value <= lookup (array assumed ascending)
                    if cell_cmp <= lookup_cmp:
                        position = index  # Keep updating to get the largest match
                    else:
                        return True
                else:  # Smallest value >= lookup (array assumed descending)
                    if cell_cmp >= lookup_cmp:
                        position = index  # Keep updating to get the smallest match
                    else:
                        return True
            return False

        # Determine if column or row array
        use_sparse_index = (
            (height == EXCEL_MAX_ROWS or width == EXCEL_MAX_COLS)
            and self.value_source is not None
            and hasattr(self.value_source, "get_sheet_values")
        )

        if use_sparse_index and height > 1 and width == 1:
            # Column array (sparse index)
            col_index, _ = self._get_sheet_value_index(target_sheet)
            for row, cell_value in col_index.get(start_col, []):
                if row < start_row:
                    continue
                if row > end_row:
                    break
                if _update_match_position(cell_value, row - start_row + 1):
                    break
        elif use_sparse_index and width > 1 and height == 1:
            # Row array (sparse index)
            _, row_index = self._get_sheet_value_index(target_sheet)
            for col, cell_value in row_index.get(start_row, []):
                if col < start_col:
                    continue
                if col > end_col:
                    break
                if _update_match_position(cell_value, col - start_col + 1):
                    break
        elif height > 1 and width == 1:
            # Column array
            for i in range(height):
                cell_row = start_row + i
                cell_value = self._get_cell_value(target_sheet, cell_row, start_col)

                if _update_match_position(cell_value, i + 1):
                    break

        else:
            # Row array (caller guarantees width > 1 and height == 1)
            for i in range(width):
                cell_col = start_col + i
                cell_value = self._get_cell_value(target_sheet, start_row, cell_col)

                if _update_match_position(cell_value, i + 1):
                    break

        if cacheable:
            self._match_scan_cache[key] = position
        return position

    def resolve_match_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve MATCH function to position integer using snapshot.

        Per design doc §3: MATCH returns position by searching lookup_array.
        Dependencies: lookup_value (if cell ref) AND lookup_array (always).

        Args:
            ast: AST of MATCH function call
            current_sheet: Current sheet context for unqualified references
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "conservative_fallback"
            - resolved_value: Integer position (1-based) if resolved
            - lookup_drivers: [lookup_value_cells, lookup_array_range]
            - notes: Position found and dependencies

        Algorithm:
            1. Parse arguments: MATCH(lookup_value, lookup_array, match_type)
            2. Resolve lookup_value using _resolve_argument()
            3. Parse lookup_array range
            4. If lookup_value resolved: Search array using match_type logic
            5. Return ResolutionResult with position and dependencies
        """
        result = ResolutionResult()

        # Check for manual override first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_value=manual_res.get("resolved_value"),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "MATCH":
            result.status = "unresolved"
            result.notes = "Not a MATCH function"
            return result

        args = ast.get("args", [])
        if len(args) < 2:
            result.status = "unresolved"
            result.notes = "MATCH requires at least 2 arguments"
            return result

        # Extract arguments
        lookup_value_arg = args[0]
        lookup_array_arg = args[1]
        match_type = 1  # Excel default: approximate match (largest value <= lookup)
        if len(args) >= 3:
            match_type_result = self._resolve_argument(args[2], current_sheet)
            if match_type_result.success and isinstance(match_type_result.value, (int, float)):
                match_type = int(match_type_result.value)

        # Resolve lookup_value
        lookup_value_result = self._resolve_argument(lookup_value_arg, current_sheet)
        lookup_value = lookup_value_result.value
        lookup_value_drivers = lookup_value_result.drivers
        lookup_value_success = lookup_value_result.success

        # Parse lookup_array
        if lookup_array_arg.get("type") != "Ref":
            result.status = "unresolved"
            result.notes = "MATCH lookup_array must be a range reference"
            return result

        lookup_array_ref = lookup_array_arg.get("ref", "")
        result.lookup_drivers = list(lookup_value_drivers) + [lookup_array_ref]

        if not lookup_value_success:
            result.status = "conservative_fallback"
            result.notes = "MATCH lookup_value could not be resolved"
            return result

        # Wildcard patterns (*/?) in exact match are not supported; fall back
        # conservatively rather than returning a literal-comparison position
        if (
            match_type == 0
            and isinstance(lookup_value, str)
            and ("*" in lookup_value or "?" in lookup_value)
        ):
            result.status = "conservative_fallback"
            result.notes = "MATCH wildcard patterns are not supported"
            return result

        # Parse lookup_array range
        try:
            parsed = parse_cell_address(lookup_array_ref)
            if not parsed:
                result.status = "unresolved"
                result.notes = f"Could not parse lookup_array: {lookup_array_ref}"
                return result

            sheet_name = parsed.get("sheet", "")
            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)
            height = parsed.get("height", 1)
            width = parsed.get("width", 1)

            # Ensure dimensions are integers
            if not isinstance(start_row, int) or not isinstance(start_col, int):
                result.status = "unresolved"
                result.notes = f"Invalid row/col types in lookup_array: {lookup_array_ref}"
                return result
            if not isinstance(height, int) or not isinstance(width, int):
                height = 1
                width = 1

            if start_row == 0 or start_col == 0:
                result.status = "unresolved"
                result.notes = f"Invalid lookup_array: {lookup_array_ref}"
                return result

            # Get worksheet
            if sheet_name:
                if sheet_name not in self._sheetnames:
                    result.status = "unresolved"
                    result.notes = f"Sheet not found: {sheet_name}"
                    return result
                target_sheet = sheet_name
            elif current_sheet and current_sheet in self._sheetnames:
                target_sheet = current_sheet
            else:
                target_sheet = self._active_sheet_name
                if not target_sheet:
                    result.status = "unresolved"
                    result.notes = "No active sheet found"
                    return result

            # Search array for matching value (memoized; see _match_scan_position)
            if not ((height > 1 and width == 1) or (width > 1 and height == 1)):
                result.status = "unresolved"
                result.notes = "MATCH lookup_array must be a single row or column"
                return result
            position = self._match_scan_position(
                target_sheet, start_row, start_col, height, width, lookup_value, match_type
            )

            if position is not None:
                result.status = "resolved"
                result.resolved_value = position
                result.notes = f"MATCH found at position {position}"
            else:
                result.status = "conservative_fallback"
                result.notes = f"MATCH value '{lookup_value}' not found in array"
                result.resolved_value = None

            return result

        except (ValueError, KeyError, IndexError, TypeError, AttributeError) as e:
            result.status = "unresolved"
            result.notes = f"MATCH resolution error: {str(e)}"
            return result

    def resolve_vlookup_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve VLOOKUP to semantic return column using snapshot.

        Per design doc §3: VLOOKUP depends on return column ONLY, not entire table.

        Args:
            ast: AST of VLOOKUP function call
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "conservative_fallback"
            - resolved_lookup_ref: A1 address of return column (e.g., "F100:F200")
            - lookup_drivers: [lookup_value, table_array]
            - notes: Explanation of resolution

        Algorithm:
            1. Check manual resolution provider first (if cell_address provided)
            2. Parse arguments: VLOOKUP(lookup_value, table_array, col_index_num, range_lookup)
            3. Extract table_array dimensions (start_col, start_row, width, height)
            4. Resolve col_index_num using _resolve_argument()
            5. If col_index_num resolved:
                - Validate: 1 <= col_index_num <= width
                - Compute return_col = start_col + (col_index_num - 1)
                - Build semantic range: sheet!{return_col}{start_row}:{return_col}{end_row}
                - Return ResolutionResult(status="resolved",
                  resolved_lookup_ref=semantic_range)
            6. If unresolvable or invalid:
                - Return ResolutionResult(status="conservative_fallback",
                  resolved_lookup_ref=table_array)
        """
        result = ResolutionResult()

        # Check manual resolution first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "VLOOKUP":
            result.status = "unresolved"
            result.notes = "Not a VLOOKUP function"
            return result

        args = ast.get("args", [])
        if len(args) < 3:
            result.status = "unresolved"
            result.notes = "VLOOKUP requires at least 3 arguments"
            return result

        # Extract arguments
        lookup_value_arg = args[0]
        table_array_arg = args[1]
        col_index_arg = args[2]

        # Get lookup_value drivers
        lookup_value_result = self._resolve_argument(lookup_value_arg, current_sheet)
        lookup_value_drivers = lookup_value_result.drivers

        # Parse table_array - handle both simple Ref and complex expressions (OFFSET, INDIRECT, etc.)
        from xl_marinade.core.strategies.index_strategies import _extract_array_dependencies

        resolved_ref, table_drivers = _extract_array_dependencies(table_array_arg)

        if not table_drivers:
            result.status = "unresolved"
            result.notes = "VLOOKUP table_array has no extractable references"
            return result

        # Use resolved_ref as primary reference, or first driver if none
        table_array_ref = resolved_ref if resolved_ref else table_drivers[0]
        result.lookup_drivers = list(lookup_value_drivers) + table_drivers

        # Parse table_array dimensions
        try:
            parsed = parse_cell_address(table_array_ref)
            if not parsed:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Could not parse table_array: {table_array_ref}"
                return result

            sheet_name = parsed.get("sheet", "")
            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)
            height = parsed.get("height", 1)
            width = parsed.get("width", 1)

            # Ensure dimensions are integers
            if not isinstance(start_row, int) or not isinstance(start_col, int):
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Invalid row/col types in table_array: {table_array_ref}"
                return result
            if not isinstance(height, int) or not isinstance(width, int):
                height = 1
                width = 1

            if start_row == 0 or start_col == 0:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Invalid table_array: {table_array_ref}"
                return result

            # Resolve col_index_num
            col_index_result = self._resolve_argument(col_index_arg, current_sheet)
            result.lookup_drivers.extend(list(col_index_result.drivers))

            if not col_index_result.success:
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = "VLOOKUP col_index_num could not be resolved; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "dynamic_col_index",
                }
                return result

            # Validate col_index is integer
            if not isinstance(col_index_result.value, (int, float)):
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = "VLOOKUP col_index_num is not numeric; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "non_numeric_col_index",
                }
                return result

            col_index_num = int(col_index_result.value)

            # Validate col_index is in valid range [1, width]
            if col_index_num < 1 or col_index_num > width:
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"VLOOKUP col_index_num {col_index_num} out of range [1, {width}]; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "col_index_out_of_range",
                }
                return result

            # Compute return column
            return_col = start_col + (col_index_num - 1)
            end_row = start_row + height - 1

            # Build semantic range
            return_col_letter = col_num_to_letter(return_col)

            if sheet_name:
                semantic_range = (
                    f"{sheet_name}!{return_col_letter}{start_row}:{return_col_letter}{end_row}"
                )
            else:
                semantic_range = f"{return_col_letter}{start_row}:{return_col_letter}{end_row}"

            result.status = "resolved"
            result.resolved_lookup_ref = semantic_range
            result.notes = (
                f"VLOOKUP resolved to column {col_index_num} (column {return_col_letter})"
            )

            return result

        except (ValueError, KeyError, IndexError, TypeError, AttributeError) as e:
            result.status = "conservative_fallback"
            result.resolved_lookup_ref = table_array_ref
            result.notes = f"VLOOKUP resolution error: {str(e)}"
            return result

    def resolve_hlookup_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve HLOOKUP to semantic return row using snapshot.

        Per design doc §3: HLOOKUP depends on return row ONLY, not entire table.
        Similar to VLOOKUP but horizontal.

        Args:
            ast: AST of HLOOKUP function call
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "conservative_fallback"
            - resolved_lookup_ref: A1 address of return row (e.g., "E6:Z6")
            - lookup_drivers: [lookup_value, table_array]
            - notes: Explanation of resolution

        Algorithm:
            1. Parse arguments: HLOOKUP(lookup_value, table_array, row_index_num, range_lookup)
            2. Extract table_array dimensions (start_col, start_row, width, height)
            3. Resolve row_index_num using _resolve_argument()
            4. If row_index_num resolved:
                - Validate: 1 <= row_index_num <= height
                - Compute return_row = start_row + (row_index_num - 1)
                - Build semantic range: sheet!{start_col}{return_row}:{end_col}{return_row}
                - Return ResolutionResult(status="resolved",
                  resolved_lookup_ref=semantic_range)
            5. If unresolvable or invalid:
                - Return ResolutionResult(status="conservative_fallback",
                  resolved_lookup_ref=table_array)
        """
        result = ResolutionResult()

        # Check for manual override first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "HLOOKUP":
            result.status = "unresolved"
            result.notes = "Not an HLOOKUP function"
            return result

        args = ast.get("args", [])
        if len(args) < 3:
            result.status = "unresolved"
            result.notes = "HLOOKUP requires at least 3 arguments"
            return result

        # Extract arguments
        lookup_value_arg = args[0]
        table_array_arg = args[1]
        row_index_arg = args[2]

        # Get lookup_value drivers
        lookup_value_result = self._resolve_argument(lookup_value_arg, current_sheet)
        lookup_value_drivers = lookup_value_result.drivers

        # Parse table_array - handle both simple Ref and complex expressions (OFFSET, INDIRECT, etc.)
        from xl_marinade.core.strategies.index_strategies import _extract_array_dependencies

        resolved_ref, table_drivers = _extract_array_dependencies(table_array_arg)

        if not table_drivers:
            result.status = "unresolved"
            result.notes = "HLOOKUP table_array has no extractable references"
            return result

        # Use resolved_ref as primary reference, or first driver if none
        table_array_ref = resolved_ref if resolved_ref else table_drivers[0]
        result.lookup_drivers = list(lookup_value_drivers) + table_drivers

        # Parse table_array dimensions
        try:
            parsed = parse_cell_address(table_array_ref)
            if not parsed:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Could not parse table_array: {table_array_ref}"
                return result

            sheet_name = parsed.get("sheet", "")
            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)
            height = parsed.get("height", 1)
            width = parsed.get("width", 1)

            # Ensure dimensions are integers
            if not isinstance(start_row, int) or not isinstance(start_col, int):
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Invalid row/col types in table_array: {table_array_ref}"
                return result
            if not isinstance(height, int) or not isinstance(width, int):
                height = 1
                width = 1

            if start_row == 0 or start_col == 0:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"Invalid table_array: {table_array_ref}"
                return result

            # Resolve row_index_num
            row_index_result = self._resolve_argument(row_index_arg, current_sheet)
            result.lookup_drivers.extend(list(row_index_result.drivers))

            if not row_index_result.success:
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = "HLOOKUP row_index_num could not be resolved; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "dynamic_row_index",
                }
                return result

            # Validate row_index is integer
            if not isinstance(row_index_result.value, (int, float)):
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = "HLOOKUP row_index_num is not numeric; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "non_numeric_row_index",
                }
                return result

            row_index_num = int(row_index_result.value)

            # Validate row_index is in valid range [1, height]
            if row_index_num < 1 or row_index_num > height:
                # Per ADR-041: Return partial_resolved with full table dependency
                result.status = "partial_resolved"
                result.resolved_lookup_ref = table_array_ref
                result.notes = f"HLOOKUP row_index_num {row_index_num} out of range [1, {height}]; using full table"
                result.partial_info = {
                    "resolution_level": "full_table",
                    "reason": "row_index_out_of_range",
                }
                return result

            # Compute return row
            return_row = start_row + (row_index_num - 1)
            end_col = start_col + width - 1

            # Build semantic range
            start_col_letter = col_num_to_letter(start_col)
            end_col_letter = col_num_to_letter(end_col)

            if sheet_name:
                semantic_range = (
                    f"{sheet_name}!{start_col_letter}{return_row}:{end_col_letter}{return_row}"
                )
            else:
                semantic_range = f"{start_col_letter}{return_row}:{end_col_letter}{return_row}"

            result.status = "resolved"
            result.resolved_lookup_ref = semantic_range
            result.notes = f"HLOOKUP resolved to row {row_index_num} (row {return_row})"

            return result

        except (ValueError, KeyError, IndexError, TypeError, AttributeError) as e:
            result.status = "conservative_fallback"
            result.resolved_lookup_ref = table_array_ref
            result.notes = f"HLOOKUP resolution error: {str(e)}"
            return result

    def resolve_index_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve INDEX to specific cell, row, or column using snapshot.

        Per design doc §3: INDEX can resolve to exact cell, partial row/col, or unresolved.

        Args:
            ast: AST of INDEX function call
            current_sheet: Current sheet context for unqualified references
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "partial_resolved" | "unresolved"
            - resolved_lookup_ref: Cell/range address
            - lookup_drivers: [array, row_num_cells, col_num_cells]
            - notes: Explanation of resolution

        Algorithm:
            1. Parse arguments: INDEX(array, row_num, [column_num])
            2. Parse array to extract dimensions
            3. Resolve row_num and column_num (aggressive resolution)
            4. If both resolved:
                - Compute exact cell: array[row_num, column_num]
                - Status: "resolved"
            5. If row resolved, not column (2D array):
                - Return entire row from array
                - Status: "partial_resolved"
            6. If column resolved, not row (2D array):
                - Return entire column from array
                - Status: "partial_resolved"
            7. If neither resolved:
                - Status: "unresolved", No edge created

        Special cases:
            - row_num=0: Return entire column (if column_num specified)
            - column_num=0: Return entire row (if row_num specified)
        """
        result = ResolutionResult()

        # Check for manual override first (same pattern as VLOOKUP)
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "INDEX":
            result.status = "unresolved"
            result.notes = "Not an INDEX function"
            return result

        args = ast.get("args", [])
        if len(args) < 2:
            result.status = "unresolved"
            result.notes = "INDEX requires at least 2 arguments"
            return result

        # Extract arguments
        array_arg = args[0]
        row_num_arg = args[1]
        column_num_arg = args[2] if len(args) >= 3 else None

        # Parse array
        if array_arg.get("type") != "Ref":
            result.status = "unresolved"
            result.notes = "INDEX array must be a range reference"
            return result

        array_ref = array_arg.get("ref", "")
        result.lookup_drivers = [array_ref]

        # Parse array dimensions
        try:
            parsed = parse_cell_address(array_ref)
            if not parsed:
                result.status = "unresolved"
                result.notes = f"Could not parse array: {array_ref}"
                return result

            sheet_name_raw = parsed.get("sheet", "")
            # If no sheet specified in reference, use active sheet
            sheet_name: str = sheet_name_raw if isinstance(sheet_name_raw, str) else ""
            if not sheet_name and current_sheet:
                sheet_name = current_sheet
            elif not sheet_name:
                sheet_name = self._active_sheet_name or ""

            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)
            height_raw = parsed.get("height", 1)
            width_raw = parsed.get("width", 1)

            # Ensure dimensions are integers
            if not isinstance(start_row, int) or not isinstance(start_col, int):
                result.status = "unresolved"
                result.notes = f"Invalid row/col types in array: {array_ref}"
                return result
            height = int(height_raw) if isinstance(height_raw, (int, float)) else 1
            width = int(width_raw) if isinstance(width_raw, (int, float)) else 1

            if start_row == 0 or start_col == 0:
                result.status = "unresolved"
                result.notes = f"Invalid array: {array_ref}"
                return result

            # Resolve row_num
            row_num_result = self._resolve_argument(row_num_arg, current_sheet)
            row_num_value = row_num_result.value
            row_num_drivers = row_num_result.drivers
            row_num_success = row_num_result.success
            result.lookup_drivers.extend(row_num_drivers)

            # Resolve column_num (if provided)
            column_num_value = None
            column_num_success = False
            if column_num_arg is not None:
                column_num_result = self._resolve_argument(column_num_arg, current_sheet)
                column_num_value = column_num_result.value
                column_num_drivers = column_num_result.drivers
                column_num_success = column_num_result.success
                result.lookup_drivers.extend(column_num_drivers)
            else:
                # For 1D arrays, column_num is implicit
                # But only if row_num was successfully resolved
                if width == 1 and row_num_success:
                    column_num_value = 1
                    column_num_success = True
                elif height == 1 and row_num_success:
                    # Row array - use row_num as column position
                    column_num_value = row_num_value
                    column_num_success = row_num_success
                    row_num_value = 1
                    row_num_success = True

            # Convert to int if successful
            if row_num_success and isinstance(row_num_value, (int, float)):
                row_num_value = int(row_num_value)
            else:
                row_num_value = None

            if column_num_success and isinstance(column_num_value, (int, float)):
                column_num_value = int(column_num_value)
            else:
                column_num_value = None

            # Handle special cases: row=0 or col=0
            # row=0 means return entire column
            # col=0 means return entire row
            if row_num_value == 0:
                row_num_value = None
            if column_num_value == 0:
                column_num_value = None

            # Decision logic based on what's resolved
            if row_num_value is not None and column_num_value is not None:
                # Both resolved - compute exact cell
                if row_num_value < 1 or row_num_value > height:
                    result.status = "unresolved"
                    result.notes = f"INDEX row_num {row_num_value} out of range [1, {height}]"
                    return result
                if column_num_value < 1 or column_num_value > width:
                    result.status = "unresolved"
                    result.notes = f"INDEX column_num {column_num_value} out of range [1, {width}]"
                    return result

                target_row = start_row + (row_num_value - 1)
                target_col = start_col + (column_num_value - 1)

                resolved_ref = format_cell_address(sheet_name, target_row, target_col)
                result.status = "resolved"
                result.resolved_lookup_ref = resolved_ref
                result.notes = f"INDEX resolved to cell [{row_num_value}, {column_num_value}]"

            elif row_num_value is not None and column_num_value is None:
                # Row resolved, column not - return entire row
                if row_num_value < 1 or row_num_value > height:
                    result.status = "unresolved"
                    result.notes = f"INDEX row_num {row_num_value} out of range [1, {height}]"
                    return result

                target_row = start_row + (row_num_value - 1)
                end_col = start_col + width - 1

                start_col_letter = col_num_to_letter(start_col)
                end_col_letter = col_num_to_letter(end_col)

                if sheet_name:
                    resolved_ref = (
                        f"{sheet_name}!{start_col_letter}{target_row}:{end_col_letter}{target_row}"
                    )
                else:
                    resolved_ref = f"{start_col_letter}{target_row}:{end_col_letter}{target_row}"

                result.status = "partial_resolved"
                result.resolved_lookup_ref = resolved_ref
                result.notes = f"INDEX resolved to row {row_num_value}"

            elif row_num_value is None and column_num_value is not None:
                # Column resolved, row not - return entire column
                if column_num_value < 1 or column_num_value > width:
                    result.status = "unresolved"
                    result.notes = f"INDEX column_num {column_num_value} out of range [1, {width}]"
                    return result

                target_col = start_col + (column_num_value - 1)
                end_row = start_row + height - 1

                target_col_letter = col_num_to_letter(target_col)

                if sheet_name:
                    resolved_ref = (
                        f"{sheet_name}!{target_col_letter}{start_row}:{target_col_letter}{end_row}"
                    )
                else:
                    resolved_ref = f"{target_col_letter}{start_row}:{target_col_letter}{end_row}"

                result.status = "partial_resolved"
                result.resolved_lookup_ref = resolved_ref
                result.notes = f"INDEX resolved to column {column_num_value}"

            else:
                # Neither resolved
                result.status = "unresolved"
                result.notes = "INDEX row_num and column_num could not be resolved"

            return result

        except (ValueError, KeyError, IndexError, TypeError, AttributeError) as e:
            result.status = "unresolved"
            result.notes = f"INDEX resolution error: {str(e)}"
            return result

    def resolve_xlookup_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve XLOOKUP to return_array using snapshot.

        Per design doc §3: XLOOKUP semantic dependency is return_array ONLY, not lookup_array.

        Args:
            ast: AST of XLOOKUP function call
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "conservative_fallback"
            - resolved_lookup_ref: A1 address of return_array
            - lookup_drivers: [lookup_value, lookup_array, return_array]
            - notes: Explanation of resolution

        Algorithm:
            1. Parse arguments: XLOOKUP(lookup_value, lookup_array, return_array, ...)
            2. Extract return_array reference
            3. Return ResolutionResult(status="resolved", resolved_lookup_ref=return_array)
            4. If parsing fails, use conservative_fallback
        """
        result = ResolutionResult()

        # Check for manual override first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "XLOOKUP":
            result.status = "unresolved"
            result.notes = "Not an XLOOKUP function"
            return result

        args = ast.get("args", [])
        if len(args) < 3:
            result.status = "unresolved"
            result.notes = "XLOOKUP requires at least 3 arguments"
            return result

        # Extract arguments
        lookup_value_arg = args[0]
        lookup_array_arg = args[1]
        return_array_arg = args[2]

        # Get lookup_value drivers
        lookup_value_result = self._resolve_argument(lookup_value_arg, current_sheet)
        lookup_value_drivers = lookup_value_result.drivers

        # Parse lookup_array
        lookup_array_ref = None
        if lookup_array_arg.get("type") == "Ref":
            lookup_array_ref = lookup_array_arg.get("ref", "")

        # Parse return_array
        if return_array_arg.get("type") != "Ref":
            result.status = "unresolved"
            result.notes = "XLOOKUP return_array must be a range reference"
            return result

        return_array_ref = return_array_arg.get("ref", "")

        # Build driver list (convert tuple to list for mutation)
        drivers_list = list(lookup_value_drivers)
        if lookup_array_ref:
            drivers_list.append(lookup_array_ref)
        drivers_list.append(return_array_ref)
        result.lookup_drivers = drivers_list

        # XLOOKUP resolution is straightforward: semantic dependency is return_array only
        try:
            parsed = parse_cell_address(return_array_ref)
            if not parsed:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = return_array_ref
                result.notes = f"Could not parse return_array: {return_array_ref}"
                return result

            # Validate it's a valid reference
            start_row = parsed.get("row", 0)
            start_col = parsed.get("col", 0)

            if start_row == 0 or start_col == 0:
                result.status = "conservative_fallback"
                result.resolved_lookup_ref = return_array_ref
                result.notes = f"Invalid return_array: {return_array_ref}"
                return result

            result.status = "resolved"
            result.resolved_lookup_ref = return_array_ref
            result.notes = f"XLOOKUP resolved to return_array: {return_array_ref}"

            return result

        except (ValueError, KeyError, IndexError, TypeError, AttributeError) as e:
            result.status = "conservative_fallback"
            result.resolved_lookup_ref = return_array_ref if return_array_ref else ""
            result.notes = f"XLOOKUP resolution error: {str(e)}"
            return result

    def resolve_choose_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve CHOOSE function to selected value argument using snapshot.

        Per design doc §3: CHOOSE(index_num, value1, value2, ...)
        - If index_num resolved: Return specific value at that position
        - If unresolvable: Return union of all value arguments (conservative_fallback)

        Args:
            ast: AST of CHOOSE function call
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "conservative_fallback"
            - resolved_lookup_ref: Cell reference of selected value (if cell ref)
            - lookup_drivers: [index_num_cells, all_value_refs]
            - notes: Explanation of resolution

        Algorithm:
            1. Parse arguments: CHOOSE(index_num, value1, [value2], ...)
            2. Resolve index_num using _resolve_argument()
            3. If index_num resolved:
                - Extract value at position index_num
                - If it's a cell reference, that's the semantic dependency
                - Status: "resolved"
            4. If index_num unresolvable:
                - Dependency: All value arguments that are cell references (union)
                - Status: "conservative_fallback"
        """
        result = ResolutionResult()

        # Check for manual override first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "CHOOSE":
            result.status = "unresolved"
            result.notes = "Not a CHOOSE function"
            return result

        args = ast.get("args", [])
        if len(args) < 2:
            result.status = "unresolved"
            result.notes = "CHOOSE requires at least 2 arguments"
            return result

        # Extract arguments
        index_num_arg = args[0]
        value_args = args[1:]

        # Resolve index_num
        index_num_result = self._resolve_argument(index_num_arg, current_sheet)
        index_num_value = index_num_result.value
        index_num_drivers = index_num_result.drivers
        index_num_success = index_num_result.success

        # Collect all value references (for conservative fallback)
        all_value_refs: list[str] = []
        for value_arg in value_args:
            if value_arg.get("type") == "Ref":
                all_value_refs.append(value_arg.get("ref", ""))

        # Build driver list (convert tuple to list for mutation)
        result.lookup_drivers = list(index_num_drivers) + all_value_refs

        if not index_num_success or not isinstance(index_num_value, (int, float)):
            # Cannot resolve index - use all value refs as conservative fallback
            result.status = "conservative_fallback"
            if all_value_refs:
                result.resolved_lookup_ref = ",".join(all_value_refs)
            result.notes = "CHOOSE index_num could not be resolved"
            return result

        index_num = int(index_num_value)

        # Validate index is in range [1, num_values]
        if index_num < 1 or index_num > len(value_args):
            result.status = "conservative_fallback"
            if all_value_refs:
                result.resolved_lookup_ref = ",".join(all_value_refs)
            result.notes = f"CHOOSE index {index_num} out of range [1, {len(value_args)}]"
            return result

        # Get the selected value argument (1-based index)
        selected_arg = value_args[index_num - 1]

        # If it's a cell reference, that's our semantic dependency
        if selected_arg.get("type") == "Ref":
            result.status = "resolved"
            result.resolved_lookup_ref = selected_arg.get("ref", "")
            result.notes = f"CHOOSE resolved to argument {index_num}"
        else:
            # Selected value is not a cell reference (literal, expression, etc.)
            # No cell dependency from selected value
            result.status = "resolved"
            result.resolved_lookup_ref = None
            result.notes = f"CHOOSE resolved to argument {index_num} (not a cell reference)"

        return result

    def resolve_address_semantic(
        self, ast: dict[str, Any], current_sheet: str = "", cell_address: str | None = None
    ) -> ResolutionResult:
        """
        Resolve ADDRESS function to computed cell address using snapshot.

        Per design doc §3: ADDRESS(row_num, column_num, [abs_num], [a1], [sheet_text])
        - If both row_num and column_num resolved: Compute cell address
        - If unresolvable: Status "unresolved", no edge created

        Args:
            ast: AST of ADDRESS function call
            cell_address: Optional cell address for manual resolution lookup

        Returns:
            ResolutionResult with:
            - status: "resolved" | "unresolved"
            - resolved_lookup_ref: Computed cell address (e.g., "C5" or "Sheet2!C5")
            - lookup_drivers: [row_num_cells, column_num_cells]
            - notes: Explanation of resolution

        Algorithm:
            1. Parse arguments: ADDRESS(row_num, column_num, ...)
            2. Resolve row_num and column_num using _resolve_argument()
            3. If both resolved:
                - Compute cell address: COL{column_num}{row_num}
                - Handle optional sheet_text (5th argument)
                - Status: "resolved"
            4. If unresolvable:
                - Status: "unresolved", no edge created
        """
        result = ResolutionResult()

        # Check for manual override first
        if self.manual_provider and cell_address:
            manual_res = self.manual_provider.get_resolution(cell_address)
            if manual_res:
                return ResolutionResult(
                    status="resolved",
                    resolved_lookup_ref=manual_res.get("resolved_ref", ""),
                    resolution_source="manual",
                    notes=manual_res.get("reason", "Manual override"),
                    partial_info={"resolution_level": "manual"},
                )

        # Validate function type
        if ast.get("type") != "Function" or ast.get("name", "").upper() != "ADDRESS":
            result.status = "unresolved"
            result.notes = "Not an ADDRESS function"
            return result

        args = ast.get("args", [])
        if len(args) < 2:
            result.status = "unresolved"
            result.notes = "ADDRESS requires at least 2 arguments"
            return result

        # Extract arguments
        row_num_arg = args[0]
        column_num_arg = args[1]
        # Optional: abs_num (arg 3), a1 (arg 4), sheet_text (arg 5)
        sheet_text_arg = args[4] if len(args) >= 5 else None

        # Resolve row_num
        row_num_result = self._resolve_argument(row_num_arg, current_sheet)
        row_num_value = row_num_result.value
        row_num_drivers = row_num_result.drivers
        row_num_success = row_num_result.success
        result.lookup_drivers = row_num_drivers.copy()

        # Resolve column_num
        column_num_result = self._resolve_argument(column_num_arg, current_sheet)
        column_num_value = column_num_result.value
        column_num_drivers = column_num_result.drivers
        column_num_success = column_num_result.success
        result.lookup_drivers.extend(column_num_drivers)

        if not (row_num_success and column_num_success):
            result.status = "unresolved"
            result.notes = "ADDRESS row_num or column_num could not be resolved"
            return result

        # Validate both are integers
        if not isinstance(row_num_value, (int, float)) or not isinstance(
            column_num_value, (int, float)
        ):
            result.status = "unresolved"
            result.notes = "ADDRESS row_num or column_num is not numeric"
            return result

        row_num = int(row_num_value)
        column_num = int(column_num_value)

        # Validate bounds
        if row_num < 1 or column_num < 1:
            result.status = "unresolved"
            result.notes = f"ADDRESS row/column out of bounds: row={row_num}, col={column_num}"
            return result

        # Resolve optional sheet_text
        sheet_name = ""
        if sheet_text_arg is not None:
            sheet_text_result = self._resolve_argument(sheet_text_arg, current_sheet)
            sheet_text_value = sheet_text_result.value
            sheet_text_drivers = sheet_text_result.drivers
            sheet_text_success = sheet_text_result.success
            result.lookup_drivers.extend(sheet_text_drivers)
            if sheet_text_success and isinstance(sheet_text_value, str):
                sheet_name = sheet_text_value

        # Compute cell address
        try:
            resolved_ref = format_cell_address(sheet_name, row_num, column_num)
            result.status = "resolved"
            result.resolved_lookup_ref = resolved_ref
            result.notes = f"ADDRESS resolved to {resolved_ref}"
        except (ValueError, TypeError) as e:
            result.status = "unresolved"
            result.notes = f"ADDRESS resolution error: {str(e)}"

        return result

    def classify_volatile_kind(self, function_name: str) -> str:
        """
        Classify volatile function by kind.

        Args:
            function_name: Function name (case-insensitive)

        Returns:
            Volatile kind classification
        """
        func_upper = function_name.upper()

        if func_upper in ("NOW", "TODAY"):
            return "time"
        elif func_upper in ("RAND", "RANDBETWEEN"):
            return "random"
        elif func_upper in ("CELL", "INFO"):
            return "workbook_meta"
        elif func_upper in ("RTD", "HYPERLINK"):
            return "external_link"
        elif func_upper in ("OFFSET", "INDIRECT"):
            return "address_computed"
        else:
            return "unknown"

    def detect_pivot_table(self, ast: dict[str, Any]) -> SpecialDataSourceInfo:
        """
        Detect GETPIVOTDATA and extract metadata.

        Args:
            ast: AST dictionary

        Returns:
            SpecialDataSourceInfo with pivot metadata

        Example:
            >>> detect_pivot_table(parse_formula("=GETPIVOTDATA(\"Sales\",A1,\"Region\",\"West\")"))
            SpecialDataSourceInfo(ref_kinds=["pivot_extract"], extras={"pivot_table": "...", ...})
        """
        info = SpecialDataSourceInfo()

        if ast.get("type") != "Function":
            return info

        func_name = ast.get("name", "").upper()
        if func_name != "GETPIVOTDATA":
            return info

        # Mark as pivot extract
        info.ref_kinds.append("pivot_extract")

        args = ast.get("args", [])
        if len(args) < 2:
            return info

        # Extract data field (first argument)
        data_field_arg = args[0]
        if data_field_arg.get("type") == "Const":
            info.extras["pivot_data_field"] = data_field_arg.get("value")

        # Extract pivot table reference (second argument)
        pivot_ref_arg = args[1]
        if pivot_ref_arg.get("type") == "Ref":
            info.extras["pivot_table_ref"] = pivot_ref_arg.get("ref")

        # Extract field/item pairs (remaining arguments in pairs)
        pivot_fields = {}
        for i in range(2, len(args), 2):
            if i + 1 < len(args):
                field_arg = args[i]
                item_arg = args[i + 1]

                if field_arg.get("type") == "Const" and item_arg.get("type") == "Const":
                    field_name = field_arg.get("value")
                    item_value = item_arg.get("value")
                    pivot_fields[field_name] = item_value

        if pivot_fields:
            info.extras["pivot_fields"] = pivot_fields

        return info

    def detect_cube_query(self, ast: dict[str, Any]) -> SpecialDataSourceInfo:
        """
        Detect cube functions (CUBEVALUE, CUBEMEMBER, etc.) and extract metadata.

        Args:
            ast: AST dictionary

        Returns:
            SpecialDataSourceInfo with cube metadata

        Example:
            >>> formula = "=CUBEVALUE(\"DataModel\",\"[Measures].[Revenue]\")"
            >>> detect_cube_query(parse_formula(formula))
            SpecialDataSourceInfo(ref_kinds=["cube_query"],
                extras={"cube_connection": "DataModel", ...})
        """
        info = SpecialDataSourceInfo()

        if ast.get("type") != "Function":
            return info

        func_name = ast.get("name", "").upper()
        if func_name not in CUBE_FUNCTIONS:
            return info

        # Mark as cube query
        info.ref_kinds.append("cube_query")

        args = ast.get("args", [])
        if len(args) < 1:
            return info

        # Extract connection (first argument)
        connection_arg = args[0]
        if connection_arg.get("type") == "Const":
            info.extras["cube_connection"] = connection_arg.get("value")

        # For CUBEVALUE, extract measure (second argument)
        if func_name == "CUBEVALUE" and len(args) >= 2:
            measure_arg = args[1]
            if measure_arg.get("type") == "Const":
                info.extras["cube_measure"] = measure_arg.get("value")

        return info


def create_resolution_engine(value_source: ValueSource | Workbook | None) -> ResolutionEngine:
    """
    Create resolution engine from workbook or value source.

    Args:
        value_source: openpyxl Workbook object or LazyValueFetcher

    Returns:
        ResolutionEngine instance
    """
    return ResolutionEngine(value_source)
