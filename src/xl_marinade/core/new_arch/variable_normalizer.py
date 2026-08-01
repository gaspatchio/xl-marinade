# ABOUTME: Post-grouping "variable normalizer" (pipeline step 6.6) — makes a
# binding correspond to one mathematical variable. Runs three gold-blind stages
# in order over the grouped binding set; gated behind --normalize-bindings
# (default off). Identity, cell_to_binding and binding_edges are recomputed
# downstream (once, after all stages), never hand-patched.
#
#   Stage 1 (H2+H9): split an over-merged binding into per-variable column/row
#     bindings while keeping a genuine 2D matrix whole. Faithful port of the
#     scorecard-validated spike (scorecard/spike_h2_h9_split.py): corpus EXACT
#     27%->39% (+51), 4 contained regressions on tablebuilder2's VBA-only catalogs.
#   Stage 2 (H4): strip a swallowed header row/column — a leading line that is
#     all-string over a non-string body. Port of scorecard/spike_h4_header.py:
#     +21 EXACT, 0 regressions. Composes with stage 1 (a split column then gets
#     its header trimmed; a kept matrix loses its header band -> EXACT).
#   Stage 3 (H5): split a fused scalar-input PANEL whose cells a VBA procedure
#     references individually (>=2 exact single-cell refs in a thin <=2-wide
#     colinear panel) into per-cell scalars. Port of scorecard/spike_h5_vba.py:
#     bsm EXACT 16->49, 0 regressions. Output matrices (read sparsely / wide) and
#     formula matrices (H9) are kept whole.
#
# All stages read ONLY the workbook's own dataflow (range_edges, MMULT consumers,
# string headers, cell dtypes, exact VBA cell refs) — never the gold answer key.
#
# STAGE-1 DISCRIMINATOR
#   A multi-column AND multi-row binding B is a genuine MATRIX -> KEEP iff ONE
#   source formula cell consumes it as a 2D object:
#     (K1) reads a block covering B AND a single-line axis (a row spanning B's
#          columns, or a column spanning B's rows) — the INDEX/MATCH signature; or
#     (K2) an MMULT consumer reads a block covering B.
#   Otherwise B is an OVER-MERGE -> SPLIT along the heterogeneous dimension:
#     split by COLUMN if columns carry distinct string headers (rows are an axis),
#     split by ROW if rows carry distinct string labels (columns are an axis).
#   A header banner alone (axis edge, no paired block read) and a VLOOKUP table
#   alone (block edge, literal column index, no header MATCH) both fail the guard.

import hashlib
import json
import sqlite3

from openpyxl.utils import get_column_letter, range_boundaries

from xl_marinade.core.bindings import compute_binding_id, compute_cells_structure_hash
from xl_marinade.core.new_arch.cell_identity import pack as pack_cell_id
from xl_marinade.core.new_arch.grouping_native import _write_binding_edges_from_cells

# Tables that carry a binding_id and must not be left pointing at a deleted
# parent. At step 6.6 only the first two (besides bindings/binding_edges) are
# populated; the rest are extracted later in the pipeline and are cleaned here
# only defensively so the pass is also safe to run on a fully-built DB (the gate).
_BINDING_ID_TABLES = (
    "cell_to_binding",
    "binding_label_candidate_cells",
    "time_index_candidates",
    "binding_time_annotations",
    "table_candidate_members",
    "formula_family_members",
)

# Cell data_types that count as textual for header detection: literal strings
# ('s') and formula-returned text ('str'). See _trim_header.is_s.
_STRING_DTYPES = frozenset({"s", "str"})

# Header-block regrid (one workbook's orientation mess): a header run must be >=_HBR_MIN_COLS
# wide over >=_HBR_MIN_ROWS data rows to be treated as a table.
_HBR_MIN_COLS = 3
_HBR_MIN_ROWS = 3


# ----------------------------------------------------------------------------
# Signal loading (gold-blind dataflow indices, built once).
# ----------------------------------------------------------------------------


def _get_guid(conn: sqlite3.Connection) -> str:
    """Read workbook_sha256 from ir_metadata. Only valid once that table exists
    (written at pipeline step 7); mid-pipeline callers pass the guid explicitly."""
    if not _table_exists(conn, "ir_metadata"):
        return ""
    row = conn.execute("SELECT value FROM ir_metadata WHERE key='workbook_sha256'").fetchone()
    return row[0] if row else ""


def _table_exists(conn: sqlite3.Connection, name: str) -> bool:
    return (
        conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
        ).fetchone()
        is not None
    )


def _load_signals(conn: sqlite3.Connection):
    """Return (edges, strs, mmult).

    edges:   {to_sheet_id: {from_cell_id: [(r1,c1,r2,c2), ...]}}  — a 2D lookup
             reads the data block AND its axis from the SAME source cell.
    strs:    {(sheet_id,row,col): text}                            — string cells.
    mmult:   set(cell_id) whose formula contains MMULT.
    """
    edges: dict[int, dict[int, list]] = {}
    for fc, tsid, r1, c1, r2, c2 in conn.execute(
        "SELECT from_cell_id, to_sheet_id, to_r1, to_c1, to_r2, to_c2 FROM range_edges"
    ):
        edges.setdefault(tsid, {}).setdefault(fc, []).append((r1, c1, r2, c2))

    # Materialise the outer query before the inner json_blobs lookups (the
    # cursor-reuse trap the spike hit: a shared cursor leaves strs with 1 entry).
    srows = conn.execute(
        "SELECT sheet_id, row, col, value_blob_id FROM cells WHERE data_type='s'"
    ).fetchall()
    strs: dict[tuple, str] = {}
    cur = conn.cursor()
    for sid, row, col, jb in srows:
        if jb is None:
            continue
        b = cur.execute("SELECT json FROM json_blobs WHERE blob_id=?", (jb,)).fetchone()
        if not b:
            continue
        try:
            v = json.loads(b[0])
            txt = v if isinstance(v, str) else str(v)
        except Exception:
            txt = b[0]
        strs[(sid, row, col)] = txt.strip()

    mmult = {
        cid for (cid,) in conn.execute("SELECT cell_id FROM cells WHERE formula_a1 LIKE '%MMULT%'")
    }
    return edges, strs, mmult


def _load_bindings_with_cells(conn: sqlite3.Connection):
    """Return list of binding dicts with their populated cells.

    Each: {binding_id, sheet_id, binding_type, label, evidence_blob_id,
           spatial_blob_id, cells: [(cell_id,row,col,formula_a1,formula_id), ...]}
    """
    meta = {}
    for bid, sid, addr, btype, label, src, eb, sb in conn.execute(
        "SELECT binding_id, sheet_id, address_a1, binding_type, label, "
        "extraction_source, evidence_blob_id, spatial_candidates_blob_id FROM bindings"
    ):
        meta[bid] = {
            "binding_id": bid,
            "sheet_id": sid,
            "address_a1": addr,
            "binding_type": btype,
            "label": label,
            "extraction_source": src,
            "evidence_blob_id": eb,
            "spatial_blob_id": sb,
            "cells": [],
        }
    for bid, cid, row, col, fa1, fid in conn.execute(
        "SELECT cb.binding_id, cb.cell_id, c.row, c.col, c.formula_a1, c.formula_id "
        "FROM cell_to_binding cb JOIN cells c ON c.cell_id = cb.cell_id"
    ):
        b = meta.get(bid)
        if b is not None:
            b["cells"].append((cid, row, col, fa1, fid))
    return list(meta.values())


# ----------------------------------------------------------------------------
# Discriminator (faithful port of spike_h2_h9_split.py).
# ----------------------------------------------------------------------------


def _is_header_text(s) -> bool:
    if not s:
        return False
    s = s.strip()
    if not s:
        return False
    try:
        float(s.replace(",", "").replace("%", ""))
        return False  # numeric -> an axis value, not a name
    except ValueError:
        return True


def _keep_matrix(sid, r1, c1, r2, c2, edges, mmult) -> bool:
    """KEEP iff some SINGLE source cell consumes B as a 2D object (K1 or K2)."""

    def covers_block(e):
        er1, ec1, er2, ec2 = e
        return er1 <= r1 and ec1 <= c1 and er2 >= r2 and ec2 >= c2 and (ec2 > ec1 and er2 > er1)

    def is_col_axis(e):  # single row spanning B's columns, at/above the block top
        er1, ec1, er2, ec2 = e
        return er1 == er2 and ec2 - ec1 >= 1 and ec1 <= c1 and ec2 >= c2 and er1 <= r1

    def is_row_axis(e):  # single col spanning B's rows, at/left of the block
        er1, ec1, er2, ec2 = e
        return ec1 == ec2 and er2 - er1 >= 1 and er1 <= r1 and er2 >= r2 and ec1 <= c1

    for fc, es in edges.get(sid, {}).items():
        if not any(covers_block(e) for e in es):
            continue
        if fc in mmult:  # K2
            return True
        if any(is_col_axis(e) or is_row_axis(e) for e in es):  # K1
            return True
    return False


def _split_groups(cells, sid, strs):
    """Return (axis, groups) where groups is a list of cell-lists to become
    sub-bindings, or (None, None) if the block cannot be confidently split.

    cells: [(cell_id,row,col,formula_a1,formula_id), ...] (the populated cells).
    Mirrors the spike: split by COLUMN when columns are distinctly named, else by
    ROW when rows are distinctly labelled.
    """
    cols = sorted({c for _, _, c, _, _ in cells})
    rows = sorted({r for _, r, _, _, _ in cells})
    c1, c2, r1, r2 = cols[0], cols[-1], rows[0], rows[-1]
    if c2 == c1 and r2 == r1:
        return None, None

    # A row-split turns each distinctly-labelled ROW into its own variable. That is
    # right for a SHORT transposed parameter stack (e.g. a handful of fund-level
    # metrics) but catastrophic for a TALL data table, whose rows are records over
    # an axis and whose row-key column is distinct-per-row by construction (e.g. a
    # Policies sheet shattering a 988-row column into 988 single-cell bindings).
    # Cap row-splits to short blocks so a record table is never row-shattered.
    _MAX_SPLIT_ROWS = 40

    def _distinct_ok(headers, n):
        named = [h for h in headers.values() if _is_header_text(h)]
        return len(set(named)) >= 2 and len(set(named)) >= 0.5 * n

    def _col_split(col_hdrs):
        groups = []
        for c in cols:
            sub = [cell for cell in cells if cell[2] == c]
            groups.append((sub, col_hdrs[c]))
        return "split_col", groups

    def _row_split(row_lbls):
        groups = []
        for r in rows:
            sub = [cell for cell in cells if cell[1] == r]
            groups.append((sub, row_lbls[r]))
        return "split_row", groups

    # --- existing behaviour first (no regression): single header row r1-1 / single
    #     label col c1-1, immediately adjacent ---
    col_hdrs = {c: strs.get((sid, r1 - 1, c)) for c in cols}
    if c2 > c1 and _distinct_ok(col_hdrs, len(cols)):
        return _col_split(col_hdrs)

    row_lbls = {r: strs.get((sid, r, c1 - 1)) for r in rows}
    if r2 > r1 and len(rows) <= _MAX_SPLIT_ROWS and _distinct_ok(row_lbls, len(rows)):
        return _row_split(row_lbls)

    # --- over-merge residual fallbacks (2026-06-29), tried only when the adjacent
    #     single-line read above found nothing. Ordered conservatively. ---

    # #3 multi-row header BAND above: concatenate up to 3 textual header rows per
    # column (e.g. a large model's "USD Up"/row.../"Curve") so distinctness carried
    # two rows up is seen when r1-1 alone is identical across columns.
    if c2 > c1 and r1 - 1 >= 1:
        band_rows = [rr for rr in range(max(1, r1 - 3), r1)]
        band = {}
        for c in cols:
            parts = [strs.get((sid, rr, c)) for rr in band_rows]
            parts = [p for p in parts if _is_header_text(p)]
            band[c] = " ".join(parts) if parts else None
        if _distinct_ok(band, len(cols)):
            return _col_split(band)

    # #2 header row immediately BELOW the block (e.g. a model's row-7 PV scalars
    # whose names sit in row 8).
    if c2 > c1:
        below = {c: strs.get((sid, r2 + 1, c)) for c in cols}
        if _distinct_ok(below, len(cols)):
            return _col_split(below)

    # #1 row-label column NOT immediately adjacent: search left past blank spacer
    # columns (e.g. one model's labels in col B with the data block at col F). Take
    # the nearest column whose labels distinctly name the rows. Capped to short blocks
    # so a tall record table's row-key column never shatters it per-row.
    if r2 > r1 and len(rows) <= _MAX_SPLIT_ROWS:
        for lc in range(c1 - 1, max(0, c1 - 1 - 8), -1):
            cand = {r: strs.get((sid, r, lc)) for r in rows}
            if _distinct_ok(cand, len(rows)):
                return _row_split(cand)

    return None, None


# ----------------------------------------------------------------------------
# Sub-binding construction.
# ----------------------------------------------------------------------------


def _build_subbinding(parent, sub_cells, label, axis, guid, sheet_name, block_bbox):
    """Build a binding INSERT row + its cell_to_binding rows for one sub-group.

    Address uses the parent block's bounding box on the FIXED dimension (matching
    the validated spike): a column split yields '<col><r1>:<col><r2>' over the
    block's full row range; a row split yields '<c1><row>:<c2><row>'.
    """
    b_r1, b_c1, b_r2, b_c2 = block_bbox
    sid = parent["sheet_id"]
    if axis == "split_col":
        col = sub_cells[0][2]
        top_row, bot_row = b_r1, b_r2
        top_col = col
        addr = (
            f"{get_column_letter(col)}{top_row}"
            if top_row == bot_row
            else f"{get_column_letter(col)}{top_row}:{get_column_letter(col)}{bot_row}"
        )
        shape_rows, shape_cols = bot_row - top_row + 1, 1
    else:  # split_row
        row = sub_cells[0][1]
        top_row, top_col = row, b_c1
        addr = (
            f"{get_column_letter(b_c1)}{row}"
            if b_c1 == b_c2
            else f"{get_column_letter(b_c1)}{row}:{get_column_letter(b_c2)}{row}"
        )
        shape_rows, shape_cols = 1, b_c2 - b_c1 + 1

    top_left_a1 = f"{get_column_letter(top_col)}{top_row}"

    # cells_structure_hash over THIS sub-group's (a1, formula) tuples.
    cell_a1 = [f"{get_column_letter(c)}{r}" for (_, r, c, _, _) in sub_cells]
    cell_formulas = {f"{get_column_letter(c)}{r}": (fa1 or "") for (_, r, c, fa1, _) in sub_cells}
    structure_hash = compute_cells_structure_hash(cell_a1, cell_formulas)

    binding_id = compute_binding_id(
        workbook_guid=guid,
        sheet=sheet_name,
        top_left_a1=top_left_a1,
        shape_rows=shape_rows,
        shape_cols=shape_cols,
        cells_structure_hash=structure_hash,
    )

    # binding_type: formula if any cell carries a formula, else constant.
    has_formula = any((fa1 or "").strip() for (_, _, _, fa1, _) in sub_cells)
    binding_type = "formula" if has_formula else "constant"

    # formula_id: spatial top-left cell's formula_id (mirror grouping's choice).
    tl_cell = min(sub_cells, key=lambda x: (x[1], x[2]))
    formula_id = tl_cell[4] if binding_type == "formula" else None

    # top_left_cell_id: the packed id of the address's top-left corner — matches
    # grouping's contract and is correct even when that corner cell is unpopulated
    # (a ragged block / sheet edge). cells.cell_id IS the packed id.
    top_left_cell_id = pack_cell_id(sid, top_row, top_col)

    row = (
        binding_id,
        sid,
        addr,
        top_left_cell_id,
        shape_rows,
        shape_cols,
        binding_type,
        formula_id,
        label if _is_header_text(label) else parent["label"],
        None,
        None,
        0,
        f"normalized:{axis}",
        parent["evidence_blob_id"],
        parent["spatial_blob_id"],
    )
    ctb = [(cid, binding_id) for (cid, _, _, _, _) in sub_cells]
    return row, ctb


# ----------------------------------------------------------------------------
# Entry point.
# ----------------------------------------------------------------------------


def _is_regen_target(source: str, shape_rows: int, shape_cols: int) -> bool:
    """Should this binding's spatial_candidates blob be regenerated on its own extent?

    (1) split_col/split_row sub-bindings (not vba_panel) — each inherits the parent's
        top-left header for every column.
    (2) 1x1 normalized scalars (header_strip / vba_panel) — each inherits the parent
        PANEL's blob (the block banner) rather than its own row label.
    (3) header_strip bindings of ANY shape — the strip SHRANK the extent (dropped the
        header row/col), so the real header is now the cell ADJACENT to the residual
        (e.g. simple age axis A2:A101 after stripping A1='age-last'). The inherited
        parent blob scanned the pre-strip extent (header inside it -> no scan_above),
        so it must be regenerated on the residual extent.
    """
    s = source or ""
    if ("split_col" in s or "split_row" in s) and "vba_panel" not in s:
        return True
    if s.startswith("normalized:") and shape_rows == 1 and shape_cols == 1:
        return True
    if "header_strip" in s:
        return True
    if "header_block_regrid" in s:
        return True  # regridded columns get named by their header (scan_above)
    return False


def _regen_split_candidates(conn, sheets, workbook_path, name_table_map=None):
    """Regenerate the spatial_candidates blob for normalized sub-bindings on their
    OWN extent, via the production evidence extractor.

    A normalized sub-binding otherwise inherits its parent's spatial blob (the parent's
    top-left/banner header). Two cases qualify:
      - split_col/split_row sub-bindings (not vba_panel): every split column would
        otherwise be named the same (e.g. all of a model's DATA block -> one label).
      - 1x1 normalized scalars (header_strip / vba_panel): each inherits the parent
        PANEL's blob — a scan_above over the label column that resolves to the block
        BANNER ('Global setting for all inputs') instead of the cell's own row label.
        Regen surfaces the scan_left row header so the labeller (score_candidate's
        short-header boost) prefers it.
    Recomputing candidates per sub-extent lets the labeller pick each binding's own
    header. (The earlier scoping — scorecard/spike_oq6_name.py — excluded vba_panel
    because vba scalars regressed under regen; that regression was a scoring gap, since
    closed by the numeric-cap + short-header rules in simple_labeller.score_candidate,
    so 1x1 scalars are now included. header_name corpus +13, band unchanged, 0 band
    regr; see scorecard/spike_name_v5.py.)

    Workbook access: REUSES the pipeline's already-open ``name_table_map`` (and its
    ``.workbook``) for named ranges + merged-cell ranges — cell values come from the
    IR DB (``build_evidence_cache_from_db``), so no workbook reopen is needed in the
    normal (pipeline) path. Only falls back to a fresh ``LazyWorkbook(workbook_path)``
    when called WITHOUT a name_table_map (standalone / tests). Merged-cell ranges are
    the one piece not persisted in the IR DB (no merged-ranges table), so a workbook
    object is still required — but the pipeline's is reused rather than re-loaded.

    No-op (returns 0) when neither a name_table_map nor a workbook_path is available
    (some gate paths) or when the evidence machinery / workbook cannot be loaded; on
    any failure the binding keeps its inherited blob. Touches ONLY
    bindings.spatial_candidates_blob_id (binding_id is structure-hashed and stable;
    extents/edges are untouched)."""
    if not workbook_path and name_table_map is None:
        return 0
    rows = conn.execute(
        "SELECT binding_id, sheet_id, address_a1, shape_rows, shape_cols, "
        "extraction_source FROM bindings"
    ).fetchall()
    targets = [r for r in rows if _is_regen_target(r[5] or "", r[3], r[4])]
    if not targets:
        return 0
    try:
        from xl_marinade.core.evidence import (
            build_evidence_cache_from_db,
            extract_evidence_for_binding,
        )
        from xl_marinade.core.lazy_workbook import LazyWorkbook
        from xl_marinade.core.names_tables import NameTableMap
    except Exception:
        return 0

    class _BC:
        __slots__ = ("binding_id", "sheet", "address_a1", "top_left_a1", "shape_rows", "shape_cols")

        def __init__(self, bid, sheet, addr, sr, sc):
            self.binding_id = bid
            self.sheet = sheet
            self.address_a1 = addr
            self.top_left_a1 = addr.split("!")[-1].split(":")[0]
            self.shape_rows = sr
            self.shape_cols = sc

    # Cache only the split targets' scan windows (compute_evidence_regions builds
    # per-binding regions, so each target's full window is still covered) — cost
    # proportional to the number of splits, not the whole workbook.
    targets_bc = [
        _BC(bid, sheets.get(sid, ""), addr, sr, sc)
        for bid, sid, addr, sr, sc, _ in targets
        if sheets.get(sid)
    ]
    try:
        cache = build_evidence_cache_from_db(conn, targets_bc)
    except Exception:
        return 0
    # Prefer the pipeline's already-open NameTableMap + workbook (no reopen). Only
    # load a fresh LazyWorkbook when called standalone (no name_table_map provided).
    if name_table_map is not None:
        ntm = name_table_map
        name_wb = name_table_map.workbook
    else:
        try:
            name_wb = LazyWorkbook(workbook_path, data_only=False, keep_vba=False)
        except Exception:
            return 0
        try:
            ntm = NameTableMap(name_wb)
        except Exception:
            ntm = None

    merged_cache: dict = {}
    n = 0
    for bid, sid, addr, sr, sc, _src in targets:
        sheet = sheets.get(sid)
        if not sheet:
            continue
        if sheet not in merged_cache:
            try:
                merged_cache[sheet] = list(name_wb[sheet].merged_cells.ranges)
            except Exception:
                merged_cache[sheet] = []
        qaddr = (
            f"'{sheet}'!{addr}"
            if any(not (ch.isalnum() or ch == "_") for ch in sheet)
            else f"{sheet}!{addr}"
        )
        # Evidence extraction AND the DB writes are guarded together so a failure on
        # one binding skips only that binding (keeps its inherited blob), per the
        # contract above. (A poisoned shared transaction would still fail downstream,
        # but per-binding skipping is the documented intent.)
        try:
            ev = extract_evidence_for_binding(
                worksheet=None,
                workbook=name_wb,
                binding_address=qaddr,
                binding_shape=(sr, sc),
                name_table_map=ntm,
                cell_value_cache=cache,
                merged_ranges=merged_cache[sheet],
            )
            blob = json.dumps(
                {
                    "label_candidates": ev.get("label_candidates", []),
                    "axis_labels": ev.get("axis_labels", []),
                }
            )
            sha = hashlib.sha256(blob.encode()).hexdigest()
            conn.execute(
                "INSERT OR IGNORE INTO json_blobs (sha256, json) VALUES (?, ?)", (sha, blob)
            )
            row = conn.execute("SELECT blob_id FROM json_blobs WHERE sha256=?", (sha,)).fetchone()
            if not row:
                continue
            conn.execute(
                "UPDATE bindings SET spatial_candidates_blob_id=? WHERE binding_id=?", (row[0], bid)
            )
        except Exception:
            continue
        n += 1
    return n


def normalize_bindings_fn(
    conn: sqlite3.Connection,
    workbook_path: str | None = None,
    workbook_sha256: str | None = None,
    name_table_map=None,
) -> dict:
    """Normalize bindings: split over-merges keeping matrices (H2+H9), strip
    swallowed headers (H4), then split VBA-referenced input panels (H5).

    Mutates the bindings / cell_to_binding / binding_edges tables in place and
    returns metrics. Reads all dataflow signals from `conn`. The binding-id guid
    is taken from `workbook_sha256` when given (the pipeline path, where the
    ir_metadata table does not yet exist at step 6.6); otherwise it is read from
    ir_metadata (the gate path, on a finished DB).
    """
    guid = workbook_sha256 if workbook_sha256 is not None else _get_guid(conn)
    sheets = {sid: name for sid, name in conn.execute("SELECT sheet_id, sheet_name FROM sheets")}

    # Ensure exact VBA cell refs exist for stage 3 (H5). Runs here (step 6.6,
    # AFTER defined_names is written) so VBA Range("name") references resolve to
    # concrete cells; idempotent, so the later wire_vba_edges step (6.9) reuses
    # these rows. No-op for non-VBA workbooks / when refs already present.
    from xl_marinade.core.vba.reference_extractor import extract_vba_references_to_table

    extract_vba_references_to_table(conn)

    edges, strs, mmult = _load_signals(conn)
    bindings = _load_bindings_with_cells(conn)

    removed: list[str] = []
    new_binding_rows: list[tuple] = []
    new_ctb_rows: list[tuple] = []
    n_split = n_kept = 0

    for b in bindings:
        cells = b["cells"]
        if not cells:
            continue
        # Idempotence: a binding this pass already produced (split_col/split_row)
        # must be terminal — never reprocessed. Without this, a column-split
        # vector with distinct left-edge row labels would row-split into scalars
        # on a second pass, so normalize(normalize(x)) != normalize(x).
        if (b["extraction_source"] or "").startswith("normalized:"):
            continue
        cols = {c for _, _, c, _, _ in cells}
        rows = {r for _, r, _, _, _ in cells}
        if len(cols) < 2 and len(rows) < 2:
            continue  # scalar / single line — nothing to split
        sid = b["sheet_id"]
        sheet_name = sheets.get(sid)
        if sheet_name is None:
            continue
        r1, r2 = min(rows), max(rows)
        c1, c2 = min(cols), max(cols)

        # H9: a genuine 2D matrix consumed as one object stays whole.
        if len(cols) >= 2 and len(rows) >= 2 and _keep_matrix(sid, r1, c1, r2, c2, edges, mmult):
            n_kept += 1
            continue

        axis, groups = _split_groups(cells, sid, strs)
        if not groups or len(groups) <= 1:
            continue

        block_bbox = (r1, c1, r2, c2)
        for sub_cells, label in groups:
            brow, bctb = _build_subbinding(b, sub_cells, label, axis, guid, sheet_name, block_bbox)
            new_binding_rows.append(brow)
            new_ctb_rows.extend(bctb)
        removed.append(b["binding_id"])
        n_split += 1

    if removed:
        _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows)

    # STAGE 2: H4 header-strip, on the post-split binding set (a split column then
    # gets its swallowed header trimmed; a kept matrix loses its header band). After
    # stripping an EDGE header/label, the residual is re-fed to the splitter so a
    # block whose headers sat at its own edge (invisible to STAGE 1's outside-the-block
    # lookup) splits per column/row — guarded by _keep_matrix so genuine 2D lookup
    # tables stay whole.
    header = _header_strip_stage(conn, guid, sheets, strs, edges, mmult)

    # STAGE 3: H5 VBA input-panel split, on the post-split/strip binding set
    # (a fused scalar-input panel whose cells VBA references individually is split
    # into per-cell scalars; output matrices and formula matrices are kept whole).
    vba = _vba_panel_stage(conn, guid, sheets, edges, mmult)

    # STAGE 4: H9 ListObject-extent — collapse the per-cell fragments of a declared
    # Excel Table (ListObject) into ONE binding addressed at the table's data_range
    # (header row excluded at source by names_tables). Runs LAST among extent
    # mutators so the result is terminal. No-op when no ListObjects exist (all gold
    # workbooks except tablebuilder2), so it cannot touch the other corpora.
    lo = (
        _listobject_extent_stage(conn, guid, sheets, name_table_map)
        if name_table_map is not None
        else {"collapsed": 0, "removed": 0, "added": 0}
    )

    # STAGE 5: header-block regrid — a header-aligned constant block cut with mixed
    # orientation by the flood-fill (a model's DATA/ERROR sheets) is re-cut into clean vertical
    # columns under the header row. Gated on the transpose-mess signature (>=2 stacked
    # horizontal row-strips), so clean tables and 2D matrices are left intact.
    hbr = _header_block_regrid_stage(conn, guid, sheets)

    # OQ-6: regenerate label candidates on each binding's OWN extent so a sub-binding
    # is named by its own header, not the parent's top-left/banner header. Covers split
    # columns AND 1x1 normalized scalars (header_strip/vba_panel) that otherwise inherit
    # the parent panel banner. Labels only — no extent/edge change. Self-guards (returns
    # 0 without a workbook_path or when no binding qualifies), so call unconditionally.
    relabeled = _regen_split_candidates(conn, sheets, workbook_path, name_table_map)

    # Recompute binding edges ONCE over the final binding set (all stages done).
    if removed or header["removed"] or vba["removed"] or lo["removed"] or hbr["removed"]:
        _rebuild_edges(conn)
        conn.commit()
    elif relabeled:
        conn.commit()

    return {
        "bindings_split": n_split,
        "bindings_kept_matrix": n_kept,
        "bindings_header_stripped": header["stripped"],
        "bindings_header_resplit": header.get("resplit", 0),
        "bindings_vba_split": vba["vba_split"],
        "bindings_listobject_collapsed": lo["collapsed"],
        "bindings_header_block_regridded": hbr["regridded"],
        "bindings_relabeled": relabeled,
        "bindings_removed": (
            len(removed) + header["removed"] + vba["removed"] + lo["removed"] + hbr["removed"]
        ),
        "bindings_added": (
            len(new_binding_rows) + header["added"] + vba["added"] + lo["added"] + hbr["added"]
        ),
    }


# ----------------------------------------------------------------------------
# Stage 2 — H4 header strip (RC3).
# ----------------------------------------------------------------------------


def _load_dtypes(conn: sqlite3.Connection) -> dict:
    return {
        (sid, r, c): d
        for sid, r, c, d in conn.execute("SELECT sheet_id, row, col, data_type FROM cells")
    }


def _trim_header(cells, sid, dtypes):
    """Return the kept cells after dropping a leading header row and/or column —
    a leading line that is all-string over a non-string body (the dtype boundary
    that marks a header). Returns None when nothing is trimmed (faithful port of
    spike_h4_header.trim_binding). Never trims to empty."""
    coords = {(c, r) for (_, r, c, _, _) in cells}
    cols = sorted({c for (_, _, c, _, _) in cells})
    rows = sorted({r for (_, r, _, _, _) in cells})
    c1, c2, r1, r2 = cols[0], cols[-1], rows[0], rows[-1]

    def is_s(r, c):
        # Both literal strings ('s') and formula-returned text ('str') are textual.
        # The spike used 's' only, which on real DBs (where 'str' is heavily used,
        # e.g. a large model has more 'str' than 's' cells) both MISSES 'str'
        # headers and — worse — wrongly strips a leading 's' line over a 'str'
        # (still-textual) body. Treating both as string fixes that false-positive
        # data loss.
        return dtypes.get((sid, r, c)) in _STRING_DTYPES

    drop_row = drop_col = drop_col_trailing = False
    if r2 > r1:
        top = [(c, r1) for c in cols if (c, r1) in coords]
        nxt = [(c, r1 + 1) for c in cols if (c, r1 + 1) in coords]
        if top and all(is_s(r, c) for c, r in top) and nxt and not all(is_s(r, c) for c, r in nxt):
            drop_row = True
    if c2 > c1:
        left = [(c1, r) for r in rows if (c1, r) in coords]
        nxt = [(c1 + 1, r) for r in rows if (c1 + 1, r) in coords]
        if (
            left
            and all(is_s(r, c) for c, r in left)
            and nxt
            and not all(is_s(r, c) for c, r in nxt)
        ):
            drop_col = True
    # Trailing annotation column: a single-row [label | value | note] block keeps a
    # tacked-on all-string cell after header-strip (e.g. a value cell whose right
    # neighbour holds a free-text note). _trim_header only peels the LEADING dimension, so
    # the note survives and the band reads band_includes_header. Drop a trailing
    # all-string column when the column to its left is NOT all-string. Gated to
    # single-row blocks: a genuine all-text VALUE column (a value-name column)
    # spans many rows, so r1==r2 excludes it.
    if c2 > c1 and r1 == r2:
        right = [(c2, r) for r in rows if (c2, r) in coords]
        prev = [(c2 - 1, r) for r in rows if (c2 - 1, r) in coords]
        if (
            right
            and all(is_s(r, c) for c, r in right)
            and prev
            and not all(is_s(r, c) for c, r in prev)
        ):
            drop_col_trailing = True
    if not (drop_row or drop_col or drop_col_trailing):
        return None
    kept = [
        cell
        for cell in cells
        if not (drop_row and cell[1] == r1)
        and not (drop_col and cell[2] == c1)
        and not (drop_col_trailing and cell[2] == c2)
    ]
    if not kept or len(kept) == len(cells):
        return None
    return kept


def _build_binding_from_cells(parent, cells, guid, sheet_name, tag):
    """Build a binding INSERT row + cell_to_binding rows from an arbitrary cell
    set, addressed by the set's OWN bounding box. `tag` is the provenance suffix
    (e.g. 'header_strip', 'vba_panel'). Used by the header-strip and VBA stages.
    """
    sid = parent["sheet_id"]
    rows = sorted({r for (_, r, _, _, _) in cells})
    cols = sorted({c for (_, _, c, _, _) in cells})
    r1, r2, c1, c2 = rows[0], rows[-1], cols[0], cols[-1]
    top_left_a1 = f"{get_column_letter(c1)}{r1}"
    addr = top_left_a1 if (r1 == r2 and c1 == c2) else f"{top_left_a1}:{get_column_letter(c2)}{r2}"
    shape_rows, shape_cols = r2 - r1 + 1, c2 - c1 + 1

    cell_a1 = [f"{get_column_letter(c)}{r}" for (_, r, c, _, _) in cells]
    cell_formulas = {f"{get_column_letter(c)}{r}": (fa1 or "") for (_, r, c, fa1, _) in cells}
    structure_hash = compute_cells_structure_hash(cell_a1, cell_formulas)
    binding_id = compute_binding_id(
        workbook_guid=guid,
        sheet=sheet_name,
        top_left_a1=top_left_a1,
        shape_rows=shape_rows,
        shape_cols=shape_cols,
        cells_structure_hash=structure_hash,
    )
    has_formula = any((fa1 or "").strip() for (_, _, _, fa1, _) in cells)
    binding_type = "formula" if has_formula else "constant"
    tl_cell = min(cells, key=lambda x: (x[1], x[2]))
    formula_id = tl_cell[4] if binding_type == "formula" else None
    top_left_cell_id = pack_cell_id(sid, r1, c1)

    psrc = parent["extraction_source"] or ""
    src = (psrc + "+" + tag) if psrc.startswith("normalized:") else "normalized:" + tag

    row = (
        binding_id,
        sid,
        addr,
        top_left_cell_id,
        shape_rows,
        shape_cols,
        binding_type,
        formula_id,
        parent["label"],
        None,
        None,
        0,
        src,
        parent["evidence_blob_id"],
        parent["spatial_blob_id"],
    )
    ctb = [(cid, binding_id) for (cid, _, _, _, _) in cells]
    return row, ctb


def _build_listobject_binding(parent, cells, rect, guid, sheet_name, tag="listobject_data_range"):
    """Like _build_binding_from_cells but addressed at an explicit RECT
    (r1,c1,r2,c2) rather than the cell set's own bbox — so a trailing blank
    column inside a ListObject data_range (e.g. a blank E2) is still covered by
    the address. structure_hash + cell_to_binding come from the real cells;
    address / shape / top_left / binding_id come from the rect. `tag` is the
    provenance suffix (default the H9 ListObject tag; the header-block regrid
    passes 'header_block_regrid')."""
    sid = parent["sheet_id"]
    r1, c1, r2, c2 = rect
    top_left_a1 = f"{get_column_letter(c1)}{r1}"
    addr = top_left_a1 if (r1 == r2 and c1 == c2) else f"{top_left_a1}:{get_column_letter(c2)}{r2}"
    shape_rows, shape_cols = r2 - r1 + 1, c2 - c1 + 1
    cell_a1 = [f"{get_column_letter(c)}{r}" for (_, r, c, _, _) in cells]
    cell_formulas = {f"{get_column_letter(c)}{r}": (fa1 or "") for (_, r, c, fa1, _) in cells}
    structure_hash = compute_cells_structure_hash(cell_a1, cell_formulas)
    binding_id = compute_binding_id(
        workbook_guid=guid,
        sheet=sheet_name,
        top_left_a1=top_left_a1,
        shape_rows=shape_rows,
        shape_cols=shape_cols,
        cells_structure_hash=structure_hash,
    )
    has_formula = any((fa1 or "").strip() for (_, _, _, fa1, _) in cells)
    binding_type = "formula" if has_formula else "constant"
    tl_cell = min(cells, key=lambda x: (x[1], x[2]))
    formula_id = tl_cell[4] if binding_type == "formula" else None
    top_left_cell_id = pack_cell_id(sid, r1, c1)
    psrc = parent["extraction_source"] or ""
    src = (psrc + "+" + tag) if psrc.startswith("normalized:") else "normalized:" + tag
    row = (
        binding_id,
        sid,
        addr,
        top_left_cell_id,
        shape_rows,
        shape_cols,
        binding_type,
        formula_id,
        parent["label"],
        None,
        None,
        0,
        src,
        parent["evidence_blob_id"],
        parent["spatial_blob_id"],
    )
    ctb = [(cid, binding_id) for (cid, _, _, _, _) in cells]
    return row, ctb


def _detect_header_blocks(text_at, data_at, maxr, maxc):
    """Yield (header_row, c1, c2, data_end) for header-aligned numeric blocks: a row
    with a contiguous run of >=_HBR_MIN_COLS text cells, over >=_HBR_MIN_ROWS rows
    that are predominantly numeric/value. text_at/data_at are sets of (row,col)."""
    blocks = []
    for r in range(1, maxr + 1):
        tcols = sorted(c for c in range(1, maxc + 1) if (r, c) in text_at)
        runs = []
        run = []
        for c in tcols:
            if run and c == run[-1] + 1:
                run.append(c)
            else:
                if len(run) >= _HBR_MIN_COLS:
                    runs.append((run[0], run[-1]))
                run = [c]
        if len(run) >= _HBR_MIN_COLS:
            runs.append((run[0], run[-1]))
        for c1, c2 in runs:
            w = c2 - c1 + 1
            de = r
            rr = r + 1
            while rr <= maxr:
                num = sum(1 for c in range(c1, c2 + 1) if (rr, c) in data_at)
                txt = sum(1 for c in range(c1, c2 + 1) if (rr, c) in text_at)
                if num == 0 and txt == 0:
                    break  # fully blank row -> block ends
                if num < max(1, (w + 1) // 2):
                    break  # not predominantly numeric -> not a data row
                de = rr
                rr += 1
            if de - r >= _HBR_MIN_ROWS:
                blocks.append((r, c1, c2, de))
    return blocks


def _header_block_regrid_stage(conn, guid, sheets) -> dict:
    """Regrid a header-aligned constant block that the constant flood-fill cut with
    MIXED orientation — some vertical fragments, some horizontal row-strips (e.g.
    a model's DATA/ERROR: a 6x10 table whose constant values made orientation ambiguous,
    cut as A4:A5 + A6 + A7:G7 + A8:H8...) — into clean vertical columns under the
    header row (header excluded).

    GATED on the transpose-mess signature: a block is only regridded when >=2 of its
    contained bindings are horizontal row-strips (span >=2 columns AND are shorter
    than the block height). A clean column table (1-col bindings) and a genuine 2D
    matrix (one block, not >=2 stacked strips) both fail the gate and are left intact
    — this is what keeps the change from shattering simple's Projection or a large
    model's correlation matrices.

    Detection reads the IR `cells` table dtypes only (no workbook reopen). Idempotent:
    after a regrid the block holds single-column bindings, which fail the row-strip
    gate on a second pass."""
    dtypes = _load_dtypes(conn)
    text_by_sid: dict = {}
    data_by_sid: dict = {}
    dims: dict = {}
    for (sid, r, c), dt in dtypes.items():
        if dt in _STRING_DTYPES:
            text_by_sid.setdefault(sid, set()).add((r, c))
        elif dt != "blank":
            data_by_sid.setdefault(sid, set()).add((r, c))
        mr, mc = dims.get(sid, (0, 0))
        dims[sid] = (max(mr, r), max(mc, c))

    bindings = _load_bindings_with_cells(conn)
    by_sid: dict = {}
    for b in bindings:
        by_sid.setdefault(b["sheet_id"], []).append(b)

    removed: list[str] = []
    new_binding_rows: list[tuple] = []
    new_ctb_rows: list[tuple] = []
    n = 0
    for sid, sheet_name in sheets.items():
        text_at = text_by_sid.get(sid)
        if not text_at:
            continue
        maxr, maxc = dims.get(sid, (0, 0))
        data_at = data_by_sid.get(sid, set())
        for hr, bc1, bc2, de in _detect_header_blocks(text_at, data_at, maxr, maxc):
            r1, r2 = hr + 1, de
            contained = [
                b
                for b in by_sid.get(sid, [])
                if b["cells"]
                and all(r1 <= row <= r2 and bc1 <= col <= bc2 for (_, row, col, _, _) in b["cells"])
            ]
            if not contained:
                continue
            # GATE 1: every contained binding must be a CONSTANT. The orientation
            # ambiguity is a constant-block phenomenon — equal literal values make the
            # flood-fill's vertical-vs-horizontal choice ambiguous. A formula block has
            # clear structure (each column a formula family) and must NOT be regridded
            # (e.g. one model's projection calc tables are 100% formula). Graceful miss.
            if any(b["binding_type"] != "constant" for b in contained):
                continue
            # GATE 2: transpose-mess signature — >=2 horizontal row-strips (span >=2
            # columns AND shorter than the block height).
            block_h = r2 - r1 + 1
            strips = sum(
                1
                for b in contained
                if len({col for (_, _, col, _, _) in b["cells"]}) >= 2
                and len({row for (_, row, _, _, _) in b["cells"]}) < block_h
            )
            if strips < 2:
                continue  # not the transpose-mess signature -> leave intact
            cells_by_col: dict = {}
            for b in contained:
                for cell in b["cells"]:
                    cells_by_col.setdefault(cell[2], []).append(cell)
            parent = max(contained, key=lambda b: len(b["cells"]))
            for c in sorted(cells_by_col):
                row, ctb = _build_listobject_binding(
                    parent,
                    cells_by_col[c],
                    (r1, c, r2, c),
                    guid,
                    sheet_name,
                    tag="header_block_regrid",
                )
                new_binding_rows.append(row)
                new_ctb_rows.extend(ctb)
            removed.extend(b["binding_id"] for b in contained)
            n += 1
    if removed:
        _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows)
    return {"regridded": n, "removed": len(removed), "added": len(new_binding_rows)}


def _listobject_extent_stage(conn, guid, sheets, name_table_map) -> dict:
    """H9: collapse the per-cell fragments inside each declared Excel Table
    (ListObject) into ONE binding addressed at the table's data_range (header
    excluded). Only bindings fully CONTAINED in the data_range rect are collapsed
    (a straddler is left alone — graceful miss, never corruption). No-op when the
    workbook has no ListObjects."""
    sid_of = {name: sid for sid, name in sheets.items()}
    bindings = _load_bindings_with_cells(conn)
    removed: list[str] = []
    new_binding_rows: list[tuple] = []
    new_ctb_rows: list[tuple] = []
    n = 0
    for t in name_table_map.get_all_tables():
        if not t.data_range or "!" not in t.data_range:
            continue  # header-only / unqualified table
        sheet, rng = t.data_range.rsplit("!", 1)
        sid = sid_of.get(sheet)
        if sid is None:
            continue
        c1, r1, c2, r2 = range_boundaries(rng)  # (min_col,min_row,max_col,max_row)
        frags = []
        data_cells = []
        # Above-edge straddlers: a binding fully within the table's COLUMN span that
        # dips ABOVE the data_range top edge (into the header row) while also holding
        # >=1 in-range value cell — the residue of a column-split that kept the header
        # row attached (e.g. tablebuilder2 catalog cols A1:A2 over data_range A2:E(n)).
        # Containment-only collapse leaves these overlapping the value region, so every
        # catalog reads band_other. Clip each to its header-only cells and fold its
        # in-range value cell into the collapsed binding. Scoped to PURE above-straddlers
        # (all cells in cols [c1,c2], none below r2) so genuine boundary-straddlers and
        # side/below overflows are left untouched (the over-merge-residual caution).
        straddlers = []
        for b in bindings:
            if b["sheet_id"] != sid or not b["cells"]:
                continue
            rows_b = [row for (_, row, _, _, _) in b["cells"]]
            cols_b = [col for (_, _, col, _, _) in b["cells"]]
            if all(r1 <= row <= r2 and c1 <= col <= c2 for (_, row, col, _, _) in b["cells"]):
                frags.append(b)
                data_cells.extend(b["cells"])
            elif (
                all(c1 <= col <= c2 for col in cols_b)
                and max(rows_b) <= r2
                and min(rows_b) < r1
                and any(r1 <= row <= r2 for row in rows_b)
            ):
                straddlers.append(b)
        # Fold straddlers' in-range value cells into the collapsed binding.
        for b in straddlers:
            data_cells.extend(c for c in b["cells"] if r1 <= c[1] <= r2)
        if not data_cells:
            continue
        # Idempotence (R1 fix): compare RECTS, not address strings — a single-cell
        # data_range is stored colon-form "G2:G2" by names_tables but a 1x1 binding
        # is addressed "G2"; a string compare would re-collapse it every pass. Only
        # idempotent when there is also nothing to clip.
        if len(frags) == 1 and not straddlers:
            try:
                fc1, fr1, fc2, fr2 = range_boundaries(frags[0]["address_a1"].split("!")[-1])
            except Exception:
                fc1 = fr1 = fc2 = fr2 = None
            if (fc1, fr1, fc2, fr2) == (c1, r1, c2, r2):
                continue
        parent = min(
            frags + straddlers,
            key=lambda b: (min(c[1] for c in b["cells"]), min(c[2] for c in b["cells"])),
        )
        row, ctb = _build_listobject_binding(parent, data_cells, (r1, c1, r2, c2), guid, sheet)
        new_binding_rows.append(row)
        new_ctb_rows.extend(ctb)
        removed.extend(b["binding_id"] for b in frags)
        # Each clipped straddler becomes a header-only residual binding (its cells
        # above the data_range top edge), addressed at its own bbox.
        for b in straddlers:
            header_cells = [c for c in b["cells"] if c[1] < r1]
            removed.append(b["binding_id"])
            if header_cells:
                hrow, hctb = _build_binding_from_cells(
                    b, header_cells, guid, sheet, "listobject_header"
                )
                new_binding_rows.append(hrow)
                new_ctb_rows.extend(hctb)
        n += 1
    if removed:
        _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows)
    return {"collapsed": n, "removed": len(removed), "added": len(new_binding_rows)}


def _header_strip_stage(conn, guid, sheets, strs=None, edges=None, mmult=None) -> dict:
    """Trim swallowed header rows/columns off the (post-split) binding set, then
    re-split the residual.

    STAGE 1's splitter looks for headers OUTSIDE the block (r1-1 / c1-1), so a block
    whose distinct headers sit at its OWN leading edge (simple Policyholder A1:F101
    headers in row 1; bsm ans_3spread A5:B12 labels in col A) is invisible to it and
    left merged. Header-strip removes that edge line, after which the per-column/per-row
    headers ARE the now-adjacent line (r1-1 / c1-1 of the residual) — so re-running
    _split_groups on the kept cells splits the residual into per-variable bindings.

    Guarded by _keep_matrix on the RESIDUAL extent: a genuine 2D lookup table whose
    header row was just stripped (simple Mortality/Premium B2:E101, consumed 2D by
    VLOOKUP over B2:E101) must stay whole, not shatter into columns. The residual extent
    is what matters — keep_matrix on the header-inclusive extent reads False because the
    consumer's covering edge excludes the header row.

    `strs`/`edges`/`mmult` are the _load_signals outputs; when omitted (older callers)
    the re-split is skipped and the stage behaves as a pure header-strip."""
    dtypes = _load_dtypes(conn)
    bindings = _load_bindings_with_cells(conn)
    removed: list[str] = []
    new_binding_rows: list[tuple] = []
    new_ctb_rows: list[tuple] = []
    n = 0
    resplit = 0
    for b in bindings:
        cells = b["cells"]
        if not cells:
            continue
        # Idempotence: never re-strip a binding this stage already produced.
        if "header_strip" in (b["extraction_source"] or ""):
            continue
        sid = b["sheet_id"]
        sheet_name = sheets.get(sid)
        if sheet_name is None:
            continue
        kept = _trim_header(cells, sid, dtypes)
        if kept is None:
            continue

        # Re-split the residual when its now-adjacent edge line gives distinct headers.
        did_split = False
        if strs is not None:
            sub_rows = sorted({r for (_, r, _, _, _) in kept})
            sub_cols = sorted({c for (_, _, c, _, _) in kept})
            if len(sub_cols) >= 2 or len(sub_rows) >= 2:
                kr1, kr2 = sub_rows[0], sub_rows[-1]
                kc1, kc2 = sub_cols[0], sub_cols[-1]
                is_matrix = (
                    len(sub_cols) >= 2
                    and len(sub_rows) >= 2
                    and _keep_matrix(sid, kr1, kc1, kr2, kc2, edges or {}, mmult or set())
                )
                if not is_matrix:
                    axis, groups = _split_groups(kept, sid, strs)
                    if groups and len(groups) > 1:
                        block_bbox = (kr1, kc1, kr2, kc2)
                        for sub_cells, label in groups:
                            brow, bctb = _build_subbinding(
                                b, sub_cells, label, axis, guid, sheet_name, block_bbox
                            )
                            new_binding_rows.append(brow)
                            new_ctb_rows.extend(bctb)
                        did_split = True
                        resplit += 1

        if not did_split:
            row, ctb = _build_binding_from_cells(b, kept, guid, sheet_name, "header_strip")
            new_binding_rows.append(row)
            new_ctb_rows.extend(ctb)
        removed.append(b["binding_id"])
        n += 1
    if removed:
        _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows)
    return {
        "stripped": n,
        "resplit": resplit,
        "removed": len(removed),
        "added": len(new_binding_rows),
    }


# ----------------------------------------------------------------------------
# Stage 3 — H5 VBA input-panel split (RC9).
# ----------------------------------------------------------------------------


def _load_vba_exact_refs(conn, sheets) -> dict:
    """Return {sheet_id: set((row, col))} for cells a VBA procedure references as
    an EXACT single cell. The program addresses these as distinct variables."""
    if not _table_exists(conn, "vba_procedure_cell_refs"):
        return {}
    name2sid = {name: sid for sid, name in sheets.items()}
    refs: dict[int, set] = {}
    for (target,) in conn.execute(
        "SELECT target FROM vba_procedure_cell_refs "
        "WHERE precision='exact' AND target_kind='cell_range'"
    ):
        if not target or "!" not in target:
            continue
        sheet, rng = target.rsplit("!", 1)
        sheet = sheet.strip("'").replace("''", "'")
        if ":" in rng:
            continue  # single-cell refs only
        try:
            c1, r1, c2, r2 = range_boundaries(rng)
        except Exception:
            continue
        if c1 != c2 or r1 != r2:
            continue
        sid = name2sid.get(sheet)
        if sid is not None:
            refs.setdefault(sid, set()).add((r1, c1))
    return refs


def _vba_panel_stage(conn, guid, sheets, edges, mmult) -> dict:
    """Split a thin scalar-input PANEL whose cells VBA references individually into
    one single-cell binding per referenced cell (+ a residual). A wide/2D block
    VBA reads sparsely (an output matrix) and a formula-consumed matrix (H9) are
    kept whole. Faithful port of spike_h5_vba.transform."""
    refs = _load_vba_exact_refs(conn, sheets)
    if not refs:
        return {"vba_split": 0, "removed": 0, "added": 0}
    bindings = _load_bindings_with_cells(conn)
    removed: list[str] = []
    new_binding_rows: list[tuple] = []
    new_ctb_rows: list[tuple] = []
    n = 0
    for b in bindings:
        cells = b["cells"]
        if len(cells) < 2:
            continue
        # Idempotence: never re-split a binding this stage already produced.
        if "vba_panel" in (b["extraction_source"] or ""):
            continue
        sid = b["sheet_id"]
        sheet_name = sheets.get(sid)
        rc = refs.get(sid)
        if not rc or sheet_name is None:
            continue
        inside = [cell for cell in cells if (cell[1], cell[2]) in rc]
        if len(inside) < 2:
            continue

        # Thin 1D scalar-input panel only (label|value): the referenced cells must
        # be colinear AND the binding thin (<=2) in the perpendicular direction.
        # A wide 2D block whose cells VBA reads sparsely (a Greeks output grid) is
        # kept whole.
        ref_cols = {c for (_, _, c, _, _) in inside}
        ref_rows = {r for (_, r, _, _, _) in inside}
        all_cols = [c for (_, _, c, _, _) in cells]
        all_rows = [r for (_, r, _, _, _) in cells]
        cspan = max(all_cols) - min(all_cols) + 1
        rspan = max(all_rows) - min(all_rows) + 1
        if len(ref_cols) == 1 and cspan <= 2:
            pass  # vertical input panel
        elif len(ref_rows) == 1 and rspan <= 2:
            pass  # horizontal input panel
        else:
            continue  # 2D / wide block -> keep

        # Never split a genuine 2D matrix (H9 guard).
        cols = set(all_cols)
        rows = set(all_rows)
        if (
            len(cols) >= 2
            and len(rows) >= 2
            and _keep_matrix(sid, min(rows), min(cols), max(rows), max(cols), edges, mmult)
        ):
            continue

        # Split: one single-cell binding per VBA-referenced cell + a residual.
        ref_set = {(cell[1], cell[2]) for cell in inside}
        residual = [cell for cell in cells if (cell[1], cell[2]) not in ref_set]
        for cell in inside:
            row, ctb = _build_binding_from_cells(b, [cell], guid, sheet_name, "vba_panel")
            new_binding_rows.append(row)
            new_ctb_rows.extend(ctb)
        if residual:
            row, ctb = _build_binding_from_cells(b, residual, guid, sheet_name, "vba_panel")
            new_binding_rows.append(row)
            new_ctb_rows.extend(ctb)
        removed.append(b["binding_id"])
        n += 1
    if removed:
        _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows)
    return {"vba_split": n, "removed": len(removed), "added": len(new_binding_rows)}


# ----------------------------------------------------------------------------
# Mutation application + edge recompute.
# ----------------------------------------------------------------------------


def _apply_binding_mutations(conn, removed, new_binding_rows, new_ctb_rows):
    """Delete parent bindings + their dependent rows, insert the replacements.
    Edges are NOT touched here — they are recomputed once after all stages."""
    # Drop dependent rows first (binding_label_candidate_cells has a real FK to
    # bindings; clean every binding_id table that exists so the DB stays
    # referentially consistent whether run mid-pipeline or on a built DB).
    for tbl in _BINDING_ID_TABLES:
        if not _table_exists(conn, tbl):
            continue
        conn.executemany(f"DELETE FROM {tbl} WHERE binding_id = ?", [(r,) for r in removed])
    # formula_families references a binding via representative_binding_id (a
    # different column, so the loop above misses it). If a representative is
    # removed, dissolve the whole family + its remaining members so no FK dangles.
    # Empty at pipeline step 6.6 (populated later at 6.85); only non-empty on a
    # finished-DB run (the gate / a re-run).
    if _table_exists(conn, "formula_families"):
        for rid in removed:
            fams = [
                f
                for (f,) in conn.execute(
                    "SELECT family_id FROM formula_families WHERE representative_binding_id = ?",
                    (rid,),
                )
            ]
            for fam in fams:
                conn.execute("DELETE FROM formula_family_members WHERE family_id = ?", (fam,))
                conn.execute("DELETE FROM formula_families WHERE family_id = ?", (fam,))
    conn.executemany("DELETE FROM bindings WHERE binding_id = ?", [(r,) for r in removed])
    # OR IGNORE: a split/trim sub-binding can be geometrically identical to a
    # SURVIVING binding (same sheet/top-left/shape/cells -> same binding_id). That
    # is the same binding, so merge into it rather than crash on the PK. Plain
    # INSERT would abort the whole pass on such a collision.
    conn.executemany(
        "INSERT OR IGNORE INTO bindings ("
        "binding_id, sheet_id, address_a1, top_left_cell_id, shape_rows, "
        "shape_cols, binding_type, formula_id, label, classification, "
        "confidence, is_orphan, extraction_source, evidence_blob_id, "
        "spatial_candidates_blob_id) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        new_binding_rows,
    )
    conn.executemany(
        "INSERT OR IGNORE INTO cell_to_binding (cell_id, binding_id) VALUES (?, ?)",
        new_ctb_rows,
    )


def _rebuild_edges(conn):
    """Recompute binding edges from the underlying cell/range edges over the final
    binding set — never re-point edges by hand (§5.3).

    _write_binding_edges_from_cells only regenerates kind='formula' edges BETWEEN
    real cell-bindings (it joins cell/range edges via cell_to_binding). Edges it
    CANNOT regenerate must be preserved:
      - non-formula edges (e.g. 'via_vba_paste'); and
      - any edge touching a VBA procedure node ('vba::...'), which is not a cell-
        binding — these are stored as kind='formula' but join a procedure to a
        cell-binding, so the cell-based rebuild would silently drop them.
    Preserve such an edge only if every NON-vba endpoint is still a live binding
    (one referencing a removed/transient binding is unrecoverable and would
    dangle). In the production pipeline (step 6.6) no VBA edges exist yet, so the
    preserve set is empty; this only matters on a finished DB (the gate / re-run)."""
    live = {bid for (bid,) in conn.execute("SELECT binding_id FROM bindings")}

    def _endpoint_ok(bid):
        return bid.startswith("vba::") or bid in live

    preserved = [
        e
        for e in conn.execute(
            "SELECT from_binding_id, to_binding_id, edge_count, kind, provenance_proc "
            "FROM binding_edges"
        ).fetchall()
        # keep only edges the cell-based rebuild won't recreate
        if (e[3] != "formula" or e[0].startswith("vba::") or e[1].startswith("vba::"))
        and _endpoint_ok(e[0])
        and _endpoint_ok(e[1])
    ]
    conn.execute("DELETE FROM binding_edges")
    _write_binding_edges_from_cells(conn)
    if preserved:
        # OR IGNORE: binding_edges' PK is (from,to) without kind, so a pair can
        # hold only one edge. On a collision the just-written formula edge wins
        # and the preserved non-formula edge is dropped. This matches production
        # semantics: VBA paste edges are wired AFTER this pass (step 6.9) with the
        # same OR IGNORE, so a formula-connected pair never gets a VBA edge there
        # either — i.e. a re-run converges to what a fresh build produces.
        conn.executemany(
            "INSERT OR IGNORE INTO binding_edges "
            "(from_binding_id, to_binding_id, edge_count, kind, provenance_proc) "
            "VALUES (?,?,?,?,?)",
            preserved,
        )
