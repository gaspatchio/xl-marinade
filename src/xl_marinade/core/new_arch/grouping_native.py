# ABOUTME: Native grouping/refinement for fast pipeline outputs.
# ABOUTME: Operates directly on fast schema tables and writes bindings to fast schema.

"""
Native Grouping/Refinement for Fast Pipeline

Consumes fast schema tables (cells/formulas/edges) to build bindings and binding
edges without the legacy adapter layer.
"""

import builtins
import json
import re
import sqlite3
import sys
from bisect import bisect_left, bisect_right
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from typing import Any

from xl_marinade.core.bindings import (
    Binding,
    _col_to_letter,
    compute_binding_id,
    compute_cells_structure_hash,
    group_cells_into_bindings,
)
from xl_marinade.core.cells import classify_dtype
from xl_marinade.core.grouping.refinement import RefinementEngine
from xl_marinade.core.grouping.rules import merge_constant_ranges
from xl_marinade.core.lazy_formulas import SheetFormulaCache
from xl_marinade.core.lazy_workbook import LazyWorkbook
from xl_marinade.core.names_tables import NameTableMap
from xl_marinade.core.new_arch.canonical_json import canonicalize_and_hash, hash_json
from xl_marinade.core.new_arch.cell_identity import pack as pack_cell_id
from xl_marinade.core.new_arch.cell_identity import unpack as unpack_cell_id
from xl_marinade.core.parser import parse_formula
from xl_marinade.core.ref_converter import parse_cell_address
from xl_marinade.core.ref_extractor import (
    expand_range_to_cells,
    extract_references_from_ast,
    is_defined_name,
)
from xl_marinade.core.reverse_index import ReverseIndex


def print(*args, **kwargs):  # noqa: A001 - keep `marinade extract` stdout clean
    """Route this module's diagnostic output to stderr so stdout stays clean for
    piping. Call sites that pass an explicit ``file=`` still win (e.g. the phase
    timers that already target stderr)."""
    kwargs.setdefault("file", sys.stderr)
    builtins.print(*args, **kwargs)


_PLACEHOLDER_TOKENS: set[str] = {
    "-",
    "–",
    "—",
    "tbc",
    "tbd",
    "na",
    "n/a",
    "n.a.",
    "n\\a",
}

_SECTION_HEADER_RE = re.compile(r"^\s*(\d+\.\s+\S|[A-Z]\.\s+\S|section\s+\S)", re.IGNORECASE)

_SCENARIOISH_TOKENS: tuple[str, ...] = (
    "base",
    "shock",
    "stress",
    "flat",
    "scenario",
    "scen",
)


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        return str(value).strip()
    return str(value).strip()


def _is_scenarioish_header_text(text: str) -> bool:
    t = text.strip().lower()
    if not t:
        return False
    if "%" in t:
        return True
    for token in _SCENARIOISH_TOKENS:
        if token in t:
            return True
    # Avoid matching "setup"/"download" etc.
    if re.search(r"(^|[^a-z])up($|[^a-z])", t):
        return True
    if re.search(r"(^|[^a-z])down($|[^a-z])", t):
        return True
    return False


def _split_scenarioish_heterogeneous_constant_bindings(
    bindings: list[Binding],
    workbook: LazyWorkbook | None,
    workbook_guid: str,
    cell_formulas: dict[str, str],
) -> list[Binding]:
    """
    Proposal #2: Split wide constant bindings into per-column bindings when headers look
    like distinct scenario labels (e.g., Base/Shock/Stress/%/Flat/Up/Down).
    """
    if not bindings or workbook is None:
        return bindings

    out: list[Binding] = []
    split_blocks = 0

    for b in bindings:
        if b.binding_type != "constant" or b.shape_cols < 2 or b.shape_rows < 2:
            out.append(b)
            continue
        if b.shape_cols > 20:
            out.append(b)
            continue
        if b.shape_rows * b.shape_cols > 10_000:
            out.append(b)
            continue

        parsed = parse_cell_address(b.address_a1)
        if not parsed:
            out.append(b)
            continue
        r1 = int(parsed.get("row", 1))
        c1 = int(parsed.get("col", 1))
        h = int(parsed.get("height", 1))
        w = int(parsed.get("width", 1))
        header_row = r1 - 1
        if header_row < 1:
            out.append(b)
            continue

        try:
            ws = workbook[b.sheet]
        except Exception:
            out.append(b)
            continue

        scenarioish_headers: list[str] = []
        for dc in range(w):
            raw = ws.cell(row=header_row, column=c1 + dc).value
            header = _normalize_header_text(raw)
            if header and _is_scenarioish_header_text(header):
                scenarioish_headers.append(header.strip().lower())

        if len(scenarioish_headers) < 2 or len(set(scenarioish_headers)) < 2:
            out.append(b)
            continue

        col_to_cells: dict[int, list[str]] = {col: [] for col in range(c1, c1 + w)}
        ok = True
        for addr in b.cells:
            p = parse_cell_address(addr)
            if not p:
                ok = False
                break
            col = int(p.get("col", 0))
            if col not in col_to_cells:
                ok = False
                break
            col_to_cells[col].append(addr.replace("$", ""))

        if not ok:
            out.append(b)
            continue

        split_bindings: list[Binding] = []
        quoted_sheet = _quote_sheet_name_for_address(b.sheet)
        for col in range(c1, c1 + w):
            cells = sorted(col_to_cells.get(col, []))
            if len(cells) != h:
                ok = False
                break
            start = f"{_col_to_letter(col)}{r1}"
            end = f"{_col_to_letter(col)}{r1 + h - 1}"
            address = f"{quoted_sheet}!{start}" if start == end else f"{quoted_sheet}!{start}:{end}"
            split_bindings.append(
                _create_binding_from_address(address, cells, workbook_guid, cell_formulas)
            )

        if not ok or not split_bindings:
            out.append(b)
            continue

        split_blocks += 1
        out.extend(split_bindings)

    if split_blocks:
        print(f"  Split scenario-ish constant blocks: {split_blocks}")

    return out


def _parse_time_index(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-6:
            return int(round(value))
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        if _is_scenarioish_header_text(s):
            return None
        if s.isdigit():
            return int(s)
        m = re.search(r"\b(\d{1,4})\b", s)
        if not m:
            return None
        return int(m.group(1))
    return None


def _looks_like_continuous_time_axis(indices: list[int | None]) -> bool:
    if not indices:
        return False
    nums = [n for n in indices if isinstance(n, int)]
    if len(nums) < 5:
        return False
    if len(nums) / len(indices) < 0.7:
        return False
    if any(b <= a for a, b in zip(nums, nums[1:], strict=False)):
        return False
    deltas = [b - a for a, b in zip(nums, nums[1:], strict=False)]
    consecutive = sum(1 for d in deltas if d == 1)
    return consecutive >= max(3, int(0.8 * len(deltas)))


def _row_vector_has_continuous_time_axis(
    cell_values: dict[tuple[int, int], Any],
    data_row: int,
    start_col: int,
    end_col: int,
) -> bool:
    header_row = data_row - 1
    if header_row < 1:
        return False
    indices: list[int | None] = []
    for c in range(start_col, end_col + 1):
        raw = cell_values.get((header_row, c))
        indices.append(_parse_time_index(raw))
    return _looks_like_continuous_time_axis(indices)


def _col_vector_has_continuous_time_axis(
    cell_values: dict[tuple[int, int], Any],
    data_col: int,
    start_row: int,
    end_row: int,
) -> bool:
    axis_col = data_col - 1
    if axis_col < 1:
        return False
    indices: list[int | None] = []
    for r in range(start_row, end_row + 1):
        raw = cell_values.get((r, axis_col))
        indices.append(_parse_time_index(raw))
    return _looks_like_continuous_time_axis(indices)


def _load_sheet_cell_values(
    conn: sqlite3.Connection,
    sheet_name: str,
) -> dict[tuple[int, int], Any]:
    """Pre-load all cell values for a sheet from the DB.

    Returns dict mapping (row, col) -> parsed value. Used by time-series
    axis checks to avoid per-cell openpyxl access (was 310s bottleneck).
    """
    import json as _json

    values: dict[tuple[int, int], Any] = {}
    try:
        rows = conn.execute(
            """
            SELECT c.row, c.col, jv.json
            FROM cells c
            JOIN sheets s ON c.sheet_id = s.sheet_id
            LEFT JOIN json_blobs jv ON c.value_blob_id = jv.blob_id
            WHERE s.sheet_name = ?
            """,
            (sheet_name,),
        ).fetchall()
        for row, col, val_json in rows:
            if val_json is not None:
                try:
                    values[(row, col)] = _json.loads(val_json)
                except (ValueError, TypeError):
                    values[(row, col)] = val_json
    except sqlite3.Error:
        pass
    return values


def _merge_continuous_time_series_formula_vectors(
    bindings: list[Binding],
    workbook: LazyWorkbook | None,
    workbook_guid: str,
    cell_formulas: dict[str, str],
    conn: sqlite3.Connection | None = None,
) -> list[Binding]:
    """
    Proposal #7: Merge contiguous formula-vector segments into a single time-series binding
    when the axis labels imply a continuous period sequence.

    Guardrails:
    - Only merge 1×N or N×1 formula bindings.
    - Require a continuous time-axis (no resets), so side-by-side scenario blocks like
      Year 1..20 repeated won't merge.
    """
    if not bindings or (workbook is None and conn is None):
        return bindings

    by_sheet: dict[str, list[Binding]] = {}
    for b in bindings:
        by_sheet.setdefault(b.sheet, []).append(b)

    out: list[Binding] = []
    merged_series = 0

    for sheet, sheet_bindings in by_sheet.items():
        # Load cell values from DB (fast) instead of openpyxl (was 310s bottleneck)
        if conn is not None:
            cell_values = _load_sheet_cell_values(conn, sheet)
        else:
            # Legacy fallback via openpyxl
            try:
                _ws = workbook[sheet]
            except Exception:
                out.extend(sheet_bindings)
                continue
            cell_values = {}
            # Note: this path is slow but preserved for backward compatibility
            continue

        parsed_rows: list[tuple[Binding, int, int, int, int]] = []
        for b in sheet_bindings:
            p = parse_cell_address(b.address_a1)
            if not p:
                continue
            r1 = int(p.get("row", 1))
            c1 = int(p.get("col", 1))
            h = int(p.get("height", 1))
            w = int(p.get("width", 1))
            parsed_rows.append((b, r1, c1, r1 + h - 1, c1 + w - 1))

        used: set[str] = set()

        # Merge row-vectors (same row, contiguous columns).
        row_vectors: dict[int, list[tuple[Binding, int, int, int]]] = {}
        for b, r1, c1, r2, c2 in parsed_rows:
            if b.binding_type != "formula" or r1 != r2:
                continue
            if (c2 - c1 + 1) < 2:
                continue
            row_vectors.setdefault(r1, []).append((b, c1, c2, r1))

        for row, segs in row_vectors.items():
            segs.sort(key=lambda t: t[1])
            i = 0
            while i < len(segs):
                b0, start_c, end_c, _ = segs[i]
                if b0.binding_id in used:
                    i += 1
                    continue
                run: list[tuple[Binding, int, int]] = [(b0, start_c, end_c)]
                run_start = start_c
                run_end = end_c

                j = i
                while j + 1 < len(segs):
                    b1, c1, c2, _ = segs[j + 1]
                    if b1.binding_id in used:
                        j += 1
                        continue
                    if c1 != run_end + 1:
                        break
                    run.append((b1, c1, c2))
                    run_end = c2
                    j += 1

                total_len = run_end - run_start + 1
                if (
                    total_len >= 5
                    and total_len <= 500
                    and _row_vector_has_continuous_time_axis(cell_values, row, run_start, run_end)
                    and len(run) > 1
                ):
                    merged_cells: list[str] = []
                    merged_ids: list[str] = []
                    for b, _c1, _c2 in run:
                        merged_cells.extend(b.cells)
                        merged_ids.append(b.binding_id)
                    for bid in merged_ids:
                        used.add(bid)
                    quoted_sheet = _quote_sheet_name_for_address(sheet)
                    start = f"{_col_to_letter(run_start)}{row}"
                    end = f"{_col_to_letter(run_end)}{row}"
                    addr = (
                        f"{quoted_sheet}!{start}:{end}"
                        if start != end
                        else f"{quoted_sheet}!{start}"
                    )
                    merged_set = sorted({c.replace("$", "") for c in merged_cells})
                    if len(merged_set) == total_len:
                        out.append(
                            _create_binding_from_address(
                                addr, merged_set, workbook_guid, cell_formulas
                            )
                        )
                        merged_series += 1
                    else:
                        for b, _c1, _c2 in run:
                            out.append(b)
                            used.add(b.binding_id)
                else:
                    for b, _c1, _c2 in run:
                        out.append(b)
                        used.add(b.binding_id)

                i = j + 1

        # Merge column-vectors (same column, contiguous rows).
        col_vectors: dict[int, list[tuple[Binding, int, int, int]]] = {}
        for b, r1, c1, r2, c2 in parsed_rows:
            if b.binding_type != "formula" or c1 != c2:
                continue
            if (r2 - r1 + 1) < 2:
                continue
            col_vectors.setdefault(c1, []).append((b, r1, r2, c1))

        for col, segs in col_vectors.items():
            segs.sort(key=lambda t: t[1])
            i = 0
            while i < len(segs):
                b0, start_r, end_r, _ = segs[i]
                if b0.binding_id in used:
                    i += 1
                    continue
                run: list[tuple[Binding, int, int]] = [(b0, start_r, end_r)]
                run_start = start_r
                run_end = end_r

                j = i
                while j + 1 < len(segs):
                    b1, r1, r2, _ = segs[j + 1]
                    if b1.binding_id in used:
                        j += 1
                        continue
                    if r1 != run_end + 1:
                        break
                    run.append((b1, r1, r2))
                    run_end = r2
                    j += 1

                total_len = run_end - run_start + 1
                if (
                    total_len >= 5
                    and total_len <= 500
                    and _col_vector_has_continuous_time_axis(cell_values, col, run_start, run_end)
                    and len(run) > 1
                ):
                    merged_cells: list[str] = []
                    merged_ids: list[str] = []
                    for b, _r1, _r2 in run:
                        merged_cells.extend(b.cells)
                        merged_ids.append(b.binding_id)
                    for bid in merged_ids:
                        used.add(bid)
                    quoted_sheet = _quote_sheet_name_for_address(sheet)
                    start = f"{_col_to_letter(col)}{run_start}"
                    end = f"{_col_to_letter(col)}{run_end}"
                    addr = (
                        f"{quoted_sheet}!{start}:{end}"
                        if start != end
                        else f"{quoted_sheet}!{start}"
                    )
                    merged_set = sorted({c.replace("$", "") for c in merged_cells})
                    if len(merged_set) == total_len:
                        out.append(
                            _create_binding_from_address(
                                addr, merged_set, workbook_guid, cell_formulas
                            )
                        )
                        merged_series += 1
                    else:
                        for b, _r1, _r2 in run:
                            out.append(b)
                            used.add(b.binding_id)
                else:
                    for b, _r1, _r2 in run:
                        out.append(b)
                        used.add(b.binding_id)

                i = j + 1

        # Add any remaining bindings not handled above.
        for b, _r1, _c1, _r2, _c2 in parsed_rows:
            if b.binding_id in used:
                continue
            out.append(b)
            used.add(b.binding_id)

        # Add any bindings that failed parsing.
        for b in sheet_bindings:
            if b.binding_id not in used:
                out.append(b)
                used.add(b.binding_id)

    if merged_series:
        print(f"  Merged continuous time-series vectors: {merged_series}")

    return out


def _iter_int_chunks(items: list[int], chunk_size: int = 900) -> Iterable[list[int]]:
    for i in range(0, len(items), chunk_size):
        yield items[i : i + chunk_size]


def _has_incoming_cell_edges(conn: sqlite3.Connection, cell_ids: list[int]) -> bool:
    if not cell_ids:
        return False
    for chunk in _iter_int_chunks(cell_ids):
        placeholders = ",".join("?" for _ in chunk)
        row = conn.execute(
            f"SELECT 1 FROM cell_edges_internal WHERE to_cell_id IN ({placeholders}) LIMIT 1",
            chunk,
        ).fetchone()
        if row:
            return True
    return False


def _binding_is_completely_empty(conn: sqlite3.Connection, cell_ids: list[int]) -> bool:
    """True if every cell in the binding has no formula AND no value."""
    if not cell_ids:
        return True
    for chunk in _iter_int_chunks(cell_ids):
        placeholders = ",".join("?" for _ in chunk)
        row = conn.execute(
            f"""
            SELECT 1 FROM cells c
            WHERE c.cell_id IN ({placeholders})
              AND (c.formula_id IS NOT NULL OR c.value_blob_id IS NOT NULL)
            LIMIT 1
            """,
            chunk,
        ).fetchone()
        if row:
            return False  # At least one cell has content
    return True


def _binding_is_placeholder_only(conn: sqlite3.Connection, cell_ids: list[int]) -> bool:
    if not cell_ids:
        return False
    saw_token = False
    for chunk in _iter_int_chunks(cell_ids):
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT jv.json
            FROM cells c
            LEFT JOIN json_blobs jv ON c.value_blob_id = jv.blob_id
            WHERE c.cell_id IN ({placeholders})
            """,
            chunk,
        ).fetchall()
        for (raw_json,) in rows:
            if raw_json is None:
                continue
            try:
                value = json.loads(raw_json)
            except Exception:
                return False
            if value is None:
                continue
            if not isinstance(value, str):
                return False
            token = value.strip().lower()
            if not token:
                continue
            saw_token = True
            if token not in _PLACEHOLDER_TOKENS:
                return False
    return saw_token


def _binding_is_section_header_only(conn: sqlite3.Connection, cell_ids: list[int]) -> bool:
    if not cell_ids:
        return False
    saw_text = False
    saw_sectionish = False
    for chunk in _iter_int_chunks(cell_ids):
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT jv.json
            FROM cells c
            LEFT JOIN json_blobs jv ON c.value_blob_id = jv.blob_id
            WHERE c.cell_id IN ({placeholders})
            """,
            chunk,
        ).fetchall()
        for (raw_json,) in rows:
            if raw_json is None:
                continue
            try:
                value = json.loads(raw_json)
            except Exception:
                return False
            if value is None:
                continue
            if not isinstance(value, str):
                return False
            text = value.strip()
            if not text:
                continue
            saw_text = True
            if _SECTION_HEADER_RE.match(text):
                saw_sectionish = True
    return saw_text and saw_sectionish


def _binding_is_consecutive_int_index(conn: sqlite3.Connection, cell_ids: list[int]) -> bool:
    if not cell_ids:
        return False
    triples: list[tuple[int, int, Any]] = []
    for chunk in _iter_int_chunks(cell_ids):
        placeholders = ",".join("?" for _ in chunk)
        triples.extend(
            conn.execute(
                f"""
                SELECT c.row, c.col, jv.json
                FROM cells c
                LEFT JOIN json_blobs jv ON c.value_blob_id = jv.blob_id
                WHERE c.cell_id IN ({placeholders})
                """,
                chunk,
            ).fetchall()
        )
    triples.sort(key=lambda t: (int(t[0]), int(t[1])))

    ints: list[int] = []
    for _, _, raw_json in triples:
        if raw_json is None:
            return False
        try:
            value = json.loads(raw_json)
        except Exception:
            return False
        if value is None:
            return False
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                rounded = int(round(value))
                if abs(value - rounded) > 1e-9:
                    return False
                value = rounded
            ints.append(int(value))
            continue
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return False
            try:
                ints.append(int(text))
                continue
            except ValueError:
                return False
        return False

    if len(ints) < 5:
        return False
    for a, b in zip(ints, ints[1:], strict=False):
        if b - a != 1:
            return False
    return True


def _create_binding_from_address(
    address_a1: str,
    cells: list[str],
    workbook_guid: str,
    cell_formulas: dict[str, str],
) -> Binding:
    parsed = parse_cell_address(address_a1)
    if not parsed:
        raise ValueError(f"Could not parse address: {address_a1}")

    sheet = str(parsed.get("sheet", ""))
    row = int(parsed.get("row", 1))
    col = int(parsed.get("col", 1))
    height = int(parsed.get("height", 1))
    width = int(parsed.get("width", 1))
    top_left_a1 = f"{_col_to_letter(col)}{row}"

    has_formula = any(bool(cell_formulas.get(c, "").strip()) for c in cells)
    binding_type = "formula" if has_formula else "constant"
    cells_structure_hash = compute_cells_structure_hash(cells, cell_formulas)
    binding_id = compute_binding_id(
        workbook_guid=workbook_guid,
        sheet=sheet.strip("'"),
        top_left_a1=top_left_a1,
        shape_rows=height,
        shape_cols=width,
        cells_structure_hash=cells_structure_hash,
    )
    return Binding(
        binding_id=binding_id,
        debug_label=f"{sheet}::{address_a1}",
        sheet=sheet.strip("'"),
        address_a1=address_a1,
        top_left_a1=top_left_a1,
        shape_rows=height,
        shape_cols=width,
        binding_type=binding_type,
        cells_structure_hash=cells_structure_hash,
        cells=sorted(cells),
    )


def _create_binding_from_cells_with_id(
    binding_id: str,
    cells: list[str],
    cell_formulas: dict[str, str],
) -> Binding | None:
    """Create a binding from explicit cells (must form a full rectangle)."""
    normalized_cells = sorted({c.replace("$", "") for c in cells})
    if not normalized_cells:
        return None

    parsed_cells = [parse_cell_address(c) for c in normalized_cells]
    if any(not p for p in parsed_cells):
        return None

    sheet_names = {str(p.get("sheet", "")) for p in parsed_cells if p}
    if len(sheet_names) != 1:
        return None
    sheet = next(iter(sheet_names))
    if not sheet:
        return None

    rows = [int(p.get("row", 0)) for p in parsed_cells if p]
    cols = [int(p.get("col", 0)) for p in parsed_cells if p]
    if not rows or not cols:
        return None

    min_r, max_r = min(rows), max(rows)
    min_c, max_c = min(cols), max(cols)
    shape_rows = max_r - min_r + 1
    shape_cols = max_c - min_c + 1
    expected = shape_rows * shape_cols
    if expected != len(normalized_cells):
        return None

    start = f"{_col_to_letter(min_c)}{min_r}"
    end = f"{_col_to_letter(max_c)}{max_r}"
    quoted_sheet = _quote_sheet_name_for_address(sheet)
    address = f"{quoted_sheet}!{start}" if start == end else f"{quoted_sheet}!{start}:{end}"
    top_left = f"{_col_to_letter(min_c)}{min_r}"
    has_formula = any(bool(cell_formulas.get(c, "").strip()) for c in normalized_cells)
    binding_type = "formula" if has_formula else "constant"
    cells_structure_hash = compute_cells_structure_hash(normalized_cells, cell_formulas)
    return Binding(
        binding_id=binding_id,
        debug_label=f"{sheet}::{address}",
        sheet=sheet,
        address_a1=address,
        top_left_a1=top_left,
        shape_rows=shape_rows,
        shape_cols=shape_cols,
        binding_type=binding_type,
        cells_structure_hash=cells_structure_hash,
        cells=normalized_cells,
    )


def _apply_init_merger_mutations(
    bindings: list[Binding],
    cell_formulas: dict[str, str],
    ir_db_path: str | None = None,
    ir_db_conn: sqlite3.Connection | None = None,
) -> list[Binding]:
    """
    Apply post-IR init-merger logic (Story 31 + Story 37) to grouped bindings.
    """
    if not bindings or (not ir_db_path and not ir_db_conn):
        return bindings

    from xl_marinade.core.labelling.init_merger import (
        create_init_merge_mutations,
        create_multi_init_merge_mutations,
        detect_init_patterns,
        detect_multi_init_patterns,
        detect_projection_init_singletons,
    )

    init_patterns = detect_init_patterns(
        bindings,
        ir_db_path=ir_db_path or "",
        overlay_db_path=None,
        ir_db_conn=ir_db_conn,
        cell_formulas=cell_formulas,
    )
    mutations = create_init_merge_mutations(init_patterns)

    consumed_by_story31 = {p.propagation_binding_id for p in init_patterns}
    consumed_by_story31 |= {p.init_binding_id for p in init_patterns}
    multi_init_candidates = detect_multi_init_patterns(
        bindings,
        ir_db_path=ir_db_path or "",
        overlay_db_path=None,
        ir_db_conn=ir_db_conn,
    )
    multi_init_candidates = [
        c for c in multi_init_candidates if c.propagation_binding_id not in consumed_by_story31
    ]
    if multi_init_candidates:
        mutations.extend(
            create_multi_init_merge_mutations(
                multi_init_candidates,
                overlay_db_path=None,
                ir_db_path=ir_db_path,
                ir_db_conn=ir_db_conn,
            )
        )

    # Track all consumed bindings from Story 31 + Story 37
    consumed_all = set(consumed_by_story31)
    for c in multi_init_candidates:
        consumed_all.add(c.propagation_binding_id)
        consumed_all.add(c.init_binding_id)

    # Story 38: Projection init singletons (no formula reference required)
    # Handles cases like Q7 (=1) above Q8:Q607 where Q8 doesn't reference Q7
    proj_init_patterns = detect_projection_init_singletons(
        bindings,
        ir_db_path=ir_db_path or "",
        already_consumed=consumed_all,
        ir_db_conn=ir_db_conn,
        cell_formulas=cell_formulas,
    )
    if proj_init_patterns:
        mutations.extend(create_init_merge_mutations(proj_init_patterns))

    if not mutations:
        return bindings

    by_id: dict[str, Binding] = {b.binding_id: b for b in bindings}
    order: list[str] = [b.binding_id for b in bindings]
    active: dict[str, bool] = {b.binding_id: True for b in bindings}
    merged_count = 0

    for mutation in mutations:
        action = str(mutation.get("action", ""))
        params = mutation.get("parameters", {})

        if action == "merge_bindings":
            source_ids = list(params.get("source_binding_ids", []))
            if not source_ids:
                continue
            if any(src_id not in by_id or not active.get(src_id, False) for src_id in source_ids):
                continue

            cell_subset_raw = params.get("cell_subset")
            cell_subset = None
            if isinstance(cell_subset_raw, list):
                cell_subset = {str(c).replace("$", "") for c in cell_subset_raw}

            merged_cells: list[str] = []
            for i, src_id in enumerate(source_ids):
                src_binding = by_id[src_id]
                src_cells = [c.replace("$", "") for c in src_binding.cells]
                if i == 0 and cell_subset is not None:
                    src_cells = [c for c in src_cells if c in cell_subset]
                merged_cells.extend(src_cells)

            merged_cells = sorted(set(merged_cells))
            if not merged_cells:
                continue

            new_binding_id = str(params.get("new_binding_id", "")).strip()
            if not new_binding_id:
                continue
            new_binding = _create_binding_from_cells_with_id(
                binding_id=new_binding_id,
                cells=merged_cells,
                cell_formulas=cell_formulas,
            )
            if new_binding is None:
                continue

            by_id[new_binding_id] = new_binding
            if new_binding_id not in order:
                order.append(new_binding_id)
            active[new_binding_id] = True
            merged_count += 1

            for i, src_id in enumerate(source_ids):
                if cell_subset is not None and i == 0:
                    continue
                active[src_id] = False

        elif action == "disable_binding":
            binding_id = str(params.get("binding_id", "")).strip()
            if binding_id:
                active[binding_id] = False

    if merged_count:
        print(f"  Applied init-merger mutations: {merged_count}")

    return [by_id[bid] for bid in order if active.get(bid, False) and bid in by_id]


def _merge_index_vectors_into_adjacent_tables(
    bindings: list[Binding],
    conn: sqlite3.Connection,
    addr_to_cell_id: dict[str, int],
    cell_formulas: dict[str, str],
    workbook_guid: str,
    workbook: LazyWorkbook | None = None,
) -> list[Binding]:
    """
    Proposal #1: merge index vectors (e.g., age/year columns) into the adjacent table they index.

    This only merges when the index is a single row/column with consecutive integer values,
    and when the neighboring binding shares the same span.
    """
    by_sheet: dict[str, list[Binding]] = {}
    for b in bindings:
        by_sheet.setdefault(b.sheet, []).append(b)

    out: list[Binding] = []
    for sheet, sheet_bindings in by_sheet.items():
        # Map from geometry to binding for quick neighbor lookup.
        parsed_rows: list[tuple[Binding, int, int, int, int]] = []
        for b in sheet_bindings:
            p = parse_cell_address(b.address_a1)
            if not p:
                continue
            r1 = int(p.get("row", 1))
            c1 = int(p.get("col", 1))
            h = int(p.get("height", 1))
            w = int(p.get("width", 1))
            parsed_rows.append((b, r1, c1, r1 + h - 1, c1 + w - 1))

        used: set[str] = set()
        for b, r1, c1, r2, c2 in parsed_rows:
            if b.binding_id in used:
                continue

            height = r2 - r1 + 1
            width = c2 - c1 + 1
            is_col_vector = width == 1 and height >= 5
            is_row_vector = height == 1 and width >= 5
            if not (is_col_vector or is_row_vector):
                out.append(b)
                used.add(b.binding_id)
                continue

            cell_ids = [addr_to_cell_id[c] for c in b.cells if c in addr_to_cell_id]
            if not cell_ids or not _binding_is_consecutive_int_index(conn, cell_ids):
                out.append(b)
                used.add(b.binding_id)
                continue

            neighbor = None
            merged_r1 = merged_c1 = merged_r2 = merged_c2 = None
            neighbor_start_col: int | None = None
            # Index column on the left of a table.
            if is_col_vector:
                for other, or1, oc1, or2, oc2 in parsed_rows:
                    if other.binding_id in used or other.binding_id == b.binding_id:
                        continue
                    # Don't merge an index vector into another 1D vector. Keep indices as standalone
                    # bindings unless the neighbor is a genuine table (width > 1).
                    if (oc2 - oc1 + 1) <= 1:
                        continue
                    if or1 == r1 and or2 == r2 and oc1 == c2 + 1:
                        neighbor = other
                        neighbor_start_col = oc1
                        merged_r1, merged_c1 = r1, c1
                        merged_r2, merged_c2 = r2, oc2
                        break
            # Index row above a table.
            if neighbor is None and is_row_vector:
                for other, or1, oc1, or2, oc2 in parsed_rows:
                    if other.binding_id in used or other.binding_id == b.binding_id:
                        continue
                    # Don't merge an index row into another 1D row vector. Keep indices as standalone
                    # bindings unless the neighbor is a genuine table (height > 1).
                    if (or2 - or1 + 1) <= 1:
                        continue
                    if oc1 == c1 and oc2 == c2 and or1 == r2 + 1:
                        neighbor = other
                        merged_r1, merged_c1 = r1, c1
                        merged_r2, merged_c2 = or2, c2
                        break

            if neighbor is None:
                out.append(b)
                used.add(b.binding_id)
                continue

            # Proposal #2 guardrail: if the adjacent binding has a scenario-ish header label
            # (e.g., "Base", "2% Flat", "Stress"), keep the index as a separate binding.
            if workbook is not None and is_col_vector and neighbor_start_col is not None and r1 > 1:
                try:
                    ws = workbook[sheet]
                    header_val = ws.cell(row=r1 - 1, column=neighbor_start_col).value
                    header = _normalize_header_text(header_val)
                    if header and _is_scenarioish_header_text(header):
                        out.append(b)
                        used.add(b.binding_id)
                        continue
                except Exception:
                    pass

            start = f"{_col_to_letter(merged_c1)}{merged_r1}"
            end = f"{_col_to_letter(merged_c2)}{merged_r2}"
            quoted_sheet = _quote_sheet_name_for_address(sheet)
            merged_address = f"{quoted_sheet}!{start}:{end}"
            # Build full visited rectangle to avoid shape/cell mismatches.
            try:
                expanded = (
                    expand_range_to_cells(merged_address, max_cells=10000, populated_cells=None)
                    or []
                )
            except Exception:
                # Fall back to union of the two cell sets.
                expanded = list({*b.cells, *neighbor.cells})
            normalized = [c.replace("$", "") for c in expanded]
            cells = [c for c in normalized if c in cell_formulas]
            expected = (merged_r2 - merged_r1 + 1) * (merged_c2 - merged_c1 + 1)
            if len(cells) != expected:
                # Only merge when the full rectangle is present in the visited cell set.
                out.append(b)
                used.add(b.binding_id)
                continue

            used.add(b.binding_id)
            used.add(neighbor.binding_id)
            out.append(
                _create_binding_from_address(merged_address, cells, workbook_guid, cell_formulas)
            )

        # Add any bindings that were not parsed (should be rare)
        for b in sheet_bindings:
            if b.binding_id not in used:
                out.append(b)
                used.add(b.binding_id)

    return out


def _is_full_column_reference(range_ref: str) -> bool:
    """
    Detect if a range reference is a full-column reference (e.g., C:C, $C:$C, C1:C1048576).

    Args:
        range_ref: A1 notation range (may include sheet name)

    Returns:
        True if this is a full-column reference, False otherwise
    """
    # Strip sheet name if present
    addr = range_ref.split("!")[-1] if "!" in range_ref else range_ref
    # Full column pattern 1: C:C or $C:$C
    if re.match(r"^\$?[A-Z]+:\$?[A-Z]+$", addr):
        return True
    # Full column pattern 2: C1:C1048576 (expanded bounds)
    # Excel max row is 1,048,576
    if re.match(r"^\$?([A-Z]+)\$?1:\$?\1\$?1048576$", addr):
        return True
    return False


def _is_full_row_reference(range_ref: str) -> bool:
    """
    Detect if a range reference is a full-row reference (e.g., 5:5, $5:$5, A5:XFD5).

    Args:
        range_ref: A1 notation range (may include sheet name)

    Returns:
        True if this is a full-row reference, False otherwise
    """
    # Strip sheet name if present
    addr = range_ref.split("!")[-1] if "!" in range_ref else range_ref
    # Full row pattern 1: 5:5 or $5:$5
    if re.match(r"^\$?\d+:\$?\d+$", addr):
        return True
    # Full row pattern 2: A5:XFD5 (expanded bounds)
    # Excel max column is XFD (16,384)
    if re.match(r"^\$?A\$?(\d+):\$?XFD\$?\1$", addr):
        return True
    return False


def _get_populated_subrange(
    range_ref: str, workbook: LazyWorkbook | None, max_cells: int = 10000
) -> str | None:
    """
    For a full-column or full-row reference, return the populated sub-range.

    Uses openpyxl's iter_rows() with max_row/max_column bounds to avoid
    iterating through all 1,048,576 rows or 16,384 columns.

    This is O(bounding_box) instead of O(1,048,576) for naive range(1, 1048577) iteration.

    Args:
        range_ref: Full-column (e.g., C:C) or full-row (e.g., 5:5) reference
        workbook: LazyWorkbook to query for populated cells
        max_cells: Maximum cells to scan (safety limit)

    Returns:
        Populated sub-range (e.g., "C41:C104") or None if no populated cells found
    """
    if not workbook:
        return None

    # Extract sheet name
    if "!" in range_ref:
        sheet_part, addr = range_ref.split("!", 1)
        sheet_name = sheet_part.strip("'").replace("''", "'")
    else:
        # No sheet name - cannot determine populated cells
        return None

    try:
        ws = workbook[sheet_name]
    except KeyError:
        return None

    # Parse the range to get column/row bounds
    addr_clean = addr.replace("$", "")

    if _is_full_column_reference(range_ref):
        # Full column: e.g., C:C or C1:C1048576
        parts = addr_clean.split(":")
        if len(parts) != 2:
            return None
        # Extract column letter (handle both C:C and C1:C1048576)
        col_letter = "".join(c for c in parts[0] if c.isalpha())

        # Find min/max populated rows in this column
        # Use openpyxl's iter_rows() with max_row bounds (not range(1, 1048577)!)
        min_row = None
        max_row = None
        row_count = 0

        # Get column index from letter
        from openpyxl.utils import column_index_from_string

        col_idx = column_index_from_string(col_letter)

        # Iterate only over bounding box (ws.max_row), not all 1M rows
        # This is O(max_row) instead of O(1,048,576)
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row or 1048576
        ):
            cell = row[0]
            if cell.value is not None:
                row_idx = cell.row
                if min_row is None:
                    min_row = row_idx
                max_row = row_idx
                row_count += 1
                if row_count >= max_cells:
                    break

        if min_row is not None and max_row is not None:
            # Return populated sub-range
            return f"{sheet_part}!{col_letter}{min_row}:{col_letter}{max_row}"

    elif _is_full_row_reference(range_ref):
        # Full row: e.g., 5:5 or A5:XFD5
        parts = addr_clean.split(":")
        if len(parts) != 2:
            return None
        # Extract row number (handle both 5:5 and A5:XFD5)
        row_num = "".join(c for c in parts[0] if c.isdigit())
        row_int = int(row_num)

        # Find min/max populated columns in this row
        # Use openpyxl's iter_rows() with max_column bounds (not range(1, 16385)!)
        min_col = None
        max_col = None
        col_count = 0

        # Iterate only over bounding box (ws.max_column), not all 16k columns
        # This is O(max_column) instead of O(16,384)
        for row in ws.iter_rows(
            min_row=row_int, max_row=row_int, min_col=1, max_col=ws.max_column or 16384
        ):
            for cell in row:
                if cell.value is not None:
                    col_letter = _col_to_letter(cell.column)
                    if min_col is None:
                        min_col = col_letter
                    max_col = col_letter
                    col_count += 1
                    if col_count >= max_cells:
                        break
            if col_count >= max_cells:
                break

        if min_col is not None and max_col is not None:
            # Return populated sub-range
            return f"{sheet_part}!{min_col}{row_num}:{max_col}{row_num}"

    return None


def _quote_sheet_name_for_address(sheet_name: str) -> str:
    """Quote sheet name using the same heuristic as fast traversal output."""
    if " " in sheet_name or "'" in sheet_name or "!" in sheet_name:
        escaped_sheet = sheet_name.replace("'", "''")
        return f"'{escaped_sheet}'"
    return sheet_name


def _unquote_sheet_name(sheet_name: str) -> str:
    """Remove Excel-style single-quote wrapping and unescape doubled quotes."""
    if sheet_name.startswith("'") and sheet_name.endswith("'") and len(sheet_name) >= 2:
        return sheet_name[1:-1].replace("''", "'")
    return sheet_name


def _normalize_ref_with_sheet(ref: str, sheet_name: str) -> str:
    """Ensure reference is sheet-qualified, preserving existing sheet formatting."""
    if "!" in ref:
        return ref
    if not sheet_name:
        return ref
    return f"{sheet_name}!{ref}"


@dataclass(frozen=True)
class VisitedCellIndex:
    """Row/column index for fast visited-cell range queries."""

    rows_sorted: list[int]
    row_to_cols: dict[int, list[int]]


def _build_visited_cell_index(cell_addresses: Iterable[str]) -> dict[str, VisitedCellIndex]:
    """Build a fast lookup index for visited cells, grouped by sheet."""
    rows_by_sheet: dict[str, dict[int, list[int]]] = {}
    for addr in cell_addresses:
        parsed = parse_cell_address(addr)
        sheet = parsed.get("sheet", "")
        row = int(parsed.get("row", 0))
        col = int(parsed.get("col", 0))
        if not sheet or row <= 0 or col <= 0:
            continue
        row_map = rows_by_sheet.setdefault(sheet, {})
        row_map.setdefault(row, []).append(col)

    visited_index: dict[str, VisitedCellIndex] = {}
    for sheet, row_map in rows_by_sheet.items():
        row_to_cols = {row: sorted(cols) for row, cols in row_map.items()}
        rows_sorted = sorted(row_to_cols.keys())
        visited_index[sheet] = VisitedCellIndex(rows_sorted=rows_sorted, row_to_cols=row_to_cols)

    return visited_index


def _build_cell_address_map(conn: sqlite3.Connection) -> dict[int, str]:
    """
    Build mapping from cell_id (integer) to A1 address (sheet-qualified).
    """
    cursor = conn.execute("""
        SELECT c.cell_id, s.sheet_name, c.a1
        FROM cells c
        JOIN sheets s ON c.sheet_id = s.sheet_id
    """)

    cell_id_to_addr = {}
    for cell_id, sheet_name, a1 in cursor:
        quoted_sheet = _quote_sheet_name_for_address(sheet_name)
        cell_id_to_addr[cell_id] = f"{quoted_sheet}!{a1}"

    return cell_id_to_addr


def _build_cell_formulas(
    conn: sqlite3.Connection,
    cell_id_to_addr: dict[int, str],
    workbook_path: str | None = None,
    use_r1c1: bool = False,
) -> dict[str, str]:
    """
    Build mapping from A1 address to formula text.

    Uses per-cell formula_a1 from the DB when available (WI-12), falling back
    to SheetFormulaCache for legacy databases without formula_a1.
    """
    # Check if cells table has formula_a1 column (WI-12)
    has_formula_a1 = False
    try:
        conn.execute("SELECT formula_a1 FROM cells LIMIT 0")
        has_formula_a1 = True
    except Exception:
        pass

    if has_formula_a1 and not use_r1c1:
        # Fast path: per-cell A1 formulas from the DB (no workbook re-read)
        cursor = conn.execute("""
            SELECT c.cell_id, c.formula_a1
            FROM cells c
        """)
        cell_formulas: dict[str, str] = {}
        for cell_id, formula_a1 in cursor:
            addr = cell_id_to_addr.get(cell_id)
            if addr:
                if formula_a1:
                    cell_formulas[addr] = (
                        f"={formula_a1}" if not formula_a1.startswith("=") else formula_a1
                    )
                else:
                    cell_formulas[addr] = ""
        return cell_formulas

    if workbook_path and not use_r1c1:
        # Legacy fallback: read workbook via SheetFormulaCache
        cell_formulas = dict.fromkeys(cell_id_to_addr.values(), "")
        formula_cells = []
        cursor = conn.execute("""
            SELECT c.cell_id
            FROM cells c
            WHERE c.formula_id IS NOT NULL
        """)
        for (cell_id,) in cursor:
            addr = cell_id_to_addr.get(cell_id)
            if addr:
                formula_cells.append(addr)
        with SheetFormulaCache(workbook_path) as formula_cache:
            for addr in formula_cells:
                formula = formula_cache.get_formula(addr)
                if formula:
                    cell_formulas[addr] = f"={formula}" if not formula.startswith("=") else formula
        return cell_formulas

    if use_r1c1:
        cursor = conn.execute("""
            SELECT c.cell_id, f.formula_r1c1
            FROM cells c
            LEFT JOIN formulas f ON c.formula_id = f.formula_id
        """)
    else:
        cursor = conn.execute("""
            SELECT c.cell_id, f.formula_a1_example
            FROM cells c
            LEFT JOIN formulas f ON c.formula_id = f.formula_id
        """)

    cell_formulas: dict[str, str] = {}
    for cell_id, formula_a1 in cursor:
        addr = cell_id_to_addr.get(cell_id)
        if addr:
            if formula_a1:
                cell_formulas[addr] = (
                    f"={formula_a1}" if not formula_a1.startswith("=") else formula_a1
                )
            else:
                cell_formulas[addr] = ""

    return cell_formulas


def _build_forward_index(
    conn: sqlite3.Connection, cell_id_to_addr: dict[int, str]
) -> dict[str, list[str]]:
    """
    Build forward dependency index (cell -> precedents) from fast schema edges.
    """
    forward_index: dict[str, list[str]] = {}

    cursor = conn.execute("SELECT sheet_id, sheet_name FROM sheets")
    sheet_id_to_name = {sheet_id: sheet_name for sheet_id, sheet_name in cursor}

    def _addr_from_cell_id(cell_id: int) -> str | None:
        addr = cell_id_to_addr.get(cell_id)
        if addr:
            return addr
        try:
            sheet_id, row, col = unpack_cell_id(cell_id)
        except Exception:
            return None
        sheet_name = sheet_id_to_name.get(sheet_id)
        if not sheet_name:
            return None
        quoted_sheet = _quote_sheet_name_for_address(sheet_name)
        return f"{quoted_sheet}!{_col_to_letter(col)}{row}"

    cursor = conn.execute("""
        SELECT from_cell_id, to_cell_id
        FROM cell_edges_internal
    """)

    for from_id, to_id in cursor:
        from_addr = _addr_from_cell_id(from_id)
        to_addr = _addr_from_cell_id(to_id)

        if from_addr and to_addr:
            if from_addr not in forward_index:
                forward_index[from_addr] = []
            forward_index[from_addr].append(to_addr)

    cursor = conn.execute("""
        SELECT re.from_cell_id, s.sheet_name, re.to_range_a1
        FROM range_edges re
        JOIN sheets s ON re.to_sheet_id = s.sheet_id
    """)

    for from_id, to_sheet, to_range_a1 in cursor:
        from_addr = cell_id_to_addr.get(from_id)

        if from_addr:
            if " " in to_sheet or "'" in to_sheet or "!" in to_sheet:
                escaped_sheet = to_sheet.replace("'", "''")
                quoted_sheet = f"'{escaped_sheet}'"
            else:
                quoted_sheet = to_sheet

            range_addr = f"{quoted_sheet}!{to_range_a1}"

            if from_addr not in forward_index:
                forward_index[from_addr] = []
            forward_index[from_addr].append(range_addr)

    return forward_index


def _build_reverse_index(forward_index: dict[str, list[str]]) -> ReverseIndex:
    """Build reverse dependency index from forward index."""
    reverse_index = ReverseIndex()

    for from_cell, precedents in forward_index.items():
        for to_cell in precedents:
            reverse_index.add_dependency(dependent=from_cell, precedent=to_cell)

    return reverse_index


def _is_constant_range(
    range_ref: str, cell_formulas: dict[str, str], formula_cache: SheetFormulaCache | None = None
) -> bool:
    """Check if a range contains only constant cells (no formulas)."""
    try:
        expanded = expand_range_to_cells(range_ref, max_cells=10000, populated_cells=None)
    except (ValueError, KeyError, AttributeError):
        return False

    if not expanded:
        return False

    for cell_addr in expanded:
        normalized = cell_addr.replace("$", "")
        if normalized in cell_formulas and cell_formulas.get(normalized):
            return False
        if formula_cache:
            formula = formula_cache.get_formula(normalized)
            if formula:
                return False

    return True


def _collect_constant_ranges(
    forward_index: dict[str, list[str]],
    cell_formulas: dict[str, str],
    name_table_map: NameTableMap | None = None,
    formula_cache: SheetFormulaCache | None = None,
    include_forward_index: bool = False,
    workbook: LazyWorkbook | None = None,
    visited_cells_by_sheet: dict[str, list[tuple[int, int]] | VisitedCellIndex] | None = None,
) -> list[dict[str, Any]]:
    """Collect constant range references from precedents."""
    refs: list[str] = []
    seen: set[str] = set()
    parsed_cache: dict[str, list[str]] = {}
    const_cache: dict[str, bool] = {}

    for cell_addr, formula in cell_formulas.items():
        if not formula:
            continue
        candidates = []
        if include_forward_index:
            candidates.extend(forward_index.get(cell_addr, []))
        cached_refs = parsed_cache.get(formula)
        if cached_refs is None:
            try:
                ast = parse_formula(formula)
                cached_refs = extract_references_from_ast(ast)
            except (ValueError, KeyError, AttributeError):
                cached_refs = []
            parsed_cache[formula] = cached_refs
        if cached_refs:
            candidates.extend(cached_refs)

        if not candidates:
            continue

        sheet_name = cell_addr.split("!")[0] if "!" in cell_addr else ""
        for ref in candidates:
            ref_part = ref.split("!")[-1] if "!" in ref else ref
            if is_defined_name(ref_part) and name_table_map:
                resolved = name_table_map.resolve_name(ref_part)
                if resolved:
                    for resolved_ref in resolved:
                        normalized_ref = _normalize_ref_with_sheet(resolved_ref, sheet_name)

                        # Detect full-column/row references and replace with populated sub-range
                        if (
                            _is_full_column_reference(normalized_ref)
                            or _is_full_row_reference(normalized_ref)
                        ) and workbook:
                            populated_subrange = _get_populated_subrange(normalized_ref, workbook)
                            if populated_subrange:
                                normalized_ref = populated_subrange
                            else:
                                # No populated cells found - skip this reference
                                continue

                        if visited_cells_by_sheet is not None:
                            visited_subrange = _get_visited_subrange(
                                normalized_ref, visited_cells_by_sheet
                            )
                            if visited_subrange is None:
                                continue
                            normalized_ref = visited_subrange

                        is_constant = const_cache.get(normalized_ref)
                        if is_constant is None:
                            is_constant = _is_constant_range(
                                normalized_ref, cell_formulas, formula_cache=formula_cache
                            )
                            const_cache[normalized_ref] = is_constant
                        if is_constant:
                            if normalized_ref not in seen:
                                seen.add(normalized_ref)
                                refs.append(normalized_ref)
                continue
            normalized_ref = _normalize_ref_with_sheet(ref, sheet_name)

            # Detect full-column/row references and replace with populated sub-range
            if (
                _is_full_column_reference(normalized_ref) or _is_full_row_reference(normalized_ref)
            ) and workbook:
                populated_subrange = _get_populated_subrange(normalized_ref, workbook)
                if populated_subrange:
                    normalized_ref = populated_subrange
                else:
                    # No populated cells found - skip this reference
                    continue

            if visited_cells_by_sheet is not None:
                visited_subrange = _get_visited_subrange(normalized_ref, visited_cells_by_sheet)
                if visited_subrange is None:
                    continue
                normalized_ref = visited_subrange

            is_constant = const_cache.get(normalized_ref)
            if is_constant is None:
                is_constant = _is_constant_range(
                    normalized_ref, cell_formulas, formula_cache=formula_cache
                )
                const_cache[normalized_ref] = is_constant
            if is_constant:
                if normalized_ref not in seen:
                    seen.add(normalized_ref)
                    refs.append(normalized_ref)

    constant_refs = [{"address": ref, "dtype": "unknown"} for ref in refs]

    if workbook and constant_refs:
        sheet_cache: dict[str, Any] = {}
        for ref_dict in constant_refs:
            try:
                cells_in_range = expand_range_to_cells(
                    ref_dict["address"], max_cells=10000, populated_cells=None
                )
            except (ValueError, KeyError, AttributeError):
                continue

            cell_types = []
            for cell_addr in cells_in_range:
                if "!" not in cell_addr:
                    continue
                sheet_part, cell_coord = cell_addr.split("!", 1)
                sheet_name = _unquote_sheet_name(sheet_part)
                if sheet_name not in sheet_cache:
                    try:
                        sheet_cache[sheet_name] = workbook[sheet_name]
                    except KeyError:
                        continue
                ws = sheet_cache.get(sheet_name)
                if not ws:
                    continue
                try:
                    cell = ws[cell_coord]
                except Exception:
                    continue
                cell_types.append({"address": cell_addr, "dtype": classify_dtype(cell)})

            if cell_types:
                ref_dict["cells"] = cell_types
                dtypes = {
                    c["dtype"]
                    for c in cell_types
                    if c["dtype"] not in ("blank", "empty", "unknown", None)
                }
                if len(dtypes) == 1:
                    ref_dict["dtype"] = next(iter(dtypes))
                else:
                    ref_dict["dtype"] = "mixed"

    return constant_refs


def _collect_constant_ranges_from_edges(
    conn: sqlite3.Connection,
    cell_formulas: dict[str, str],
    visited_cells_by_sheet: dict[str, list[tuple[int, int]] | VisitedCellIndex],
) -> list[dict[str, Any]]:
    """Collect constant ranges from range_edges, scoped to visited cells."""
    constant_refs: list[dict[str, Any]] = []
    seen: set[str] = set()

    cursor = conn.execute("""
        SELECT DISTINCT s.sheet_name, re.to_range_a1
        FROM range_edges re
        JOIN sheets s ON re.to_sheet_id = s.sheet_id
    """)

    for sheet_name, to_range_a1 in cursor:
        quoted_sheet = _quote_sheet_name_for_address(sheet_name)
        address = f"{quoted_sheet}!{to_range_a1}"
        visited_subrange = _get_visited_subrange(address, visited_cells_by_sheet)
        if visited_subrange is None:
            continue
        if visited_subrange in seen:
            continue
        if not _is_constant_range(visited_subrange, cell_formulas, formula_cache=None):
            continue
        seen.add(visited_subrange)
        constant_refs.append({"address": visited_subrange, "dtype": "unknown"})

    return constant_refs


def _scan_visited_cells_in_range(
    row: int, col: int, r2: int, c2: int, candidates: list[tuple[int, int]] | VisitedCellIndex
) -> tuple[int, int, int, int, int] | None:
    """Find visited cells within a range and return bounds + count."""
    if isinstance(candidates, VisitedCellIndex):
        rows = candidates.rows_sorted
        start_idx = bisect_left(rows, row)
        end_idx = bisect_right(rows, r2)
        if start_idx == end_idx:
            return None
        min_r = max_r = None
        min_c = max_c = None
        total_count = 0
        for row_val in rows[start_idx:end_idx]:
            cols = candidates.row_to_cols.get(row_val)
            if not cols:
                continue
            left = bisect_left(cols, col)
            right = bisect_right(cols, c2)
            if left == right:
                continue
            total_count += right - left
            min_c = cols[left] if min_c is None else min(min_c, cols[left])
            max_c = cols[right - 1] if max_c is None else max(max_c, cols[right - 1])
            if min_r is None:
                min_r = row_val
            max_r = row_val
        if total_count == 0 or min_r is None or min_c is None or max_r is None or max_c is None:
            return None
        return min_r, max_r, min_c, max_c, total_count

    in_range = [(r, c) for r, c in candidates if row <= r <= r2 and col <= c <= c2]
    if not in_range:
        return None

    min_r = min(r for r, _ in in_range)
    max_r = max(r for r, _ in in_range)
    min_c = min(c for _, c in in_range)
    max_c = max(c for _, c in in_range)
    return min_r, max_r, min_c, max_c, len(in_range)


def _get_visited_subrange(
    address: str, visited_cells_by_sheet: dict[str, list[tuple[int, int]] | VisitedCellIndex]
) -> str | None:
    """Shrink a range to the visited subrange if it forms a full rectangle."""
    parsed = parse_cell_address(address)
    sheet = parsed.get("sheet", "")
    row = int(parsed.get("row", 0))
    col = int(parsed.get("col", 0))
    height = int(parsed.get("height", 1))
    width = int(parsed.get("width", 1))
    if not sheet or row <= 0 or col <= 0:
        return None

    r2 = row + height - 1
    c2 = col + width - 1
    candidates = visited_cells_by_sheet.get(sheet, [])
    if not candidates:
        return None

    scan = _scan_visited_cells_in_range(row, col, r2, c2, candidates)
    if scan is None:
        return None
    min_r, max_r, min_c, max_c, total_count = scan
    expected_count = (max_r - min_r + 1) * (max_c - min_c + 1)
    if total_count != expected_count:
        return None

    start = f"{_col_to_letter(min_c)}{min_r}"
    end = f"{_col_to_letter(max_c)}{max_r}"
    quoted_sheet = _quote_sheet_name_for_address(sheet)
    return f"{quoted_sheet}!{start}:{end}"


def _filter_constant_refs_to_visited(
    constant_refs: list[dict[str, Any]],
    visited_cells_by_sheet: dict[str, list[tuple[int, int]] | VisitedCellIndex],
) -> list[dict[str, Any]]:
    """Keep only constant ranges that intersect visited cells.

    If visited cells within a range form a full contiguous rectangle, shrink the
    range to the visited subrange to avoid upstream-only ranges.
    """
    filtered = []
    for ref_dict in constant_refs:
        address = ref_dict.get("address")
        if not address:
            continue
        parsed = parse_cell_address(address)
        sheet = parsed.get("sheet", "")
        row = int(parsed.get("row", 0))
        col = int(parsed.get("col", 0))
        height = int(parsed.get("height", 1))
        width = int(parsed.get("width", 1))
        if not sheet or row <= 0 or col <= 0:
            continue
        r2 = row + height - 1
        c2 = col + width - 1
        candidates = visited_cells_by_sheet.get(sheet, [])
        if not candidates:
            continue

        scan = _scan_visited_cells_in_range(row, col, r2, c2, candidates)
        if scan is None:
            continue

        min_r, max_r, min_c, max_c, total_count = scan
        expected_count = (max_r - min_r + 1) * (max_c - min_c + 1)
        if total_count == expected_count:
            start = f"{_col_to_letter(min_c)}{min_r}"
            end = f"{_col_to_letter(max_c)}{max_r}"
            quoted_sheet = _quote_sheet_name_for_address(sheet)
            ref_dict = {**ref_dict, "address": f"{quoted_sheet}!{start}:{end}"}

        filtered.append(ref_dict)
    return filtered


def _collect_constant_cells_by_sheet(
    cell_formulas: dict[str, str], grouped_cell_addresses: set[str]
) -> dict[str, dict[tuple[int, int], str]]:
    """Collect ungrouped constant cells organized by sheet and coordinate."""
    constant_cells_by_sheet: dict[str, dict[tuple[int, int], str]] = {}

    for addr, formula in cell_formulas.items():
        if formula:
            continue
        normalized = addr.replace("$", "")
        if normalized in grouped_cell_addresses:
            continue
        parsed = parse_cell_address(normalized)
        sheet = parsed.get("sheet", "")
        row = int(parsed.get("row", 0))
        col = int(parsed.get("col", 0))
        if not sheet or row <= 0 or col <= 0:
            continue
        constant_cells_by_sheet.setdefault(sheet, {})[(row, col)] = normalized

    return constant_cells_by_sheet


def _split_component_into_rectangles(
    component: list[tuple[int, int]],
) -> list[tuple[int, int, int, int]]:
    """Split a non-rectangular component into maximal row-span rectangles."""
    if not component:
        return []

    row_to_cols: dict[int, list[int]] = {}
    for row, col in component:
        row_to_cols.setdefault(row, []).append(col)

    rectangles: list[tuple[int, int, int, int]] = []
    active: dict[tuple[int, int], tuple[int, int]] = {}

    for row in sorted(row_to_cols):
        cols = sorted(row_to_cols[row])
        if not cols:
            continue

        runs: list[tuple[int, int]] = []
        start = prev = cols[0]
        for col in cols[1:]:
            if col == prev + 1:
                prev = col
                continue
            runs.append((start, prev))
            start = prev = col
        runs.append((start, prev))

        current_keys: set[tuple[int, int]] = set()
        for start_col, end_col in runs:
            key = (start_col, end_col)
            if key in active:
                min_row, max_row = active[key]
                if max_row == row - 1:
                    active[key] = (min_row, row)
                else:
                    rectangles.append((min_row, max_row, start_col, end_col))
                    active[key] = (row, row)
            else:
                active[key] = (row, row)
            current_keys.add(key)

        for key in list(active.keys()):
            if key not in current_keys:
                min_row, max_row = active.pop(key)
                rectangles.append((min_row, max_row, key[0], key[1]))

    for key, (min_row, max_row) in active.items():
        rectangles.append((min_row, max_row, key[0], key[1]))

    rectangles.sort()
    return rectangles


def _build_constant_bindings_from_cells(
    constant_cells_by_sheet: dict[str, dict[tuple[int, int], str]],
    workbook_guid: str,
    cell_formulas: dict[str, str],
) -> list[Binding]:
    """Group contiguous constant cells into rectangular bindings."""
    bindings: list[Binding] = []

    for sheet, positions_to_addr in constant_cells_by_sheet.items():
        remaining = set(positions_to_addr.keys())
        quoted_sheet = _quote_sheet_name_for_address(sheet)

        while remaining:
            start = min(remaining)
            stack = [start]
            component: list[tuple[int, int]] = []

            while stack:
                current = stack.pop()
                if current not in remaining:
                    continue
                remaining.remove(current)
                component.append(current)
                row, col = current
                for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    neighbor = (row + dr, col + dc)
                    if neighbor in remaining:
                        stack.append(neighbor)

            min_r = min(r for r, _ in component)
            max_r = max(r for r, _ in component)
            min_c = min(c for _, c in component)
            max_c = max(c for _, c in component)
            expected_count = (max_r - min_r + 1) * (max_c - min_c + 1)

            if len(component) != expected_count:
                rectangles = _split_component_into_rectangles(component)
                if rectangles:
                    for min_r, max_r, min_c, max_c in rectangles:
                        cells = []
                        for r in range(min_r, max_r + 1):
                            for c in range(min_c, max_c + 1):
                                addr = positions_to_addr.get((r, c))
                                if addr:
                                    cells.append(addr)
                        expected_rect = (max_r - min_r + 1) * (max_c - min_c + 1)
                        if len(cells) != expected_rect:
                            for addr in sorted(cells):
                                bindings.append(
                                    _create_constant_binding(
                                        address_a1=addr,
                                        cells=[addr],
                                        workbook_guid=workbook_guid,
                                        cell_formulas=cell_formulas,
                                    )
                                )
                            continue
                        start_addr = f"{_col_to_letter(min_c)}{min_r}"
                        end_addr = f"{_col_to_letter(max_c)}{max_r}"
                        if start_addr == end_addr:
                            address_a1 = f"{quoted_sheet}!{start_addr}"
                        else:
                            address_a1 = f"{quoted_sheet}!{start_addr}:{end_addr}"
                        bindings.append(
                            _create_constant_binding(
                                address_a1=address_a1,
                                cells=cells,
                                workbook_guid=workbook_guid,
                                cell_formulas=cell_formulas,
                            )
                        )
                else:
                    for row, col in sorted(component):
                        addr = positions_to_addr[(row, col)]
                        bindings.append(
                            _create_constant_binding(
                                address_a1=addr,
                                cells=[addr],
                                workbook_guid=workbook_guid,
                                cell_formulas=cell_formulas,
                            )
                        )
                continue

            address_a1 = (
                f"{quoted_sheet}!{_col_to_letter(min_c)}{min_r}:{_col_to_letter(max_c)}{max_r}"
            )
            cells = [positions_to_addr[pos] for pos in component]
            bindings.append(
                _create_constant_binding(
                    address_a1=address_a1,
                    cells=cells,
                    workbook_guid=workbook_guid,
                    cell_formulas=cell_formulas,
                )
            )

    return bindings


def _create_constant_binding(
    address_a1: str, cells: list[str], workbook_guid: str, cell_formulas: dict[str, str]
) -> Binding:
    """Create a Binding object for constant range."""
    if "!" in address_a1:
        sheet_name = address_a1.split("!")[0]
        if sheet_name.startswith("'") and sheet_name.endswith("'"):
            sheet_name = sheet_name[1:-1]
    else:
        sheet_name = ""

    parsed = parse_cell_address(address_a1)
    if not parsed:
        raise ValueError(f"Could not parse address: {address_a1}")

    row = int(parsed.get("row", 1))
    col = int(parsed.get("col", 1))
    height = int(parsed.get("height", 1))
    width = int(parsed.get("width", 1))

    top_left_a1 = f"{_col_to_letter(col)}{row}"

    cells_structure_hash = compute_cells_structure_hash(cells, cell_formulas)
    binding_id = compute_binding_id(
        workbook_guid=workbook_guid,
        sheet=sheet_name,
        top_left_a1=top_left_a1,
        shape_rows=height,
        shape_cols=width,
        cells_structure_hash=cells_structure_hash,
    )

    return Binding(
        binding_id=binding_id,
        debug_label=f"{sheet_name}::{address_a1}",
        sheet=sheet_name,
        address_a1=address_a1,
        top_left_a1=top_left_a1,
        shape_rows=height,
        shape_cols=width,
        binding_type="constant",
        cells_structure_hash=cells_structure_hash,
        cells=sorted(cells),
    )


def _classify_bindings(bindings: list[Binding], cell_formulas: dict[str, str]) -> list[Binding]:
    """Classify bindings as 'formula' or 'constant' based on cell contents."""
    for binding in bindings:
        has_formula = False
        for cell_addr in binding.cells:
            formula = cell_formulas.get(cell_addr, "")
            if formula and formula.strip():
                has_formula = True
                break

        if not has_formula:
            binding.binding_type = "constant"

    return bindings


def _write_bindings_to_db(
    conn: sqlite3.Connection,
    bindings: list[Binding],
    cell_id_to_addr: dict[int, str],
    workbook_sha256: str,
    spatial_candidates_by_binding: dict[str, dict[str, Any]] | None = None,
) -> None:
    """
    Write bindings to fast schema database.
    """
    addr_to_cell_id = {addr: cell_id for cell_id, addr in cell_id_to_addr.items()}

    cursor = conn.execute("SELECT sheet_id, sheet_name FROM sheets")
    sheet_name_to_id = {sheet_name: sheet_id for sheet_id, sheet_name in cursor}

    def _resolve_sheet(sheet_name: str) -> tuple[int | None, str | None]:
        sheet_id = sheet_name_to_id.get(sheet_name)
        if sheet_id is not None:
            return sheet_id, sheet_name
        lowered = sheet_name.lower()
        for existing_name, existing_id in sheet_name_to_id.items():
            if existing_name.lower() == lowered:
                return existing_id, existing_name
        return None, None

    def _value_text(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            text = value.strip()
            return text if text else None
        return str(value)

    def _infer_data_type_from_value(value: Any) -> str:
        if value is None or value == "":
            return "blank"
        if isinstance(value, bool):
            return "b"
        if isinstance(value, (int, float)):
            return "n"
        return "s"

    binding_rows: list[tuple[Any, ...]] = []
    cell_to_binding_rows: list[tuple[int, str]] = []
    label_candidate_rows: list[tuple[Any, ...]] = []
    binding_batch_size = 20_000
    cell_to_binding_batch_size = 100_000
    has_label_candidate_table = (
        conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='binding_label_candidate_cells'"
        ).fetchone()
        is not None
    )

    spatial_candidates_by_binding = spatial_candidates_by_binding or {}
    spatial_json_by_sha: dict[str, str] = {}
    spatial_sha_by_binding: dict[str, str] = {}
    pending_cells_by_address: dict[str, tuple[int, int, int, int, str, str | None, str]] = {}
    pending_value_json_by_sha: dict[str, str] = {}
    label_cells_by_binding: dict[str, list[tuple[Any, ...]]] = {}
    for binding in bindings:
        spatial_obj = spatial_candidates_by_binding.get(binding.binding_id, {})
        canonical_json, sha = canonicalize_and_hash(spatial_obj)
        spatial_sha_by_binding[binding.binding_id] = sha
        spatial_json_by_sha.setdefault(sha, canonical_json)

        candidates = (
            spatial_obj.get("label_candidates", []) if isinstance(spatial_obj, dict) else []
        )
        if not isinstance(candidates, list):
            continue

        label_rows: list[tuple[Any, ...]] = []
        for candidate in candidates:
            if not isinstance(candidate, dict):
                continue
            candidate_type = str(candidate.get("type") or "")
            candidate_address = str(candidate.get("address") or "")
            candidate_cells = candidate.get("cells", [])
            if not isinstance(candidate_cells, list):
                continue
            for candidate_cell in candidate_cells:
                cell_addr_raw: str | None = None
                cell_value: Any = None
                if isinstance(candidate_cell, dict):
                    raw_addr = candidate_cell.get("address")
                    if isinstance(raw_addr, str):
                        cell_addr_raw = raw_addr
                    cell_value = candidate_cell.get("value")
                elif isinstance(candidate_cell, str):
                    cell_addr_raw = candidate_cell

                if not cell_addr_raw:
                    continue

                parse_target = (
                    cell_addr_raw if "!" in cell_addr_raw else f"{binding.sheet}!{cell_addr_raw}"
                )
                parsed = parse_cell_address(parse_target)
                row = int(parsed.get("row", 0) or 0)
                col = int(parsed.get("col", 0) or 0)
                if row <= 0 or col <= 0:
                    continue

                parsed_sheet = str(parsed.get("sheet") or binding.sheet)
                parsed_sheet_id, canonical_sheet = _resolve_sheet(parsed_sheet)
                if parsed_sheet_id is None or canonical_sheet is None:
                    continue

                a1 = f"{_col_to_letter(col)}{row}"
                cell_address = f"{canonical_sheet}!{a1}"
                cell_id = addr_to_cell_id.get(cell_address)
                if cell_id is None:
                    cell_id = pack_cell_id(parsed_sheet_id, row, col)
                    value_sha: str | None = None
                    if cell_value is not None and cell_value != "":
                        value_json, value_sha = canonicalize_and_hash(cell_value)
                        pending_value_json_by_sha.setdefault(value_sha, value_json)
                    existing = pending_cells_by_address.get(cell_address)
                    if existing is None or (existing[5] is None and value_sha is not None):
                        pending_cells_by_address[cell_address] = (
                            cell_id,
                            parsed_sheet_id,
                            row,
                            col,
                            a1,
                            value_sha,
                            _infer_data_type_from_value(cell_value),
                        )
                    addr_to_cell_id[cell_address] = cell_id

                label_rows.append(
                    (
                        binding.binding_id,
                        candidate_type,
                        candidate_address,
                        cell_address,
                        parsed_sheet_id,
                        row,
                        col,
                        _value_text(cell_value),
                    )
                )

        if label_rows:
            label_cells_by_binding[binding.binding_id] = label_rows

    if spatial_json_by_sha:
        conn.executemany(
            "INSERT OR IGNORE INTO json_blobs (sha256, json) VALUES (?, ?)",
            list(spatial_json_by_sha.items()),
        )

    pending_value_blob_id_by_sha: dict[str, int] = {}
    if pending_value_json_by_sha:
        conn.executemany(
            "INSERT OR IGNORE INTO json_blobs (sha256, json) VALUES (?, ?)",
            list(pending_value_json_by_sha.items()),
        )
        placeholders = ",".join("?" for _ in pending_value_json_by_sha)
        pending_blob_rows = conn.execute(
            f"SELECT sha256, blob_id FROM json_blobs WHERE sha256 IN ({placeholders})",
            list(pending_value_json_by_sha.keys()),
        ).fetchall()
        pending_value_blob_id_by_sha = {row_sha: blob_id for row_sha, blob_id in pending_blob_rows}

    if pending_cells_by_address:
        pending_rows = sorted(
            pending_cells_by_address.values(),
            key=lambda row: (row[1], row[2], row[0]),
        )
        conn.executemany(
            """
            INSERT OR IGNORE INTO cells (
                cell_id, sheet_id, row, col, a1, formula_id, formula_a1,
                format_blob_id, value_blob_id,
                data_type, is_array_formula, is_spilled, spilled_from_cell_id
            ) VALUES (?, ?, ?, ?, ?, NULL, NULL, NULL, ?, ?, 0, 0, NULL)
            """,
            [
                (
                    cell_id,
                    sheet_id,
                    row,
                    col,
                    a1,
                    pending_value_blob_id_by_sha.get(value_sha) if value_sha else None,
                    data_type,
                )
                for (cell_id, sheet_id, row, col, a1, value_sha, data_type) in pending_rows
            ],
        )

    evidence_sha = hash_json({})
    conn.execute(
        "INSERT OR IGNORE INTO json_blobs (sha256, json) VALUES (?, ?)", (evidence_sha, "{}")
    )

    spatial_blob_ids: dict[str, int] = {}
    if spatial_json_by_sha:
        placeholders = ",".join("?" for _ in spatial_json_by_sha)
        rows = conn.execute(
            f"SELECT sha256, blob_id FROM json_blobs WHERE sha256 IN ({placeholders})",
            list(spatial_json_by_sha.keys()),
        ).fetchall()
        spatial_blob_ids = {sha: blob_id for sha, blob_id in rows}

    cursor = conn.execute("SELECT blob_id FROM json_blobs WHERE sha256 = ?", (evidence_sha,))
    evidence_blob_id = cursor.fetchone()[0]

    def _flush_bindings() -> None:
        if not binding_rows:
            return
        try:
            conn.executemany(
                """
                INSERT INTO bindings (
                    binding_id, sheet_id, address_a1, top_left_cell_id,
                    shape_rows, shape_cols, binding_type, formula_id,
                    label, classification, confidence, is_orphan,
                    extraction_source, evidence_blob_id, spatial_candidates_blob_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                binding_rows,
            )
        except sqlite3.IntegrityError as e:
            raise ValueError(f"Failed to insert bindings (duplicate binding_id?): {e}") from e
        binding_rows.clear()

    def _flush_cell_to_binding() -> None:
        if not cell_to_binding_rows:
            return
        try:
            conn.executemany(
                """
                INSERT INTO cell_to_binding (cell_id, binding_id)
                VALUES (?, ?)
            """,
                cell_to_binding_rows,
            )
        except sqlite3.IntegrityError as e:
            raise ValueError(f"Failed to insert cell_to_binding mappings: {e}") from e
        cell_to_binding_rows.clear()

    for binding in bindings:
        mapped_cell_ids = []
        for cell_addr in binding.cells:
            cell_id = addr_to_cell_id.get(cell_addr)
            if cell_id:
                mapped_cell_ids.append(cell_id)

        if not mapped_cell_ids:
            print(f"Warning: Skipping binding with no mapped cells: {binding.debug_label}")
            continue

        sheet_id = sheet_name_to_id.get(binding.sheet)
        if not sheet_id:
            for name, sid in sheet_name_to_id.items():
                if name.lower() == binding.sheet.lower():
                    sheet_id = sid
                    break

        if not sheet_id:
            print(f"Warning: Sheet not found for binding: {binding.sheet}")
            continue

        if " " in binding.sheet or "'" in binding.sheet or "!" in binding.sheet:
            escaped_sheet = binding.sheet.replace("'", "''")
            quoted_sheet = f"'{escaped_sheet}'"
        else:
            quoted_sheet = binding.sheet

        top_left_addr = f"{quoted_sheet}!{binding.top_left_a1}"
        top_left_cell_id = addr_to_cell_id.get(top_left_addr)

        if not top_left_cell_id:
            parsed = parse_cell_address(top_left_addr)
            row = int(parsed.get("row", 0))
            col = int(parsed.get("col", 0))
            if row > 0 and col > 0:
                top_left_cell_id = pack_cell_id(sheet_id, row, col)
            else:
                print(f"Warning: Top-left cell not found for binding: {top_left_addr}")
                continue

        formula_id = None
        if binding.binding_type == "formula" and binding.cells:
            # binding.cells is lex-sorted by A1 string; pick the spatial top-left
            # (min row, then min col) so a lex-leading blank cell doesn't leave
            # formula_id NULL when the range actually carries a formula.
            def _row_col(addr: str) -> tuple[int, int]:
                parsed = parse_cell_address(addr)
                return (int(parsed.get("row", 0) or 0), int(parsed.get("col", 0) or 0))

            first_cell_addr = min(binding.cells, key=_row_col)
            first_cell_id = addr_to_cell_id.get(first_cell_addr)

            if first_cell_id:
                cursor = conn.execute(
                    """
                    SELECT f.formula_id, f.formula_r1c1
                    FROM cells c
                    JOIN formulas f ON c.formula_id = f.formula_id
                    WHERE c.cell_id = ?
                """,
                    (first_cell_id,),
                )
                row = cursor.fetchone()
                if row:
                    formula_id = row[0]

        if "!" in binding.address_a1:
            address_a1_no_sheet = binding.address_a1.split("!", 1)[1]
        else:
            address_a1_no_sheet = binding.address_a1

        binding_rows.append(
            (
                binding.binding_id,
                sheet_id,
                address_a1_no_sheet,
                top_left_cell_id,
                binding.shape_rows,
                binding.shape_cols,
                binding.binding_type,
                formula_id,
                None,
                None,
                None,
                0,
                "grouping",
                evidence_blob_id,
                spatial_blob_ids.get(
                    spatial_sha_by_binding.get(binding.binding_id, ""), evidence_blob_id
                ),
            )
        )

        for cell_id in mapped_cell_ids:
            cell_to_binding_rows.append((cell_id, binding.binding_id))

        binding_label_rows = label_cells_by_binding.get(binding.binding_id, [])
        if binding_label_rows:
            label_candidate_rows.extend(binding_label_rows)

        if len(binding_rows) >= binding_batch_size:
            _flush_bindings()
        if len(cell_to_binding_rows) >= cell_to_binding_batch_size:
            _flush_cell_to_binding()

    _flush_bindings()
    _flush_cell_to_binding()

    conn.execute("""
        DELETE FROM bindings
        WHERE binding_id NOT IN (
            SELECT DISTINCT binding_id FROM cell_to_binding
        )
    """)

    if has_label_candidate_table and label_candidate_rows:
        label_candidate_rows.sort(
            key=lambda row: (row[0], row[4], row[5], row[6], row[1], row[2], row[3])
        )
        conn.executemany(
            """
            INSERT OR IGNORE INTO binding_label_candidate_cells (
                binding_id, candidate_type, candidate_address, cell_address,
                sheet_id, row, col, value_text
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            label_candidate_rows,
        )

    conn.commit()


def _write_binding_edges_to_db(
    conn: sqlite3.Connection,
    binding_edges: list[Any],
) -> None:
    """Write binding-level edges to fast schema database."""
    edge_rows = []

    for edge in binding_edges:
        edge_rows.append((edge.from_binding_id, edge.to_binding_id, 1))

    if edge_rows:
        try:
            conn.executemany(
                """
                INSERT INTO binding_edges (from_binding_id, to_binding_id, edge_count)
                VALUES (?, ?, ?)
            """,
                edge_rows,
            )
        except sqlite3.IntegrityError as e:
            raise ValueError(f"Failed to insert binding edges: {e}") from e

    conn.commit()


def _disjoint_boxes(
    rects: list[tuple[int, int, int, int]],
) -> list[tuple[int, int, int, int]]:
    """Decompose overlapping (r1, c1, r2, c2) rects into disjoint boxes.

    Column-boundary sweep: for each column strip between consecutive distinct
    column boundaries, merge the row intervals of the rects spanning that
    strip. Rolling-window rect-sets (thousands of one-row-shifted windows over
    the same columns) collapse to a handful of boxes, so a union cardinality
    becomes a few disjoint COUNTs instead of a scan of every overlapping rect.
    """
    if len(rects) == 1:
        return list(rects)
    bounds = sorted({c for _, c1, _, c2 in rects for c in (c1, c2 + 1)})
    boxes: list[tuple[int, int, int, int]] = []
    for cs, ce in zip(bounds, bounds[1:], strict=False):
        intervals = sorted((r1, r2) for r1, c1, r2, c2 in rects if c1 <= cs and c2 >= ce - 1)
        if not intervals:
            continue
        merged_r1, merged_r2 = intervals[0]
        for r1, r2 in intervals[1:]:
            if r1 <= merged_r2 + 1:
                merged_r2 = max(merged_r2, r2)
            else:
                boxes.append((merged_r1, cs, merged_r2, ce - 1))
                merged_r1, merged_r2 = r1, r2
        boxes.append((merged_r1, cs, merged_r2, ce - 1))
    return boxes


def _union_membership(
    boxes: list[tuple[int, int, int, int]],
) -> Callable[[int, int], bool]:
    """O(log n) point-in-union test over the disjoint boxes of one rect union.

    The boxes come from ``_disjoint_boxes``: column strips never overlap and
    row intervals within a strip are disjoint, so two bisections decide
    membership.
    """
    col_strips = sorted({(c1, c2) for _, c1, _, c2 in boxes})
    strip_starts = [c1 for c1, _ in col_strips]
    rows_by_strip: dict[tuple[int, int], list[tuple[int, int]]] = {}
    for r1, c1, r2, c2 in boxes:
        rows_by_strip.setdefault((c1, c2), []).append((r1, r2))
    for intervals in rows_by_strip.values():
        intervals.sort()

    def contains(row: int, col: int) -> bool:
        i = bisect_right(strip_starts, col) - 1
        if i < 0:
            return False
        c1, c2 = col_strips[i]
        if col > c2:
            return False
        intervals = rows_by_strip[(c1, c2)]
        j = bisect_right(intervals, (row, 1 << 40)) - 1
        return j >= 0 and intervals[j][0] <= row <= intervals[j][1]

    return contains


def _write_binding_edges_from_cells(conn: sqlite3.Connection) -> int:
    """Collapse cell and range edges into binding edges using SQL joins.

    The range-edges → binding-edges step enumerates populated cells per
    DISTINCT target rect (not per range_edge, which took 16+ hours on a
    large workbook, nor per (binding, rect), which took 367s on a
    lookup-dense model after INDEX/XLOOKUP argument refs landed). See the
    per-sheet loop for the exact-equivalence argument.
    """
    if not conn:
        raise RuntimeError("Database not open")

    import sys as _sys
    import time as _t

    def _log(msg: str, t0: float) -> None:
        print(f"    [edges] {msg} ({_t.perf_counter() - t0:.2f}s)", file=_sys.stderr, flush=True)

    with conn:
        # Pre-compute cell→binding lookup (one row per cell, picks any binding)
        t0 = _t.perf_counter()
        conn.execute("""
            CREATE TEMP TABLE IF NOT EXISTS _cell_binding_lookup (
                cell_id INTEGER PRIMARY KEY,
                binding_id TEXT NOT NULL
            )
        """)
        conn.execute("DELETE FROM _cell_binding_lookup")
        conn.execute("""
            INSERT INTO _cell_binding_lookup (cell_id, binding_id)
            SELECT cell_id, MIN(binding_id)
            FROM cell_to_binding
            GROUP BY cell_id
        """)
        _log("built cell→binding lookup", t0)

        # Cell edges → binding edges via temp lookup (faster than double WITHOUT ROWID join)
        # DISTINCT removed: INSERT OR IGNORE handles duplicates via PK, avoiding expensive sort
        t0 = _t.perf_counter()
        conn.execute("""
            INSERT OR IGNORE INTO binding_edges (from_binding_id, to_binding_id, edge_count)
            SELECT
                lk_from.binding_id,
                lk_to.binding_id,
                1
            FROM cell_edges_internal cei
            JOIN _cell_binding_lookup lk_from ON cei.from_cell_id = lk_from.cell_id
            JOIN _cell_binding_lookup lk_to ON cei.to_cell_id = lk_to.cell_id
            WHERE lk_from.binding_id != lk_to.binding_id
        """)
        _log("collapsed internal cell edges", t0)

        # Range edges → binding edges:
        # 1. Dedup at source: many cells in the same from-binding share the same
        #    range target. Collapse (from_binding, to_rect) up front. For
        #    a large workbook this compresses 4.6M rows to ~1.2M.
        # 2. For each range, enumerate the *populated* cells in its target rect
        #    via the cells table and its (sheet_id, row, col) index, then look
        #    up each cell's bindings via cell_to_binding. This scales with the
        #    number of populated cells in the range, not the number of bindings
        #    on the target sheet — a dramatic win for whole-column refs on
        #    sheets with many bindings (where the R*Tree approach matched
        #    every binding on the sheet).
        t0 = _t.perf_counter()
        conn.execute("DROP TABLE IF EXISTS _binding_range_edges")
        conn.execute("""
            CREATE TEMP TABLE _binding_range_edges AS
            SELECT
                lk.binding_id AS from_binding_id,
                re.to_sheet_id,
                re.to_r1, re.to_c1, re.to_r2, re.to_c2,
                MAX(CASE WHEN re.provenance = 'resolved_from_cache' THEN 1 ELSE 0 END) AS is_dynamic
            FROM range_edges re
            JOIN _cell_binding_lookup lk ON re.from_cell_id = lk.cell_id
            GROUP BY lk.binding_id, re.to_sheet_id, re.to_r1, re.to_c1, re.to_r2, re.to_c2
        """)
        # Index to drive the per-sheet filter in the join below.
        conn.execute("CREATE INDEX _bre_sheet ON _binding_range_edges(to_sheet_id)")
        _bre_count = conn.execute("SELECT COUNT(*) FROM _binding_range_edges").fetchone()[0]
        _log(f"deduped range edges per source binding ({_bre_count:,} rows)", t0)

        target_sheets = [
            row[0]
            for row in conn.execute(
                "SELECT DISTINCT to_sheet_id FROM _binding_range_edges ORDER BY to_sheet_id"
            )
        ]
        _log(
            f"per-sheet cell-enumeration join starting ({len(target_sheets)} sheets)",
            _t.perf_counter(),
        )

        # Stage range-derived binding pairs (the expensive populated-cell join runs
        # once, here) so we can apply kind/breadth and the anchor de-suppression
        # afterwards without repeating the join. breadth = populated target cells
        # read; is_dynamic = 1 if any contributing range edge is an Issue #1
        # cache-resolved (snapshot-specific) edge.
        conn.execute("DROP TABLE IF EXISTS _range_pairs")
        conn.execute("""
            CREATE TEMP TABLE _range_pairs (
                from_binding_id TEXT NOT NULL,
                to_binding_id TEXT NOT NULL,
                breadth INTEGER NOT NULL,
                is_dynamic INTEGER NOT NULL
            )
        """)
        total_sheet_time = 0.0
        from collections import defaultdict as _defaultdict

        for sheet_id in target_sheets:
            sheet_t0 = _t.perf_counter()
            # Enumerate each DISTINCT rect's populated cells ONCE, not once per
            # referencing binding. Transparent-lookup argument refs (INDEX/XLOOKUP
            # table args) give hundreds of bindings identical or row-shifted rects
            # over the same tables; the naive per-binding join re-enumerated every
            # rect once per referencing binding. Three exact-equivalent reductions:
            #   1. _rect_breadths: per-(rect, to_binding) populated-cell counts,
            #      aggregated in rect chunks — counts are stored, cell rows never
            #      are. Issue #7: the previous _rect_cells temp table stored one
            #      row per (rect × populated cell); rolling-window rects (each
            #      row/column referencing a one-step-shifted window) are DISTINCT
            #      but almost fully overlapping, so one sheet of a 2.3M-formula
            #      forecast model (361,743 distinct rects) spilled >15 GB of
            #      SQLite temp and died disk-full. Chunking by estimated cell
            #      count caps the GROUP BY's sort spill regardless of rect shape.
            #   2. a pair fed by ONE rect takes the precomputed rect breadth;
            #   3. a pair fed by SEVERAL rects needs the cell count of the rect
            #      UNION (rects overlap, so per-rect breadths cannot be summed):
            #      the rect-set is decomposed into disjoint boxes and the target
            #      binding's own cells — few, vs unions spanning whole sheets —
            #      are bisected against them; pairs sharing a (to_binding,
            #      rect-set) get one union count per group.
            rects = conn.execute(
                """
                SELECT DISTINCT to_r1, to_c1, to_r2, to_c2
                FROM _binding_range_edges WHERE to_sheet_id = ?
                ORDER BY to_r1, to_c1, to_r2, to_c2
                """,
                (sheet_id,),
            ).fetchall()
            sheet_pop = conn.execute(
                "SELECT COUNT(*) FROM cells WHERE sheet_id = ?", (sheet_id,)
            ).fetchone()[0]
            conn.execute("DROP TABLE IF EXISTS _rect_breadths")
            conn.execute(
                """
                CREATE TEMP TABLE _rect_breadths (
                    to_r1 INTEGER, to_c1 INTEGER, to_r2 INTEGER, to_c2 INTEGER,
                    to_binding_id TEXT NOT NULL,
                    breadth INTEGER NOT NULL
                )
                """
            )
            chunk_budget = 5_000_000
            chunk: list[tuple[int, int, int, int]] = []
            chunk_est = 0
            chunks: list[list[tuple[int, int, int, int]]] = []
            for rect in rects:
                r1, c1, r2, c2 = rect
                est = min((r2 - r1 + 1) * (c2 - c1 + 1), sheet_pop)
                if chunk and chunk_est + est > chunk_budget:
                    chunks.append(chunk)
                    chunk, chunk_est = [], 0
                chunk.append(rect)
                chunk_est += est
            if chunk:
                chunks.append(chunk)
            for chunk in chunks:
                conn.execute("DROP TABLE IF EXISTS _chunk_rects")
                conn.execute(
                    "CREATE TEMP TABLE _chunk_rects "
                    "(to_r1 INTEGER, to_c1 INTEGER, to_r2 INTEGER, to_c2 INTEGER)"
                )
                conn.executemany("INSERT INTO _chunk_rects VALUES (?, ?, ?, ?)", chunk)
                conn.execute(
                    """
                    INSERT INTO _rect_breadths
                    SELECT r.to_r1, r.to_c1, r.to_r2, r.to_c2,
                           ctb.binding_id, COUNT(DISTINCT c.cell_id)
                    FROM _chunk_rects r
                    JOIN cells c
                      ON c.sheet_id = ?
                     AND c.row BETWEEN r.to_r1 AND r.to_r2
                     AND c.col BETWEEN r.to_c1 AND r.to_c2
                    JOIN cell_to_binding ctb ON ctb.cell_id = c.cell_id
                    GROUP BY r.to_r1, r.to_c1, r.to_r2, r.to_c2, ctb.binding_id
                    """,
                    (sheet_id,),
                )
            conn.execute("DROP TABLE IF EXISTS _chunk_rects")
            conn.execute(
                "CREATE INDEX _rect_breadths_i ON _rect_breadths"
                "(to_r1, to_c1, to_r2, to_c2, to_binding_id)"
            )
            # (from_binding, to_binding, rect) rows carrying the rect's breadth.
            conn.execute("DROP TABLE IF EXISTS _pair_rects")
            conn.execute(
                """
                CREATE TEMP TABLE _pair_rects AS
                SELECT bre.from_binding_id, rp.to_binding_id,
                       rp.to_r1, rp.to_c1, rp.to_r2, rp.to_c2,
                       rp.breadth, bre.is_dynamic
                FROM _binding_range_edges bre
                JOIN _rect_breadths rp
                  ON rp.to_r1 = bre.to_r1 AND rp.to_c1 = bre.to_c1
                 AND rp.to_r2 = bre.to_r2 AND rp.to_c2 = bre.to_c2
                WHERE bre.to_sheet_id = ?
                  AND bre.from_binding_id != rp.to_binding_id
                """,
                (sheet_id,),
            )
            # Single-rect pairs: the rect-level breadth IS the pair breadth.
            cur = conn.execute(
                """
                INSERT INTO _range_pairs (from_binding_id, to_binding_id, breadth, is_dynamic)
                SELECT from_binding_id, to_binding_id, MAX(breadth), MAX(is_dynamic)
                FROM _pair_rects
                GROUP BY from_binding_id, to_binding_id
                HAVING COUNT(*) = 1
                """
            )
            inserted = cur.rowcount if cur.rowcount >= 0 else 0
            # Multi-rect pairs: one distinct-cell union per (to_binding, rect-set).
            multi = conn.execute(
                """
                SELECT from_binding_id, to_binding_id, to_r1, to_c1, to_r2, to_c2, is_dynamic
                FROM _pair_rects
                WHERE (from_binding_id, to_binding_id) IN (
                    SELECT from_binding_id, to_binding_id FROM _pair_rects
                    GROUP BY from_binding_id, to_binding_id HAVING COUNT(*) > 1
                )
                ORDER BY from_binding_id, to_binding_id, to_r1, to_c1, to_r2, to_c2
                """
            ).fetchall()
            pair_rects: dict[tuple[str, str], list[tuple[int, int, int, int]]] = _defaultdict(list)
            pair_dyn: dict[tuple[str, str], int] = _defaultdict(int)
            for fb, tb, r1, c1, r2, c2, dyn in multi:
                pair_rects[(fb, tb)].append((r1, c1, r2, c2))
                pair_dyn[(fb, tb)] = max(pair_dyn[(fb, tb)], dyn)
            union_breadth: dict[tuple[str, tuple[tuple[int, int, int, int], ...]], int] = {}
            contains_cache: dict[
                tuple[tuple[int, int, int, int], ...], Callable[[int, int], bool]
            ] = {}
            tb_cells_cache: dict[str, list[tuple[int, int]]] = {}
            for (fb, tb), rect_set in pair_rects.items():
                key = (tb, tuple(rect_set))
                if key not in union_breadth:
                    # Populated cells of to_binding across the rect UNION — the
                    # rects overlap, so this cannot be summed from per-rect
                    # breadths. A rolling-window pair can carry tens of
                    # thousands of overlapping rects, and no cell rows were
                    # materialized, so: decompose the rect-set into disjoint
                    # boxes, fetch the target binding's own cells once
                    # (bindings are small; unions can span whole sheets), and
                    # bisect each cell against the boxes.
                    rects_key = tuple(rect_set)
                    if rects_key not in contains_cache:
                        contains_cache[rects_key] = _union_membership(_disjoint_boxes(rect_set))
                    contains = contains_cache[rects_key]
                    if tb not in tb_cells_cache:
                        # CROSS JOIN pins the join order: drive from the
                        # binding's few cells, not the sheet's many — the
                        # planner otherwise scans the whole sheet per target
                        # binding.
                        tb_cells_cache[tb] = conn.execute(
                            """
                            SELECT c.row, c.col
                            FROM cell_to_binding ctb
                            CROSS JOIN cells c ON c.cell_id = ctb.cell_id
                            WHERE ctb.binding_id = ? AND c.sheet_id = ?
                            """,
                            (tb, sheet_id),
                        ).fetchall()
                    union_breadth[key] = sum(
                        1 for row, col in tb_cells_cache[tb] if contains(row, col)
                    )
                conn.execute(
                    "INSERT INTO _range_pairs VALUES (?, ?, ?, ?)",
                    (fb, tb, union_breadth[key], pair_dyn[(fb, tb)]),
                )
                inserted += 1
            conn.execute("DROP TABLE IF EXISTS _pair_rects")
            conn.execute("DROP TABLE IF EXISTS _rect_breadths")
            dur = _t.perf_counter() - sheet_t0
            total_sheet_time += dur
            sheet_name_row = conn.execute(
                "SELECT sheet_name FROM sheets WHERE sheet_id = ?", (sheet_id,)
            ).fetchone()
            sheet_name = sheet_name_row[0] if sheet_name_row else f"sheet_id={sheet_id}"
            print(
                f"    [edges]   sheet {sheet_name}: +{inserted:,} range pairs in {dur:.2f}s",
                file=_sys.stderr,
                flush=True,
            )

        _log(
            f"per-sheet cell-enumeration join done (sum={total_sheet_time:.2f}s)",
            _t.perf_counter() - total_sheet_time,
        )

        # Range-only pairs become binding edges. Cell-backed pairs already exist as
        # 'formula' (the cell arm ran first); INSERT OR IGNORE leaves those alone.
        conn.execute("""
            INSERT OR IGNORE INTO binding_edges (from_binding_id, to_binding_id, edge_count, kind)
            SELECT from_binding_id, to_binding_id, breadth,
                   CASE WHEN is_dynamic = 1 THEN 'range_dynamic' ELSE 'range_static' END
            FROM _range_pairs
        """)
        # Anchor de-suppression (Issue #2): an Issue #1 cache-resolved range often
        # also has a single-cell INDIRECT anchor that lands in cell_edges_internal,
        # so the cell arm wrote a 'formula' edge for the same pair and won the PK,
        # hiding the snapshot-specific nature. Upgrade those to 'cell_dynamic' so
        # consumers can still caveat the lineage. Static cell refs (is_dynamic=0)
        # are never upgraded.
        conn.execute("""
            UPDATE binding_edges
            SET kind = 'cell_dynamic'
            WHERE kind = 'formula'
              AND (from_binding_id, to_binding_id) IN (
                  SELECT from_binding_id, to_binding_id FROM _range_pairs WHERE is_dynamic = 1
              )
        """)

        conn.execute("DROP TABLE IF EXISTS _cell_binding_lookup")
        conn.execute("DROP TABLE IF EXISTS _binding_range_edges")
        conn.execute("DROP TABLE IF EXISTS _range_pairs")

    cursor = conn.execute("SELECT COUNT(*) FROM binding_edges")
    return cursor.fetchone()[0]


def run_grouping_on_fast_output(
    db_path: str | None = None,
    workbook_sha256: str = "",
    conn: sqlite3.Connection | None = None,
    workbook_path: str | None = None,
    ir_db_path: str | None = None,
    workbook: "LazyWorkbook | None" = None,
) -> dict[str, Any]:
    """
    Run native grouping/refinement logic on fast pipeline outputs.

    Args:
        db_path: Path to fast schema SQLite database (if conn not provided)
        workbook_sha256: Workbook GUID for binding_id computation
        conn: Existing SQLite connection (if provided, db_path is ignored)
        workbook_path: Workbook path for optional constant-range validation

    Returns:
        Dictionary with grouping metrics.
    """
    if conn is None:
        if db_path is None:
            raise ValueError("Either db_path or conn must be provided")
        conn = sqlite3.connect(db_path)
        close_conn = True
    else:
        close_conn = False

    conn.execute("BEGIN")

    try:
        import sys as _sys
        import time as _t

        try:
            import psutil as _psutil

            _proc = _psutil.Process()
        except Exception:
            _proc = None
        _phase_times: dict[str, float] = {}
        _phase_rss: dict[str, tuple[float, float]] = {}
        _phase_start = _t.perf_counter()
        _phase_rss_start = _proc.memory_info().rss / (1024 * 1024) if _proc else 0.0

        def _rss_mb() -> float:
            if _proc:
                try:
                    return _proc.memory_info().rss / (1024 * 1024)
                except Exception:
                    return 0.0
            return 0.0

        def _mark(name: str) -> None:
            now = _t.perf_counter()
            nonlocal _phase_start, _phase_rss_start
            dur = now - _phase_start
            _phase_times[name] = dur
            rss_now = _rss_mb()
            _phase_rss[name] = (_phase_rss_start, rss_now)
            print(
                f"  [phase] {name:42s} {dur:8.2f}s  rss {_phase_rss_start:7.0f}->{rss_now:7.0f} MB "
                f"(delta {rss_now - _phase_rss_start:+7.0f})",
                file=_sys.stderr,
                flush=True,
            )
            _phase_start = now
            _phase_rss_start = rss_now

        print("Building grouping data structures...", flush=True)

        cell_id_to_addr = _build_cell_address_map(conn)
        print(f"  Cell address map: {len(cell_id_to_addr)} cells")
        addr_to_cell_id = {addr: cell_id for cell_id, addr in cell_id_to_addr.items()}
        _mark("build_cell_address_map")

        cell_formulas = _build_cell_formulas(
            conn, cell_id_to_addr, workbook_path=workbook_path, use_r1c1=not bool(workbook_path)
        )
        print(f"  Cell formulas: {len(cell_formulas)} cells")
        _mark("build_cell_formulas")

        # Canonical R1C1 signatures keyed by cell address. Sourced from the
        # formulas table (PK=formula_r1c1) so cells sharing a formula_id are
        # guaranteed to share a signature. The grouper prefers this over
        # re-deriving R1C1 from per-cell A1 strings, which has a converter
        # bug for mixed absolute/relative range refs that singletonises rows.
        cell_signatures: dict[str, str] = {}
        sig_cursor = conn.execute("""
            SELECT c.cell_id, f.formula_r1c1
            FROM cells c JOIN formulas f ON c.formula_id = f.formula_id
        """)
        for cell_id, r1c1 in sig_cursor:
            addr = cell_id_to_addr.get(cell_id)
            if addr and r1c1:
                cell_signatures[addr] = r1c1
        print(f"  Canonical signatures: {len(cell_signatures)} cells")
        _mark("build_cell_signatures")

        forward_index = _build_forward_index(conn, cell_id_to_addr)
        print(f"  Forward index: {len(forward_index)} cells with precedents")

        reverse_index = _build_reverse_index(forward_index)
        print("  Reverse index built")
        _mark("build_indices")

        cell_addresses = list(cell_id_to_addr.values())
        print(f"  Total cell addresses: {len(cell_addresses)}")

        # SheetFormulaCache no longer needed — _build_cell_formulas uses
        # per-cell formula_a1 from the DB (WI-12). Keep formula_cache=None
        # for _is_constant_range fallback compatibility.
        formula_cache = None
        name_table_map = None
        name_wb = workbook  # WI-11: reuse shared workbook from pipeline
        _close_name_wb = False
        if name_wb is not None:
            try:
                name_table_map = NameTableMap(name_wb)
            except Exception:
                name_table_map = None
        elif workbook_path:
            try:
                name_wb = LazyWorkbook(workbook_path, data_only=False, keep_vba=False)
                _close_name_wb = True
                name_table_map = NameTableMap(name_wb)
            except Exception:
                name_table_map = None
        visited_cells_by_sheet = _build_visited_cell_index(cell_id_to_addr.values())

        constant_refs = _collect_constant_ranges_from_edges(
            conn=conn, cell_formulas=cell_formulas, visited_cells_by_sheet=visited_cells_by_sheet
        )
        if constant_refs:
            print(f"  Constant range refs: {len(constant_refs)}")

        def validate_merge(merged_range: dict[str, Any]) -> bool:
            try:
                return _is_constant_range(
                    merged_range["address"], cell_formulas, formula_cache=formula_cache
                )
            except Exception:
                return False

        merged_constant_ranges = merge_constant_ranges(constant_refs, validator=validate_merge)
        constant_bindings = []
        grouped_cell_addresses: set[str] = set()
        for merged_range in merged_constant_ranges:
            try:
                cells_in_range = expand_range_to_cells(
                    merged_range["address"], max_cells=10000, populated_cells=None
                )
            except (ValueError, KeyError, AttributeError):
                continue
            if not cells_in_range:
                continue
            normalized_cells = [c.replace("$", "") for c in cells_in_range]
            constant_bindings.append(
                _create_constant_binding(
                    address_a1=merged_range["address"],
                    cells=normalized_cells,
                    workbook_guid=workbook_sha256,
                    cell_formulas=cell_formulas,
                )
            )
            for cell in normalized_cells:
                grouped_cell_addresses.add(cell)

        extra_constant_cells_by_sheet = _collect_constant_cells_by_sheet(
            cell_formulas=cell_formulas, grouped_cell_addresses=grouped_cell_addresses
        )
        extra_constant_bindings = _build_constant_bindings_from_cells(
            constant_cells_by_sheet=extra_constant_cells_by_sheet,
            workbook_guid=workbook_sha256,
            cell_formulas=cell_formulas,
        )
        if extra_constant_bindings:
            print(f"  Extra constant bindings: {len(extra_constant_bindings)}")
            constant_bindings.extend(extra_constant_bindings)
            for binding in extra_constant_bindings:
                grouped_cell_addresses.update(binding.cells)

        # Proposal #2: Split scenario-ish constant blocks by column (avoid merging Base/Stress/etc).
        if workbook_path and name_wb and constant_bindings:
            constant_bindings = _split_scenarioish_heterogeneous_constant_bindings(
                bindings=constant_bindings,
                workbook=name_wb,
                workbook_guid=workbook_sha256,
                cell_formulas=cell_formulas,
            )

        # Proposal #3: Exclude placeholder-only dummy ranges (e.g., columns of "-"/"TBC").
        if constant_bindings:
            kept: list[Binding] = []
            filtered_placeholders = 0
            filtered_headers = 0
            for binding in constant_bindings:
                cell_count = binding.shape_rows * binding.shape_cols
                cell_ids = [addr_to_cell_id[c] for c in binding.cells if c in addr_to_cell_id]
                if cell_count > 1:
                    if (
                        cell_ids
                        and _binding_is_placeholder_only(conn, cell_ids)
                        and not _has_incoming_cell_edges(conn, cell_ids)
                    ):
                        filtered_placeholders += 1
                        continue
                    # Proposal #5: omit section headers (e.g., "1. Something") that are not real variables.
                    if (
                        cell_count <= 10
                        and cell_ids
                        and _binding_is_section_header_only(conn, cell_ids)
                        and not _has_incoming_cell_edges(conn, cell_ids)
                    ):
                        filtered_headers += 1
                        continue
                kept.append(binding)
            if filtered_placeholders:
                print(f"  Filtered placeholder-only bindings: {filtered_placeholders}")
            if filtered_headers:
                print(f"  Filtered section-header bindings: {filtered_headers}")
            constant_bindings = kept

        # Filter out completely empty bindings — ranges where no cell has a
        # value or formula.  These are artifacts of SUMIF/VLOOKUP range
        # references pointing at empty cells.
        if constant_bindings:
            kept_nonempty: list[Binding] = []
            filtered_empty = 0
            for binding in constant_bindings:
                cell_ids = [addr_to_cell_id[c] for c in binding.cells if c in addr_to_cell_id]
                if cell_ids and _binding_is_completely_empty(conn, cell_ids):
                    filtered_empty += 1
                    continue
                kept_nonempty.append(binding)
            if filtered_empty:
                print(f"  Filtered completely-empty bindings: {filtered_empty}")
            constant_bindings = kept_nonempty

        _mark("constant_range_collection")

        print("\nRunning grouping...")
        cells_to_bind = []
        skipped_count = 0
        for cell_addr in cell_addresses:
            normalized_cell = cell_addr.replace("$", "")
            if normalized_cell in grouped_cell_addresses:
                skipped_count += 1
                continue
            cells_to_bind.append(cell_addr)
        if skipped_count:
            print(f"  Filtered {skipped_count} cells already in constant bindings")

        bindings = group_cells_into_bindings(
            cells=cells_to_bind,
            cell_formulas=cell_formulas,
            workbook_guid=workbook_sha256,
            formula_is_r1c1=not bool(workbook_path),
            cell_signatures=cell_signatures,
        )
        print(f"  Initial bindings: {len(bindings)}")
        _mark("group_cells_into_bindings")

        bindings = _classify_bindings(bindings, cell_formulas)
        bindings.extend(constant_bindings)
        bindings = _merge_index_vectors_into_adjacent_tables(
            bindings=bindings,
            conn=conn,
            addr_to_cell_id=addr_to_cell_id,
            cell_formulas=cell_formulas,
            workbook_guid=workbook_sha256,
            workbook=name_wb,
        )
        formula_count = sum(1 for b in bindings if b.binding_type == "formula")
        constant_count = sum(1 for b in bindings if b.binding_type == "constant")
        print(f"  Formula bindings: {formula_count}")
        print(f"  Constant bindings: {constant_count}")

        _mark("classify_and_merge_index_vectors")

        print("\nRunning refinement...")
        refinement_engine = RefinementEngine(
            workbook_guid=workbook_sha256, cell_formulas=cell_formulas
        )

        refined_bindings = refinement_engine.refine_bindings(
            bindings=bindings, reverse_index=reverse_index, forward_index=forward_index
        )
        print(f"  Refined bindings: {len(refined_bindings)}")
        _mark("refine_bindings")

        # Proposal #7: Merge contiguous formula-vector segments into a single binding when the
        # axis labels indicate a continuous time-series.
        if refined_bindings:
            refined_bindings = _merge_continuous_time_series_formula_vectors(
                bindings=refined_bindings,
                workbook=name_wb,
                workbook_guid=workbook_sha256,
                cell_formulas=cell_formulas,
                conn=conn,
            )

        _mark("merge_time_series")

        # Apply Story 31/37/38 post-IR init-merger logic directly in fast extraction output.
        # WI-6: Pass connection directly — init_merger is read-only (confirmed: zero
        # INSERT/UPDATE/DELETE in 1531 lines). No snapshot copy needed.
        refined_bindings = _apply_init_merger_mutations(
            bindings=refined_bindings,
            cell_formulas=cell_formulas,
            ir_db_conn=conn,
        )

        _mark("apply_init_merger")

        spatial_candidates_by_binding: dict[str, dict[str, Any]] | None = None
        if workbook_path and name_wb:
            try:
                from xl_marinade.core.evidence import (
                    build_evidence_cache_from_db,
                    extract_evidence_for_binding,
                )

                # WI-7: Build evidence cache from SQLite instead of re-reading
                # the workbook via openpyxl. This eliminates the dominant cost
                # for large models (80+ min for a large workbook).
                print("  Building evidence cache from DB...", file=sys.stderr)
                evidence_cache = build_evidence_cache_from_db(conn, refined_bindings)
                print(
                    f"  Cached {len(evidence_cache)} cells for evidence extraction", file=sys.stderr
                )

                # Pre-load merged cell ranges per sheet once (avoids per-binding
                # openpyxl access — was a large-workbook bottleneck at 120+ min).
                merged_ranges_by_sheet: dict[str, list] = {}
                sheets_seen: set[str] = set()
                for binding in refined_bindings:
                    if binding.sheet not in sheets_seen:
                        sheets_seen.add(binding.sheet)
                        try:
                            ws = name_wb[binding.sheet]
                            merged_ranges_by_sheet[binding.sheet] = list(ws.merged_cells.ranges)
                        except Exception:
                            merged_ranges_by_sheet[binding.sheet] = []
                print(f"  Pre-loaded merged cells for {len(sheets_seen)} sheets", file=sys.stderr)

                spatial_candidates_by_binding = {}
                for binding in refined_bindings:
                    evidence = extract_evidence_for_binding(
                        worksheet=None,
                        workbook=name_wb,
                        binding_address=binding.address_a1,
                        binding_shape=(binding.shape_rows, binding.shape_cols),
                        name_table_map=name_table_map,
                        cell_value_cache=evidence_cache,
                        merged_ranges=merged_ranges_by_sheet.get(binding.sheet),
                    )
                    spatial_candidates_by_binding[binding.binding_id] = {
                        "label_candidates": evidence.get("label_candidates", []),
                        "axis_labels": evidence.get("axis_labels", []),
                    }
            except Exception as e:
                print(f"Warning: Evidence extraction failed: {e}", file=sys.stderr)
                spatial_candidates_by_binding = None

        _mark("evidence_extraction")

        print("\nWriting bindings to database...")
        _write_bindings_to_db(
            conn=conn,
            bindings=refined_bindings,
            cell_id_to_addr=cell_id_to_addr,
            workbook_sha256=workbook_sha256,
            spatial_candidates_by_binding=spatial_candidates_by_binding,
        )
        print(f"  Bindings written: {len(refined_bindings)}")

        _mark("write_bindings_to_db")

        print("\nCollapsing cell edges to binding edges...")
        binding_edges_total = _write_binding_edges_from_cells(conn)
        print(f"  Binding edges: {binding_edges_total}")

        if formula_cache:
            formula_cache.close()
        if name_wb and _close_name_wb:
            name_wb.close()

        _mark("write_binding_edges")
        conn.commit()

        # Print profiling summary
        print(f"\n{'=' * 80}", file=sys.stderr)
        print("GROUPING SUB-PHASE TIMING + MEMORY", file=sys.stderr)
        print(f"{'=' * 80}", file=sys.stderr)
        _total = sum(_phase_times.values())
        for _name, _dur in sorted(_phase_times.items(), key=lambda x: -x[1]):
            _pct = (_dur / _total * 100) if _total else 0
            _before, _after = _phase_rss.get(_name, (0.0, 0.0))
            print(
                f"  {_name:42s} {_dur:8.1f}s  ({_pct:5.1f}%)  "
                f"rss {_before:7.0f}->{_after:7.0f} MB  ({_after - _before:+7.0f})",
                file=sys.stderr,
            )
        print(f"  {'TOTAL':42s} {_total:8.1f}s", file=sys.stderr)
        print(f"{'=' * 80}", file=sys.stderr)

        return {
            "bindings_total": len(refined_bindings),
            "bindings_formula": sum(1 for b in refined_bindings if b.binding_type == "formula"),
            "bindings_constant": sum(1 for b in refined_bindings if b.binding_type == "constant"),
            "binding_edges_total": binding_edges_total,
        }

    except Exception as e:
        conn.rollback()
        print(f"Error during grouping: {e}", flush=True)
        raise

    finally:
        if close_conn:
            conn.close()
