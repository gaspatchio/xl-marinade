# ABOUTME: Evidence extraction for bindings - label candidates and axis invariants per IR Spec §7.
# ABOUTME: Harvests variable-name candidates (7-cell scan, names, tables) and axis invariants.

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import TYPE_CHECKING, Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.names_tables import NameTableMap
from xl_marinade.core.ref_converter import parse_cell_address, quote_sheet_name

if TYPE_CHECKING:
    from xl_marinade.core.bindings import Binding


def _safe_cell_value_to_str(value: Any) -> str:
    """
    Safely convert openpyxl cell value to string for evidence extraction.

    Handles special cases like ArrayFormula objects that aren't directly str-convertible.

    Args:
        value: Cell value from openpyxl

    Returns:
        String representation safe for JSON serialization
    """
    if value is None:
        return ""

    # Handle ArrayFormula objects (convert to formula string)
    if isinstance(value, ArrayFormula):
        # ArrayFormula has a .text attribute with the formula string
        return str(value.text) if hasattr(value, "text") else str(value)

    # For other types, use str()
    return str(value)


def _is_numeric_literal(text: str) -> bool:
    """True if `text` is a bare number (a VALUE, not a header). Used to gate the
    scan_below fallback so it never treats a numeric cell below a scalar as a label."""
    if text is None:
        return False
    s = str(text).strip().replace(",", "").replace("%", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


@dataclass
class CandidateCell:
    """
    Detailed cell information for label candidates.

    Attributes:
        address: A1 address
        value: Typed value (not just string)
        formula: Formula string (if present)
        dtype: Data type classification
    """

    address: str
    value: Any
    formula: str | None
    dtype: str


@dataclass
class LabelCandidate:
    """
    Label candidate from scan, names, or tables.

    Attributes:
        type: Candidate type (e.g., "scan_left", "scan_above", "named_exact")
        address: A1 address or range of candidate source
        literals: List of text values (legacy/simple view)
        cells: List of detailed CandidateCell objects (rich view) - NEW
        format_tokens: Format token dictionary
        merged_span: Merged cell span {rows, cols} or None
    """

    type: str
    address: str
    literals: list[str]
    format_tokens: dict[str, Any]
    cells: list[CandidateCell] = field(default_factory=list)
    merged_span: dict[str, int] | None = None


@dataclass
class AxisInvariant:
    """
    Axis label candidate with computed invariants.

    Attributes:
        axis: "rows" or "columns"
        candidate: Candidate metadata with values and invariants
    """

    axis: str
    candidate: dict[str, Any]


def extract_evidence_for_binding(
    worksheet: Worksheet | None,
    workbook: Workbook,
    binding_address: str,
    binding_shape: tuple[int, int],
    name_table_map: NameTableMap | None = None,
    cell_value_cache: dict[str, Any] | None = None,
    merged_ranges: list | None = None,
) -> dict[str, Any]:
    """
    Extract evidence (label candidates + axis invariants) for a binding.

    Per IR Spec §7: Harvest deterministic variable-name candidates and axis label
    candidates for each binding, computing invariants.

    Implements unified 7-cell scan logic:
    - 1D Row (1xN): Scan LEFT
    - 1D Column (Mx1): Scan ABOVE
    - 2D (MxN): Scan LEFT and ABOVE

    Also includes explicit metadata (Named Ranges, Tables, Sheet Names).

    Args:
        worksheet: Worksheet containing binding
        workbook: Workbook object (for name/table lookups)
        binding_address: A1 range address (e.g., "B2:D2" or "B2")
        binding_shape: (rows, cols) tuple
        name_table_map: Optional name/table map for name candidate lookup

    Returns:
        Dictionary with:
            - "label_candidates": List of LabelCandidate dicts
            - "axis_labels": List of AxisInvariant dicts
    """
    label_candidates = []
    axis_labels = []

    # Parse binding address to get coordinates
    parsed = _parse_binding_range(binding_address)
    if not parsed:
        return {"label_candidates": [], "axis_labels": []}

    sheet_name, top_row, top_col, bottom_row, bottom_col = parsed
    rows, cols = binding_shape

    # 1. Add Sheet Name as high-priority candidate (if non-generic)
    sheet_candidate = _create_sheet_name_candidate(sheet_name)
    if sheet_candidate:
        label_candidates.append(sheet_candidate)

    # 2. Determine Scan Directions based on geometry
    scan_directions = []

    is_row_vector = rows == 1 and cols > 1
    is_col_vector = cols == 1 and rows > 1

    if is_row_vector:
        scan_directions.append("left")
    elif is_col_vector:
        scan_directions.append("left")
        scan_directions.append("above")
    else:
        # 2D or 1x1: Scan both
        scan_directions.append("left")
        scan_directions.append("above")

    # 3. Execute Scans (7-cell context)
    left_candidate = None
    above_candidate = None
    if "left" in scan_directions:
        left_candidate = _scan_vector(
            worksheet,
            start_row=top_row,
            start_col=top_col,
            direction="left",
            length=7,
            cell_value_cache=cell_value_cache,
            sheet_name=sheet_name,
        )
        if left_candidate:
            label_candidates.append(left_candidate)

    if "above" in scan_directions:
        above_candidate = _scan_vector(
            worksheet,
            start_row=top_row,
            start_col=top_col,
            direction="above",
            length=7,
            cell_value_cache=cell_value_cache,
            sheet_name=sheet_name,
        )
        if above_candidate:
            label_candidates.append(above_candidate)

    # 3b. scan_below (name rank 4): a 1x1 scalar whose describing header sits in the
    # row BELOW it (e.g. a model's PV summary scalars — value in row 7, header
    # 'Investment Margin' in row 8). Gated tightly: only a single-cell binding for which
    # NEITHER the left nor above scan found a usable non-numeric text header, AND the
    # immediately-below cell is genuine text — so normal top/left-headered layouts
    # (where the cell below holds data) never grab a value as a label.
    if rows == 1 and cols == 1:
        # Gate on the IMMEDIATE neighbour only (literals[0] = the adjacent cell, scans
        # are ordered closest-first). A header sits right next to its value; a stray text
        # cell several rows up (a blank immediate-above, then 'x' at distance 3) must NOT
        # block scan_below — that left a model's PV scalars unnamed.
        def _immediate_text(cand):
            return bool(
                cand
                and cand.literals
                and (cand.literals[0] or "").strip()
                and not _is_numeric_literal(cand.literals[0])
            )

        if not _immediate_text(left_candidate) and not _immediate_text(above_candidate):
            below_candidate = _scan_vector(
                worksheet,
                start_row=top_row,
                start_col=top_col,
                direction="below",
                length=1,
                cell_value_cache=cell_value_cache,
                sheet_name=sheet_name,
            )
            if (
                below_candidate
                and below_candidate.literals
                and (below_candidate.literals[0] or "").strip()
                and not _is_numeric_literal(below_candidate.literals[0])
            ):
                label_candidates.append(below_candidate)

    # 4. explicit metadata (Named Ranges / Tables)
    if name_table_map:
        name_candidates = _scan_named_ranges(
            name_table_map, sheet_name, top_row, top_col, bottom_row, bottom_col
        )
        label_candidates.extend(name_candidates)

        table_candidates = _scan_table_headers(
            name_table_map, worksheet, sheet_name, top_row, top_col, bottom_row, bottom_col
        )
        label_candidates.extend(table_candidates)

    # 5. Merged cell candidates (keep for now as they handle span logic well)
    merged_candidates = _scan_merged_cells(
        worksheet,
        top_row,
        top_col,
        bottom_row,
        bottom_col,
        cell_value_cache,
        merged_ranges=merged_ranges,
        sheet_name=sheet_name,
    )
    label_candidates.extend(merged_candidates)

    # 6. Extract axis label candidates (Compatibility with old logic, but using scans?)
    # Logic: If we are 1xN, the "axis" is columns, and labels are typically above.
    # But our main scan for 1xN is LEFT (for the variable name).
    # Axis labels are DIFFERENT from Variable Name.
    # Variable Name: "Revenue" (left of row)
    # Axis Labels: "2020", "2021" (above row)

    # We need to preserve the specific axis-detection logic:
    # 1xN -> Check Row Above for axis labels
    if rows == 1 and cols > 1:
        # We need to look ABOVE specifically for axis labels, even if main scan is LEFT
        axis_candidate = _scan_row_segment(
            worksheet, top_row - 1, top_col, bottom_col, cell_value_cache
        )
        if axis_candidate:
            axis_invariants = _compute_axis_invariants_from_cells(axis_candidate.cells, "columns")
            if axis_invariants:
                axis_labels.append(
                    {
                        "axis": "columns",
                        "candidate": {
                            "type": "row_above_1",  # Keep legacy type for axis logic?
                            "address": axis_candidate.address,
                            **axis_invariants,
                        },
                    }
                )
                # Proposal #6: When row vectors have no usable left label (report layouts),
                # treat the row-above segment as a label candidate fallback.
                if not left_candidate or not any(
                    (lit or "").strip() for lit in left_candidate.literals
                ):
                    label_candidates.append(
                        LabelCandidate(
                            type="row_above_label_fallback",
                            address=axis_candidate.address,
                            literals=axis_candidate.literals,
                            cells=axis_candidate.cells,
                            format_tokens=axis_candidate.format_tokens,
                            merged_span=axis_candidate.merged_span,
                        )
                    )

    # Mx1 -> Check Col Left for axis labels
    if cols == 1 and rows > 1:
        # We need to look LEFT specifically for axis labels, even if main scan is ABOVE
        axis_candidate = _scan_col_segment(
            worksheet, top_col - 1, top_row, bottom_row, cell_value_cache
        )
        if axis_candidate:
            axis_invariants = _compute_axis_invariants_from_cells(axis_candidate.cells, "rows")
            if axis_invariants:
                axis_labels.append(
                    {
                        "axis": "rows",
                        "candidate": {
                            "type": "col_left_1",
                            "address": axis_candidate.address,
                            **axis_invariants,
                        },
                    }
                )

    # Sort for determinism
    label_candidates.sort(key=lambda c: (c.type, c.address))
    axis_labels.sort(key=lambda a: a["axis"])

    # Convert dataclasses to dicts for return
    return {
        "label_candidates": [
            {
                "type": c.type,
                "address": c.address,
                "literals": c.literals,
                "format_tokens": c.format_tokens,
                "merged_span": c.merged_span,
                "cells": [
                    {
                        "address": cell.address,
                        "value": cell.value,
                        "formula": cell.formula,
                        "dtype": cell.dtype,
                    }
                    for cell in c.cells
                ],
            }
            for c in label_candidates
        ],
        "axis_labels": axis_labels,
    }


def _create_sheet_name_candidate(sheet_name: str) -> LabelCandidate | None:
    """
    Create a label candidate from sheet name if non-generic.

    Filters out generic/default sheet names that don't provide semantic value:
    - Sheet1, Sheet2, Sheet3, etc.
    - Generic terms: Input, Output, Data, Sheet, Inputs, Outputs

    Args:
        sheet_name: Sheet name from binding

    Returns:
        LabelCandidate if sheet name is meaningful, None if generic/filtered
    """
    if not sheet_name or not sheet_name.strip():
        return None

    # Normalize for comparison
    normalized = sheet_name.strip().lower()

    # Blacklist of generic sheet names (case-insensitive)
    generic_names = {
        "sheet1",
        "sheet2",
        "sheet3",
        "sheet4",
        "sheet5",
        "sheet 1",
        "sheet 2",
        "sheet 3",
        "sheet 4",
        "sheet 5",
        "input",
        "inputs",
        "output",
        "outputs",
        "data",
        "sheet",
        "temp",
        "working",
        "calculation",
        "calculations",
    }

    # Also filter pattern "Sheet" followed by number
    if re.match(r"^sheet\s*\d+$", normalized):
        return None

    if normalized in generic_names:
        return None

    # Non-generic sheet name - create high-confidence candidate
    return LabelCandidate(
        type="sheet_name",
        address=sheet_name,  # Use sheet name as address
        literals=[sheet_name],
        format_tokens={},
        cells=[],  # Sheet names don't have associated cells
        merged_span=None,
    )


def _parse_binding_range(address: str) -> tuple[str, int, int, int, int] | None:
    """
    Parse binding address to coordinates.
    """
    if ":" in address:
        parts = address.split(":")
        if len(parts) != 2:
            return None
        top_parsed = parse_cell_address(parts[0])
        bottom_parsed = parse_cell_address(parts[1])
        return (
            top_parsed.get("sheet", ""),
            top_parsed.get("row", 0),
            top_parsed.get("col", 0),
            bottom_parsed.get("row", 0),
            bottom_parsed.get("col", 0),
        )
    else:
        parsed = parse_cell_address(address)
        return (
            parsed.get("sheet", ""),
            parsed.get("row", 0),
            parsed.get("col", 0),
            parsed.get("row", 0),
            parsed.get("col", 0),
        )


def _get_cell_data_cached(
    worksheet: Worksheet | None,
    row: int,
    col: int,
    cell_value_cache: dict[str, Any] | None,
    sheet_name: str | None = None,
) -> CandidateCell:
    """
    Get cell data, using cache if available to avoid slow worksheet access.

    Args:
        worksheet: Worksheet containing cell (can be None if cache is provided)
        row: Row (1-indexed)
        col: Column (1-indexed)
        cell_value_cache: Optional cache mapping "Sheet!A1" -> value
        sheet_name: Sheet name for cache key (uses worksheet.title if not provided)

    Returns:
        CandidateCell with value (and formula/dtype if not using cache)
    """
    from openpyxl.utils import get_column_letter

    coord = f"{get_column_letter(col)}{row}"
    _sheet = sheet_name or (worksheet.title if worksheet else "")
    cell_addr = f"{_sheet}!{coord}"

    # Try cache first (fast path)
    if cell_value_cache and cell_addr in cell_value_cache:
        return CandidateCell(
            address=coord,
            value=cell_value_cache[cell_addr],
            formula=None,  # Not available from cache
            dtype="cached",  # Marker that this came from cache
        )

    # Cache miss handling depends on whether we have a cache at all
    if cell_value_cache is not None:
        # Cache was provided but cell wasn't in it - means cell is empty or outside bbox
        # Return empty cell immediately (fast) rather than accessing worksheet (slow)
        # This is safe because build_evidence_cache_for_bindings populates the full bbox
        return CandidateCell(address=coord, value=None, formula=None, dtype="blank")

    # No cache provided - fall back to worksheet access (slow path)
    cell = worksheet.cell(row, col)
    return _extract_rich_cell_data(cell)


def _scan_vector(
    worksheet: Worksheet | None,
    start_row: int,
    start_col: int,
    direction: str,
    length: int = 7,
    cell_value_cache: dict[str, Any] | None = None,
    sheet_name: str | None = None,
) -> LabelCandidate | None:
    """
    Scan a vector of cells in a specific direction for label context.
    Records ALL cells in the path (including empty ones).

    Args:
        worksheet: Worksheet to scan
        start_row: Starting row (1-indexed)
        start_col: Starting column (1-indexed)
        direction: "left" or "above"
        length: Number of cells to scan
        cell_value_cache: Optional cache of cell values to avoid slow worksheet access
    """
    cells_data = []
    literals = []

    curr_row, curr_col = start_row, start_col

    # Determine scan range boundaries
    if direction == "left":
        # Scan left from start_col-1 down to start_col-length
        # Stop at col 1
        end_col = max(1, start_col - length)
        scan_range_cols = range(start_col - 1, end_col - 1, -1)  # Iterate backwards (closest first)

        if start_col <= 1:
            return None  # At edge

        _sheet = sheet_name or (worksheet.title if worksheet else "")
        for col in scan_range_cols:
            c_data = _get_cell_data_cached(
                worksheet, curr_row, col, cell_value_cache, sheet_name=_sheet
            )
            cells_data.append(c_data)
            literals.append(_safe_cell_value_to_str(c_data.value))

        # Calculate address range for result
        # Range is from end_col to start_col-1 (left-to-right standard notation)
        actual_start_col = scan_range_cols[-1]  # The furthest left column visited
        actual_end_col = scan_range_cols[0]  # The closest column (start_col - 1)
        address = _format_address_range(
            _sheet, curr_row, actual_start_col, curr_row, actual_end_col
        )

        return LabelCandidate(
            type="scan_left",
            address=address,
            literals=literals,  # Note: literals are ordered closest-to-furthest here based on scan loop?
            # Actually standard JSON usually expects document order (left-to-right).
            # Let's re-order cells to be document order (Left -> Right) for consistency with address
            cells=sorted(
                cells_data, key=lambda c: _parse_binding_range(c.address)[2]
            ),  # Sort by col index
            format_tokens={},
            merged_span=None,
        )

    elif direction == "above":
        # Scan above from start_row-1 down to start_row-length
        # Stop at row 1
        end_row = max(1, start_row - length)
        scan_range_rows = range(start_row - 1, end_row - 1, -1)  # Backwards (closest first)

        if start_row <= 1:
            return None  # At edge

        _sheet = sheet_name or (worksheet.title if worksheet else "")
        for row in scan_range_rows:
            c_data = _get_cell_data_cached(
                worksheet, row, curr_col, cell_value_cache, sheet_name=_sheet
            )
            cells_data.append(c_data)
            literals.append(_safe_cell_value_to_str(c_data.value))

        actual_start_row = scan_range_rows[-1]  # Topmost row
        actual_end_row = scan_range_rows[0]  # Bottommost row
        address = _format_address_range(
            _sheet, actual_start_row, curr_col, actual_end_row, curr_col
        )

        return LabelCandidate(
            type="scan_above",
            address=address,
            literals=literals,
            cells=sorted(
                cells_data, key=lambda c: _parse_binding_range(c.address)[1]
            ),  # Sort by row index (Top -> Bottom)
            format_tokens={},
            merged_span=None,
        )

    elif direction == "below":
        # Scan below from start_row+1 down to start_row+length (closest first). Mirror
        # of "above"; used only for the gated scan_below fallback (name rank 4).
        scan_range_rows = range(start_row + 1, start_row + length + 1)
        _sheet = sheet_name or (worksheet.title if worksheet else "")
        for row in scan_range_rows:
            c_data = _get_cell_data_cached(
                worksheet, row, curr_col, cell_value_cache, sheet_name=_sheet
            )
            cells_data.append(c_data)
            literals.append(_safe_cell_value_to_str(c_data.value))

        address = _format_address_range(
            _sheet, scan_range_rows[0], curr_col, scan_range_rows[-1], curr_col
        )
        return LabelCandidate(
            type="scan_below",
            address=address,
            literals=literals,
            cells=cells_data,
            format_tokens={},
            merged_span=None,
        )

    return None


def _scan_row_segment(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    cell_value_cache: dict[str, Any] | None = None,
) -> LabelCandidate | None:
    """Helper to scan a specific horizontal segment (for axis logic).

    PERFORMANCE: Uses cell_value_cache to avoid slow worksheet.cell() calls.
    """
    if row < 1:
        return None
    cells_data = []
    literals = []

    for col in range(start_col, end_col + 1):
        c_data = _get_cell_data_cached(worksheet, row, col, cell_value_cache)
        cells_data.append(c_data)
        literals.append(_safe_cell_value_to_str(c_data.value))

    if not any(literals):
        return None

    address = _format_address_range(worksheet.title, row, start_col, row, end_col)
    return LabelCandidate(
        type="row_segment", address=address, literals=literals, cells=cells_data, format_tokens={}
    )


def _scan_col_segment(
    worksheet: Worksheet,
    col: int,
    start_row: int,
    end_row: int,
    cell_value_cache: dict[str, Any] | None = None,
) -> LabelCandidate | None:
    """Helper to scan a specific vertical segment (for axis logic).

    PERFORMANCE: Uses cell_value_cache to avoid slow worksheet.cell() calls.
    """
    if col < 1:
        return None
    cells_data = []
    literals = []

    for row in range(start_row, end_row + 1):
        c_data = _get_cell_data_cached(worksheet, row, col, cell_value_cache)
        cells_data.append(c_data)
        literals.append(_safe_cell_value_to_str(c_data.value))

    if not any(literals):
        return None

    address = _format_address_range(worksheet.title, start_row, col, end_row, col)
    return LabelCandidate(
        type="col_segment", address=address, literals=literals, cells=cells_data, format_tokens={}
    )


def _extract_rich_cell_data(cell) -> CandidateCell:
    """Extract rich data from a cell.

    Handles both regular Cell objects and EmptyCell objects (from read-only mode).
    """
    from openpyxl.utils import get_column_letter

    value, dtype = _extract_cell_value_and_dtype(cell)

    # Extract formula if present
    formula = None
    if hasattr(cell, "data_type") and cell.data_type == "f":
        # OpenPyXL stores formula in .value if it's not cached, or we need to access it specifically?
        # Actually for data_type='f', .value IS the formula string usually, unless loaded data_only=True
        # Assuming data_only=False (formulas present)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            formula = cell.value
        elif isinstance(cell.value, ArrayFormula):
            formula = cell.value.text

    # Handle both Cell (has coordinate) and EmptyCell (needs row/column)
    if hasattr(cell, "coordinate") and cell.coordinate:
        address = cell.coordinate
    elif hasattr(cell, "row") and hasattr(cell, "column"):
        address = f"{get_column_letter(cell.column)}{cell.row}"
    else:
        # Fallback - shouldn't happen but safe default
        address = "??"

    return CandidateCell(address=address, value=value, formula=formula, dtype=dtype)


def _scan_named_ranges(
    name_table_map: NameTableMap,
    sheet: str,
    top_row: int,
    top_col: int,
    bottom_row: int,
    bottom_col: int,
) -> list[LabelCandidate]:
    """
    Find named ranges that intersect with binding.
    """
    candidates = []
    all_names = name_table_map.get_all_names()

    for name_info in all_names:
        if name_info.is_external:
            continue

        for name_range in name_info.ranges:
            parsed = _parse_binding_range(name_range)
            if not parsed:
                continue

            name_sheet, name_top_row, name_top_col, name_bottom_row, name_bottom_col = parsed

            if name_sheet != sheet:
                continue

            relationship = _classify_range_relationship(
                top_row,
                top_col,
                bottom_row,
                bottom_col,
                name_top_row,
                name_top_col,
                name_bottom_row,
                name_bottom_col,
            )

            if relationship:
                candidates.append(
                    LabelCandidate(
                        type=relationship,
                        address=name_range,
                        literals=[name_info.name],
                        format_tokens={},
                        merged_span=None,
                        cells=[],  # Explicit names don't have "cells" in the same way
                    )
                )

    return candidates


def _classify_range_relationship(
    b_top: int,
    b_left: int,
    b_bottom: int,
    b_right: int,
    n_top: int,
    n_left: int,
    n_bottom: int,
    n_right: int,
) -> str | None:
    """
    Classify relationship between binding (b) and name range (n).
    """
    if b_top == n_top and b_left == n_left and b_bottom == n_bottom and b_right == n_right:
        return "named_exact"

    if n_top <= b_top and n_left <= b_left and n_bottom >= b_bottom and n_right >= b_right:
        return "named_superset"

    if b_top <= n_top and b_left <= n_left and b_bottom >= n_bottom and b_right >= n_right:
        return "named_subset"

    return None


def _scan_table_headers(
    name_table_map: NameTableMap,
    worksheet: Worksheet,
    sheet: str,
    top_row: int,
    top_col: int,
    bottom_row: int,
    bottom_col: int,
) -> list[LabelCandidate]:
    """
    Find table headers that intersect with binding.
    """
    candidates = []
    all_tables = name_table_map.get_all_tables()

    for table_info in all_tables:
        if table_info.sheet != sheet:
            continue

        parsed = _parse_binding_range(table_info.range)
        if not parsed:
            continue

        _, table_top, table_left, table_bottom, table_right = parsed

        if not _ranges_overlap(
            top_row,
            top_col,
            bottom_row,
            bottom_col,
            table_top,
            table_left,
            table_bottom,
            table_right,
        ):
            continue

        header_parsed = _parse_binding_range(table_info.header_row)
        if header_parsed:
            _, header_row, header_left, _, header_right = header_parsed

            if _ranges_overlap(
                top_row,
                top_col,
                bottom_row,
                bottom_col,
                header_row,
                header_left,
                header_row,
                header_right,
            ):
                candidates.append(
                    LabelCandidate(
                        type="table_header_row",
                        address=table_info.header_row,
                        literals=table_info.columns,
                        format_tokens={},
                        merged_span=None,
                        cells=[],
                    )
                )

        if top_row == table_top and bottom_row == table_bottom:
            matching_columns = []
            for col in range(top_col, bottom_col + 1):
                col_index = col - table_left
                if 0 <= col_index < len(table_info.columns):
                    matching_columns.append(table_info.columns[col_index])

            if matching_columns:
                candidates.append(
                    LabelCandidate(
                        type="table_column_headers",
                        address=table_info.header_row,  # Using header row as address
                        literals=matching_columns,
                        format_tokens={},
                        merged_span=None,
                        cells=[],
                    )
                )

    return candidates


def _ranges_overlap(
    a_top: int,
    a_left: int,
    a_bottom: int,
    a_right: int,
    b_top: int,
    b_left: int,
    b_bottom: int,
    b_right: int,
) -> bool:
    """Check if two ranges overlap."""
    return not (a_bottom < b_top or a_top > b_bottom or a_right < b_left or a_left > b_right)


def _scan_merged_cells(
    worksheet: Worksheet | None,
    top_row: int,
    top_col: int,
    bottom_row: int,
    bottom_col: int,
    cell_value_cache: dict[str, Any] | None = None,
    merged_ranges: list | None = None,
    sheet_name: str | None = None,
) -> list[LabelCandidate]:
    """
    Find merged cells touching binding margins.

    PERFORMANCE: Uses O(n) single-pass algorithm instead of O(n×m) nested loops.
    Previous O(n²) implementation caused 1.5M+ iterations for large workbooks.
    Also uses cell_value_cache to avoid slow worksheet.cell() calls.
    When merged_ranges is pre-loaded, avoids per-binding openpyxl worksheet access.
    See Observation 16 in SPRINT4_OBSERVATIONS.md for details.
    """
    candidates = []
    seen_addresses: set[str] = set()

    # Define margin positions (row above, column left)
    check_row_above = top_row - 1 if top_row > 1 else None
    check_col_left = top_col - 1 if top_col > 1 else None

    # Use pre-loaded merged ranges if available, otherwise read from worksheet
    ranges_to_scan = merged_ranges
    if ranges_to_scan is None:
        if worksheet is None:
            return []
        ranges_to_scan = worksheet.merged_cells.ranges

    _sheet = sheet_name or (worksheet.title if worksheet else "")

    # Single pass through merged cells - O(n) instead of O(n × binding_size)
    for merged_range in ranges_to_scan:
        # Check if merged range touches the row ABOVE the binding
        # Condition: merged range contains check_row AND overlaps with binding's columns
        if check_row_above is not None:
            if (
                merged_range.min_row <= check_row_above <= merged_range.max_row
                and merged_range.max_col >= top_col
                and merged_range.min_col <= bottom_col
            ):
                address = _format_address_range(
                    _sheet,
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col,
                )
                if address not in seen_addresses:
                    seen_addresses.add(address)
                    merged_cell_data = _get_cell_data_cached(
                        worksheet,
                        merged_range.min_row,
                        merged_range.min_col,
                        cell_value_cache,
                        sheet_name=_sheet,
                    )
                    if merged_cell_data.value:
                        candidates.append(
                            LabelCandidate(
                                type="merged_above",
                                address=address,
                                literals=[_safe_cell_value_to_str(merged_cell_data.value)],
                                format_tokens={},
                                merged_span={
                                    "rows": merged_range.max_row - merged_range.min_row + 1,
                                    "cols": merged_range.max_col - merged_range.min_col + 1,
                                },
                                cells=[],  # TODO: Extract merged cell content as rich cell?
                            )
                        )

        # Check if merged range touches the column LEFT of the binding
        # Condition: merged range contains check_col AND overlaps with binding's rows
        if check_col_left is not None:
            if (
                merged_range.min_col <= check_col_left <= merged_range.max_col
                and merged_range.max_row >= top_row
                and merged_range.min_row <= bottom_row
            ):
                address = _format_address_range(
                    _sheet,
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col,
                )
                if address not in seen_addresses:
                    seen_addresses.add(address)
                    merged_cell_data = _get_cell_data_cached(
                        worksheet,
                        merged_range.min_row,
                        merged_range.min_col,
                        cell_value_cache,
                        sheet_name=_sheet,
                    )
                    if merged_cell_data.value:
                        candidates.append(
                            LabelCandidate(
                                type="merged_left",
                                address=address,
                                literals=[_safe_cell_value_to_str(merged_cell_data.value)],
                                format_tokens={},
                                merged_span={
                                    "rows": merged_range.max_row - merged_range.min_row + 1,
                                    "cols": merged_range.max_col - merged_range.min_col + 1,
                                },
                                cells=[],
                            )
                        )

    return candidates


def _compute_axis_invariants(
    worksheet: Worksheet,
    axis_address: str,
    axis_direction: str,
) -> dict[str, Any] | None:
    """
    Compute invariants for axis labels (Legacy entry point, parses address).
    """
    parsed = _parse_binding_range(axis_address)
    if not parsed:
        return None

    _, top_row, top_col, bottom_row, bottom_col = parsed

    cells = []
    if axis_direction == "rows":
        for row in range(top_row, bottom_row + 1):
            cells.append(worksheet.cell(row, top_col))
    else:
        for col in range(top_col, bottom_col + 1):
            cells.append(worksheet.cell(top_row, col))

    # Convert to CandidateCell for shared logic
    rich_cells = [_extract_rich_cell_data(c) for c in cells]
    return _compute_axis_invariants_from_cells(rich_cells, axis_direction)


def _compute_axis_invariants_from_cells(
    cells: list[CandidateCell],
    axis_direction: str,
) -> dict[str, Any] | None:
    """
    Compute invariants from pre-extracted rich cells.
    """
    values = [c.value for c in cells]
    display_texts = [_safe_cell_value_to_str(c.value) for c in cells]
    dtype_set = {c.dtype for c in cells}

    monotonic = _compute_monotonicity(values)
    step_kind = _compute_step_kind(values, list(dtype_set))
    distinct_count = len({str(v) for v in values})
    duplicates_present = distinct_count < len(values)

    return {
        "values_snapshot": values,
        "display_texts": display_texts,
        "dtype_set": sorted(dtype_set),
        "monotonic": monotonic,
        "step_kind": step_kind,
        "distinct_count": distinct_count,
        "duplicates_present": duplicates_present,
    }


def _extract_cell_value_and_dtype(cell) -> tuple[Any, str]:
    """
    Extract cell value and data type classification.
    """
    if cell.value is None:
        return (None, "blank")

    if isinstance(cell.value, ArrayFormula):
        formula_str = str(cell.value.text) if hasattr(cell.value, "text") else str(cell.value)
        return (formula_str, "array_formula")

    if isinstance(cell.value, bool):
        return (cell.value, "boolean")

    if isinstance(cell.value, (int, float)):
        if cell.number_format and any(d in cell.number_format.lower() for d in ["d", "m", "y"]):
            return (cell.value, "date")
        return (cell.value, "number")

    if isinstance(cell.value, datetime):
        return (cell.value.isoformat(), "date")

    if isinstance(cell.value, str):
        if cell.value.startswith("#"):
            return (cell.value, "error")
        return (cell.value, "text")

    return (str(cell.value), "unknown")


def _compute_monotonicity(values: list[Any]) -> str:
    """
    Compute monotonicity of numeric sequence.
    """
    numeric_values = [v for v in values if v is not None and isinstance(v, (int, float))]

    if len(numeric_values) < 2:
        return "none"

    if all(v == numeric_values[0] for v in numeric_values):
        return "constant"

    increasing = all(
        numeric_values[i] <= numeric_values[i + 1] for i in range(len(numeric_values) - 1)
    )
    if increasing:
        return "increasing"

    decreasing = all(
        numeric_values[i] >= numeric_values[i + 1] for i in range(len(numeric_values) - 1)
    )
    if decreasing:
        return "decreasing"

    return "none"


def _compute_step_kind(values: list[Any], dtypes: list[str]) -> dict[str, Any] | None:
    """
    Compute step kind for numeric sequences.
    """
    numeric_values = [v for v in values if v is not None and isinstance(v, (int, float))]

    if len(numeric_values) < 2:
        return None

    diffs = [numeric_values[i + 1] - numeric_values[i] for i in range(len(numeric_values) - 1)]

    if all(abs(d - diffs[0]) < 1e-9 for d in diffs):
        step_value = diffs[0]

        if "date" in dtypes:
            return {"kind": "constant_days", "days": int(round(step_value))}
        else:
            return {"kind": "constant_step", "step": step_value}

    return None


def _format_address_range(
    sheet: str, top_row: int, left_col: int, bottom_row: int, right_col: int
) -> str:
    """Format A1 range address."""

    def col_to_letter(col: int) -> str:
        result = ""
        while col > 0:
            col -= 1
            result = chr(col % 26 + ord("A")) + result
            col //= 26
        return result

    top_left = f"{col_to_letter(left_col)}{top_row}"
    quoted_sheet = quote_sheet_name(sheet)

    if top_row == bottom_row and left_col == right_col:
        return f"{quoted_sheet}!{top_left}"

    bottom_right = f"{col_to_letter(right_col)}{bottom_row}"
    return f"{quoted_sheet}!{top_left}:{bottom_right}"


def compute_evidence_regions(
    bindings: list,
) -> tuple[dict[str, list[tuple[int, int, int, int]]], dict[str, tuple[int, int, int, int]]]:
    """Compute per-sheet bounding boxes for evidence extraction.

    Each binding needs: 7 cells left + 7 cells above its top-left,
    plus axis labels capped at MAX_AXIS_EXTENT=20.

    Returns:
        (regions_by_sheet, bbox_by_sheet) where:
        - regions_by_sheet maps sheet -> list of (min_row, min_col, max_row, max_col)
        - bbox_by_sheet maps sheet -> merged bounding box for the sheet
    """
    MAX_AXIS_EXTENT = 20

    regions_by_sheet: dict[str, list[tuple[int, int, int, int]]] = {}
    bbox_by_sheet: dict[str, tuple[int, int, int, int]] = {}

    for binding in bindings:
        sheet = binding.sheet
        parsed = parse_cell_address(binding.top_left_a1)
        if not parsed:
            continue

        top_row = parsed.get("row", 1)
        top_col = parsed.get("col", 1)

        scan_min_row = max(1, top_row - 7)
        scan_min_col = max(1, top_col - 7)
        # +1 row so a 1x1 scalar's scan_below (name rank 4) always has its
        # immediately-below header cell cached. Without it, scan_below only fired when a
        # neighbouring binding's region happened to cover that cell — which left some
        # scalars' below-header cells uncached (so they fell back to the sheet name).
        scan_max_row = min(top_row + binding.shape_rows - 1, top_row + MAX_AXIS_EXTENT - 1) + 1
        scan_max_col = min(top_col + binding.shape_cols - 1, top_col + MAX_AXIS_EXTENT - 1)

        region = (scan_min_row, scan_min_col, scan_max_row, scan_max_col)

        if sheet not in regions_by_sheet:
            regions_by_sheet[sheet] = [region]
            bbox_by_sheet[sheet] = region
        else:
            regions_by_sheet[sheet].append(region)
            curr = bbox_by_sheet[sheet]
            bbox_by_sheet[sheet] = (
                min(curr[0], scan_min_row),
                min(curr[1], scan_min_col),
                max(curr[2], scan_max_row),
                max(curr[3], scan_max_col),
            )

    return regions_by_sheet, bbox_by_sheet


def build_evidence_cache_from_db(
    conn,
    bindings: list,
) -> dict[str, Any]:
    """Build evidence cache from SQLite IR database instead of re-reading the workbook.

    Replaces build_evidence_cache_for_bindings for the fast pipeline path.
    Queries the cells/json_blobs tables that already contain all extracted data.

    Uses an SQL JOIN against a temp regions table to let SQLite do the
    point-in-rectangle filtering via the (sheet_id, row, col) index. The prior
    implementation fetched every cell in the union bbox per sheet and did a
    linear Python-side scan over all regions for each cell, which was
    O(cells × regions) and minutes-to-hours on large workbooks
    (a large workbook: ~800k cells × ~10k regions on one sheet = 8B comparisons).

    Args:
        conn: SQLite connection to the IR build database
        bindings: List of Binding objects to extract evidence for

    Returns:
        Dict mapping "Sheet!A1" -> cell value for evidence-relevant cells
    """
    import json as _json

    if not bindings:
        return {}

    regions_by_sheet, _bbox_by_sheet = compute_evidence_regions(bindings)
    if not regions_by_sheet:
        return {}

    # sheet_name -> sheet_id
    sheet_id_map = {
        name: sid for sid, name in conn.execute("SELECT sheet_id, sheet_name FROM sheets")
    }

    region_rows: list[tuple[int, int, int, int, int]] = []
    for sheet_name, regions in regions_by_sheet.items():
        sid = sheet_id_map.get(sheet_name)
        if sid is None:
            continue
        for r1, c1, r2, c2 in regions:
            region_rows.append((sid, r1, c1, r2, c2))

    if not region_rows:
        return {}

    conn.execute("DROP TABLE IF EXISTS temp_evidence_regions")
    conn.execute(
        """
        CREATE TEMP TABLE temp_evidence_regions (
            sheet_id INTEGER NOT NULL,
            r1 INTEGER NOT NULL,
            c1 INTEGER NOT NULL,
            r2 INTEGER NOT NULL,
            c2 INTEGER NOT NULL
        )
        """
    )
    conn.executemany(
        "INSERT INTO temp_evidence_regions(sheet_id, r1, c1, r2, c2) VALUES (?, ?, ?, ?, ?)",
        region_rows,
    )
    conn.execute("CREATE INDEX temp_evidence_regions_sheet ON temp_evidence_regions(sheet_id)")

    cache: dict[str, Any] = {}
    try:
        query = """
            SELECT DISTINCT s.sheet_name, c.a1, jv.json
            FROM temp_evidence_regions r
            JOIN cells c
              ON c.sheet_id = r.sheet_id
             AND c.row BETWEEN r.r1 AND r.r2
             AND c.col BETWEEN r.c1 AND r.c2
            JOIN sheets s ON c.sheet_id = s.sheet_id
            JOIN json_blobs jv ON c.value_blob_id = jv.blob_id
        """
        for sheet_name, a1, value_json in conn.execute(query):
            try:
                value = _json.loads(value_json)
            except (ValueError, TypeError):
                value = value_json
            if value is not None:
                cache[f"{sheet_name}!{a1}"] = value
    finally:
        conn.execute("DROP TABLE IF EXISTS temp_evidence_regions")

    return cache


def build_evidence_cache_for_bindings(workbook: Any, bindings: list["Binding"]) -> dict[str, Any]:
    """
    Pre-load cells needed for evidence extraction using minimal bounding boxes.

    Strategy: Compute one bounding box per sheet that covers all binding scan regions,
    then filter to only cache cells that are actually in a binding's scan region.
    This gives O(1) worksheet iterations per sheet while keeping cache size minimal.

    Performance: O(sheets × bbox_cells) where bbox_cells is the merged scan regions

    Args:
        workbook: Workbook object (LazyWorkbook or standard openpyxl Workbook)
        bindings: List of Binding objects to extract evidence for

    Returns:
        Dict mapping "Sheet!A1" -> cell value for evidence-relevant cells only
    """
    from openpyxl.utils import get_column_letter

    if not bindings:
        return {}

    # 1. Collect required regions per sheet
    # Each binding needs: 7 cells left + 7 cells above its top-left
    # Plus axis labels (capped at 20 cells for large bindings)
    MAX_AXIS_EXTENT = 20

    # Maps sheet -> list of (min_row, min_col, max_row, max_col) regions
    regions_by_sheet: dict[str, list[tuple[int, int, int, int]]] = {}
    # Maps sheet -> (overall_min_row, overall_min_col, overall_max_row, overall_max_col)
    bbox_by_sheet: dict[str, tuple[int, int, int, int]] = {}

    for binding in bindings:
        sheet = binding.sheet
        parsed = parse_cell_address(binding.top_left_a1)
        if not parsed:
            continue

        top_row = parsed.get("row", 1)
        top_col = parsed.get("col", 1)

        # Scan region for this binding
        scan_min_row = max(1, top_row - 7)
        scan_min_col = max(1, top_col - 7)
        # +1 row so a 1x1 scalar's scan_below header cell is cached (see the
        # compute_evidence_regions note above — keeps the two region builders in sync).
        scan_max_row = min(top_row + binding.shape_rows - 1, top_row + MAX_AXIS_EXTENT - 1) + 1
        scan_max_col = min(top_col + binding.shape_cols - 1, top_col + MAX_AXIS_EXTENT - 1)

        region = (scan_min_row, scan_min_col, scan_max_row, scan_max_col)

        if sheet not in regions_by_sheet:
            regions_by_sheet[sheet] = []
            bbox_by_sheet[sheet] = region
        else:
            regions_by_sheet[sheet].append(region)
            curr = bbox_by_sheet[sheet]
            bbox_by_sheet[sheet] = (
                min(curr[0], scan_min_row),
                min(curr[1], scan_min_col),
                max(curr[2], scan_max_row),
                max(curr[3], scan_max_col),
            )
        regions_by_sheet[sheet].append(region)

    # 2. Build cache by streaming each sheet's bounding box once
    cache: dict[str, Any] = {}

    for sheet, bbox in bbox_by_sheet.items():
        try:
            ws = workbook[sheet]
        except (KeyError, ValueError):
            continue

        min_row, min_col, max_row, max_col = bbox
        regions = regions_by_sheet[sheet]

        # Stream cells in the bounding box
        try:
            for row in ws.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                for cell in row:
                    # Get cell coordinates
                    if hasattr(cell, "coordinate") and cell.coordinate:
                        coord = cell.coordinate
                    elif hasattr(cell, "row") and hasattr(cell, "column"):
                        coord = f"{get_column_letter(cell.column)}{cell.row}"
                    else:
                        continue

                    # Get row/col indices
                    if hasattr(cell, "row") and hasattr(cell, "column"):
                        cell_row, cell_col = cell.row, cell.column
                    else:
                        # Parse from coordinate
                        from openpyxl.utils import column_index_from_string, coordinate_from_string

                        col_str, cell_row = coordinate_from_string(coord)
                        cell_col = column_index_from_string(col_str)

                    # Check if this cell is in ANY binding's region
                    in_region = False
                    for r_min_row, r_min_col, r_max_row, r_max_col in regions:
                        if (
                            r_min_row <= cell_row <= r_max_row
                            and r_min_col <= cell_col <= r_max_col
                        ):
                            in_region = True
                            break

                    if in_region:
                        value = cell.value
                        if isinstance(value, ArrayFormula):
                            value = str(value.text) if hasattr(value, "text") else str(value)
                        elif isinstance(value, datetime):
                            value = value.isoformat()
                        cache[f"{sheet}!{coord}"] = value
        except Exception:
            continue

    return cache
