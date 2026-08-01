# Integration test for VBA UDF extraction (Story 10)

from pathlib import Path

import pytest
from openpyxl import load_workbook

from xl_marinade.core.vba_parser import extract_udfs_from_workbook


def test_extract_udfs_from_sprint2_workbook():
    """
    Integration test: Extract UDFs from real .xlsm workbook.

    This test validates the end-to-end VBA extraction using the
    sprint2_unified_use_case_workbook.xlsm test fixture.
    """
    workbook_path = (
        Path(__file__).parent.parent.parent
        / "test_spreadsheets"
        / "sprint2_unified_use_case_workbook.xlsm"
    )

    if not workbook_path.exists():
        pytest.skip(f"Test workbook not found: {workbook_path}")

    # Load workbook with VBA
    wb = load_workbook(workbook_path, keep_vba=True)

    # Extract UDFs
    udfs = extract_udfs_from_workbook(wb)

    # Verify we extracted the expected UDFs
    assert len(udfs) == 2, f"Expected 2 UDFs, got {len(udfs)}"

    udf_names = {udf.name for udf in udfs}
    assert "SampleUdf" in udf_names, "SampleUdf not found"
    assert "NonVolatileUdf" in udf_names, "NonVolatileUdf not found"

    # Verify SampleUdf metadata
    sample_udf = next(udf for udf in udfs if udf.name == "SampleUdf")
    assert sample_udf.module == "Module1.bas"
    assert sample_udf.param_count == 3
    assert sample_udf.param_names == ["a", "b", "c"]
    assert sample_udf.declared_volatile is True, (
        "SampleUdf should be volatile (Application.Volatile)"
    )
    assert len(sample_udf.source_hash) == 64, "Source hash should be SHA-256 (64 hex chars)"

    # Verify NonVolatileUdf metadata
    non_volatile_udf = next(udf for udf in udfs if udf.name == "NonVolatileUdf")
    assert non_volatile_udf.module == "Module1.bas"
    assert non_volatile_udf.param_count == 1
    assert non_volatile_udf.param_names == ["x"]
    assert non_volatile_udf.declared_volatile is False, "NonVolatileUdf should NOT be volatile"
    assert len(non_volatile_udf.source_hash) == 64


def test_extract_udfs_from_xlsx_returns_empty():
    """
    Test that .xlsx files (without VBA) return empty list.
    """
    workbook_path = (
        Path(__file__).parent.parent.parent
        / "test_spreadsheets"
        / "sprint2_unified_use_case_workbook.xlsx"
    )

    if not workbook_path.exists():
        pytest.skip(f"Test workbook not found: {workbook_path}")

    # Load .xlsx workbook (no VBA)
    wb = load_workbook(workbook_path, keep_vba=False)

    # Extract UDFs - should be empty
    udfs = extract_udfs_from_workbook(wb)

    assert len(udfs) == 0, "xlsx files should have no UDFs"


def test_udf_source_hash_determinism():
    """
    Test that UDF source hashes are deterministic across multiple extractions.
    """
    workbook_path = (
        Path(__file__).parent.parent.parent
        / "test_spreadsheets"
        / "sprint2_unified_use_case_workbook.xlsm"
    )

    if not workbook_path.exists():
        pytest.skip(f"Test workbook not found: {workbook_path}")

    # Extract twice
    wb1 = load_workbook(workbook_path, keep_vba=True)
    udfs1 = extract_udfs_from_workbook(wb1)

    wb2 = load_workbook(workbook_path, keep_vba=True)
    udfs2 = extract_udfs_from_workbook(wb2)

    # Build maps for comparison
    hash_map1 = {udf.name: udf.source_hash for udf in udfs1}
    hash_map2 = {udf.name: udf.source_hash for udf in udfs2}

    # Hashes should be identical
    assert hash_map1 == hash_map2, "UDF source hashes should be deterministic"
