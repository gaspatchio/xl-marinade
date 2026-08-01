"""Performance regression tests for IR extraction.

These tests ensure that critical performance issues discovered during debugging
are caught before deployment. Tests use real workbook operations and set
strict time limits to catch O(n²) regressions.
"""

import time

import pytest
from openpyxl import Workbook, load_workbook

from xl_marinade.core.lazy_workbook import LazyWorkbook
from xl_marinade.core.names_tables import NameTableMap
from xl_marinade.core.topology import compute_sheet_topology


class TestLazyWorksheetPerformance:
    """Test LazyWorksheet doesn't regress to O(n²) with cache thrashing."""

    def test_random_access_large_cache(self, tmp_path):
        """Test random access pattern with 25K+ cells completes quickly.

        Regression test for LazyWorksheet LRU cache thrashing issue (2024-12-22).
        With MAX_CACHE_SIZE=5000, random access caused cache thrashing and O(n²).
        With MAX_CACHE_SIZE=1M, random access should be fast.
        """
        # Create workbook with 30K cells (1000 rows × 30 cols)
        wb = Workbook()
        ws = wb.active
        ws.title = "LargeSheet"

        # Populate with formulas
        for row in range(1, 1001):
            for col in range(1, 31):
                ws.cell(row, col, value=f"=A1+{row}+{col}")

        # Save and load as LazyWorkbook
        test_file = tmp_path / "large_wb.xlsx"
        wb.save(test_file)

        lazy_wb = LazyWorkbook(test_file)
        lazy_ws = lazy_wb["LargeSheet"]

        # Random access pattern (simulates BFS traversal order)
        import random

        from openpyxl.utils import get_column_letter

        coords = [
            f"{get_column_letter(col + 1)}{row}" for row in range(1, 1001) for col in range(30)
        ]
        random.shuffle(coords)

        # Access first 10K cells in random order
        # This should complete in < 5 seconds
        start = time.time()
        for coord in coords[:10000]:
            cell = lazy_ws.get_cell(coord)
            _ = cell.value
        elapsed = time.time() - start

        # Should be fast (< 5 seconds for 10K accesses)
        assert elapsed < 5.0, f"Random access took {elapsed:.1f}s (cache thrashing?)"


class TestSheetTopologyPerformance:
    """Test sheet topology computation doesn't iterate entire bounding box."""

    def test_large_bounding_box_fast(self, tmp_path):
        """Test topology computation with large bounding box completes quickly.

        Regression test for compute_sheet_topology bounding box iteration (2024-12-22).
        Old code iterated entire bounding box (1000 rows × 40 cols = 40K cells) calling
        worksheet.cell() for each, causing O(n²). New code uses only visited cells.
        """
        # Create workbook with sparse large range
        wb = Workbook()
        ws = wb.active
        ws.title = "SparseSheet"

        # Sparse cells spanning large bounding box (rows 1-1000, cols A-Z)
        from openpyxl.utils import get_column_letter

        visited_cells = []
        for row in [1, 500, 1000]:
            for col in [1, 13, 26]:  # A, M, Z
                coord = f"{get_column_letter(col)}{row}"
                ws[coord] = f"=A1+{row}"
                visited_cells.append(f"SparseSheet!{coord}")

        # Save and reload (to use real workbook)
        test_file = tmp_path / "sparse_wb.xlsx"
        wb.save(test_file)

        loaded_wb = load_workbook(test_file)
        loaded_ws = loaded_wb["SparseSheet"]

        # Compute topology - should be fast (< 1 second)
        # Old code would iterate 1000 rows × 26 cols = 26K cells
        # New code only processes 9 visited cells
        start = time.time()
        topology = compute_sheet_topology(loaded_ws, visited_cells, margin=5, name_table_map=None)
        elapsed = time.time() - start

        assert topology is not None
        assert "bbox" in topology
        assert "cells" in topology

        # Should complete in < 1 second (was hanging for minutes)
        assert elapsed < 1.0, f"Topology computation took {elapsed:.1f}s (bbox iteration?)"

        # Verify it only processed visited cells, not entire bbox
        assert len(topology["cells"]) == len(visited_cells), (
            "Should only contain visited cells, not entire bounding box"
        )


class TestNameTableMapPerformance:
    """Test NameTableMap doesn't cause performance issues."""

    def test_large_name_table_initialization(self, tmp_path):
        """Test NameTableMap initialization with many names/tables is fast."""
        # Create workbook with 100 defined names
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        ws = wb.active

        for i in range(100):
            # Use modern API instead of deprecated create_named_range
            defn = DefinedName(name=f"Name_{i}", attr_text=f"'{ws.title}'!$A${i + 1}")
            wb.defined_names[f"Name_{i}"] = defn

        test_file = tmp_path / "many_names.xlsx"
        wb.save(test_file)

        loaded_wb = load_workbook(test_file)

        # NameTableMap initialization should be fast
        start = time.time()
        name_map = NameTableMap(loaded_wb)
        elapsed = time.time() - start

        assert elapsed < 1.0, f"NameTableMap init took {elapsed:.1f}s"
        assert len(name_map.get_all_names()) == 100


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
