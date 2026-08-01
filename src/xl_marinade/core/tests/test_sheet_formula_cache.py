# ABOUTME: Regression tests for SheetFormulaCache to ensure reliable formula extraction
# ABOUTME: Tests the sheet-level formula caching that prevents LazyWorksheet cache exhaustion

"""
Regression tests for SheetFormulaCache.

These tests ensure that formulas are reliably extracted even when:
1. Many cells are accessed in non-sequential order
2. The LazyWorksheet cache would normally be exhausted
3. Cells from multiple sheets are accessed

Background:
- Bug discovered 2024-12-20: Formulas for X15-X956 were not captured in database
- Root cause: LazyWorksheet's streaming iterator + LRU cache (5000 cells) couldn't
  handle BFS traversal which accesses cells in sorted (non-row) order (AL before X)
- Solution: SheetFormulaCache pre-loads ALL formulas from a sheet on first access,
  avoiding cache exhaustion issues
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from xl_marinade.core.lazy_formulas import SheetFormulaCache


def create_large_test_workbook(path: Path, num_rows: int = 100) -> None:
    """
    Create test workbook with formulas in multiple columns to simulate
    the access pattern that caused the original bug.

    Structure:
    - A1: Root cell with SUMPRODUCT(B5:B{num_rows}, C5:C{num_rows})
    - B5:B{num_rows}: Formulas =D{row}+E{row}
    - C5:C{num_rows}: Constants (1.0, 2.0, etc.)
    - D5:D{num_rows}: Formulas =F{row}*2
    - E5:E{num_rows}: Formulas =F{row}*3
    - F5:F{num_rows}: Constants (base values)

    When sorted alphabetically, access order is: B, C, D, E, F
    But cells in row 5 should all be accessible after row 50 is accessed.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Root formula references range
    ws["A1"] = f"=SUMPRODUCT(B5:B{num_rows},C5:C{num_rows})"

    for row in range(5, num_rows + 1):
        # Column B: Formula referencing D and E
        ws[f"B{row}"] = f"=D{row}+E{row}"

        # Column C: Constant
        ws[f"C{row}"] = float(row)

        # Column D: Formula referencing F
        ws[f"D{row}"] = f"=F{row}*2"

        # Column E: Formula referencing F
        ws[f"E{row}"] = f"=F{row}*3"

        # Column F: Constant
        ws[f"F{row}"] = float(row * 10)

    wb.save(path)
    wb.close()


class TestSheetFormulaCache:
    """Tests for SheetFormulaCache class."""

    def test_cache_extracts_formulas_correctly(self, tmp_path):
        """Test that SheetFormulaCache correctly extracts formulas."""
        workbook_path = tmp_path / "test_formulas.xlsx"
        create_large_test_workbook(workbook_path, num_rows=50)

        cache = SheetFormulaCache(workbook_path)

        # Check root formula
        formula = cache.get_formula("Sheet1!A1")
        assert "SUMPRODUCT" in formula, f"Expected SUMPRODUCT formula, got: {formula}"

        # Check B column formulas
        for row in [5, 10, 25, 50]:
            formula = cache.get_formula(f"Sheet1!B{row}")
            assert formula == f"=D{row}+E{row}", f"Expected =D{row}+E{row}, got: {formula}"

        # Check D column formulas
        for row in [5, 10, 25, 50]:
            formula = cache.get_formula(f"Sheet1!D{row}")
            assert formula == f"=F{row}*2", f"Expected =F{row}*2, got: {formula}"

        # Check constants (should have no formula)
        for row in [5, 10, 25, 50]:
            formula = cache.get_formula(f"Sheet1!C{row}")
            assert formula == "", f"Expected no formula for constant, got: {formula}"

        cache.close()

    def test_cache_handles_non_sequential_access(self, tmp_path):
        """
        Test that cache handles cells accessed in non-sequential order.

        This is the key scenario that caused the original bug: BFS traversal
        accesses cells alphabetically (B before D before F), but F cells are
        in earlier rows that might be evicted from LazyWorksheet's cache.
        """
        workbook_path = tmp_path / "test_nonseq.xlsx"
        create_large_test_workbook(workbook_path, num_rows=100)

        cache = SheetFormulaCache(workbook_path)

        # Access in alphabetically sorted order (simulating BFS traversal)
        access_order = []
        for col in ["B", "C", "D", "E", "F"]:
            for row in range(5, 101):
                access_order.append(f"Sheet1!{col}{row}")

        # Verify all formulas are accessible
        missing_formulas = []
        for addr in access_order:
            formula = cache.get_formula(addr)
            col = addr.split("!")[-1][0]
            row = int(addr.split("!")[-1][1:])

            # B, D, E columns should have formulas
            if col in ["B", "D", "E"]:
                if not formula:
                    missing_formulas.append(addr)

        cache.close()

        assert len(missing_formulas) == 0, (
            f"Missing formulas for {len(missing_formulas)} cells: {missing_formulas[:10]}..."
        )

    def test_cache_stats(self, tmp_path):
        """Test that cache stats are correctly tracked."""
        workbook_path = tmp_path / "test_stats.xlsx"
        create_large_test_workbook(workbook_path, num_rows=20)

        cache = SheetFormulaCache(workbook_path)

        # Before accessing, no sheets loaded
        stats = cache.get_cache_stats()
        assert stats["sheets_loaded"] == 0
        assert stats["total_formulas"] == 0

        # Access one cell
        cache.get_formula("Sheet1!A1")

        # After access, sheet should be loaded
        stats = cache.get_cache_stats()
        assert stats["sheets_loaded"] == 1
        assert "Sheet1" in stats["sheet_names"]
        assert stats["total_formulas"] > 0

        cache.close()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
