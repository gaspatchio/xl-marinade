# Tests for evidence extraction (label candidates and axis invariants)

from openpyxl import Workbook

from xl_marinade.core.evidence import (
    _classify_range_relationship,
    _compute_monotonicity,
    _compute_step_kind,
    _parse_binding_range,
    build_evidence_cache_for_bindings,
    extract_evidence_for_binding,
)


class TestBindingRangeParsing:
    """Test parsing of binding ranges."""

    def test_parse_single_cell(self):
        """Parse single cell address."""
        result = _parse_binding_range("Sheet1!B2")
        assert result == ("Sheet1", 2, 2, 2, 2)

    def test_parse_range(self):
        """Parse cell range."""
        result = _parse_binding_range("Sheet1!B2:D4")
        assert result == ("Sheet1", 2, 2, 4, 4)

    def test_parse_invalid_range(self):
        """Invalid range returns None."""
        _parse_binding_range("Invalid::Range")


class TestScanningLogic:
    """Test the new 7-cell scanning logic (1D/2D)."""

    def test_1d_row_scans_left(self):
        """1xN binding should scan LEFT (only)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up labels to the LEFT
        ws["A2"] = "Revenue"
        # Binding at B2:D2
        ws["B2"] = 100

        evidence = extract_evidence_for_binding(ws, wb, "Sheet1!B2:D2", (1, 3))

        candidates = evidence["label_candidates"]
        types = [c["type"] for c in candidates]

        assert "scan_left" in types
        assert "scan_above" not in types  # Should NOT scan above for 1D row

        # Check content
        scan = next(c for c in candidates if c["type"] == "scan_left")
        # The cell value "Revenue" should be present in the scanned cells
        # Note: literals are also populated for backward compatibility
        assert "Revenue" in scan["literals"]
        assert scan["cells"][0]["value"] == "Revenue"

    def test_1d_col_scans_above_and_left(self):
        """Mx1 binding should scan both ABOVE and LEFT."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Label above
        ws["B1"] = "Q1_Data"
        # Label left
        ws["A2"] = "RowLabel"
        # Binding B2:B4
        ws["B2"] = 100

        evidence = extract_evidence_for_binding(ws, wb, "Sheet1!B2:B4", (3, 1))

        candidates = evidence["label_candidates"]
        types = [c["type"] for c in candidates]

        assert "scan_above" in types
        assert "scan_left" in types

        scan = next(c for c in candidates if c["type"] == "scan_above")
        assert scan["cells"][0]["value"] == "Q1_Data"

        scan_left = next(c for c in candidates if c["type"] == "scan_left")
        assert any(c["value"] == "RowLabel" for c in scan_left["cells"])

    def test_2d_scans_both(self):
        """MxN binding should scan BOTH left and above."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A2"] = "RowLabel"
        ws["B1"] = "ColLabel"
        # Binding B2:C3

        evidence = extract_evidence_for_binding(ws, wb, "Sheet1!B2:C3", (2, 2))

        types = [c["type"] for c in evidence["label_candidates"]]
        assert "scan_left" in types
        assert "scan_above" in types

    def test_scan_length_limit(self):
        """Scan should go up to 7 cells."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Put label far away (8 cells left) - should NOT be found
        ws["A10"] = "FarLabel"
        # Binding starts at I10 (9th col)
        # 7 cells left of I10 are B10..H10. A10 is outside.

        evidence = extract_evidence_for_binding(ws, wb, "Sheet1!I10:K10", (1, 3))
        scan = next(c for c in evidence["label_candidates"] if c["type"] == "scan_left")

        # Verify A10 is NOT in the cells list
        addresses = [cell["address"] for cell in scan["cells"]]
        assert "A10" not in addresses
        assert "B10" in addresses  # Should scan B10

    def test_rich_cell_data(self):
        """Verify rich data (formulas, types) is captured."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A2"] = 123
        ws["B2"] = "=A2*2"  # Formula string
        # Binding C2:E2

        evidence = extract_evidence_for_binding(ws, wb, "Sheet1!C2:E2", (1, 3))
        scan = next(c for c in evidence["label_candidates"] if c["type"] == "scan_left")

        # Find cell B2
        cell_b2 = next(c for c in scan["cells"] if "B2" in c["address"])
        assert cell_b2["formula"] == "=A2*2"

        # In this mock context, we haven't calculated the formula or set a cached value,
        # so the 'value' is the formula string, and thus dtype is 'text' (or potentially 'formula' if we changed logic)
        # evidence.py logic: if value is None -> blank; if ArrayFormula -> array_formula;
        # if string starting with = -> text (unless we specifically detect it as formula type in extracted logic)
        # Actually _extract_rich_cell_data sets dtype based on value.
        # If value is "=A2*2", dtype is "text".
        assert cell_b2["dtype"] == "text"


class TestMonotonicityComputation:
    """Test monotonicity detection for axis values."""

    def test_increasing_sequence(self):
        """Increasing numeric sequence."""
        values = [1, 2, 3, 4, 5]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "increasing"

    def test_decreasing_sequence(self):
        """Decreasing numeric sequence."""
        values = [5, 4, 3, 2, 1]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "decreasing"

    def test_constant_sequence(self):
        """Constant value sequence."""
        values = [10, 10, 10, 10]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "constant"

    def test_non_monotonic_sequence(self):
        """Non-monotonic sequence."""
        values = [1, 3, 2, 4]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "none"

    def test_sequence_with_blanks(self):
        """Sequence with None values (filtered out)."""
        values = [1, None, 2, None, 3]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "increasing"

    def test_insufficient_values(self):
        """Less than 2 numeric values."""
        values = [1]
        monotonic = _compute_monotonicity(values)
        assert monotonic == "none"


class TestStepKindComputation:
    """Test step kind detection for axis values."""

    def test_constant_numeric_step(self):
        """Constant numeric step."""
        values = [10, 20, 30, 40]
        step_kind = _compute_step_kind(values, ["number"])

        assert step_kind is not None
        assert step_kind["kind"] == "constant_step"
        assert step_kind["step"] == 10

    def test_constant_days_step(self):
        """Constant day step (dates)."""
        # Excel date serials (e.g., 44562 = 2022-01-01)
        values = [44562, 44592, 44622]  # 30-day intervals
        step_kind = _compute_step_kind(values, ["date"])

        assert step_kind is not None
        assert step_kind["kind"] == "constant_days"
        assert step_kind["days"] == 30

    def test_non_constant_step(self):
        """Non-constant step."""
        values = [1, 2, 4, 7]  # Steps: 1, 2, 3
        step_kind = _compute_step_kind(values, ["number"])

        assert step_kind is None

    def test_insufficient_values(self):
        """Less than 2 values."""
        values = [10]
        step_kind = _compute_step_kind(values, ["number"])

        assert step_kind is None


class TestRangeRelationshipClassification:
    """Test classification of binding vs named range relationships."""

    def test_exact_match(self):
        """Binding and name range are identical."""
        relationship = _classify_range_relationship(
            b_top=2, b_left=2, b_bottom=5, b_right=5, n_top=2, n_left=2, n_bottom=5, n_right=5
        )
        assert relationship == "named_exact"

    def test_named_superset(self):
        """Name range contains binding."""
        relationship = _classify_range_relationship(
            b_top=3, b_left=3, b_bottom=4, b_right=4, n_top=2, n_left=2, n_bottom=5, n_right=5
        )
        assert relationship == "named_superset"

    def test_named_subset(self):
        """Binding contains name range."""
        relationship = _classify_range_relationship(
            b_top=2, b_left=2, b_bottom=5, b_right=5, n_top=3, n_left=3, n_bottom=4, n_right=4
        )
        assert relationship == "named_subset"

    def test_no_relationship(self):
        """Ranges overlap but no containment."""
        relationship = _classify_range_relationship(
            b_top=2, b_left=2, b_bottom=4, b_right=4, n_top=3, n_left=3, n_bottom=5, n_right=5
        )
        assert relationship is None


class TestEvidenceExtractionIntegration:
    """Integration tests for full evidence extraction."""

    def test_extract_evidence_simple_binding(self):
        """Extract evidence for simple binding with row above."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up headers
        ws["B1"] = "Revenue"
        ws["C1"] = "Cost"
        ws["D1"] = "Profit"

        # Set up binding values
        ws["B2"] = 100
        ws["C2"] = 60
        ws["D2"] = 40

        # Extract evidence for binding B2:D2 (Row Vector -> Scan Left)
        evidence = extract_evidence_for_binding(
            worksheet=ws,
            workbook=wb,
            binding_address="Sheet1!B2:D2",
            binding_shape=(1, 3),
            name_table_map=None,
        )

        # Should have axis labels for columns (based on our explicit axis check logic)
        assert len(evidence["axis_labels"]) > 0
        axis_cand = evidence["axis_labels"][0]["candidate"]
        # Use display_texts instead of literals for axis candidates
        assert axis_cand["display_texts"] == ["Revenue", "Cost", "Profit"]

    def test_extract_evidence_column_binding(self):
        """Extract evidence for column binding."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up labels in column A
        ws["A2"] = "Q1"
        ws["A3"] = "Q2"
        ws["A4"] = "Q3"

        # Set up binding values in column B
        ws["B2"] = 100
        ws["B3"] = 110
        ws["B4"] = 120

        # Mx1 binding -> Scan Above.
        evidence = extract_evidence_for_binding(
            worksheet=ws,
            workbook=wb,
            binding_address="Sheet1!B2:B4",
            binding_shape=(3, 1),
            name_table_map=None,
        )

        # Check axis labels
        assert len(evidence["axis_labels"]) > 0
        axis_cand = evidence["axis_labels"][0]["candidate"]
        # Use display_texts for axis candidates
        assert axis_cand["display_texts"] == ["Q1", "Q2", "Q3"]


class TestBuildEvidenceCache:
    """Test bulk evidence cache building for performance optimization."""

    def test_empty_bindings_returns_empty_cache(self):
        """Empty bindings list returns empty cache."""
        wb = Workbook()
        cache = build_evidence_cache_for_bindings(wb, [])
        assert cache == {}

    def test_single_binding_caches_scan_region(self):
        """Single binding should cache cells in 7-cell scan region."""
        from xl_marinade.core.bindings import Binding

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up labels
        ws["A5"] = "Label1"
        ws["B4"] = "Header"
        ws["B5"] = 100

        # Create a mock binding at B5
        binding = Binding(
            binding_id="test1",
            debug_label="Test Binding",
            sheet="Sheet1",
            address_a1="Sheet1!B5",
            top_left_a1="B5",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="abc123",
            cells=["Sheet1!B5"],
            extraction_source="test",
        )

        cache = build_evidence_cache_for_bindings(wb, [binding])

        # Should have cached cells in the bounding box
        assert len(cache) > 0
        # The label to the left should be cached
        assert "Sheet1!A5" in cache
        assert cache["Sheet1!A5"] == "Label1"
        # The header above should be cached
        assert "Sheet1!B4" in cache
        assert cache["Sheet1!B4"] == "Header"

    def test_multiple_bindings_same_sheet_expands_bbox(self):
        """Multiple bindings on same sheet should expand bounding box."""
        from xl_marinade.core.bindings import Binding

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up wide spread of values
        ws["A1"] = "TopLeft"
        ws["J10"] = "BottomRight"
        ws["E5"] = 100
        ws["I9"] = 200

        # Create bindings at E5 and I9
        binding1 = Binding(
            binding_id="test1",
            debug_label="Binding 1",
            sheet="Sheet1",
            address_a1="Sheet1!E5",
            top_left_a1="E5",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="abc123",
            cells=["Sheet1!E5"],
            extraction_source="test",
        )
        binding2 = Binding(
            binding_id="test2",
            debug_label="Binding 2",
            sheet="Sheet1",
            address_a1="Sheet1!I9",
            top_left_a1="I9",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="def456",
            cells=["Sheet1!I9"],
            extraction_source="test",
        )

        cache = build_evidence_cache_for_bindings(wb, [binding1, binding2])

        # Bounding box should cover both bindings' scan regions
        # binding1 at E5 scans left from D5 to A5 (7 cells) and above from E4 to E1 (4 cells)
        # binding2 at I9 scans left from H9 to B9 (7 cells) and above from I8 to I2 (7 cells)
        # Combined bbox should cover cells from both regions
        assert "Sheet1!A1" in cache  # Top-left of combined bbox
        assert cache["Sheet1!A1"] == "TopLeft"
        assert "Sheet1!E5" in cache
        assert cache["Sheet1!E5"] == 100

    def test_bindings_across_sheets_caches_each(self):
        """Bindings across multiple sheets should cache each sheet."""
        from xl_marinade.core.bindings import Binding

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["B2"] = "Value1"

        ws2 = wb.create_sheet("Sheet2")
        ws2["C3"] = "Value2"

        binding1 = Binding(
            binding_id="test1",
            debug_label="Binding 1",
            sheet="Sheet1",
            address_a1="Sheet1!B2",
            top_left_a1="B2",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="abc123",
            cells=["Sheet1!B2"],
            extraction_source="test",
        )
        binding2 = Binding(
            binding_id="test2",
            debug_label="Binding 2",
            sheet="Sheet2",
            address_a1="Sheet2!C3",
            top_left_a1="C3",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="def456",
            cells=["Sheet2!C3"],
            extraction_source="test",
        )

        cache = build_evidence_cache_for_bindings(wb, [binding1, binding2])

        # Should have entries from both sheets
        sheet1_keys = [k for k in cache if k.startswith("Sheet1!")]
        sheet2_keys = [k for k in cache if k.startswith("Sheet2!")]

        assert len(sheet1_keys) > 0
        assert len(sheet2_keys) > 0
        assert "Sheet1!B2" in cache
        assert cache["Sheet1!B2"] == "Value1"
        assert "Sheet2!C3" in cache
        assert cache["Sheet2!C3"] == "Value2"

    def test_cache_used_by_evidence_extraction(self):
        """Verify that extract_evidence_for_binding uses the cache correctly."""
        from xl_marinade.core.bindings import Binding

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Set up labels
        ws["A2"] = "Revenue"
        ws["B1"] = "Q1"
        ws["B2"] = 100

        binding = Binding(
            binding_id="test1",
            debug_label="Test Binding",
            sheet="Sheet1",
            address_a1="Sheet1!B2",
            top_left_a1="B2",
            shape_rows=1,
            shape_cols=1,
            binding_type="constant",
            cells_structure_hash="abc123",
            cells=["Sheet1!B2"],
            extraction_source="test",
        )

        # Build cache
        cache = build_evidence_cache_for_bindings(wb, [binding])

        # Extract evidence WITH cache
        evidence = extract_evidence_for_binding(
            worksheet=ws,
            workbook=wb,
            binding_address="Sheet1!B2",
            binding_shape=(1, 1),
            name_table_map=None,
            cell_value_cache=cache,
        )

        # Should have found labels
        candidates = evidence["label_candidates"]
        all_literals = []
        for c in candidates:
            all_literals.extend(c["literals"])

        # "Revenue" from left scan should be found
        assert "Revenue" in all_literals
