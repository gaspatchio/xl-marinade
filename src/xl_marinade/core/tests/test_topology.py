"""Tests for topology module (Story 9)."""

from openpyxl import Workbook

from xl_marinade.core.topology import (
    BoundingBox,
    compute_bounding_box,
    compute_sheet_structure_hash,
    compute_sheet_topologies,
    compute_sheet_topology,
    create_sheet_structure_hash_entries,
    serialize_cell_topology,
)


def test_bounding_box_expand():
    """Test bounding box expansion with margin."""
    bbox = BoundingBox(min_row=3, max_row=5, min_col=2, max_col=4)

    # Expand by margin 5
    expanded = bbox.expand(margin=5)

    assert expanded.min_row == 1  # max(1, 3-5) = 1 (clamped)
    assert expanded.max_row == 10  # 5+5 = 10
    assert expanded.min_col == 1  # max(1, 2-5) = 1 (clamped)
    assert expanded.max_col == 9  # 4+5 = 9


def test_bounding_box_to_a1_range():
    """Test conversion of bounding box to A1 range."""
    bbox = BoundingBox(min_row=1, max_row=10, min_col=1, max_col=6)

    a1_range = bbox.to_a1_range("Sheet1")

    assert a1_range == "Sheet1!A1:F10"


def test_compute_bounding_box():
    """Test bounding box computation from visited cells."""
    visited_cells = ["Sheet1!B3", "Sheet1!D5", "Sheet1!C4"]

    bbox = compute_bounding_box(visited_cells)

    assert bbox is not None
    assert bbox.min_row == 3
    assert bbox.max_row == 5
    assert bbox.min_col == 2  # Column B
    assert bbox.max_col == 4  # Column D


def test_compute_bounding_box_single_cell():
    """Test bounding box for single cell."""
    visited_cells = ["Sheet1!B3"]

    bbox = compute_bounding_box(visited_cells)

    assert bbox is not None
    assert bbox.min_row == 3
    assert bbox.max_row == 3
    assert bbox.min_col == 2
    assert bbox.max_col == 2


def test_compute_bounding_box_empty():
    """Test bounding box for empty list."""
    bbox = compute_bounding_box([])

    assert bbox is None


def test_serialize_cell_topology_blank():
    """Test cell topology serialization for blank cell."""
    wb = Workbook()
    ws = wb.active

    # Blank cell
    cell_data = serialize_cell_topology(ws, 1, 1, name_table_map=None)

    assert cell_data["address"] == "A1"
    assert cell_data["has_formula"] is False
    assert cell_data["dtype"] == "blank"
    assert cell_data["value_snapshot"] is None
    assert cell_data["merged_span"] is None
    assert cell_data["in_table"] is None
    assert cell_data["in_defined_names"] == []


def test_serialize_cell_topology_with_value():
    """Test cell topology serialization for cell with value."""
    wb = Workbook()
    ws = wb.active
    ws["B2"] = 42

    cell_data = serialize_cell_topology(ws, 2, 2, name_table_map=None)

    assert cell_data["address"] == "B2"
    assert cell_data["has_formula"] is False
    assert cell_data["dtype"] == "number"
    assert cell_data["value_snapshot"] == 42


def test_serialize_cell_topology_with_formula():
    """Test cell topology serialization for cell with formula."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["B1"] = "=A1*2"

    cell_data = serialize_cell_topology(ws, 1, 2, name_table_map=None)

    assert cell_data["address"] == "B1"
    assert cell_data["has_formula"] is True
    assert cell_data["dtype"] == "number"  # Default for formulas


def test_compute_sheet_topology():
    """Test sheet topology computation with visited cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some data
    ws["B3"] = 10
    ws["C3"] = 20
    ws["D3"] = 30

    visited_cells = ["Sheet1!B3", "Sheet1!C3", "Sheet1!D3"]

    topology = compute_sheet_topology(ws, visited_cells, margin=2)

    assert topology is not None
    assert "cells" in topology
    assert "regions" in topology

    # Bounding box: B3:D3 → min_row=3, max_row=3, min_col=2, max_col=4
    # Expanded by 2: min_row=1, max_row=5, min_col=0→1 (clamped), max_col=6
    # Result: A1:F5 = 6 cols × 5 rows = 30 cells
    # NOTE: Implementation changed to only store visited cells for performance (see topology.py)
    # So we expect only the 3 visited cells, not the full dense 30 cells
    assert len(topology["cells"]) == 3
    assert set(topology["cells"]) == {"B3", "C3", "D3"}

    # Verify regions structure
    assert "table_regions" in topology["regions"]
    assert "defined_name_regions" in topology["regions"]
    assert "merged_label_regions" in topology["regions"]
    assert "row_nonblank_runs" in topology["regions"]
    assert "col_nonblank_runs" in topology["regions"]


def test_compute_sheet_topology_empty():
    """Test sheet topology with no visited cells."""
    wb = Workbook()
    ws = wb.active

    topology = compute_sheet_topology(ws, [], margin=5)

    assert topology is None


def test_compute_sheet_structure_hash():
    """Test structure hash computation for sheet topology."""
    # Create two identical topologies
    topology1 = {
        "cells": [{"address": "A1", "dtype": "number", "value_snapshot": 10}],
        "regions": {
            "table_regions": [],
            "defined_name_regions": [],
            "merged_label_regions": [],
            "row_nonblank_runs": [],
            "col_nonblank_runs": [],
        },
    }

    topology2 = {
        "cells": [{"address": "A1", "dtype": "number", "value_snapshot": 10}],
        "regions": {
            "table_regions": [],
            "defined_name_regions": [],
            "merged_label_regions": [],
            "row_nonblank_runs": [],
            "col_nonblank_runs": [],
        },
    }

    hash1 = compute_sheet_structure_hash(topology1)
    hash2 = compute_sheet_structure_hash(topology2)

    # Hashes should be identical
    assert hash1 == hash2
    assert len(hash1) == 64  # SHA-256 hex = 64 chars
    assert hash1.islower()  # Lowercase hex


def test_compute_sheet_structure_hash_different():
    """Test structure hash changes when topology changes."""
    topology1 = {
        "cells": [{"address": "A1", "dtype": "number", "value_snapshot": 10}],
        "regions": {},
    }

    topology2 = {
        "cells": [{"address": "A1", "dtype": "number", "value_snapshot": 20}],
        "regions": {},
    }

    hash1 = compute_sheet_structure_hash(topology1)
    hash2 = compute_sheet_structure_hash(topology2)

    # Hashes should be different
    assert hash1 != hash2


def test_compute_sheet_topologies_multiple_sheets():
    """Test topology computation for multiple sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 10

    ws2 = wb.create_sheet("Sheet2")
    ws2["B2"] = 20

    visited_cells_by_sheet = {"Sheet1": ["Sheet1!A1"], "Sheet2": ["Sheet2!B2"]}

    topologies = compute_sheet_topologies(wb, visited_cells_by_sheet, margin=1)

    assert len(topologies) == 2
    assert "Sheet1" in topologies
    assert "Sheet2" in topologies


def test_create_sheet_structure_hash_entries():
    """Test creation of structure hash entries."""
    topologies = {
        "Sheet1": {"cells": [{"address": "A1", "value_snapshot": 10}], "regions": {}},
        "Sheet2": {"cells": [{"address": "B2", "value_snapshot": 20}], "regions": {}},
    }

    entries = create_sheet_structure_hash_entries(topologies)

    # Should have 2 entries (one per sheet)
    assert len(entries) == 2

    # Entries should be sorted by sheet name
    assert entries[0]["hash_key"] == "Sheet1"
    assert entries[1]["hash_key"] == "Sheet2"

    # All should have hash_type="sheets"
    for entry in entries:
        assert entry["hash_type"] == "sheets"
        assert len(entry["hash_value"]) == 64  # SHA-256 hex
        assert entry["hash_value"].islower()


def test_determinism_bounding_box():
    """Test that bounding box computation is deterministic."""
    # Same cells in different order
    cells1 = ["Sheet1!D5", "Sheet1!B3", "Sheet1!C4"]
    cells2 = ["Sheet1!B3", "Sheet1!C4", "Sheet1!D5"]

    bbox1 = compute_bounding_box(cells1)
    bbox2 = compute_bounding_box(cells2)

    assert bbox1 == bbox2


def test_determinism_structure_hash():
    """Test that structure hash is deterministic."""
    # Create topology with unsorted data
    topology = {
        "cells": [{"address": "B2", "value": 20}, {"address": "A1", "value": 10}],
        "regions": {"table_regions": [{"name": "Table2"}, {"name": "Table1"}]},
    }

    # Hash twice
    hash1 = compute_sheet_structure_hash(topology)
    hash2 = compute_sheet_structure_hash(topology)

    # Should be identical
    assert hash1 == hash2


def test_margin_expansion_example():
    """Test margin expansion example from story acceptance criteria."""
    # Visited cells at Sheet1!B3:D5
    visited_cells = []
    for row in range(3, 6):  # Rows 3-5
        for col_letter in ["B", "C", "D"]:
            visited_cells.append(f"Sheet1!{col_letter}{row}")

    bbox = compute_bounding_box(visited_cells)

    # bbox = B3:D5
    assert bbox.min_row == 3
    assert bbox.max_row == 5
    assert bbox.min_col == 2  # B
    assert bbox.max_col == 4  # D

    # Expanded by M=5 → A1:F10 (clamped to min 1)
    expanded = bbox.expand(margin=5)
    assert expanded.min_row == 1  # max(1, 3-5) = 1
    assert expanded.max_row == 10  # 5+5 = 10
    assert expanded.min_col == 1  # max(1, 2-5) = 1 (clamped)
    assert expanded.max_col == 9  # 4+5 = 9

    # Check A1 range (note: spec says A1:F10, but F is column 6, not 9)
    # Actually bbox is B(2):D(4), expanded by 5 → (2-5=min1):(4+5=9)
    # So expanded is col 1 to col 9 (A to I)
    # The spec example may be illustrative, not exact
    # Let's verify the A1 notation
    a1_range = expanded.to_a1_range("Sheet1")
    assert a1_range == "Sheet1!A1:I10"  # A(1) to I(9), rows 1-10
