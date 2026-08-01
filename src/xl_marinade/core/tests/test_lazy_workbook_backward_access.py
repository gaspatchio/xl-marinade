# ABOUTME: Regression tests for LazyWorksheet backward access bug
# ABOUTME: Ensures headers can be read after BFS traversal has moved forward

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from xl_marinade.core.lazy_workbook import LazyWorkbook


class TestLazyWorksheetBackwardAccess:
    """
    Test LazyWorksheet can access cells BEFORE the current iterator position.

    Bug Context: During IR extraction, BFS traversal accesses data cells (row 15+),
    then evidence extraction tries to scan headers (row 13). The streaming iterator
    had already passed row 13, causing it to return empty cells.

    Fix: Added fallback in get_cell() to use iter_rows() when main iterator has
    passed the target row.
    """

    @pytest.fixture
    def test_workbook_path(self):
        """Create a test workbook with headers on row 13, data starting row 15."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Add headers on row 13 (like the real Calculations sheet)
        ws["A13"] = "Column A Header"
        ws["B13"] = "Column B Header"
        ws["C13"] = "Column C Header"

        # Add data starting row 15 (with gap between headers and data)
        for row in range(15, 25):
            ws[f"A{row}"] = f"A{row} value"
            ws[f"B{row}"] = f"B{row} value"
            ws[f"C{row}"] = f"C{row} value"

        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)
            wb.save(temp_path)

        yield temp_path

        # Cleanup
        temp_path.unlink()

    def test_backward_access_after_forward_iteration(self, test_workbook_path):
        """
        Test that accessing row 13 works after iterator has moved to row 15+.

        This reproduces the bug scenario:
        1. BFS traversal accesses rows 15-24 (forward iteration)
        2. Evidence extraction tries to read row 13 (backward access)
        """
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["TestSheet"]

        # Phase 1: Simulate BFS traversal (access rows 15-24)
        for row in range(15, 25):
            cell_a = ws.cell(row, 1)  # Column A
            assert cell_a.value == f"A{row} value", f"Row {row} should have data"

        # At this point, the streaming iterator has moved past row 13
        # This is where the bug occurred - row 13 would return None

        # Phase 2: Simulate evidence extraction (scan row 13 for headers)
        cell_a13 = ws.cell(13, 1)
        cell_b13 = ws.cell(13, 2)
        cell_c13 = ws.cell(13, 3)

        # CRITICAL: These should return the header values, not None
        assert cell_a13.value == "Column A Header", "Should read header after forward iteration"
        assert cell_b13.value == "Column B Header", "Should read header after forward iteration"
        assert cell_c13.value == "Column C Header", "Should read header after forward iteration"

    def test_backward_access_multiple_rows(self, test_workbook_path):
        """
        Test scanning multiple rows backward (like evidence scan_above does).
        """
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["TestSheet"]

        # Access data rows
        for row in range(15, 25):
            ws.cell(row, 2)  # Column B

        # Scan backward from row 14 down to row 8 (7 cells, like evidence extraction)
        scanned_values = []
        for row in [14, 13, 12, 11, 10, 9, 8]:
            cell = ws.cell(row, 2)  # Column B
            scanned_values.append(cell.value)

        # Row 13 should have the header, others should be None/empty
        assert scanned_values[0] is None, "Row 14 is empty"
        assert scanned_values[1] == "Column B Header", "Row 13 has header"
        assert all(v is None for v in scanned_values[2:]), "Rows 8-12 are empty"

    def test_forward_access_still_works(self, test_workbook_path):
        """
        Verify that normal forward access still works efficiently.
        """
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["TestSheet"]

        # Access cells in forward order (normal traversal pattern)
        for row in range(1, 25):
            cell = ws.cell(row, 1)
            if row == 13:
                assert cell.value == "Column A Header"
            elif row >= 15:
                assert cell.value == f"A{row} value"
            else:
                assert cell.value is None

    def test_random_access_after_cache_filled(self, test_workbook_path):
        """
        Test that random access works even after cache is full.

        LazyWorksheet has MAX_CACHE_SIZE = 5000. This test fills the cache
        with forward iteration, then tests backward access.
        """
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["TestSheet"]

        # Fill cache with forward access (rows 15-24, columns A-C = 30 cells)
        for row in range(15, 25):
            for col in range(1, 4):
                ws.cell(row, col)

        # Cache should have ~30 cells now
        assert ws.cache_size > 0

        # Access row 13 (before cached rows)
        cell_a13 = ws.cell(13, 1)
        assert cell_a13.value == "Column A Header"

        # Access row 1 (far before cached rows)
        cell_a1 = ws.cell(1, 1)
        assert cell_a1.value is None  # Empty cell

    def test_interleaved_forward_backward_access(self, test_workbook_path):
        """
        Test interleaved forward and backward access pattern.
        """
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["TestSheet"]

        # Access row 20 (forward)
        assert ws.cell(20, 1).value == "A20 value"

        # Access row 13 (backward)
        assert ws.cell(13, 1).value == "Column A Header"

        # Access row 18 (forward again, but before 20)
        assert ws.cell(18, 1).value == "A18 value"

        # Access row 10 (backward again)
        assert ws.cell(10, 1).value is None

        # Access row 24 (forward)
        assert ws.cell(24, 1).value == "A24 value"


class TestLazyWorksheetGetItemBackwardAccess:
    """Test that __getitem__ (ws['A13']) also handles backward access."""

    @pytest.fixture
    def test_workbook_path(self):
        """Create a test workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A13"] = "Header"
        ws["A15"] = "Data"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_path = Path(f.name)
            wb.save(temp_path)

        yield temp_path
        temp_path.unlink()

    def test_dict_access_backward(self, test_workbook_path):
        """Test ws['A13'] syntax works with backward access."""
        lazy_wb = LazyWorkbook(test_workbook_path, data_only=False)
        ws = lazy_wb["Sheet1"]

        # Access forward
        assert ws["A15"].value == "Data"

        # Access backward
        assert ws["A13"].value == "Header"
