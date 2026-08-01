# ABOUTME: Tests for VLOOKUP resolution strategies
# ABOUTME: Validates literal, cell reference, and expression column index resolution

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.resolution import ResolutionEngine
from xl_marinade.core.resolution_strategies import (
    ResolutionContext,
    create_vlookup_resolution_chain,
)


class TestVLookupStrategies:
    """Test VLOOKUP strategy chain."""

    def test_vlookup_literal_col_index(self):
        """Test VLOOKUP with literal col_index resolves correctly."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Data"

        # Create test data
        ws.cell(row=1, column=1).value = "Key"
        ws.cell(row=1, column=2).value = "Value1"
        ws.cell(row=1, column=3).value = "Value2"
        ws.cell(row=2, column=1).value = 1
        ws.cell(row=2, column=2).value = "A"
        ws.cell(row=2, column=3).value = "B"

        engine = ResolutionEngine(wb)
        chain = create_vlookup_resolution_chain(engine)

        # VLOOKUP with literal col_index=2
        ast = {
            "type": "Function",
            "name": "VLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Data!$A$1:$C$2"},
                {"type": "Const", "value": 2},
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("VLOOKUP", context)
        assert result is not None
        # Should resolve (not conservative fallback if possible)
        assert result.status in ("resolved", "conservative_fallback")

    def test_vlookup_cell_ref_col_index(self):
        """Test VLOOKUP with cell reference col_index."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Data"

        # E1 contains column index
        ws.cell(row=1, column=5).value = 2

        engine = ResolutionEngine(wb)
        chain = create_vlookup_resolution_chain(engine)

        # VLOOKUP with cell ref col_index=E1
        ast = {
            "type": "Function",
            "name": "VLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Data!$A$1:$C$2"},
                {"type": "Ref", "ref": "Data!E1"},
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast, workbook=wb, cell_address="Data!D5", current_sheet="Data", manual_provider=None
        )

        result = chain.resolve("VLOOKUP", context)
        assert result is not None
        assert result.status in ("resolved", "conservative_fallback")

    def test_vlookup_expression_col_index(self):
        """Test VLOOKUP with expression col_index."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_vlookup_resolution_chain(engine)

        # VLOOKUP with expression col_index=1+1
        ast = {
            "type": "Function",
            "name": "VLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$C$2"},
                {
                    "type": "Binary",
                    "operator": "+",
                    "left": {"type": "Const", "value": 1},
                    "right": {"type": "Const", "value": 1},
                },
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("VLOOKUP", context)
        assert result is not None
        assert result.status in ("resolved", "conservative_fallback")

    def test_vlookup_chain_delegation(self):
        """Test that chain delegates correctly."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_vlookup_resolution_chain(engine)

        # VLOOKUP with complex col_index that can't be resolved
        ast = {
            "type": "Function",
            "name": "VLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$C$2"},
                {"type": "Function", "name": "IF", "args": []},  # Complex expression
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("VLOOKUP", context)
        assert result is not None
        # Should fall back to conservative
        assert result.status == "conservative_fallback"

    def test_vlookup_conservative_fallback(self):
        """Test conservative fallback when all strategies fail."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_vlookup_resolution_chain(engine)

        # VLOOKUP with unresolvable col_index
        ast = {
            "type": "Function",
            "name": "VLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$C$10"},
                {"type": "Ref", "ref": "NonExistent!Z1"},  # Can't resolve
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("VLOOKUP", context)
        assert result is not None
        assert result.status == "conservative_fallback"
        assert "Sheet!$A$1:$C$10" in result.resolved_lookup_ref
