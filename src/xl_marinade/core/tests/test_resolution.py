# ABOUTME: Tests for snapshot-only resolution engine
# ABOUTME: Validates volatile/lookup resolution and special data source detection

from openpyxl import Workbook

from xl_marinade.core.parser import parse_formula
from xl_marinade.core.resolution import (
    ArgumentResolutionResult,
    ResolutionEngine,
    create_resolution_engine,
)


class TestVolatileDetection:
    """Test volatile function detection"""

    def test_detect_offset(self):
        """Test OFFSET detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=OFFSET(A1,1,1)")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "OFFSET" in funcs

    def test_detect_indirect(self):
        """Test INDIRECT detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=INDIRECT("A1")')

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "INDIRECT" in funcs

    def test_detect_now(self):
        """Test NOW detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=NOW()")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "NOW" in funcs

    def test_detect_today(self):
        """Test TODAY detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=TODAY()")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "TODAY" in funcs

    def test_detect_rand(self):
        """Test RAND detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=RAND()")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "RAND" in funcs

    def test_detect_multiple_volatile(self):
        """Test multiple volatile functions in one formula"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=OFFSET(A1,1,1)+NOW()")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert is_volatile
        assert "OFFSET" in funcs
        assert "NOW" in funcs

    def test_non_volatile(self):
        """Test non-volatile formula"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=SUM(A1:A10)")

        is_volatile, funcs = engine.detect_volatile(ast)

        assert not is_volatile
        assert len(funcs) == 0


class TestOffsetResolution:
    """Test OFFSET function resolution"""

    def test_offset_single_cell_resolved(self):
        """Test OFFSET with literal offsets resolves to single cell"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("A1", 1, 1, current_sheet="Sheet1")

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet1!B2"
        assert result.volatile_kind == "address_computed"
        assert "A1" in result.volatile_drivers

    def test_offset_with_height_width(self):
        """Test OFFSET with explicit height and width"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("A1", 0, 0, height=3, width=2, current_sheet="Sheet1")

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet1!A1:B3"

    def test_offset_context_dependent(self):
        """Test OFFSET with None parameters is context_dependent"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("A1", None, 1, current_sheet="Sheet1")

        assert result.status == "context_dependent"
        assert "context-dependent" in result.notes.lower()

    def test_offset_out_of_bounds(self):
        """Test OFFSET with negative result"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("A1", -5, 0, current_sheet="Sheet1")

        assert result.status == "unresolved"
        assert "out of bounds" in result.notes.lower()

    def test_offset_with_qualified_base(self):
        """Test OFFSET with sheet-qualified base reference"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("Sheet2!C5", 2, 3)

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet2!F7"


class TestIndirectResolution:
    """Test INDIRECT function resolution"""

    def test_indirect_literal_resolved(self):
        """Test INDIRECT with literal text resolves"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_indirect("A1", current_sheet="Sheet1")

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet1!A1"
        assert result.volatile_kind == "address_computed"

    def test_indirect_qualified_ref(self):
        """Test INDIRECT with sheet-qualified reference"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_indirect("Sheet2!B5")

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet2!B5"

    def test_indirect_context_dependent(self):
        """Test INDIRECT with None argument is context_dependent"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_indirect(None, current_sheet="Sheet1")

        assert result.status == "context_dependent"
        assert "context-dependent" in result.notes.lower()

    def test_indirect_invalid_ref(self):
        """Test INDIRECT with invalid reference text"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_indirect("InvalidRef!@#", current_sheet="Sheet1")

        assert result.status == "unresolved"


class TestIndexMatchResolution:
    """Test INDEX/MATCH resolution"""

    def test_index_match_column_resolved(self):
        """Test INDEX/MATCH with column lookup resolves"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:A10", 3, current_sheet="Sheet1")

        assert result.status == "resolved"
        assert result.resolved_lookup_ref == "Sheet1!A3"
        assert "A1:A10" in result.lookup_drivers

    def test_index_match_row_resolved(self):
        """Test INDEX/MATCH with row lookup resolves"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:E1", 4, current_sheet="Sheet1")

        assert result.status == "resolved"
        assert result.resolved_lookup_ref == "Sheet1!D1"

    def test_index_match_context_dependent(self):
        """Test INDEX/MATCH with None position is context_dependent"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:A10", None, current_sheet="Sheet1")

        assert result.status == "context_dependent"

    def test_index_match_out_of_range(self):
        """Test INDEX/MATCH with position out of range"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:A5", 10, current_sheet="Sheet1")

        assert result.status == "unresolved"
        assert "out of range" in result.notes.lower()

    def test_index_match_2d_array(self):
        """Test INDEX/MATCH with 2D array is context_dependent"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:E10", 3, current_sheet="Sheet1")

        assert result.status == "context_dependent"
        assert "2D array" in result.notes


class TestVolatileKindClassification:
    """Test volatile kind classification"""

    def test_classify_time_functions(self):
        """Test time-based volatile functions"""
        engine = ResolutionEngine(Workbook())

        assert engine.classify_volatile_kind("NOW") == "time"
        assert engine.classify_volatile_kind("TODAY") == "time"

    def test_classify_random_functions(self):
        """Test random volatile functions"""
        engine = ResolutionEngine(Workbook())

        assert engine.classify_volatile_kind("RAND") == "random"
        assert engine.classify_volatile_kind("RANDBETWEEN") == "random"

    def test_classify_workbook_meta(self):
        """Test workbook metadata volatile functions"""
        engine = ResolutionEngine(Workbook())

        assert engine.classify_volatile_kind("CELL") == "workbook_meta"
        assert engine.classify_volatile_kind("INFO") == "workbook_meta"

    def test_classify_external_link(self):
        """Test external link volatile functions"""
        engine = ResolutionEngine(Workbook())

        assert engine.classify_volatile_kind("RTD") == "external_link"
        assert engine.classify_volatile_kind("HYPERLINK") == "external_link"

    def test_classify_address_computed(self):
        """Test address-computing volatile functions"""
        engine = ResolutionEngine(Workbook())

        assert engine.classify_volatile_kind("OFFSET") == "address_computed"
        assert engine.classify_volatile_kind("INDIRECT") == "address_computed"


class TestPivotTableDetection:
    """Test GETPIVOTDATA detection and metadata extraction"""

    def test_detect_getpivotdata(self):
        """Test GETPIVOTDATA detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=GETPIVOTDATA("Sales",A1,"Region","West")')

        info = engine.detect_pivot_table(ast)

        assert "pivot_extract" in info.ref_kinds
        assert info.extras.get("pivot_data_field") == "Sales"
        assert info.extras.get("pivot_table_ref") == "A1"
        assert info.extras.get("pivot_fields") == {"Region": "West"}

    def test_detect_getpivotdata_multiple_fields(self):
        """Test GETPIVOTDATA with multiple field/item pairs"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=GETPIVOTDATA("Revenue",PivotTable1,"Year","2024","Product","Widget")')

        info = engine.detect_pivot_table(ast)

        assert "pivot_extract" in info.ref_kinds
        assert info.extras.get("pivot_fields") == {"Year": "2024", "Product": "Widget"}

    def test_non_pivot_function(self):
        """Test non-pivot function returns empty info"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=SUM(A1:A10)")

        info = engine.detect_pivot_table(ast)

        assert len(info.ref_kinds) == 0
        assert len(info.extras) == 0


class TestCubeQueryDetection:
    """Test cube function detection and metadata extraction"""

    def test_detect_cubevalue(self):
        """Test CUBEVALUE detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=CUBEVALUE("DataModel","[Measures].[Revenue]")')

        info = engine.detect_cube_query(ast)

        assert "cube_query" in info.ref_kinds
        assert info.extras.get("cube_connection") == "DataModel"
        assert info.extras.get("cube_measure") == "[Measures].[Revenue]"

    def test_detect_cubemember(self):
        """Test CUBEMEMBER detection"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=CUBEMEMBER("DataModel","[Product].[Category].[Electronics]")')

        info = engine.detect_cube_query(ast)

        assert "cube_query" in info.ref_kinds
        assert info.extras.get("cube_connection") == "DataModel"

    def test_non_cube_function(self):
        """Test non-cube function returns empty info"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=VLOOKUP(A1,B1:C10,2,FALSE)")

        info = engine.detect_cube_query(ast)

        assert len(info.ref_kinds) == 0


class TestFactoryFunction:
    """Test factory function"""

    def test_create_resolution_engine(self):
        """Test create_resolution_engine factory"""
        wb = Workbook()
        engine = create_resolution_engine(wb)

        assert isinstance(engine, ResolutionEngine)
        assert engine.value_source == wb


class TestDeterminism:
    """Test deterministic behavior"""

    def test_offset_deterministic(self):
        """Test OFFSET resolution is deterministic"""
        engine = ResolutionEngine(Workbook())

        result1 = engine.resolve_offset("A1", 1, 1, current_sheet="Sheet1")
        result2 = engine.resolve_offset("A1", 1, 1, current_sheet="Sheet1")

        assert result1.status == result2.status
        assert result1.resolved_volatile_ref == result2.resolved_volatile_ref

    def test_indirect_deterministic(self):
        """Test INDIRECT resolution is deterministic"""
        engine = ResolutionEngine(Workbook())

        result1 = engine.resolve_indirect("B5", current_sheet="Sheet1")
        result2 = engine.resolve_indirect("B5", current_sheet="Sheet1")

        assert result1.status == result2.status
        assert result1.resolved_volatile_ref == result2.resolved_volatile_ref

    def test_detection_deterministic(self):
        """Test volatile detection is deterministic"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula("=OFFSET(A1,1,1)+NOW()")

        is_volatile1, funcs1 = engine.detect_volatile(ast)
        is_volatile2, funcs2 = engine.detect_volatile(ast)

        assert is_volatile1 == is_volatile2
        assert funcs1 == funcs2


class TestEdgeCases:
    """Test edge cases and error handling"""

    def test_offset_invalid_base(self):
        """Test OFFSET with invalid base reference"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_offset("InvalidRef", 1, 1)

        assert result.status == "unresolved"

    def test_index_match_invalid_array(self):
        """Test INDEX/MATCH with invalid array"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("InvalidRange", 1)

        assert result.status == "unresolved"

    def test_index_match_negative_position(self):
        """Test INDEX/MATCH with negative position"""
        engine = ResolutionEngine(Workbook())

        result = engine.resolve_index_match("A1:A10", -1)

        assert result.status == "unresolved"

    def test_pivot_detection_empty_args(self):
        """Test GETPIVOTDATA with no arguments"""
        engine = ResolutionEngine(Workbook())
        ast = {"type": "Function", "name": "GETPIVOTDATA", "args": []}

        info = engine.detect_pivot_table(ast)

        assert "pivot_extract" in info.ref_kinds
        # Should not crash, just return partial info


class TestConcreteExamples:
    """Test concrete examples from story specification"""

    def test_offset_example(self):
        """Test: =OFFSET(A1,1,1) with A1=Sheet1!B5 → resolves to Sheet1!C6"""
        engine = ResolutionEngine(Workbook())

        # Simulating A1 contains Sheet1!B5 as base reference
        result = engine.resolve_offset("Sheet1!B5", 1, 1)

        assert result.status == "resolved"
        assert result.resolved_volatile_ref == "Sheet1!C6"

    def test_getpivotdata_example(self):
        """Test: =GETPIVOTDATA("Sales",PivotTable1,"Region","West")"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=GETPIVOTDATA("Sales",PivotTable1,"Region","West")')

        info = engine.detect_pivot_table(ast)

        assert "pivot_extract" in info.ref_kinds
        assert info.extras.get("pivot_table_ref") == "PivotTable1"
        assert info.extras.get("pivot_fields") == {"Region": "West"}

    def test_cubevalue_example(self):
        """Test: =CUBEVALUE("DataModel","[Measures].[Revenue]")"""
        engine = ResolutionEngine(Workbook())
        ast = parse_formula('=CUBEVALUE("DataModel","[Measures].[Revenue]")')

        info = engine.detect_cube_query(ast)

        assert "cube_query" in info.ref_kinds
        assert info.extras.get("cube_connection") == "DataModel"
        assert info.extras.get("cube_measure") == "[Measures].[Revenue]"


class TestArgumentResolutionResult:
    """Test ArgumentResolutionResult dataclass"""

    def test_argument_resolution_result_dataclass(self):
        """Test ArgumentResolutionResult dataclass structure"""
        result = ArgumentResolutionResult(
            value=5,
            drivers=("A1",),
            success=True,
            failure_reason=None,
            attempted_strategies=("literal",),
        )

        assert result.value == 5
        assert result.drivers == ("A1",)
        assert result.success is True
        assert result.failure_reason is None
        assert result.attempted_strategies == ("literal",)

    def test_resolve_argument_returns_dataclass(self):
        """Test that _resolve_argument returns ArgumentResolutionResult"""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = 42

        engine = ResolutionEngine(wb)

        # Test with const node
        const_node = {"type": "Const", "value": 10}
        result = engine._resolve_argument(const_node)

        assert isinstance(result, ArgumentResolutionResult)
        assert result.value == 10
        assert result.success is True
        assert result.drivers == ()

        # Test with cell ref node
        ref_node = {"type": "Ref", "ref": "Sheet!A1"}
        result = engine._resolve_argument(ref_node, current_sheet="Sheet")

        assert isinstance(result, ArgumentResolutionResult)
        assert result.value == 42
        assert result.success is True
        assert result.drivers == ("Sheet!A1",)

    def test_failure_reason_populated(self):
        """Test that failure_reason is populated on failure"""
        engine = ResolutionEngine(Workbook())

        # Test with invalid ref (non-existent sheet)
        ref_node = {"type": "Ref", "ref": "NonExistent!A1"}
        result = engine._resolve_argument(ref_node)

        assert result.success is False
        assert result.failure_reason is not None
        assert "Sheet not found" in result.failure_reason

        # Test with unresolvable node type
        unknown_node = {"type": "UnknownType"}
        result = engine._resolve_argument(unknown_node)

        assert result.success is False
        assert result.failure_reason is not None
        assert "Unresolvable node type" in result.failure_reason


class TestMatchScanCacheScope:
    """The MATCH-scan memo must not serve stale positions from a mutable source."""

    def _match_ast(self) -> dict:
        return {
            "type": "Function",
            "name": "MATCH",
            "args": [
                {"type": "Const", "value": 30},
                {"type": "Ref", "ref": "Data!$A$1:$A$3"},
                {"type": "Const", "value": 0},
            ],
        }

    def test_live_workbook_mutation_is_visible(self):
        from openpyxl import Workbook

        from xl_marinade.core.resolution import ResolutionEngine

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        for i, v in enumerate([10, 20, 30], start=1):
            ws.cell(row=i, column=1, value=v)

        engine = ResolutionEngine(wb)
        first = engine.resolve_match_semantic(self._match_ast(), "Data")
        assert first.status == "resolved"
        assert first.resolved_value == 3

        ws.cell(row=1, column=1, value=30)  # 30 now first in the array
        second = engine.resolve_match_semantic(self._match_ast(), "Data")
        assert second.status == "resolved"
        assert second.resolved_value == 1, (
            "stale memoized position served from a mutated live Workbook"
        )

    def test_snapshot_source_scans_are_memoized(self):
        from xl_marinade.core.resolution import ResolutionEngine

        class _Snapshot:
            sheetnames = ["Data"]
            active_sheet = "Data"

            def __init__(self) -> None:
                self.reads = 0

            def get_value_at(self, sheet: str, coord: str):
                self.reads += 1
                return {"A1": 10, "A2": 20, "A3": 30}.get(coord)

        src = _Snapshot()
        engine = ResolutionEngine(src)
        assert engine.resolve_match_semantic(self._match_ast(), "Data").resolved_value == 3
        reads_after_first = src.reads
        assert engine.resolve_match_semantic(self._match_ast(), "Data").resolved_value == 3
        assert src.reads == reads_after_first, "identical scan re-read the snapshot"
