# ABOUTME: Tests for HLOOKUP resolution strategies
# ABOUTME: Validates literal, cell reference, and expression row index resolution

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.resolution import ResolutionEngine
from xl_marinade.core.resolution_strategies import (
    ResolutionContext,
    create_hlookup_resolution_chain,
)


class TestHLookupStrategies:
    """Test HLOOKUP strategy chain."""

    def test_hlookup_literal_row_index(self):
        """Test HLOOKUP with literal row_index resolves correctly."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Data"

        # Create test data (horizontal layout)
        ws.cell(row=1, column=1).value = "Key"
        ws.cell(row=2, column=1).value = "Value1"
        ws.cell(row=3, column=1).value = "Value2"
        ws.cell(row=1, column=2).value = 1
        ws.cell(row=2, column=2).value = "A"
        ws.cell(row=3, column=2).value = "B"

        engine = ResolutionEngine(wb)
        chain = create_hlookup_resolution_chain(engine)

        # HLOOKUP with literal row_index=2
        ast = {
            "type": "Function",
            "name": "HLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Data!$A$1:$B$3"},
                {"type": "Const", "value": 2},
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("HLOOKUP", context)
        assert result is not None
        assert result.status in ("resolved", "conservative_fallback")

    def test_hlookup_cell_ref_row_index(self):
        """Test HLOOKUP with cell reference row_index."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Data"

        # E1 contains row index
        ws.cell(row=1, column=5).value = 2

        engine = ResolutionEngine(wb)
        chain = create_hlookup_resolution_chain(engine)

        # HLOOKUP with cell ref row_index=E1
        ast = {
            "type": "Function",
            "name": "HLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Data!$A$1:$B$3"},
                {"type": "Ref", "ref": "Data!E1"},
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast, workbook=wb, cell_address="Data!D5", current_sheet="Data", manual_provider=None
        )

        result = chain.resolve("HLOOKUP", context)
        assert result is not None
        assert result.status in ("resolved", "conservative_fallback")

    def test_hlookup_expression_row_index(self):
        """Test HLOOKUP with expression row_index."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_hlookup_resolution_chain(engine)

        # HLOOKUP with expression row_index=1+1
        ast = {
            "type": "Function",
            "name": "HLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$B$3"},
                {
                    "type": "Binary",
                    "operator": "+",
                    "left": {"type": "Const", "value": 1},
                    "right": {"type": "Const", "value": 1},
                },
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("HLOOKUP", context)
        assert result is not None
        assert result.status in ("resolved", "conservative_fallback")

    def test_hlookup_chain_delegation(self):
        """Test that chain delegates correctly."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_hlookup_resolution_chain(engine)

        # HLOOKUP with complex row_index that can't be resolved
        ast = {
            "type": "Function",
            "name": "HLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$B$3"},
                {"type": "Function", "name": "IF", "args": []},  # Complex expression
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("HLOOKUP", context)
        assert result is not None
        assert result.status == "conservative_fallback"

    def test_hlookup_conservative_fallback(self):
        """Test conservative fallback when all strategies fail."""
        wb = Workbook()
        engine = ResolutionEngine(wb)
        chain = create_hlookup_resolution_chain(engine)

        # HLOOKUP with unresolvable row_index
        ast = {
            "type": "Function",
            "name": "HLOOKUP",
            "args": [
                {"type": "Ref", "ref": "A1"},
                {"type": "Ref", "ref": "Sheet!$A$1:$B$10"},
                {"type": "Ref", "ref": "NonExistent!Z1"},  # Can't resolve
                {"type": "Const", "value": False},
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet!D5",
            current_sheet="Sheet",
            manual_provider=None,
        )

        result = chain.resolve("HLOOKUP", context)
        assert result is not None
        assert result.status == "conservative_fallback"
        assert "Sheet!$A$1:$B$10" in result.resolved_lookup_ref
