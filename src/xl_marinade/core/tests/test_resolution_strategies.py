# ABOUTME: Tests for resolution strategy framework and INDEX strategies
# ABOUTME: Validates chain delegation, partial resolution, and UC-S3-07 regression

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.resolution import ResolutionEngine, ResolutionResult
from xl_marinade.core.resolution_strategies import (
    ConservativeFallbackStrategy,
    ManualResolutionStrategy,
    ResolutionChain,
    ResolutionContext,
    create_index_resolution_chain,
)
from xl_marinade.core.strategies.index_strategies import (
    IndexPartialColumnStrategy,
)


class TestResolutionChainFramework:
    """Test core strategy chain framework behavior."""

    def test_resolution_chain_delegation(self):
        """Test that chain delegates to strategies in order."""

        # Create a mock strategy that always returns None (delegates)
        class AlwaysDelegateStrategy:
            def can_handle(self, func_name: str) -> bool:
                return True

            def try_resolve(self, context: ResolutionContext) -> ResolutionResult | None:
                return None

        # Create a mock strategy that returns a result
        class SuccessStrategy:
            def can_handle(self, func_name: str) -> bool:
                return True

            def try_resolve(self, context: ResolutionContext) -> ResolutionResult | None:
                return ResolutionResult(status="resolved", notes="Success strategy")

        # Chain with delegate first, then success
        chain = ResolutionChain(
            [
                AlwaysDelegateStrategy(),
                SuccessStrategy(),
            ]
        )

        wb = Workbook()
        context = ResolutionContext(
            ast={"type": "Function", "name": "TEST"},
            workbook=wb,
            cell_address="Sheet1!A1",
            current_sheet="Sheet1",
            manual_provider=None,
        )

        result = chain.resolve("TEST", context)

        # Should get result from second strategy
        assert result.status == "resolved"
        assert "Success strategy" in result.notes

    def test_resolution_chain_no_match(self):
        """Test chain returns unresolved when all strategies return None."""

        class AlwaysDelegateStrategy:
            def can_handle(self, func_name: str) -> bool:
                return True

            def try_resolve(self, context: ResolutionContext) -> ResolutionResult | None:
                return None

        chain = ResolutionChain([AlwaysDelegateStrategy()])

        wb = Workbook()
        context = ResolutionContext(
            ast={"type": "Function", "name": "TEST"},
            workbook=wb,
            cell_address="Sheet1!A1",
            current_sheet="Sheet1",
            manual_provider=None,
        )

        result = chain.resolve("TEST", context)

        # Should return unresolved status
        assert result.status == "unresolved"


class TestManualResolutionStrategy:
    """Test manual resolution strategy."""

    def test_manual_strategy_returns_result_when_override_exists(self):
        """Test manual override is returned when available."""
        from xl_marinade.core.manual_resolution import ManualResolutionProvider

        # Create manual provider with override
        manual_provider = ManualResolutionProvider()
        manual_provider.resolutions = {
            "Sheet1!A1": {"resolved_ref": "Sheet1!B1:B10", "reason": "Test override"}
        }

        strategy = ManualResolutionStrategy(manual_provider)

        wb = Workbook()
        context = ResolutionContext(
            ast={"type": "Function", "name": "INDEX"},
            workbook=wb,
            cell_address="Sheet1!A1",
            current_sheet="Sheet1",
            manual_provider=manual_provider,
        )

        result = strategy.try_resolve(context)

        assert result is not None
        assert result.status == "resolved"
        assert result.resolved_lookup_ref == "Sheet1!B1:B10"
        assert result.resolution_source == "manual"

    def test_manual_strategy_returns_none_when_no_override(self):
        """Test delegation when no manual override exists."""
        from xl_marinade.core.manual_resolution import ManualResolutionProvider

        manual_provider = ManualResolutionProvider()
        manual_provider.resolutions = {}
        strategy = ManualResolutionStrategy(manual_provider)

        wb = Workbook()
        context = ResolutionContext(
            ast={"type": "Function", "name": "INDEX"},
            workbook=wb,
            cell_address="Sheet1!A1",
            current_sheet="Sheet1",
            manual_provider=manual_provider,
        )

        result = strategy.try_resolve(context)

        # Should delegate to next strategy
        assert result is None

    def test_manual_strategy_handles_all_lookup_functions(self):
        """Test manual strategy supports all lookup functions."""
        from xl_marinade.core.manual_resolution import ManualResolutionProvider

        manual_provider = ManualResolutionProvider()
        manual_provider.resolutions = {"Sheet1!A1": {"resolved_ref": "Sheet1!B1"}}

        strategy = ManualResolutionStrategy(manual_provider)

        # Should handle all these functions
        for func_name in ["INDEX", "VLOOKUP", "HLOOKUP", "XLOOKUP", "MATCH", "CHOOSE", "ADDRESS"]:
            assert strategy.can_handle(func_name)


class TestConservativeFallbackStrategy:
    """Test conservative fallback strategy."""

    def test_conservative_fallback_returns_full_table(self):
        """Test fallback returns entire table as dependency."""
        strategy = ConservativeFallbackStrategy()

        wb = Workbook()
        context = ResolutionContext(
            ast={
                "type": "Function",
                "name": "INDEX",
                "args": [
                    {"type": "Ref", "ref": "Sheet1!A1:D10"},
                    {"type": "Const", "value": 5},
                    {"type": "Const", "value": 2},
                ],
            },
            workbook=wb,
            cell_address="Sheet1!Z1",
            current_sheet="Sheet1",
            manual_provider=None,
        )

        result = strategy.try_resolve(context)

        assert result is not None
        assert result.status == "conservative_fallback"
        assert result.resolved_lookup_ref == "Sheet1!A1:D10"
        assert "Sheet1!A1:D10" in result.lookup_drivers


class TestIndexPartialColumnStrategy:
    """Test INDEX partial column resolution strategy."""

    def test_partial_column_resolves_static_match(self):
        """Test partial column resolution for static MATCH."""
        # Create workbook with header row
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Data"

        # Create resolution engine and strategy
        from xl_marinade.core.resolution import ResolutionEngine

        engine = ResolutionEngine(wb)
        strategy = IndexPartialColumnStrategy(engine)

        # Header row with "Payment Frequency" at position 3
        ws.cell(row=1, column=1).value = "Asset Name"
        ws.cell(row=1, column=2).value = "Starting Value"
        ws.cell(row=1, column=3).value = "Payment Frequency"
        ws.cell(row=1, column=4).value = "Regular Payment"

        # Data rows
        for i in range(2, 11):
            ws.cell(row=i, column=1).value = f"Asset {i - 1}"
            ws.cell(row=i, column=2).value = 1000 * (i - 1)
            ws.cell(row=i, column=3).value = "Monthly"
            ws.cell(row=i, column=4).value = 100 * (i - 1)

        # AST for INDEX(Data!A2:D10, MATCH(Sheet2!B5, Data!A2:A10, 0), MATCH("Payment Frequency", Data!A1:D1, 0))
        ast = {
            "type": "Function",
            "name": "INDEX",
            "args": [
                {"type": "Ref", "ref": "Data!$A$2:$D$10"},
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Ref", "ref": "Sheet2!$B5"},
                        {"type": "Ref", "ref": "Data!$A$2:$A$10"},
                        {"type": "Const", "value": 0},
                    ],
                },
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Const", "value": "Payment Frequency"},
                        {"type": "Ref", "ref": "Data!$A$1:$D$1"},
                        {"type": "Const", "value": 0},
                    ],
                },
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet1!C5",
            current_sheet="Sheet1",
            manual_provider=None,
        )

        result = strategy.try_resolve(context)

        assert result is not None
        assert result.status == "partial_resolved"
        # Should resolve to column C (position 3)
        assert "Data!$C$2:$C$10" in result.resolved_lookup_ref
        assert result.partial_info["resolution_level"] == "column_range"
        assert result.partial_info["column_position"] == 3
        assert result.partial_info["row_driver"] == "Sheet2!$B5"

    def test_partial_column_returns_none_for_dynamic_column(self):
        """Test delegation when column is also dynamic."""
        wb = Workbook()

        # Create resolution engine and strategy
        from xl_marinade.core.resolution import ResolutionEngine

        engine = ResolutionEngine(wb)
        strategy = IndexPartialColumnStrategy(engine)

        # Both row and column are dynamic (cell references)
        ast = {
            "type": "Function",
            "name": "INDEX",
            "args": [
                {"type": "Ref", "ref": "Data!$A$2:$D$10"},
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Ref", "ref": "Sheet2!$B5"},
                        {"type": "Ref", "ref": "Data!$A$2:$A$10"},
                        {"type": "Const", "value": 0},
                    ],
                },
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Ref", "ref": "Sheet2!$C5"},  # Dynamic column too!
                        {"type": "Ref", "ref": "Data!$A$1:$D$1"},
                        {"type": "Const", "value": 0},
                    ],
                },
            ],
        }

        context = ResolutionContext(
            ast=ast,
            workbook=wb,
            cell_address="Sheet1!C5",
            current_sheet="Sheet1",
            manual_provider=None,
        )

        result = strategy.try_resolve(context)

        # Should delegate - column is dynamic
        assert result is None


class TestIndexChainIntegration:
    """Test complete INDEX resolution chain."""

    def test_index_chain_resolves_all_patterns(self):
        """Test that chain handles different INDEX patterns."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "Test"

        # Simple data
        ws.cell(row=1, column=1).value = "Header1"
        ws.cell(row=1, column=2).value = "Header2"
        ws.cell(row=2, column=1).value = 10
        ws.cell(row=2, column=2).value = 20

        engine = ResolutionEngine(wb, manual_provider=None)
        chain = create_index_resolution_chain(engine, manual_provider=None)

        # Test case: Static column MATCH, dynamic row
        ast = {
            "type": "Function",
            "name": "INDEX",
            "args": [
                {"type": "Ref", "ref": "Test!$A$2:$B$10"},
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Ref", "ref": "Test!$C1"},
                        {"type": "Ref", "ref": "Test!$A$2:$A$10"},
                        {"type": "Const", "value": 0},
                    ],
                },
                {
                    "type": "Function",
                    "name": "MATCH",
                    "args": [
                        {"type": "Const", "value": "Header2"},
                        {"type": "Ref", "ref": "Test!$A$1:$B$1"},
                        {"type": "Const", "value": 0},
                    ],
                },
            ],
        }

        context = ResolutionContext(
            ast=ast, workbook=wb, cell_address="Test!D5", current_sheet="Test", manual_provider=None
        )

        result = chain.resolve("INDEX", context)

        # Should NOT be unresolved - at least partial or conservative fallback
        assert result.status in ("resolved", "partial_resolved", "conservative_fallback")
        assert result.resolved_lookup_ref is not None


class TestBackwardCompatibility:
    """Test that existing code still works."""

    def test_direct_resolve_index_semantic_still_works(self):
        """Verify direct method calls (without chain) still work."""
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.cell(row=1, column=1).value = 10
        ws.cell(row=1, column=2).value = 20

        engine = ResolutionEngine(wb)

        # Direct call (old API) should still work
        ast = {
            "type": "Function",
            "name": "INDEX",
            "args": [
                {"type": "Ref", "ref": "Sheet!$A$1:$B$1"},
                {"type": "Const", "value": 1},
                {"type": "Const", "value": 2},
            ],
        }

        result = engine.resolve_index_semantic(ast=ast, current_sheet="Sheet")

        # Should return valid result
        assert result is not None
        assert hasattr(result, "status")


# Run tests with: pytest xl_marinade.core/tests/test_resolution_strategies.py -v
