# ABOUTME: Sheet-level formula cache for reliable formula extraction from large workbooks.
# ABOUTME: Pre-loads ALL formulas from a sheet on first access, avoiding streaming iterator issues.

from pathlib import Path

from openpyxl import load_workbook

from xl_marinade.core.formula_utils import extract_formula_string


def strip_sheet_quotes(sheet_name: str) -> str:
    """Remove surrounding single quotes from sheet name if present."""
    if sheet_name.startswith("'") and sheet_name.endswith("'"):
        return sheet_name[1:-1]
    return sheet_name


class SheetFormulaCache:
    """
    Sheet-level formula cache for reliable formula extraction from large workbooks.

    Architecture Rationale:
    -----------------------
    LazyWorksheet uses a streaming iterator + LRU cache (5000 cells). When traversing
    large models (27K+ cells) with non-sequential access patterns (BFS sorts cells
    alphabetically: AL before X), the cache fills up and evicts older cells. Since
    the streaming iterator can't rewind, those cells become inaccessible.

    Solution: Pre-load ALL formulas from a sheet when first accessed, similar to
    how LazyValueFetcher pre-loads values. This gives:
    - O(n) initial scan per sheet
    - O(1) subsequent lookups
    - Predictable memory (bounded by sheets visited)
    - No cache eviction issues

    This is consistent with the dual-load strategy used elsewhere in the pipeline.
    """

    def __init__(self, workbook_path: Path | str):
        """
        Initialize formula cache.

        Args:
            workbook_path: Path to Excel workbook
        """
        self.workbook_path = Path(workbook_path)
        self._wb = None

        # Cache: {sheet_name: {cell_coord: formula}}
        # Sheets are loaded lazily on first access
        self._sheet_formulas: dict[str, dict[str, str]] = {}
        self._loaded_sheets: set[str] = set()

    def _ensure_workbook_loaded(self):
        """Lazy-load workbook on first access."""
        if self._wb is None:
            self._wb = load_workbook(
                self.workbook_path,
                data_only=False,  # We need formulas
                read_only=True,  # Read-only for efficiency
                keep_vba=False,  # Don't need VBA
            )

    def _load_sheet_formulas(self, sheet_name: str) -> None:
        """
        Pre-load ALL formulas from a sheet into cache.

        This scans the entire sheet once and caches all formulas.
        Subsequent lookups are O(1).

        Args:
            sheet_name: Name of the sheet to load
        """
        if sheet_name in self._loaded_sheets:
            return  # Already loaded

        self._ensure_workbook_loaded()

        if sheet_name not in self._wb.sheetnames:
            self._loaded_sheets.add(sheet_name)  # Mark as "loaded" (empty)
            self._sheet_formulas[sheet_name] = {}
            return

        ws = self._wb[sheet_name]
        formulas: dict[str, str] = {}

        # Scan entire sheet, collecting formulas
        # iter_rows() creates a fresh iterator, so no cache issues.
        # Use bounds to avoid iterating through 1M+ rows for sheets with large dimensions.
        # This is the same fix as _get_populated_subrange() in grouping_native.py.
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell is None:
                    continue

                # Check if cell has a formula
                if cell.data_type == "f" and cell.value:
                    try:
                        formula = extract_formula_string(cell)
                        if formula:
                            # Store by coordinate (e.g., "A1", "X15")
                            coord = cell.coordinate
                            formulas[coord] = formula
                    except Exception:
                        pass  # Skip cells with extraction errors

        self._sheet_formulas[sheet_name] = formulas
        self._loaded_sheets.add(sheet_name)

    def get_formula(self, cell_address: str) -> str:
        """
        Get formula for a cell address.

        Args:
            cell_address: Full cell address (e.g., "Sheet1!A1")

        Returns:
            Formula string if cell has formula, empty string otherwise
        """
        try:
            # Parse address
            if "!" not in cell_address:
                return ""

            sheet_name, cell_coord = cell_address.split("!", 1)
            sheet_name = strip_sheet_quotes(sheet_name)

            # Ensure sheet is loaded into cache
            self._load_sheet_formulas(sheet_name)

            # O(1) lookup from cache
            return self._sheet_formulas.get(sheet_name, {}).get(cell_coord, "")

        except (KeyError, AttributeError, ValueError):
            return ""

    def is_sheet_loaded(self, sheet_name: str) -> bool:
        """Check if a sheet has been loaded into cache."""
        return strip_sheet_quotes(sheet_name) in self._loaded_sheets

    def get_cache_stats(self) -> dict:
        """Get statistics about the cache for debugging."""
        return {
            "sheets_loaded": len(self._loaded_sheets),
            "sheet_names": list(self._loaded_sheets),
            "total_formulas": sum(len(f) for f in self._sheet_formulas.values()),
            "formulas_per_sheet": {
                name: len(formulas) for name, formulas in self._sheet_formulas.items()
            },
        }

    def close(self):
        """Close the workbook and clear cache."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None
        self._sheet_formulas.clear()
        self._loaded_sheets.clear()

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()


# Backwards compatibility alias
LazyFormulaFetcher = SheetFormulaCache
