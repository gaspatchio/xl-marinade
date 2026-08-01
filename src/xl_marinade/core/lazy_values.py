# ABOUTME: Lazy value fetcher for memory-efficient evaluated value access.
# ABOUTME: Loads sheet values on-demand instead of entire workbook upfront.

import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class LazyValueFetcher:
    """On-demand cell value loader with sheet-level caching."""

    DEFAULT_MAX_CACHED_SHEETS: int = 10  # Default maximum sheets to cache (increased from 3)

    def __init__(self, workbook_path: Path, max_cached_sheets: int | None = None):
        """Initialize lazy value fetcher.

        Args:
            workbook_path: Path to Excel workbook
            max_cached_sheets: Maximum sheets to cache (None = unlimited, 0 = no caching)
        """
        self.workbook_path = workbook_path
        self._sheet_caches: dict[str, dict[str, Any]] = {}
        self._loaded_sheets: set[str] = set()
        self._sheet_order: list[str] = []  # Track load order for FIFO eviction
        self._wb: Any = None  # Cached workbook instance for value fetching

        # Pre-load metadata (separate workbook instance - closes immediately)
        try:
            wb = load_workbook(workbook_path, read_only=True)
            try:
                self.sheetnames = list(wb.sheetnames)
                self.active_sheet = wb.active.title if wb.active else None
            finally:
                wb.close()
        except (OSError, PermissionError, zipfile.BadZipFile) as e:
            print(f"Warning: Failed to load workbook metadata: {e}", file=sys.stderr)
            self.sheetnames = []
            self.active_sheet = None

        # Set max_cached_sheets: None = unlimited, explicit value = use it, default = DEFAULT_MAX_CACHED_SHEETS
        if max_cached_sheets is None:
            # Unlimited caching (no eviction)
            self.max_cached_sheets = None
        else:
            self.max_cached_sheets = max(0, max_cached_sheets)  # Ensure non-negative

    def get_value(self, cell_addr: str) -> Any:
        """Get evaluated value for cell, loading sheet if needed.

        Args:
            cell_addr: Cell address (e.g., "Sheet1!A1")

        Returns:
            Evaluated value or None if not found
        """
        sheet, coord = self._parse_address(cell_addr)
        if not sheet:
            return None

        return self.get_value_at(sheet, coord)

    def get_value_at(self, sheet: str, coord: str) -> Any:
        """Get evaluated value for cell by sheet and coordinate.

        Args:
            sheet: Sheet name
            coord: Cell coordinate (e.g., "A1")

        Returns:
            Evaluated value or None if not found
        """
        if sheet not in self._loaded_sheets:
            self._load_sheet_values(sheet)

        return self._sheet_caches.get(sheet, {}).get(coord)

    def _parse_address(self, cell_addr: str) -> tuple[str | None, str]:
        """Parse cell address into sheet and coordinate.

        Args:
            cell_addr: Full cell address (e.g., "'Sheet 1'!A1" or "Sheet1!A1")

        Returns:
            Tuple of (sheet_name, coordinate). sheet_name is None if no sheet part.
        """
        if "!" not in cell_addr:
            return None, cell_addr

        sheet_part, coord = cell_addr.rsplit("!", 1)

        # Handle quoted sheet names
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_part = sheet_part[1:-1]

        return sheet_part, coord

    def _ensure_workbook_open(self) -> bool:
        """Ensure workbook is open for value fetching. Returns True if available."""
        if self._wb is not None:
            return True
        try:
            self._wb = load_workbook(self.workbook_path, data_only=True, read_only=True)
            return True
        except (OSError, PermissionError, zipfile.BadZipFile) as e:
            print(f"Warning: Failed to open workbook for values: {e}", file=sys.stderr)
            return False

    def _load_sheet_values(self, sheet_name: str) -> None:
        """Load values for a single sheet into cache.

        Uses read-only mode for memory efficiency.
        Evicts oldest sheet if at capacity before loading.
        Reuses a single workbook instance for all sheet loads.

        Args:
            sheet_name: Name of sheet to load
        """
        # Evict oldest sheet if at limit (only if max_cached_sheets is set)
        if self.max_cached_sheets is not None and len(self._sheet_order) >= self.max_cached_sheets:
            oldest = self._sheet_order[0]  # Don't pop yet, evict_sheet does that
            self.evict_sheet(oldest)

        try:
            # Reuse cached workbook instance (avoids repeated load_workbook calls)
            if not self._ensure_workbook_open():
                self._sheet_caches[sheet_name] = {}
                self._loaded_sheets.add(sheet_name)
                self._sheet_order.append(sheet_name)
                return

            wb = self._wb

            if sheet_name not in wb.sheetnames:
                self._sheet_caches[sheet_name] = {}
                self._loaded_sheets.add(sheet_name)
                self._sheet_order.append(sheet_name)
                return

            ws = wb[sheet_name]
            cache = {}

            # Use iter_rows with bounds for efficiency.
            # Without bounds, iter_rows() iterates through the entire bounding box
            # which can be 1M+ rows for sheets with large dimensions.
            # This is the same fix as _get_populated_subrange() in grouping_native.py.
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1

            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    # We only cache non-None values to save memory
                    if cell.value is not None:
                        cache[cell.coordinate] = cell.value

            self._sheet_caches[sheet_name] = cache
            self._loaded_sheets.add(sheet_name)
            self._sheet_order.append(sheet_name)

        except (OSError, PermissionError, zipfile.BadZipFile, KeyError) as e:
            print(f"Warning: Failed to load values for {sheet_name}: {e}", file=sys.stderr)
            self._sheet_caches[sheet_name] = {}
            self._loaded_sheets.add(sheet_name)
            self._sheet_order.append(sheet_name)

    def get_sheet_values(self, sheet_name: str) -> dict[str, Any]:
        """Get all values for a sheet, loading if needed.

        Args:
            sheet_name: Name of sheet

        Returns:
            Dictionary of {coordinate: value}
        """
        if sheet_name not in self._loaded_sheets:
            self._load_sheet_values(sheet_name)
        return self._sheet_caches.get(sheet_name, {})

    def evict_sheet(self, sheet_name: str) -> None:
        """Remove sheet from cache to free memory.

        Sheet will be reloaded on next access.

        Args:
            sheet_name: Name of sheet to evict
        """
        self._sheet_caches.pop(sheet_name, None)
        self._loaded_sheets.discard(sheet_name)
        if sheet_name in self._sheet_order:
            self._sheet_order.remove(sheet_name)

    def close(self) -> None:
        """Clear caches and release resources."""
        self._sheet_caches.clear()
        self._loaded_sheets.clear()
        self._sheet_order.clear()
        if self._wb is not None:
            try:
                self._wb.close()
            except Exception:
                pass
            self._wb = None

    def __enter__(self) -> "LazyValueFetcher":
        """Context manager entry."""
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Context manager exit - close workbook."""
        self.close()
