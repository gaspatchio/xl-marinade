# ABOUTME: Cell serialization and edge writing for IR extraction.
# ABOUTME: Handles cell records per IR Spec §3.2, §9 including dynamic
# ABOUTME: arrays, ref_kinds, and extras.

import json
import sqlite3
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.formula_utils import extract_formula_string
from xl_marinade.core.parser import parse_formula
from xl_marinade.core.ref_converter import convert_formula_to_r1c1, parse_cell_address
from xl_marinade.core.ref_extractor import extract_references_from_ast
from xl_marinade.core.resolution import ResolutionEngine
from xl_marinade.core.vba_parser import (
    UDFMetadata,
    detect_udf_calls_in_formula,
    is_formula_volatile_due_to_udfs,
)

if TYPE_CHECKING:
    from xl_marinade.core.bindings import Binding, BindingEdge
    from xl_marinade.core.levels_cycles import Cycle, LevelAssignment


def _evaluate_ast_node(
    node: dict[str, Any] | None, workbook: Workbook, sheet_name: str
) -> Any | None:
    """
    Evaluate AST node to literal value or None if context-dependent.

    Handles:
    - Number nodes → return numeric value
    - String nodes → return string value
    - Ref nodes → look up cell value, return if literal, None if formula
    - Other nodes → return None (not evaluable from snapshot)

    Args:
        node: AST node dict from parser
        workbook: openpyxl Workbook for cell lookups
        sheet_name: Current sheet name for relative refs

    Returns:
        Literal value (int, float, str) or None if context-dependent
    """
    if not node or not isinstance(node, dict):
        return None

    node_type = node.get("type")

    if node_type == "Const":
        return node.get("value")

    if node_type == "Ref":
        # Look up cell reference
        ref = node.get("ref", "")
        # Parse sheet-qualified ref (e.g., "Sheet1!A1" or "A1")
        if "!" in ref:
            ref_sheet, ref_coord = ref.split("!", 1)
        else:
            ref_sheet = sheet_name
            ref_coord = ref

        try:
            ws = workbook[ref_sheet]
            cell = ws[ref_coord]

            # If cell has formula, it's context-dependent
            if cell.data_type == "f":
                return None

            # Return literal value
            return cell.value
        except (KeyError, AttributeError):
            # Cell doesn't exist or invalid ref
            return None

    # Other node types (Binary, Function, etc.) not evaluable from snapshot
    return None


def _find_function_node(ast: dict[str, Any] | None, func_name: str) -> dict[str, Any] | None:
    """Recursively find function node by name in AST"""
    if not isinstance(ast, dict):
        return None

    if ast.get("type") == "Function" and ast.get("name", "").upper() == func_name.upper():
        return ast

    # Recurse into children
    for key in ("left", "right", "operand"):
        child = ast.get(key)
        if child:
            result = _find_function_node(child, func_name)
            if result:
                return result

    # Recurse into function args
    for arg in ast.get("args", []):
        result = _find_function_node(arg, func_name)
        if result:
            return result

    return None


def _extract_offset_args(
    ast: dict[str, Any], workbook: Workbook, sheet_name: str
) -> tuple[str, int | None, int | None, int | None, int | None]:
    """
    Extract OFFSET function arguments from AST.

    OFFSET(base_ref, rows, cols, [height], [width])

    Returns evaluated arguments or None for context-dependent parameters.

    Args:
        ast: AST node for entire formula (will find OFFSET function)
        workbook: openpyxl Workbook for evaluating cell references
        sheet_name: Current sheet name for relative refs

    Returns:
        Tuple of (base_ref, row_offset, col_offset, height, width)
        - base_ref: String (A1 notation)
        - Others: int or None if context-dependent
    """
    # Find OFFSET function node
    offset_node = _find_function_node(ast, "OFFSET")
    if not offset_node:
        return ("", None, None, None, None)

    args = offset_node.get("args", [])
    if len(args) < 3:
        return ("", None, None, None, None)

    # Extract base_ref (must be a Ref node)
    base_node = args[0]
    if base_node.get("type") != "Ref":
        return ("", None, None, None, None)

    base_ref = base_node.get("ref", "")
    if "!" not in base_ref:
        base_ref = f"{sheet_name}!{base_ref}"

    # Extract row_offset (evaluate to int or None)
    row_val = _evaluate_ast_node(args[1], workbook, sheet_name)
    row_offset = int(row_val) if isinstance(row_val, (int, float)) else None

    # Extract col_offset (evaluate to int or None)
    col_val = _evaluate_ast_node(args[2], workbook, sheet_name)
    col_offset = int(col_val) if isinstance(col_val, (int, float)) else None

    # Extract height (optional, evaluate to int or None)
    height = None
    if len(args) > 3:
        height_val = _evaluate_ast_node(args[3], workbook, sheet_name)
        height = int(height_val) if isinstance(height_val, (int, float)) else None

    # Extract width (optional, evaluate to int or None)
    width = None
    if len(args) > 4:
        width_val = _evaluate_ast_node(args[4], workbook, sheet_name)
        width = int(width_val) if isinstance(width_val, (int, float)) else None

    return (base_ref, row_offset, col_offset, height, width)


def _extract_indirect_args(ast: dict[str, Any], workbook: Workbook, sheet_name: str) -> str | None:
    """
    Extract INDIRECT function argument from AST.

    INDIRECT(ref_text)

    Returns evaluated address string or None if context-dependent.

    Args:
        ast: AST node for entire formula (will find INDIRECT function)
        workbook: openpyxl Workbook for evaluating cell references
        sheet_name: Current sheet name for relative refs

    Returns:
        Address string (e.g., "B2", "Sheet1!A1") or None if context-dependent
    """
    # Find INDIRECT function node
    indirect_node = _find_function_node(ast, "INDIRECT")
    if not indirect_node:
        return None

    args = indirect_node.get("args", [])
    if len(args) < 1:
        return None

    # Evaluate first argument to string
    address_val = _evaluate_ast_node(args[0], workbook, sheet_name)

    if isinstance(address_val, str):
        return address_val

    return None


@dataclass
class CellRecord:
    """
    Complete cell record per IR Spec §3.2.

    Attributes:
        cell_address_a1: Cell address in A1 notation (e.g., "Sheet1!B5")
        formula_a1: Formula in A1 notation (empty string if value-only)
        formula_r1c1: Formula in R1C1 notation (empty string if value-only)
        ast_structural: AST dictionary (None if no formula)
        dtype: Data type classification
        value_snapshot: Snapshot value (JSON-compatible)
        evaluated_value: Actual evaluated result from Excel (from data_only=True load)
        format_tokens: Format token dictionary
        udf_calls: List of UDF call dictionaries
        protection: Protection dictionary {locked, hidden_formula}
        spilled_from: A1 address of spiller cell (None if not spilled)
        resolution: Resolution metadata dictionary
        ref_kinds: List of reference kind strings
        extras: Extras metadata dictionary
    """

    cell_address_a1: str
    formula_a1: str
    formula_r1c1: str
    ast_structural: dict[str, Any] | None
    dtype: str
    value_snapshot: Any
    evaluated_value: Any | None
    format_tokens: dict[str, Any]
    udf_calls: list[dict[str, str]]
    protection: dict[str, bool]
    spilled_from: str | None
    resolution: dict[str, Any]
    ref_kinds: list[str]
    broken_refs: list[str]
    extras: dict[str, Any]


def classify_dtype(cell: Cell) -> str:
    """
    Classify cell data type per IR Spec §3.2.

    Args:
        cell: openpyxl Cell object

    Returns:
        dtype classification: "number"|"date"|"text"|"boolean"|"blank"|"error"|"unknown"
    """
    if cell.value is None:
        return "blank"

    # Check for error
    if isinstance(cell.value, str) and cell.value.startswith("#"):
        return "error"

    # Check for boolean
    if isinstance(cell.value, bool):
        return "boolean"

    # Check for number
    if isinstance(cell.value, int | float):
        # Check if it's a date (basic heuristic based on number format)
        if cell.number_format and (
            "d" in cell.number_format.lower()
            or "m" in cell.number_format.lower()
            or "y" in cell.number_format.lower()
        ):
            return "date"
        return "number"

    # Check for text
    if isinstance(cell.value, str):
        return "text"

    return "unknown"


def extract_value_snapshot(cell: Cell) -> Any:
    """
    Extract value snapshot for cell per IR Spec §3.2.

    Returns JSON-compatible value:
    - Numbers: float (with dot-decimal)
    - Strings: string
    - Booleans: boolean
    - Blank: null
    - Errors: error string

    Args:
        cell: openpyxl Cell object

    Returns:
        JSON-compatible value
    """
    if cell.value is None:
        return None

    # Ensure numbers are float for JSON compatibility
    if isinstance(cell.value, int | float):
        return float(cell.value)

    # Return strings, booleans as-is
    if isinstance(cell.value, str | bool):
        return cell.value

    # Fallback: convert to string
    return str(cell.value)


def extract_format_tokens(cell: Cell) -> dict[str, Any]:
    """
    Extract format tokens from cell per IR Spec §3.2.

    Returns dictionary with:
    - kind: "number"|"currency"|"percent"|"date"|"text"|"custom"
    - currency: currency symbol/code (if applicable)
    - percent: boolean (if percent format)
    - thousands_sep: boolean (if thousands separator)
    - decimals: int (number of decimal places, if applicable)
    - date_code: string (date format code fragments, if applicable)

    Args:
        cell: openpyxl Cell object

    Returns:
        Format tokens dictionary
    """
    tokens: dict[str, Any] = {
        "kind": "number",
        "currency": None,
        "percent": False,
        "thousands_sep": False,
        "decimals": 0,
        "date_code": None,
    }

    # Get number format
    number_format = cell.number_format or "General"

    # Classify kind
    if "%" in number_format:
        tokens["kind"] = "percent"
        tokens["percent"] = True
    elif any(c in number_format.lower() for c in ["d", "m", "y", "h", "s"]):
        tokens["kind"] = "date"
        tokens["date_code"] = number_format
    elif "$" in number_format or "€" in number_format or "£" in number_format:
        tokens["kind"] = "currency"
        # Extract currency symbol
        for sym in ["$", "€", "£", "¥"]:
            if sym in number_format:
                tokens["currency"] = sym
                break
    elif "@" in number_format:
        tokens["kind"] = "text"
    elif number_format != "General":
        tokens["kind"] = "custom"

    # Check for thousands separator
    if "," in number_format or "#,##0" in number_format:
        tokens["thousands_sep"] = True

    # Extract decimal places (count zeros after decimal point)
    if "." in number_format:
        parts = number_format.split(".")
        if len(parts) > 1:
            # Count consecutive zeros/hashes after decimal
            decimal_part = parts[1]
            decimals = 0
            for char in decimal_part:
                if char in ("0", "#"):
                    decimals += 1
                elif char in (";", "]", ")"):
                    break
            tokens["decimals"] = decimals

    return tokens


def extract_protection(cell: Cell) -> dict[str, bool]:
    """
    Extract protection settings from cell per IR Spec §3.2.

    Args:
        cell: openpyxl Cell object

    Returns:
        Protection dictionary: {locked: bool, hidden_formula: bool}
    """
    protection = {"locked": False, "hidden_formula": False}

    if cell.protection:
        protection["locked"] = bool(cell.protection.locked)
        protection["hidden_formula"] = bool(cell.protection.hidden)

    return protection


def classify_ref_kinds(
    cell: Cell,
    ast: dict[str, Any] | None,
    resolution: dict[str, Any],
    special_data_sources: dict[str, Any],
) -> list[str]:
    """
    Classify cell reference kinds per IR Spec §3.2.

    Possible kinds:
    - "cell_reference": Normal cell/range reference
    - "dynamic_resolved": Volatile function resolved
    - "dynamic_unresolved": Volatile function unresolved
    - "pivot_extract": GETPIVOTDATA
    - "query_table": Query table reference
    - "cube_query": Cube function
    - "whatif_table": Data table
    - "spill_owner": Dynamic array spiller
    - "external_ref": External workbook reference

    Args:
        cell: openpyxl Cell object
        ast: AST dictionary (None if no formula)
        resolution: Resolution metadata
        special_data_sources: Special data source info from resolution

    Returns:
        List of reference kind strings (sorted)
    """
    kinds = set()

    if not ast:
        # Value-only cell - no ref_kinds
        return []

    # Check for spill owner (dynamic array)
    if (
        hasattr(cell, "data_type")
        and cell.data_type == "f"
        and cell.value
        and str(cell.value).startswith("=")
        and special_data_sources.get("is_spiller")
    ):
        kinds.add("spill_owner")

    # Check for volatile functions
    if resolution.get("is_volatile"):
        if resolution.get("status") == "resolved":
            kinds.add("dynamic_resolved")
        elif resolution.get("status") == "context_dependent":
            kinds.add("dynamic_unresolved")

    # Check for special data sources
    for kind in special_data_sources.get("ref_kinds", []):
        kinds.add(kind)

    # Check for external references
    if _has_external_ref(ast):
        kinds.add("external_ref")

    # Default: normal cell reference (if has references but no special kinds)
    if not kinds and _has_cell_ref(ast):
        kinds.add("cell_reference")

    # Sort for determinism
    return sorted(kinds)


def _has_cell_ref(node: dict[str, Any] | None) -> bool:
    """Check if AST contains cell/range references"""
    if not node or not isinstance(node, dict):
        return False

    if node.get("type") == "Ref":
        return True

    # Recurse
    for key in ("left", "right", "operand"):
        if _has_cell_ref(node.get(key)):
            return True

    # Check function args
    return any(_has_cell_ref(arg) for arg in node.get("args", []))


def _has_external_ref(node: dict[str, Any] | None) -> bool:
    """Check if AST contains external workbook references"""
    if not node or not isinstance(node, dict):
        return False

    if node.get("type") == "Ref":
        ref = node.get("ref", "")
        # External ref starts with [workbook]
        if ref.startswith("["):
            return True

    # Recurse
    for key in ("left", "right", "operand"):
        if _has_external_ref(node.get(key)):
            return True

    # Check function args
    return any(_has_external_ref(arg) for arg in node.get("args", []))


def _find_broken_refs(
    ast: dict[str, Any] | None, workbook: Workbook, current_sheet: str
) -> list[str]:
    """
    Find references to non-existent sheets in AST.

    Args:
        ast: Formula AST
        workbook: Workbook object (to check sheetnames)
        current_sheet: Name of current sheet

    Returns:
        List of reference strings pointing to missing sheets
    """
    if not ast:
        return []

    refs = extract_references_from_ast(ast)
    broken = []

    for ref in refs:
        # Check for sheet qualification
        if "!" in ref:
            sheet_part = ref.split("!")[0]
            # Strip single quotes if present
            if sheet_part.startswith("'") and sheet_part.endswith("'"):
                sheet_part = sheet_part[1:-1]

            # Check existence (case-sensitive usually, but Excel is case-insensitive)
            # We'll check case-insensitive match against sheetnames
            sheet_exists = False
            for existing_sheet in workbook.sheetnames:
                if existing_sheet.lower() == sheet_part.lower():
                    sheet_exists = True
                    break

            if not sheet_exists:
                broken.append(ref)

    return sorted(broken)


def build_extras(
    cell_address: str,
    defined_names: list[str],
    table_ref: str | None,
    special_data_sources: dict[str, Any],
) -> dict[str, Any]:
    """
    Build extras dictionary per IR Spec §3.2.

    Args:
        cell_address: Cell address (for lookups)
        defined_names: List of defined names containing this cell
        table_ref: Table reference (if cell is in table)
        special_data_sources: Special data source metadata

    Returns:
        Extras dictionary with all fields (nulls for unused)
    """
    extras = {
        "defined_name": None,
        "table_ref": None,
        "block_id": None,
        "pivot_table": None,
        "pivot_fields": None,
        "query_connection": None,
        "cube_connection": None,
        "cube_measure": None,
        "whatif_inputs": None,
    }

    # Defined name: use first if multiple
    if defined_names:
        extras["defined_name"] = defined_names[0]

    # Table reference
    if table_ref:
        extras["table_ref"] = table_ref

    # Special data sources
    for key in (
        "pivot_table",
        "pivot_fields",
        "query_connection",
        "cube_connection",
        "cube_measure",
        "whatif_inputs",
    ):
        if key in special_data_sources:
            extras[key] = special_data_sources[key]

    return extras


def serialize_cell(
    worksheet: Worksheet,
    cell_address_a1: str,
    resolution_engine: ResolutionEngine | None = None,
    defined_names: list[str] | None = None,
    table_ref: str | None = None,
    is_spilled: bool = False,
    spilled_from: str | None = None,
    udf_map: dict[str, UDFMetadata] | None = None,
    evaluated_value: Any | None = None,
    formula_a1: str | None = None,
) -> CellRecord:
    """
    Serialize a cell to CellRecord per IR Spec §3.2.

    Args:
        worksheet: Worksheet containing cell
        cell_address_a1: Full A1 address (e.g., "Sheet1!B5")
        resolution_engine: Resolution engine for volatile/lookup resolution (optional)
        defined_names: List of defined names containing this cell
        table_ref: Table reference string (if in table)
        is_spilled: True if this is a spill cell (not spiller)
        spilled_from: A1 address of spiller cell (if is_spilled)
        udf_map: Mapping from uppercase UDF name to UDFMetadata (optional)
        evaluated_value: Evaluated value from data_only=True workbook (optional)
        formula_a1: Pre-extracted formula string (optional, for memory-efficient serialization)

    Returns:
        CellRecord with all fields populated
    """
    # Parse address to get sheet-local cell reference
    parsed = parse_cell_address(cell_address_a1)
    if not parsed:
        raise ValueError(f"Invalid cell address: {cell_address_a1}")

    row = parsed.get("row", 0)
    col = parsed.get("col", 0)

    if row == 0 or col == 0:
        raise ValueError(f"Invalid cell coordinates in address: {cell_address_a1}")

    # Get cell from worksheet
    cell = worksheet.cell(row=row, column=col)

    # Extract basic info
    dtype = classify_dtype(cell)
    value_snapshot = extract_value_snapshot(cell)
    format_tokens = extract_format_tokens(cell)
    protection = extract_protection(cell)

    # Formula and AST handling
    formula_r1c1 = ""
    ast_structural = None

    # Handle spill cells vs formula cells
    if is_spilled:
        # Spill cell: no formula, just value
        formula_a1 = ""
        formula_r1c1 = ""
        ast_structural = None
    else:
        # Use pre-extracted formula if provided (memory-efficient streaming mode)
        # Otherwise extract from cell (legacy/compatibility mode)
        if formula_a1 is None:
            formula_a1 = extract_formula_string(cell)
        # else: use provided formula_a1

        if formula_a1:
            # Formula cell (regular or array formula)
            # Parse formula to AST
            try:
                ast_structural = parse_formula(formula_a1)
            except Exception:
                # If parsing fails, set AST to None
                ast_structural = None

            # Convert to R1C1
            try:
                formula_r1c1 = convert_formula_to_r1c1(formula_a1, row, col)
            except Exception:
                # If conversion fails, use empty string
                formula_r1c1 = ""
        else:
            # No formula - constant cell
            formula_r1c1 = ""
            ast_structural = None

    # Resolution metadata (empty if no formula)
    resolution_meta = {
        "is_volatile": False,
        "volatile_kinds": [],
        "status": "resolved",
        "resolved_volatile_ref": None,
        "volatile_drivers": [],
        "resolved_lookup_ref": None,
        "lookup_drivers": [],
        "notes": "",
    }

    special_data_sources: dict[str, Any] = {}

    if ast_structural and resolution_engine:
        # Check for volatile functions
        is_volatile, volatile_funcs = resolution_engine.detect_volatile(ast_structural)
        resolution_meta["is_volatile"] = is_volatile
        resolution_meta["volatile_kinds"] = volatile_funcs

        # Get current sheet name from cell_address_a1
        sheet_name = cell_address_a1.split("!")[0] if "!" in cell_address_a1 else ""

        # Actually resolve volatile functions
        if "OFFSET" in volatile_funcs:
            base_ref, row_offset, col_offset, height, width = _extract_offset_args(
                ast_structural, worksheet.parent, sheet_name
            )
            if base_ref:  # Valid extraction
                result = resolution_engine.resolve_offset(
                    base_ref, row_offset, col_offset, height, width, sheet_name
                )
                resolution_meta["status"] = result.status
                resolution_meta["resolved_volatile_ref"] = result.resolved_volatile_ref
                resolution_meta["volatile_drivers"] = result.volatile_drivers
                resolution_meta["notes"] = result.notes

        elif "INDIRECT" in volatile_funcs:
            address = _extract_indirect_args(ast_structural, worksheet.parent, sheet_name)
            result = resolution_engine.resolve_indirect(address, sheet_name)
            resolution_meta["status"] = result.status
            resolution_meta["resolved_volatile_ref"] = result.resolved_volatile_ref
            resolution_meta["volatile_drivers"] = result.volatile_drivers
            resolution_meta["notes"] = result.notes

        # Check for special data sources
        pivot_info = resolution_engine.detect_pivot_table(ast_structural)
        if pivot_info.ref_kinds:
            special_data_sources["ref_kinds"] = pivot_info.ref_kinds
            special_data_sources.update(pivot_info.extras)

        cube_info = resolution_engine.detect_cube_query(ast_structural)
        if cube_info.ref_kinds:
            if "ref_kinds" not in special_data_sources:
                special_data_sources["ref_kinds"] = []
            special_data_sources["ref_kinds"].extend(cube_info.ref_kinds)
            special_data_sources.update(cube_info.extras)

    # Classify ref_kinds
    ref_kinds = classify_ref_kinds(cell, ast_structural, resolution_meta, special_data_sources)

    # Detect broken references (non-existent sheets)
    broken_refs = _find_broken_refs(ast_structural, worksheet.parent, worksheet.title)

    # Build extras
    extras = build_extras(cell_address_a1, defined_names or [], table_ref, special_data_sources)

    # UDF calls - detect from formula if udf_map provided
    udf_calls: list[dict[str, str]] = []
    if formula_a1 and udf_map:
        udf_calls = detect_udf_calls_in_formula(formula_a1, udf_map)

        # Update volatile flag if formula calls volatile UDFs
        if (
            udf_calls
            and is_formula_volatile_due_to_udfs(udf_calls, udf_map)
            and "UDF_VOLATILE" not in resolution_meta.get("volatile_kinds", [])
        ):
            resolution_meta.setdefault("volatile_kinds", []).append("UDF_VOLATILE")

    return CellRecord(
        cell_address_a1=cell_address_a1,
        formula_a1=formula_a1,
        formula_r1c1=formula_r1c1,
        ast_structural=ast_structural,
        dtype=dtype,
        value_snapshot=value_snapshot,
        evaluated_value=evaluated_value,
        format_tokens=format_tokens,
        udf_calls=udf_calls,
        protection=protection,
        spilled_from=spilled_from,
        resolution=resolution_meta,
        ref_kinds=ref_kinds,
        broken_refs=broken_refs,
        extras=extras,
    )


def is_spiller_cell(cell: Cell) -> bool:
    """
    Detect if cell is a dynamic array spiller.

    Checks for Excel dynamic array functions (SEQUENCE, FILTER, etc.)
    in formula text. openpyxl may not expose full dynamic array metadata,
    so detection uses formula pattern matching.

    Args:
        cell: openpyxl Cell object

    Returns:
        True if cell is a dynamic array spiller, False otherwise

    Determinism: Always returns same result for same cell state.
    """
    # Check if cell has formula
    if not (hasattr(cell, "data_type") and cell.data_type == "f"):
        return False

    formula = str(cell.value) if cell.value else ""
    if not formula.startswith("="):
        return False

    # Check for known dynamic array functions
    # Make uppercase for case-insensitive matching
    formula_upper = formula.upper()
    dynamic_functions = [
        "SEQUENCE(",
        "FILTER(",
        "SORT(",
        "UNIQUE(",
        "SORTBY(",
        "RANDARRAY(",
        "XLOOKUP(",  # XLOOKUP can spill
    ]

    return any(func in formula_upper for func in dynamic_functions)


def _detect_spill_range_for_cell(worksheet: Worksheet, spiller_cell: Cell) -> list[str]:
    """
    Detect spill range for a single spiller cell.

    Detects 2D spill extent: scans both rows (down) and columns (right)
    from spiller cell. Stops when encountering formula, blank, or edge.

    Algorithm (deterministic):
    1. Find row extent: scan down until formula/blank/edge
    2. Find column extent: scan right until formula/blank/edge
    3. Return all cells in the resulting rectangle (excluding spiller)

    Returns list of cell coordinates (e.g., ["A2", "A3", "B2", "B3"])
    Coordinates are in canonical order (row ASC, col ASC within row).

    Determinism: Always scans in same order for same spiller position.
    """
    spill_coords = []
    spiller_row, spiller_col = spiller_cell.row, spiller_cell.column

    # Step 1: Determine row extent (how many rows does spill cover?)
    row_extent = 1  # Spiller itself counts as 1 row
    row = spiller_row + 1
    while row <= worksheet.max_row:
        cell = worksheet.cell(row, spiller_col)
        # Stop if: has formula, is None/empty
        if hasattr(cell, "data_type") and cell.data_type == "f":
            break
        if cell.value is None:
            break
        row_extent += 1
        row += 1

    # Step 2: Determine column extent (how many columns does spill cover?)
    col_extent = 1  # Spiller itself counts as 1 column
    col = spiller_col + 1
    while col <= worksheet.max_column:
        cell = worksheet.cell(spiller_row, col)
        # Stop if: has formula, is None/empty
        if hasattr(cell, "data_type") and cell.data_type == "f":
            break
        if cell.value is None:
            break
        col_extent += 1
        col += 1

    # Step 3: Generate all spill cell coordinates (excluding spiller)
    # Iterate in canonical order: row by row, col by col within row
    for r in range(spiller_row, spiller_row + row_extent):
        for c in range(spiller_col, spiller_col + col_extent):
            # Skip the spiller cell itself
            if r == spiller_row and c == spiller_col:
                continue
            cell = worksheet.cell(r, c)
            spill_coords.append(cell.coordinate)

    return spill_coords


def detect_spill_ranges(worksheet: Worksheet) -> dict[str, str]:
    """
    Detect all spill cells and map them to their spiller cells.

    Scans worksheet for dynamic array spillers, then identifies cells
    in their spill ranges. Since openpyxl may not expose spill metadata,
    uses heuristic: cells below/right of spiller with values but no formulas
    are likely spill cells.

    Args:
        worksheet: openpyxl Worksheet object

    Returns:
        Dictionary mapping spill cell addresses to spiller cell addresses.
        Example: {"Sheet1!A2": "Sheet1!A1", "Sheet1!A3": "Sheet1!A1"}

    Determinism: Iterates cells in canonical order (row ASC, col ASC).
    """
    spill_map = {}

    # Find all spiller cells first
    spillers = []
    for row in worksheet.iter_rows():
        for cell in row:
            if is_spiller_cell(cell):
                spillers.append(cell)

    # For each spiller, detect its spill range
    for spiller in spillers:
        spiller_addr = f"{worksheet.title}!{spiller.coordinate}"
        spill_cells = _detect_spill_range_for_cell(worksheet, spiller)

        for spill_cell_coord in spill_cells:
            spill_addr = f"{worksheet.title}!{spill_cell_coord}"
            spill_map[spill_addr] = spiller_addr

    return spill_map


def _serialize_value(value):
    """
    Recursively serialize a value for JSON, converting datetime objects to ISO strings.

    Args:
        value: Value to serialize (can be datetime, dict, list, etc.)

    Returns:
        JSON-serializable value with datetime objects converted to ISO strings
    """
    from datetime import date, datetime, time

    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    elif isinstance(value, dict):
        return {k: _serialize_value(v) for k, v in value.items()}
    elif isinstance(value, (list, tuple)):
        return [_serialize_value(item) for item in value]
    else:
        return value


def cell_record_to_dict(record: CellRecord) -> dict[str, Any]:
    """
    Convert CellRecord to dictionary for database insertion.

    All JSON fields are serialized with sorted keys per ADR-011.

    Args:
        record: CellRecord to convert

    Returns:
        Dictionary with all fields ready for database insertion
    """
    return {
        "cell_address_a1": record.cell_address_a1,
        "formula_a1": record.formula_a1,
        "formula_r1c1": record.formula_r1c1,
        "ast_structural": (
            json.dumps(record.ast_structural, sort_keys=True) if record.ast_structural else None
        ),
        "dtype": record.dtype,
        "value_snapshot": json.dumps(_serialize_value(record.value_snapshot), sort_keys=True),
        "evaluated_value": (
            json.dumps(_serialize_value(record.evaluated_value), sort_keys=True)
            if record.evaluated_value is not None
            else None
        ),
        "format_tokens_json": json.dumps(record.format_tokens, sort_keys=True),
        "udf_calls_json": json.dumps(record.udf_calls, sort_keys=True),
        "protection_locked": record.protection["locked"],
        "protection_hidden_formula": record.protection["hidden_formula"],
        "spilled_from": record.spilled_from,
        "resolution_json": json.dumps(record.resolution, sort_keys=True),
        "ref_kinds_json": json.dumps(record.ref_kinds, sort_keys=True),
        "broken_refs_json": json.dumps(record.broken_refs, sort_keys=True),
        "extras_json": json.dumps(record.extras, sort_keys=True),
    }


def write_cells_to_db(
    conn: sqlite3.Connection, cell_records: list[CellRecord], cell_to_binding: dict[str, list[str]]
) -> None:
    """
    Write cell records to database.

    Args:
        conn: SQLite connection
        cell_records: List of CellRecord objects to write
        cell_to_binding: Mapping from cell address to list of binding_ids

    Determinism: Writes cells in order provided (caller must sort).
    """
    cursor = conn.cursor()

    for record in cell_records:
        # Get binding_id for this cell
        bindings = cell_to_binding.get(record.cell_address_a1, [])
        # Use primary binding (first in sorted list) or empty string
        binding_id = bindings[0] if bindings else ""

        # Convert record to dict
        cell_dict = cell_record_to_dict(record)

        # Insert into cells table
        cursor.execute(
            """
            INSERT INTO cells (
                binding_id,
                cell_address_a1,
                formula_a1,
                formula_r1c1,
                ast_structural,
                dtype,
                value_snapshot,
                evaluated_value,
                format_tokens_json,
                udf_calls_json,
                protection_locked,
                protection_hidden_formula,
                spilled_from,
                resolution_json,
                ref_kinds_json,
                broken_refs_json,
                extras_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                binding_id,
                cell_dict["cell_address_a1"],
                cell_dict["formula_a1"],
                cell_dict["formula_r1c1"],
                cell_dict["ast_structural"],
                cell_dict["dtype"],
                cell_dict["value_snapshot"],
                cell_dict["evaluated_value"],
                cell_dict["format_tokens_json"],
                cell_dict["udf_calls_json"],
                cell_dict["protection_locked"],
                cell_dict["protection_hidden_formula"],
                cell_dict["spilled_from"],
                cell_dict["resolution_json"],
                cell_dict["ref_kinds_json"],
                cell_dict["broken_refs_json"],
                cell_dict["extras_json"],
            ),
        )


def write_cells_to_db_batch(
    conn: sqlite3.Connection,
    cell_records: list[CellRecord],
    cell_to_binding: dict[str, list[str]],
    batch_size: int = 1000,
) -> int:
    """Write cell records to database in batches with transactions.

    Args:
        conn: SQLite connection
        cell_records: List of CellRecord objects to write
        cell_to_binding: Mapping from cell address to list of binding_ids
        batch_size: Number of cells per transaction batch

    Returns:
        Total number of cells written

    Note: Uses explicit transactions for performance. Each batch
    is committed separately to limit memory usage.
    """
    total_written = 0

    for i in range(0, len(cell_records), batch_size):
        batch = cell_records[i : i + batch_size]

        conn.execute("BEGIN TRANSACTION")
        try:
            # Prepare all cell data for batch insert (much faster than individual inserts)
            cell_tuples = []
            for record in batch:
                bindings = cell_to_binding.get(record.cell_address_a1, [])
                binding_id = bindings[0] if bindings else ""
                cell_dict = cell_record_to_dict(record)
                cell_tuples.append(
                    (
                        binding_id,
                        cell_dict["cell_address_a1"],
                        cell_dict["formula_a1"],
                        cell_dict["formula_r1c1"],
                        cell_dict["ast_structural"],
                        cell_dict["dtype"],
                        cell_dict["value_snapshot"],
                        cell_dict["evaluated_value"],
                        cell_dict["format_tokens_json"],
                        cell_dict["udf_calls_json"],
                        cell_dict["protection_locked"],
                        cell_dict["protection_hidden_formula"],
                        cell_dict["spilled_from"],
                        cell_dict["resolution_json"],
                        cell_dict["ref_kinds_json"],
                        cell_dict["broken_refs_json"],
                        cell_dict["extras_json"],
                    )
                )

            # Batch insert all cells in this batch
            conn.executemany(
                """
                INSERT INTO cells (
                    binding_id,
                    cell_address_a1,
                    formula_a1,
                    formula_r1c1,
                    ast_structural,
                    dtype,
                    value_snapshot,
                    evaluated_value,
                    format_tokens_json,
                    udf_calls_json,
                    protection_locked,
                    protection_hidden_formula,
                    spilled_from,
                    resolution_json,
                    ref_kinds_json,
                    broken_refs_json,
                    extras_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                cell_tuples,
            )

            conn.execute("COMMIT")
            total_written += len(batch)

        except Exception as e:
            conn.execute("ROLLBACK")
            raise ValueError(f"Batch write failed at offset {i}: {e}")

    return total_written


def write_edges_to_db(conn: sqlite3.Connection, edges: list[tuple[str, str]]) -> None:
    """
    Write cell-level edges to database.

    Args:
        conn: SQLite connection
        edges: List of (from_cell, to_cell) tuples

    Determinism: Writes edges in order provided (caller must sort).
    """
    cursor = conn.cursor()

    for from_cell, to_cell in edges:
        # Check if edge already exists
        cursor.execute(
            """
            SELECT 1 FROM cell_level_edges
            WHERE from_cell = ? AND to_cell = ?
        """,
            (from_cell, to_cell),
        )

        if cursor.fetchone() is None:
            # Insert edge
            cursor.execute(
                """
                INSERT INTO cell_level_edges (from_cell, to_cell, to_external)
                VALUES (?, ?, NULL)
            """,
                (from_cell, to_cell),
            )


def write_bindings_to_db(
    conn: sqlite3.Connection,
    bindings: list["Binding"],
    workbook: Workbook | None = None,
    name_table_map: Any | None = None,
    cell_value_cache: dict[str, Any] | None = None,
) -> None:
    """
    Write bindings to database with evidence extraction.

    Args:
        conn: SQLite connection
        bindings: List of Binding objects
        workbook: Optional workbook for evidence extraction
        name_table_map: Optional name/table map
        cell_value_cache: Optional pre-built cache of cell values to avoid slow worksheet access
        bindings: List of Binding objects
        workbook: Optional Workbook object for evidence extraction
        name_table_map: Optional NameTableMap for evidence extraction

    Determinism: Writes bindings in order provided (caller must sort).

    Note: If workbook/name_table_map provided, extracts evidence
    (label candidates + axis invariants) per IR Spec §7. Otherwise writes
    empty evidence structures.
    """
    from xl_marinade.core.evidence import extract_evidence_for_binding

    cursor = conn.cursor()

    # Prepare all binding data first, then batch insert (much faster)
    binding_tuples = []
    for binding in bindings:
        # Extract evidence if context provided
        if workbook:
            # Get worksheet for this binding's sheet
            worksheet = workbook[binding.sheet]
            evidence = extract_evidence_for_binding(
                worksheet=worksheet,
                workbook=workbook,
                binding_address=binding.address_a1,
                binding_shape=(binding.shape_rows, binding.shape_cols),
                name_table_map=name_table_map,
                cell_value_cache=cell_value_cache,
            )
            # Combine label_candidates and axis_labels into single structure
            combined_evidence = {
                "label_candidates": evidence.get("label_candidates", []),
                "axis_labels": evidence.get("axis_labels", []),
            }
            label_candidates_json = json.dumps(combined_evidence, sort_keys=True)
        else:
            label_candidates_json = json.dumps(
                {"label_candidates": [], "axis_labels": []}, sort_keys=True
            )

        binding_tuples.append(
            (
                binding.binding_id,
                binding.debug_label,
                binding.sheet,
                binding.address_a1,
                binding.top_left_a1,
                binding.shape_rows,
                binding.shape_cols,
                binding.binding_type,
                binding.cells_structure_hash,
                label_candidates_json,
                json.dumps({}, sort_keys=True),  # Empty relationships for now
                binding.extraction_source,
                binding.spatial_candidates_json,
            )
        )

    # Batch insert all bindings
    cursor.executemany(
        """
        INSERT INTO bindings (
            binding_id,
            debug_label,
            sheet,
            address_a1,
            top_left_a1,
            shape_rows,
            shape_cols,
            binding_type,
            cells_structure_hash,
            label_candidates_json,
            relationships_json,
            extraction_source,
            spatial_candidates_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """,
        binding_tuples,
    )


def write_binding_edges_to_db(conn: sqlite3.Connection, binding_edges: list["BindingEdge"]) -> None:
    """
    Write binding-level edges to database.

    Args:
        conn: SQLite connection
        binding_edges: List of BindingEdge objects

    Determinism: Writes edges in order provided (caller must sort).
    """
    cursor = conn.cursor()

    # Use executemany for batch insert (much faster than individual inserts)
    edge_tuples = [(edge.from_binding_id, edge.to_binding_id) for edge in binding_edges]
    cursor.executemany(
        """
        INSERT INTO binding_level_edges (from_binding_id, to_binding_id)
        VALUES (?, ?)
    """,
        edge_tuples,
    )


def write_structure_hashes_to_db(
    conn: sqlite3.Connection, structure_hash_entries: list[dict[str, str]]
) -> None:
    """
    Write structure hashes to database.

    Args:
        conn: SQLite connection
        structure_hash_entries: List of hash entries with hash_type, hash_key, hash_value

    Determinism: Writes entries in order provided (caller must sort).
    """
    cursor = conn.cursor()

    for entry in structure_hash_entries:
        cursor.execute(
            """
            INSERT OR REPLACE INTO structure_hashes (hash_type, hash_key, hash_value)
            VALUES (?, ?, ?)
        """,
            (entry["hash_type"], entry["hash_key"], entry["hash_value"]),
        )


def write_levels_to_db(conn: sqlite3.Connection, levels: list["LevelAssignment"]) -> None:
    """
    Write topological level assignments to database.

    Args:
        conn: SQLite connection
        levels: List of LevelAssignment objects from compute_levels()

    Writes to: levels table (level, binding_id)

    Determinism: Writes levels in order provided (caller must sort).
    """
    cursor = conn.cursor()

    for level_assignment in levels:
        cursor.execute(
            """
            INSERT INTO levels (level, binding_id)
            VALUES (?, ?)
        """,
            (level_assignment.level, level_assignment.binding_id),
        )


def write_cycles_to_db(conn: sqlite3.Connection, cycles: list["Cycle"]) -> None:
    """
    Write cycle (SCC) information to database.

    Args:
        conn: SQLite connection
        cycles: List of Cycle objects from detect_cycles()

    Writes to: cycles table (cycle_id, ord, binding_id)

    ord is the position within the cycle (0-indexed),
    providing stable ordering within each SCC.

    Determinism: Writes cycles in order provided (caller must sort).
    """
    cursor = conn.cursor()

    for cycle in cycles:
        for cycle_order, binding_id in enumerate(cycle.bindings):
            cursor.execute(
                """
                INSERT INTO cycles (cycle_id, ord, binding_id)
                VALUES (?, ?, ?)
            """,
                (cycle.cycle_id, cycle_order, binding_id),
            )


def write_sheet_topologies_to_db(
    conn: sqlite3.Connection, topologies: dict[str, dict[str, Any]]
) -> None:
    """
    Write sheet topology data to database.

    Args:
        conn: SQLite connection
        topologies: Dictionary mapping sheet name to topology data
    """
    import json

    cursor = conn.cursor()
    for sheet_name in sorted(topologies.keys()):
        topology = topologies[sheet_name]
        bbox = topology.get("bbox", {})

        # Serialize topology to JSON
        topology_json = json.dumps(topology, sort_keys=True, ensure_ascii=True)

        cursor.execute(
            """
            INSERT INTO sheet_topology (
                sheet_name,
                topology_json,
                bbox_min_row,
                bbox_max_row,
                bbox_min_col,
                bbox_max_col
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """,
            (
                sheet_name,
                topology_json,
                bbox.get("min_row", 0),
                bbox.get("max_row", 0),
                bbox.get("min_col", 0),
                bbox.get("max_col", 0),
            ),
        )
