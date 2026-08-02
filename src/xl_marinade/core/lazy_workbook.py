# ABOUTME: Lazy workbook wrapper for memory-efficient Excel file access.
# ABOUTME: Uses openpyxl read-only mode with cached random cell access.

from __future__ import annotations

import posixpath
import xml.etree.ElementTree as ET
from collections import OrderedDict
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import cell as cell_utils
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedNameDict
from openpyxl.worksheet.table import Table, TableColumn
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class PopulatedCells:
    """Precomputed set of populated cells in a sheet for sparse range expansion.

    Stores cells that have formulas or non-empty values, enabling O(1) lookup
    and deterministic iteration without expanding entire range bounding boxes.

    PERFORMANCE: Includes spatial_index for O(log N + cells_in_range) range queries
    instead of O(all_cells) linear scan.

    Attributes:
        formula_cells: Set of cell coordinates with formulas (e.g., {'A1', 'B2'})
        value_cells: Set of cell coordinates with non-empty values (e.g., {'C3', 'D4'})
        all_cells_sorted: Sorted list of all populated cells in row-major order (deterministic)
        spatial_index: List of (coord, (row, col)) tuples for binary search (precomputed)
    """

    formula_cells: set[str]
    value_cells: set[str]
    all_cells_sorted: list[str]
    spatial_index: list[tuple[str, tuple[int, int]]]  # Precomputed for binary search


class LazyWorksheet:
    """Worksheet wrapper providing random cell access over streaming iterator.

    Caches cells as they're accessed to enable coordinate-based lookups
    while maintaining memory efficiency of read-only mode.

    Attributes:
        title: Sheet name
        max_row: Maximum row with data
        max_column: Maximum column with data
        populated_cells: Precomputed set of populated cells (lazy-loaded on first access)
    """

    MAX_CACHE_SIZE: int = 1_000_000  # Effectively unbounded - eviction causes O(n²) perf regression

    def __init__(
        self, ws: Worksheet, parent: Any, title: str, tables: dict[str, Table] | None = None
    ):
        """Initialize lazy worksheet from openpyxl worksheet.

        Args:
            ws: openpyxl read-only worksheet
            parent: Parent workbook (LazyWorkbook)
            title: Sheet name
            tables: Dictionary of tables for this sheet (name -> Table)
        """
        self._ws = ws
        self.parent = parent
        self.title = title
        self.max_row = ws.max_row
        self.max_column = ws.max_column
        self._cache: OrderedDict[str, Cell] = OrderedDict()

        # Initialize iterator
        self._row_iterator = ws.iter_rows()

        # Initialize tables (passed from parent if available, else try fallback)
        self._tables: dict[str, Table] = tables if tables is not None else getattr(ws, "tables", {})
        self.defined_names: DefinedNameDict | None = getattr(ws, "defined_names", None)

        # Mock merged_cells for read-only mode compatibility
        # Read-only mode does not support merged cells, so we provide an empty list
        class MockMergedCells:
            ranges = []

        self.merged_cells = MockMergedCells()

        # Populated cells cache (lazy-loaded on first access)
        self._populated_cells: PopulatedCells | None = None

    @property
    def cache_size(self) -> int:
        """Current number of cells in cache."""
        return len(self._cache)

    def get_cell(self, coord: str) -> Cell:
        """Get cell by A1 coordinate, using cache or iterating if needed.

        Args:
            coord: Cell coordinate (e.g., "A1", "B5")

        Returns:
            Cell object (returns empty Cell if not found)
        """
        # Check cache
        if coord in self._cache:
            # Move to end (most recently used)
            self._cache.move_to_end(coord)
            return self._cache[coord]

        try:
            # Parse coordinate
            xy = cell_utils.coordinate_from_string(coord)
            col_idx = cell_utils.column_index_from_string(xy[0])
            target_row = xy[1]
        except ValueError:
            # If invalid coordinate, strictly raise or return None?
            # Cell() requires valid row/col.
            # Let's assume valid A1.
            return self._create_empty_cell(coord)

        # If we need to iterate
        try:
            for row in self._row_iterator:
                # Attempt to determine row index from cells
                row_idx = -1
                for c in row:
                    if hasattr(c, "row"):
                        row_idx = c.row
                        break

                # If we can't find row index from cells, we skip logic relying on it
                if row_idx == -1:
                    continue

                # Iterate cells in row
                # In read-only mode, rows start from min_column
                min_col = getattr(self._ws, "min_column", 1)

                found_in_this_row = None

                for i, cell in enumerate(row):
                    current_col_idx = min_col + i

                    # Determine coordinate
                    if hasattr(cell, "coordinate"):
                        self._cache[cell.coordinate] = cell
                        if len(self._cache) > self.MAX_CACHE_SIZE:
                            self._cache.popitem(last=False)  # Remove oldest (LRU)

                        if cell.coordinate == coord:
                            found_in_this_row = cell
                    else:
                        # Calculate coordinate for EmptyCell
                        col_letter = get_column_letter(current_col_idx)
                        current_coord = f"{col_letter}{row_idx}"

                        # Create dummy cell to replace EmptyCell so it has coordinate/parent
                        dummy_cell = Cell(self, row=row_idx, column=current_col_idx, value=None)

                        self._cache[current_coord] = dummy_cell
                        if len(self._cache) > self.MAX_CACHE_SIZE:
                            self._cache.popitem(last=False)  # Remove oldest (LRU)

                        if current_coord == coord:
                            found_in_this_row = dummy_cell

                if found_in_this_row:
                    self._cache.move_to_end(coord)  # Mark as recently used
                    return found_in_this_row

                # Stop if we've gone past the target row
                if row_idx > target_row:
                    break

        except StopIteration:
            pass

        # Fallback: If main iterator has passed target row, use iter_rows() to fetch it
        # iter_rows() creates a NEW iterator from the underlying worksheet
        try:
            for row_tuple in self._ws.iter_rows(
                min_row=target_row, max_row=target_row, min_col=col_idx, max_col=col_idx
            ):
                for cell in row_tuple:
                    if hasattr(cell, "coordinate") and cell.coordinate == coord:
                        self._cache[coord] = cell
                        if len(self._cache) > self.MAX_CACHE_SIZE:
                            self._cache.popitem(last=False)
                        return cell
                    elif hasattr(cell, "value") and cell.value is not None:
                        # Cell found but coordinate doesn't match (shouldn't happen for single cell)
                        self._cache[coord] = cell
                        if len(self._cache) > self.MAX_CACHE_SIZE:
                            self._cache.popitem(last=False)
                        return cell
        except Exception:
            pass  # Fall through to empty cell creation

        # Not found - create empty cell
        empty_cell = self._create_empty_cell(coord)
        self._cache[coord] = empty_cell
        if len(self._cache) > self.MAX_CACHE_SIZE:
            self._cache.popitem(last=False)  # Remove oldest (LRU)
        return empty_cell

    def _create_empty_cell(self, coord: str) -> Cell:
        """Create an empty cell for missing data."""
        try:
            xy = cell_utils.coordinate_from_string(coord)
            col_idx = cell_utils.column_index_from_string(xy[0])
            row_idx = xy[1]
            return Cell(self, row=row_idx, column=col_idx, value=None)
        except Exception:
            # Fallback for invalid coords
            return Cell(self, row=1, column=1, value=None)

    def __getitem__(self, key: str) -> Any:
        """Get cell or range by coordinate using dictionary syntax (ws['A1'] or ws['A1:B2'])."""
        if isinstance(key, str) and ":" in key:
            # Range access
            try:
                min_col, min_row, max_col, max_row = cell_utils.range_boundaries(key)
                # Use iter_rows to get cells.
                # Note: iter_rows yields tuple of cells. We collect them into tuple of tuples.
                rows = []
                for row in self.iter_rows(
                    min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
                ):
                    rows.append(row)
                return tuple(rows)
            except ValueError:
                # Fallback if invalid range
                pass

        return self.get_cell(key)

    def iter_rows(
        self,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
    ) -> Iterator[tuple[Cell, ...]]:
        """Iterate rows, populating cache as we go.

        Yields:
            Tuples of cells for each row
        """
        # We delegate to the underlying worksheet's iter_rows if specific bounds are requested
        # AND we haven't consumed the main iterator past that point.
        # However, read-only iter_rows is a generator.

        # Simpler approach: Create a NEW iterator from the underlying ws for this specific request
        # caching results as we see them.

        # Note: calling iter_rows on read-only worksheet multiple times might re-read file?
        # openpyxl read-only mode supports multiple iterators but it re-reads the XML.

        for row in self._ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            # Cache cells we see
            for cell in row:
                if hasattr(cell, "coordinate"):
                    self._cache[cell.coordinate] = cell
                    if len(self._cache) > self.MAX_CACHE_SIZE:
                        self._cache.popitem(last=False)
            yield row

    def cell(self, row: int, column: int) -> Cell:
        """Get cell by row and column index (1-based)."""
        col_letter = cell_utils.get_column_letter(column)
        coord = f"{col_letter}{row}"
        return self.get_cell(coord)

    @property
    def populated_cells(self) -> PopulatedCells:
        """Get precomputed populated cells for this sheet (lazy-loaded on first access).

        Returns:
            PopulatedCells object with formula cells, value cells, and sorted list
        """
        if self._populated_cells is None:
            self._populated_cells = self._precompute_populated_cells()
        return self._populated_cells

    def _precompute_populated_cells(self) -> PopulatedCells:
        """Precompute set of populated cells in this sheet.

        Scans sheet once to identify all cells with formulas.
        Returns cells in deterministic row-major order (A1, B1, C1, ..., A2, B2, ...).

        This enables sparse range expansion: instead of expanding A1:ZZ10000 to 260k cells,
        we intersect with populated cells and return only cells that actually exist.

        CRITICAL: We only include FORMULA cells, not all cells with values.
        Including all value cells causes traversal regression (69k cells instead of 27k)
        because iter_rows() iterates over the entire bounding box, not just populated cells.

        Performance: O(bounding_box) scan (unavoidable with openpyxl), O(1) lookup via set membership.

        Returns:
            PopulatedCells with formula_cells, value_cells (empty), and all_cells_sorted
        """
        import sys

        try:
            formula_cells = set()

            # Single pass through sheet using iter_rows
            # WARNING: iter_rows() iterates over ENTIRE BOUNDING BOX (max_row × max_column),
            # not just populated cells. This is an openpyxl limitation.
            # We filter to only formula cells to avoid traversal regression.
            for row in self._ws.iter_rows():
                for cell in row:
                    if not hasattr(cell, "coordinate"):
                        continue

                    coord = cell.coordinate

                    # Check for formula (data_type == 'f' or value starts with '=')
                    if (
                        hasattr(cell, "data_type")
                        and cell.data_type == "f"
                        or (
                            hasattr(cell, "value")
                            and isinstance(cell.value, str)
                            and cell.value.startswith("=")
                        )
                    ):
                        formula_cells.add(coord)

            # Sort formula cells deterministically (row-major order: A1, B1, C1, ..., A2, B2, ...)
            def sort_key(coord: str) -> tuple[int, int]:
                """Extract (row, col) for sorting."""
                try:
                    xy = cell_utils.coordinate_from_string(coord)
                    col_idx = cell_utils.column_index_from_string(xy[0])
                    row_idx = xy[1]
                    return (row_idx, col_idx)
                except ValueError:
                    return (999999, 999999)  # Invalid coords sort last

            all_cells_sorted = sorted(formula_cells, key=sort_key)

            # PERFORMANCE: Precompute spatial index for O(log N + cells_in_range) lookups
            # Build list of (coord, (row, col)) tuples for binary search in range queries
            spatial_index = [(coord, sort_key(coord)) for coord in all_cells_sorted]

            return PopulatedCells(
                formula_cells=formula_cells,
                value_cells=set(),  # Empty - we only track formula cells
                all_cells_sorted=all_cells_sorted,
                spatial_index=spatial_index,
            )

        except Exception as e:
            print(
                f"WARNING: Failed to precompute populated cells for sheet '{self.title}': {e}",
                file=sys.stderr,
            )
            # Return empty PopulatedCells to allow graceful fallback to legacy path
            return PopulatedCells(
                formula_cells=set(), value_cells=set(), all_cells_sorted=[], spatial_index=[]
            )


class LazyWorkbook:
    """Memory-efficient workbook wrapper using read-only mode.

    Provides same interface as openpyxl Workbook but with lazy loading.
    Must be used as context manager or explicitly closed.
    """

    def __init__(self, path: Path, data_only: bool = False, keep_vba: bool = False):
        """Load workbook in read-only mode.

        Args:
            path: Path to Excel file
            data_only: If True, load evaluated values instead of formulas
            keep_vba: Ignored in read-only mode (VBA not accessible)
        """
        self.path = path
        self.data_only = data_only
        self._wb = load_workbook(filename=path, read_only=True, data_only=data_only)
        self._sheets: dict[str, LazyWorksheet] = {}

        # Load tables separately because read-only mode skips them
        self._tables_cache = self._load_tables_fast()

    def _load_tables_fast(self) -> dict[str, dict[str, Table]]:
        """Fast table extraction using direct zip access."""
        tables_by_sheet: dict[str, dict[str, Table]] = {}
        try:
            # We can access the zip file from the read-only workbook
            # internal attribute is _archive (ZipFile)
            if hasattr(self._wb, "_archive"):
                archive = self._wb._archive
            else:
                print("DEBUG: No _archive attribute")
                return {}

            # XML Namespaces
            ns = {
                "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
            }

            # 1. Parse workbook.xml to map rId -> sheet name
            try:
                wb_xml = archive.read("xl/workbook.xml")
                root = ET.fromstring(wb_xml)

                sheet_map = {}  # rId -> name
                for sheet in root.findall(".//main:sheet", ns):
                    name = sheet.get("name")
                    rid = sheet.get(f"{{{ns['r']}}}id")
                    sheet_map[rid] = name
            except Exception:
                return {}

            # 2. Parse workbook.xml.rels to map rId -> target filename
            try:
                rels_xml = archive.read("xl/_rels/workbook.xml.rels")
                rels_root = ET.fromstring(rels_xml)

                rid_to_target = {}
                for rel in rels_root.findall(".//rel:Relationship", ns):
                    rid = rel.get("Id")
                    target = rel.get("Target")
                    rid_to_target[rid] = target
            except Exception:
                return {}

            # 3. Process each sheet
            for rid, name in sheet_map.items():
                if rid not in rid_to_target:
                    continue

                target = rid_to_target[rid]
                # Target is usually "worksheets/sheet1.xml", relative to xl/

                if target.startswith("/"):
                    path = target[1:]
                else:
                    path = f"xl/{target}"

                # print(f"DEBUG: Processing sheet {name}, path {path}")

                # Split dir and filename
                if "/" in path:
                    folder, filename = path.rsplit("/", 1)
                    rels_path = f"{folder}/_rels/{filename}.rels"
                else:
                    continue

                # Check if rels exists
                if rels_path not in archive.namelist():
                    continue

                # Parse sheet rels
                try:
                    sheet_rels = archive.read(rels_path)
                    s_rels_root = ET.fromstring(sheet_rels)

                    sheet_tables = {}

                    for rel in s_rels_root.findall(".//rel:Relationship", ns):
                        if "relationships/table" in rel.get("Type"):
                            table_target = rel.get("Target")
                            # Resolve table path
                            table_path = posixpath.normpath(posixpath.join(folder, table_target))
                            # Zip paths shouldn't start with /
                            if table_path.startswith("/"):
                                table_path = table_path[1:]

                            # print(f"DEBUG: Found table ref {table_target} -> {table_path}")

                            # Parse table
                            if table_path in archive.namelist():
                                table_xml = archive.read(table_path)
                                t_root = ET.fromstring(table_xml)

                                t_name = t_root.get("displayName") or t_root.get("name")
                                t_ref = t_root.get("ref")

                                # Parse columns
                                cols = []
                                col_node = t_root.find("main:tableColumns", ns)
                                if col_node is not None:
                                    for col in col_node.findall("main:tableColumn", ns):
                                        cols.append(
                                            TableColumn(id=int(col.get("id")), name=col.get("name"))
                                        )

                                # Create Table object
                                tab = Table(displayName=t_name, ref=t_ref)
                                tab.tableColumns = cols
                                sheet_tables[t_name] = tab
                                # print(f"DEBUG: Loaded table {t_name}")

                    if sheet_tables:
                        tables_by_sheet[name] = sheet_tables

                except Exception:
                    continue

            return tables_by_sheet

        except Exception:
            return {}

    @property
    def sheetnames(self) -> list[str]:
        """Return list of sheet names."""
        return self._wb.sheetnames

    @property
    def worksheets(self) -> Iterator[LazyWorksheet]:
        """Return iterator of worksheets (for NameTableMap iteration).

        Mirrors openpyxl's ``Workbook.worksheets``, which excludes chartsheets —
        they have no cell grid, and ``LazyWorksheet`` requires one (``max_row``).
        """
        for ws in self._wb.worksheets:
            yield self[ws.title]

    @property
    def active(self) -> LazyWorksheet:
        """Return active worksheet (for ResolutionEngine fallback)."""
        if self._wb.active:
            return self[self._wb.active.title]
        return self[self.sheetnames[0]]

    @property
    def defined_names(self) -> DefinedNameDict:
        """Return workbook-scoped defined names (for NameTableMap)."""
        return self._wb.defined_names

    def __getitem__(self, name: str) -> LazyWorksheet:
        """Get worksheet by name."""
        if name not in self._sheets:
            if name not in self.sheetnames:
                raise KeyError(f"Worksheet {name} does not exist.")

            # Pass extracted tables to worksheet
            tables = self._tables_cache.get(name)
            self._sheets[name] = LazyWorksheet(self._wb[name], self, name, tables=tables)

        return self._sheets[name]

    def close(self) -> None:
        """Close workbook and release file handle."""
        self._wb.close()

    def __enter__(self) -> LazyWorkbook:
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit - ensures close is called."""
        self.close()

    def extract_vba_if_present(self) -> list[Any]:
        """Extract VBA UDFs, loading separately if needed for read-only mode.

        Returns:
            List of UDF metadata, or empty list if no VBA
        """
        # Read-only mode generally doesn't support VBA access.
        # We check if we need to do a separate load.
        # Check for xlsm extension or just try loading.

        path_str = str(self.path)
        if not path_str.lower().endswith((".xlsm", ".xlam", ".xlsb")):
            return []

        # We need to perform a separate load to get VBA
        # We use keep_vba=True, data_only=False, read_only=False (standard load)
        # But we only need it briefly.

        from xl_marinade.core.vba_parser import extract_udfs_from_workbook

        try:
            # We assume extract_udfs_from_workbook works on a standard Workbook object
            # We load it minimally if possible, but standard load is full.
            # To avoid memory spike, we open, extract, and close immediately.
            # NOTE: This defeats the purpose of memory optimization IF we do it fully.
            # However, for VBA we might not have a choice if we want UDFs.
            # But wait, does openpyxl allow loading JUST vbaProject.bin? No.

            # Optimization: Use zipfile to check for vbaProject.bin first?
            # extract_udfs_from_workbook takes a Workbook object.

            # Let's try to load just the VBA part if possible, or accept the hit.
            # Given the constraints, we will do a separate standard load just for VBA.
            # This is "extract_vba_if_present".

            temp_wb = load_workbook(self.path, read_only=False, keep_vba=True, data_only=False)
            udfs = extract_udfs_from_workbook(temp_wb)
            temp_wb.close()
            return udfs

        except Exception:
            return []

    def __getattr__(self, name: str) -> Any:
        """Delegate missing attributes to underlying workbook."""
        return getattr(self._wb, name)
