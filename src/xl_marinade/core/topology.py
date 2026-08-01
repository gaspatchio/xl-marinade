# ABOUTME: Sheet topology computation with bounding box expansion and margin.
# ABOUTME: Computes per-sheet topology for visited regions per IR Spec §8 and ADR-012.

import datetime
import hashlib
import json
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from xl_marinade.core.names_tables import NameTableMap
from xl_marinade.core.ref_converter import parse_cell_address


@dataclass
class BoundingBox:
    """Bounding box for visited cells."""

    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def expand(self, margin: int = 5) -> "BoundingBox":
        """
        Expand bounding box by margin in all directions.

        Args:
            margin: Number of rows/cols to expand (default: 5 per ADR-012).

        Returns:
            New BoundingBox with expanded bounds (clamped to min 1).
        """
        return BoundingBox(
            min_row=max(1, self.min_row - margin),
            max_row=self.max_row + margin,
            min_col=max(1, self.min_col - margin),
            max_col=self.max_col + margin,
        )

    def to_a1_range(self, sheet: str) -> str:
        """
        Convert bounding box to A1 range notation.

        Args:
            sheet: Sheet name.

        Returns:
            A1 range string (e.g., "Sheet1!A1:F10").
        """
        top_left = f"{get_column_letter(self.min_col)}{self.min_row}"
        bottom_right = f"{get_column_letter(self.max_col)}{self.max_row}"
        return f"{sheet}!{top_left}:{bottom_right}"


def compute_bounding_box(visited_cells: list[str]) -> BoundingBox | None:
    """
    Compute minimal bounding box around visited cells.

    Args:
        visited_cells: List of cell addresses (e.g., ["Sheet1!B3", "Sheet1!D5"]).

    Returns:
        BoundingBox or None if no cells.
    """
    if not visited_cells:
        return None

    min_row = float("inf")
    max_row = float("-inf")
    min_col = float("inf")
    max_col = float("-inf")

    for cell_addr in visited_cells:
        try:
            # Parse row and column
            parsed = parse_cell_address(cell_addr)
            row = parsed["row"]
            col = parsed["col"]

            # Skip cells with row=0, col=0 (invalid/external refs)
            if row == 0 or col == 0:
                continue

            min_row = min(min_row, row)
            max_row = max(max_row, row)
            min_col = min(min_col, col)
            max_col = max(max_col, col)
        except Exception:
            # Skip invalid addresses
            continue

    if min_row == float("inf"):
        return None

    return BoundingBox(
        min_row=int(min_row), max_row=int(max_row), min_col=int(min_col), max_col=int(max_col)
    )


def serialize_cell_topology(
    worksheet: Worksheet, row: int, col: int, name_table_map: NameTableMap | None = None
) -> dict[str, Any]:
    """
    Serialize topology for a single cell per IR Spec §8.

    Args:
        worksheet: Worksheet containing cell.
        row: 1-indexed row.
        col: 1-indexed column.
        name_table_map: Optional name/table map for memberships.

    Returns:
        Dictionary with cell topology data.
    """
    cell = worksheet.cell(row, col)
    coord = cell.coordinate
    full_addr = f"{worksheet.title}!{coord}"

    # Basic properties
    has_formula = cell.data_type == "f"

    # Data type classification
    dtype = "blank"
    if cell.value is not None:
        if cell.data_type == "n":
            dtype = "number"
        elif cell.data_type == "b":
            dtype = "boolean"
        elif cell.data_type == "s":
            dtype = "text"
        elif cell.data_type == "f":
            # Formula cell - check if error
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                dtype = "error"
            else:
                dtype = "number"  # Default for formulas
        elif cell.data_type == "d":
            dtype = "date"
        elif cell.data_type == "e":
            dtype = "error"
        else:
            dtype = "unknown"

    # Value snapshot
    # Note: cell.value can be ArrayFormula object for array formulas
    # Convert to string for JSON serialization
    value_snapshot = cell.value
    if hasattr(value_snapshot, "__class__") and value_snapshot.__class__.__name__ == "ArrayFormula":
        # ArrayFormula object - convert to string representation
        value_snapshot = str(value_snapshot)
    elif isinstance(value_snapshot, (datetime.datetime, datetime.date)):
        value_snapshot = value_snapshot.isoformat()

    # Display text (formatted value)
    display_text = str(cell.value) if cell.value is not None else ""

    # Format tokens (simplified)
    format_tokens = {
        "kind": "general",
        "currency": None,
        "percent": False,
        "thousands_sep": False,
        "decimals": None,
        "date_code": None,
    }

    # Detect currency and percent from number format
    if cell.number_format:
        nf = cell.number_format.upper()
        if "$" in nf or "¤" in nf:
            format_tokens["kind"] = "currency"
            format_tokens["currency"] = "USD"  # Simplified
        elif "%" in nf:
            format_tokens["kind"] = "percent"
            format_tokens["percent"] = True
        elif "," in nf:
            format_tokens["thousands_sep"] = True

        # Date detection
        if any(code in nf for code in ["D", "M", "Y", "H", "S"]):
            format_tokens["kind"] = "date"
            format_tokens["date_code"] = cell.number_format

    # Merged span
    merged_span = None
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            merged_span = {
                "rows": merged_range.max_row - merged_range.min_row + 1,
                "cols": merged_range.max_col - merged_range.min_col + 1,
            }
            break

    # Table membership
    in_table = None
    if name_table_map:
        table_ref = name_table_map.get_cell_table_ref(full_addr)
        if table_ref:
            parts = table_ref.split(".")
            if len(parts) == 2:
                in_table = {
                    "name": parts[0],
                    "header": False,  # Simplified - data cells only
                    "column_name": parts[1],
                }

    # Defined names membership
    in_defined_names = []
    if name_table_map:
        in_defined_names = name_table_map.get_cell_defined_names(full_addr)

    # Style (simplified)
    style = {
        "bold": cell.font.bold if cell.font else False,
        "italic": cell.font.italic if cell.font else False,
        "underline": bool(cell.font.underline) if cell.font else False,
    }

    # Protection
    protection = {
        "locked": cell.protection.locked if cell.protection else True,
        "hidden_formula": cell.protection.hidden if cell.protection else False,
    }

    return {
        "address": coord,
        "has_formula": has_formula,
        "dtype": dtype,
        "value_snapshot": value_snapshot,
        "display_text": display_text,
        "format_tokens": format_tokens,
        "merged_span": merged_span,
        "in_table": in_table,
        "in_defined_names": in_defined_names,
        "style": style,
        "protection": protection,
    }


def extract_regions(
    worksheet: Worksheet, bbox: BoundingBox, name_table_map: NameTableMap | None = None
) -> dict[str, list[dict[str, Any]]]:
    """
    Extract topology regions per IR Spec §8.

    Args:
        worksheet: Worksheet to extract from.
        bbox: Bounding box defining the extraction region.
        name_table_map: Optional name/table map.

    Returns:
        Dictionary with region lists.
    """
    regions: dict[str, list[dict[str, Any]]] = {
        "table_regions": [],
        "defined_name_regions": [],
        "merged_label_regions": [],
        "row_nonblank_runs": [],
        "col_nonblank_runs": [],
    }

    # Table regions (if name_table_map available)
    if name_table_map:
        for table_info in name_table_map.get_all_tables():
            if table_info.sheet == worksheet.title:
                # Check if table intersects bounding box
                # Simplified: add all tables on the sheet
                regions["table_regions"].append(
                    {
                        "name": table_info.name,
                        "address": table_info.range,
                        "header_row_address": table_info.header_row,
                    }
                )

    # Defined name regions
    if name_table_map:
        for name_info in name_table_map.get_all_names():
            if name_info.is_external:
                continue

            # Check if any range overlaps with worksheet
            for range_str in name_info.ranges:
                if range_str.startswith(f"{worksheet.title}!"):
                    regions["defined_name_regions"].append(
                        {"name": name_info.name, "address": range_str}
                    )
                    break

    # Merged label regions
    for merged_range in worksheet.merged_cells.ranges:
        # Check if merged range overlaps with bbox
        if (
            merged_range.min_row <= bbox.max_row
            and merged_range.max_row >= bbox.min_row
            and merged_range.min_col <= bbox.max_col
            and merged_range.max_col >= bbox.min_col
        ):
            # Get text from top-left cell
            top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
            text = str(top_left_cell.value) if top_left_cell.value else ""

            regions["merged_label_regions"].append(
                {
                    "address": str(merged_range),
                    "text": text,
                    "merged_span": {
                        "rows": merged_range.max_row - merged_range.min_row + 1,
                        "cols": merged_range.max_col - merged_range.min_col + 1,
                    },
                }
            )

    # Row and column nonblank runs - DISABLED for performance
    #
    # Computing row/col nonblank runs requires iterating over the entire bounding box
    # (e.g., 1000 rows × 40 cols = 40K cells) and accessing each cell via worksheet.cell().
    # With LazyWorksheet, this causes O(n²) performance regression.
    #
    # These runs provide context for label detection but are not critical for IR extraction.
    # The cell data is already in the `cells` table and can be analyzed separately if needed.
    #
    # TODO: If runs are needed, compute from `cells` table data instead of worksheet access.

    return regions


def compute_sheet_topology(
    worksheet: Worksheet,
    visited_cells: list[str],
    margin: int = 5,
    name_table_map: NameTableMap | None = None,
) -> dict[str, Any] | None:
    """
    Compute sheet topology for visited cells per IR Spec §8 and ADR-012.

    Args:
        worksheet: Worksheet to compute topology for.
        visited_cells: List of visited cell addresses on this sheet.
        margin: Bounding box expansion margin (default: 5 per ADR-012).
        name_table_map: Optional name/table map for region extraction.

    Returns:
        Topology dictionary or None if no visited cells.
    """
    if not visited_cells:
        return None

    # Compute bounding box
    bbox = compute_bounding_box(visited_cells)
    if not bbox:
        return None

    # Expand by margin
    expanded_bbox = bbox.expand(margin)

    # Note: We do NOT clamp to worksheet.max_row/max_column because those return
    # the highest row/column WITH DATA, not sheet bounds. Clamping would defeat
    # the purpose of margin expansion. The span may include blanks beyond data range.

    # OPTIMIZATION: Skip cell serialization entirely. Cell data is already stored in
    # the `cells` table. Re-serializing 27K cells here causes O(n) × worksheet access
    # which is extremely slow with LazyWorksheet. The bounding box + regions is sufficient
    # for sheet topology; cell data can be joined from the cells table if needed.
    #
    # We store only the list of visited cell addresses (not their full serialization)
    # to maintain the topology structure without the performance cost.
    cells = [addr.split("!", 1)[1] if "!" in addr else addr for addr in visited_cells]

    # Extract regions (fast - just iterates over tables/names, not cells)
    regions = extract_regions(worksheet, expanded_bbox, name_table_map)

    return {
        "cells": cells,  # Just cell coordinates, not full serialization
        "regions": regions,
        "bbox": {
            "min_row": expanded_bbox.min_row,
            "max_row": expanded_bbox.max_row,
            "min_col": expanded_bbox.min_col,
            "max_col": expanded_bbox.max_col,
        },
        "cell_count": len(visited_cells),  # For structure hash stability
    }


def compute_sheet_structure_hash(topology: dict[str, Any]) -> str:
    """
    Compute structure hash for sheet topology.

    Per IR Spec §12: hash of formulas + names + tables for the sheet.
    We hash the topology JSON with sorted keys for determinism.

    Args:
        topology: Sheet topology dictionary.

    Returns:
        SHA-256 hash (lowercase hex, 64 chars).
    """
    # Serialize topology with sorted keys
    topology_json = json.dumps(topology, sort_keys=True, ensure_ascii=True)

    # Compute SHA-256 hash
    hash_obj = hashlib.sha256(topology_json.encode("utf-8"))
    return hash_obj.hexdigest()


def compute_sheet_topologies(
    workbook: Workbook,
    visited_cells_by_sheet: dict[str, list[str]],
    margin: int = 5,
    name_table_map: NameTableMap | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Compute topologies for all sheets with visited cells.

    Args:
        workbook: Workbook object.
        visited_cells_by_sheet: Dict mapping sheet name to list of visited cell addresses.
        margin: Bounding box expansion margin (default: 5 per ADR-012).
        name_table_map: Optional name/table map.

    Returns:
        Dictionary mapping sheet name to topology data.
    """
    topologies = {}

    for sheet_name, visited_cells in sorted(visited_cells_by_sheet.items()):
        if not visited_cells:
            continue

        worksheet = workbook[sheet_name]
        topology = compute_sheet_topology(
            worksheet, visited_cells, margin=margin, name_table_map=name_table_map
        )

        if topology:
            topologies[sheet_name] = topology

    return topologies


def create_sheet_structure_hash_entries(
    topologies: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    """
    Create structure_hashes table entries for sheet topologies.

    Args:
        topologies: Dict mapping sheet name to topology data.

    Returns:
        List of dicts with hash_type, hash_key, hash_value for structure_hashes table.
    """
    entries = []

    for sheet_name in sorted(topologies.keys()):
        topology = topologies[sheet_name]
        structure_hash = compute_sheet_structure_hash(topology)

        entries.append(
            {"hash_type": "sheets", "hash_key": sheet_name, "hash_value": structure_hash}
        )

    return entries
